"""Parse and import the MASTER EXPENSES Excel sheet into verified transactions."""

from __future__ import annotations

import calendar
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from bson import ObjectId

from tahmeed.db.connection import get_db
from tahmeed.models.transaction import Transaction, pack_money
from tahmeed.services.description_mapping_service import (
    get_mappings_for_descriptions,
    normalize_description,
    save_mapping,
)
from tahmeed.services.excel_dates import parse_excel_date
from tahmeed.services.daily_import_service import parse_amount


_MONTH_LABELS = [
    "", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


@dataclass
class MasterImportRow:
    """One parsed row from the master Excel sheet."""

    serial: Optional[int]
    date: datetime
    description: str
    truck_number: str
    lpo_do: str
    do_number: str
    memo: str
    notes: str
    amount: float
    currency: str
    receipt_raw: str
    ownership: str
    approver: str
    category_id: Optional[ObjectId] = None
    category_name: Optional[str] = None
    amount_usd: Optional[float] = None


@dataclass
class MasterImportPreview:
    source_filename: str
    rows: List[MasterImportRow] = field(default_factory=list)
    unmapped: Dict[str, int] = field(default_factory=dict)  # description_key -> row count
    skipped: int = 0


def _cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _parse_date(val) -> Optional[datetime]:
    parsed = parse_excel_date(val)
    if parsed is None:
        return None
    return parsed.replace(hour=0, minute=0, second=0, microsecond=0)


def _normalize_receipt(val: str) -> str:
    s = (val or "").strip().lower()
    if not s:
        return "pending"
    if "no receipt" in s or s in {"missing", "no", "n/a"}:
        return "missing"
    if "received" in s or s in {"yes", "receipt", "rcvd"}:
        return "received"
    return "pending"


def _month_label(dt: datetime) -> str:
    return f"{_MONTH_LABELS[dt.month]} {str(dt.year)[-2:]}"


def parse_master_expenses_excel(
    path: str | Path,
    *,
    sheet_name: Optional[str] = None,
) -> Tuple[List[MasterImportRow], int]:
    """Return parsed rows and a count of skipped blank rows."""
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    skipped = 0
    rows: List[MasterImportRow] = []
    try:
        if sheet_name is None:
            for candidate in wb.sheetnames:
                if "MASTER EXPENSE" in candidate.upper():
                    sheet_name = candidate
                    break
            sheet_name = sheet_name or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                skipped += 1
                continue
            desc = _cell_str(row[3]) if len(row) > 3 else ""
            if not desc:
                skipped += 1
                continue
            dt = _parse_date(row[1] if len(row) > 1 else None)
            if dt is None:
                skipped += 1
                continue

            # Positional TZS (11) / USD (12). Short rows without col 12 are
            # treated as TZS-only — same as sheets that omit the USD column.
            tzs = parse_amount(row[11] if len(row) > 11 else None)
            usd = parse_amount(row[12]) if len(row) > 12 else None
            amount, amount_usd, currency = pack_money(tzs, usd)

            serial = None
            try:
                if row[0] is not None:
                    serial = int(float(row[0]))
            except (TypeError, ValueError):
                pass

            memo = _cell_str(row[9]) if len(row) > 9 else ""
            notes = _cell_str(row[10]) if len(row) > 10 else ""
            rows.append(MasterImportRow(
                serial=serial,
                date=dt,
                description=desc,
                truck_number=_cell_str(row[4]) if len(row) > 4 else "",
                lpo_do=_cell_str(row[5]) if len(row) > 5 else "",
                do_number=_cell_str(row[6]) if len(row) > 6 else "",
                memo=memo,
                notes=notes,
                amount=amount,
                currency=currency,
                amount_usd=amount_usd,
                receipt_raw=_cell_str(row[13]) if len(row) > 13 else "",
                ownership=_cell_str(row[14]) if len(row) > 14 else "",
                approver=_cell_str(row[15]) if len(row) > 15 else "",
            ))
    finally:
        wb.close()
    return rows, skipped


async def preview_master_import(path: str | Path) -> MasterImportPreview:
    """Parse the workbook and resolve description mappings."""
    path = Path(path)
    parsed, skipped = parse_master_expenses_excel(path)
    unique_descs = list({r.description for r in parsed})
    mappings = await get_mappings_for_descriptions(unique_descs)

    unmapped: Dict[str, int] = {}
    for row in parsed:
        key = normalize_description(row.description)
        mapping = mappings.get(key)
        if mapping:
            row.category_id = mapping.category_id
            row.category_name = mapping.category_name
        else:
            unmapped[key] = unmapped.get(key, 0) + 1

    return MasterImportPreview(
        source_filename=path.name,
        rows=parsed,
        unmapped=unmapped,
        skipped=skipped,
    )


async def apply_mapping_to_preview(
    preview: MasterImportPreview,
    description_key: str,
    category_id: ObjectId,
    category_name: str,
) -> None:
    """Assign an item to all preview rows sharing this description key."""
    display = description_key
    for row in preview.rows:
        if normalize_description(row.description) == description_key:
            display = row.description
            break
    await save_mapping(display, category_id, category_name)
    for row in preview.rows:
        if normalize_description(row.description) == description_key:
            row.category_id = category_id
            row.category_name = category_name
    preview.unmapped.pop(description_key, None)


def _build_memo(row: MasterImportRow) -> str:
    parts = []
    if row.memo:
        parts.append(row.memo)
    if row.notes:
        parts.append(row.notes)
    return " | ".join(parts)


def _row_to_transaction(
    row: MasterImportRow,
    *,
    upload_id: str,
    source_filename: str,
    verified_by: Optional[ObjectId] = None,
) -> Transaction:
    return Transaction(
        date=row.date,
        description=row.description,
        truck_number=row.truck_number,
        amount=row.amount,
        currency=row.currency,
        amount_usd=row.amount_usd,
        category_id=row.category_id,
        category_name=row.category_name,
        category_confidence=1.0 if row.category_name else 0.0,
        item=row.category_name or "",
        lpo_do=row.lpo_do,
        do_number=row.do_number,
        memo=_build_memo(row),
        receipt_status=_normalize_receipt(row.receipt_raw),
        notes_flag=bool(row.notes),
        ownership=row.ownership,
        approver=row.approver,
        verified=True,
        verified_by=verified_by,
        verified_at=datetime.utcnow(),
        month=_month_label(row.date),
        year=row.date.year,
        created_at=datetime.utcnow(),
    )


async def commit_master_import(
    preview: MasterImportPreview,
    *,
    verified_by: Optional[ObjectId] = None,
    skip_duplicates: bool = True,
) -> dict:
    """Insert all preview rows as verified master transactions.

    Soft pre-check skips known serials; a unique index on
    ``(master_import_source, master_serial)`` makes retries after a partial
    insert safe (duplicate keys are counted, not raised).
    """
    from tahmeed.db.import_idempotency import ensure_import_indexes, insert_many_idempotent

    if preview.unmapped:
        raise ValueError(
            f"{len(preview.unmapped)} description(s) still need an item mapping."
        )

    upload_id = str(uuid.uuid4())
    db = get_db()
    await ensure_import_indexes()
    inserted = 0
    duplicates = 0
    batch: List[dict] = []

    existing_serials: set[int] = set()
    if skip_duplicates:
        serials = [r.serial for r in preview.rows if r.serial is not None]
        if serials:
            cursor = db.transactions.find(
                {
                    "master_import_source": preview.source_filename,
                    "master_serial": {"$in": serials},
                },
                {"master_serial": 1},
            )
            docs = await cursor.to_list(length=None)
            existing_serials = {d["master_serial"] for d in docs if "master_serial" in d}

    for row in preview.rows:
        if skip_duplicates and row.serial is not None and row.serial in existing_serials:
            duplicates += 1
            continue
        tx = _row_to_transaction(
            row,
            upload_id=upload_id,
            source_filename=preview.source_filename,
            verified_by=verified_by,
        )
        doc = tx.to_doc()
        doc["master_import_id"] = upload_id
        doc["master_import_source"] = preview.source_filename
        if row.serial is not None:
            doc["master_serial"] = row.serial
        batch.append(doc)
        if len(batch) >= 500:
            n, d = await insert_many_idempotent(db.transactions, batch)
            inserted += n
            duplicates += d
            batch.clear()

    if batch:
        n, d = await insert_many_idempotent(db.transactions, batch)
        inserted += n
        duplicates += d

    result = {
        "inserted": inserted,
        "duplicates_skipped": duplicates,
        "upload_id": upload_id,
        "source": preview.source_filename,
    }
    try:
        from tahmeed.services.audit_service import record_event

        await record_event(
            "import.master",
            actor_id=verified_by,
            entity_type="import_batch",
            upload_id=upload_id,
            details=result,
        )
    except Exception:
        pass
    return result
