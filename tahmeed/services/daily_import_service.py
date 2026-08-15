"""Parse daily MATUMIZI-style Excel sheets into staged cashier transactions."""

from __future__ import annotations

import asyncio
import re
import uuid
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import openpyxl
from bson import ObjectId

from tahmeed.db.connection import get_db
from tahmeed.models.transaction import Transaction
from tahmeed.services.description_mapping_service import (
    get_mappings_for_descriptions,
    normalize_description,
    save_mapping,
)
from tahmeed.services.excel_dates import parse_excel_date

# Header aliases → logical field (first match wins per column)
_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "serial": ("s/no", "sno", "s no", "serial", "#"),
    "date": ("date",),
    "description": ("description", "desc"),
    "truck": ("truck no", "truck no.", "truck", "truck number"),
    "lpo": ("lpo nos", "lpo nos.", "lpo", "lpo/do"),
    "do": ("do no", "do no.", "do number"),
    "memo": ("memo",),
    "notes": ("notes", "ref_float", "ref float"),
    "tzs": ("tzs",),
    "usd": ("usd", "us$", "dollar"),
    "receipt": ("receipt status", "receipt", "rcpt"),
    "ownership": ("ownership", "own"),
    "approver": ("apr by", "approver", "apr"),
}

_RCPT_NORM = {
    "received": "received",
    "receipt": "received",
    "1": "received",
    "yes": "received",
    "rcvd": "received",
    "missing": "missing",
    "pending": "pending",
    "0": "pending",
    "no receipt": "no_receipt",
    "no_receipt": "no_receipt",
    "none": "no_receipt",
    "n/a": "no_receipt",
    "na": "no_receipt",
}


@dataclass
class DailyImportRow:
    serial: Optional[int]
    date: datetime
    description: str
    truck_number: str
    lpo_do: str
    do_number: str
    memo: str
    notes: str
    amount: float
    currency: str
    receipt_status: str
    ownership: str
    approver: str
    category_id: Optional[ObjectId] = None
    category_name: Optional[str] = None
    skipped_item: bool = False  # user chose Skip — leave without item


class DailyImportCancelled(Exception):
    """User aborted reading the daily Excel file."""


@dataclass
class DailyImportPreview:
    source_filename: str
    source_path: str
    rows: List[DailyImportRow] = field(default_factory=list)
    unmapped: Dict[str, int] = field(default_factory=dict)  # key -> count
    skipped_blank: int = 0
    primary_date: Optional[date] = None
    detected_dates: List[date] = field(default_factory=list)
    date_counts: Dict[date, int] = field(default_factory=dict)
    date_majority_clear: bool = True
    outlier_count: int = 0
    sheet_name: str = ""
    upload_id: str = field(default_factory=lambda: str(uuid.uuid4()))
    force_primary_date: bool = False
    flag_date_discrepancy: bool = False


def cell_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


def normalize_receipt(val: str) -> str:
    """Map Excel receipt text to stored status.

    Daily files use only RECEIPT / NO RECEIPT (shown uppercase in the grid).
    Stored keys remain received / no_receipt for compatibility.
    """
    s = " ".join((val or "").strip().lower().split())
    if not s:
        return "pending"
    if s in _RCPT_NORM:
        return _RCPT_NORM[s]
    if "no receipt" in s:
        return "no_receipt"
    if s == "receipt" or "received" in s:
        return "received"
    return "pending"


def parse_amount(val) -> Optional[float]:
    """Parse Excel/number/paste amounts; '(1,000)' → -1000."""
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.upper() in {"NONE", "N/A", "-", "—"}:
        return None
    # Excel formula leftovers
    if s.startswith("="):
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()
    s = s.replace(",", "").replace(" ", "").replace("TZS", "").replace("USD", "")
    s = s.replace("Tsh", "").replace("tsh", "")
    if s.startswith("-"):
        negative = True
        s = s[1:]
    elif s.startswith("+"):
        s = s[1:]
    s = re.sub(r"[^\d.]", "", s)
    if not s or s.count(".") > 1:
        return None
    try:
        amount = float(s)
    except ValueError:
        return None
    return -amount if negative else amount


def parse_date_value(val) -> Optional[datetime]:
    """Parse Excel datetime, serial number, or common date string → midnight datetime."""
    parsed = parse_excel_date(val)
    if parsed is None:
        return None
    return parsed.replace(hour=0, minute=0, second=0, microsecond=0)


def detect_date_from_name(text: str) -> Optional[date]:
    """Pull DD-MM-YYYY / DD/MM/YYYY from filename or sheet title."""
    if not text:
        return None
    m = re.search(r"(\d{1,2})[-./](\d{1,2})[-./](\d{4})", text)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def _norm_header(val) -> str:
    return " ".join(cell_str(val).lower().replace(".", " ").split())


def _map_headers(header_row: List) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(header_row or []):
        h = _norm_header(cell)
        if not h:
            continue
        for field, aliases in _HEADER_ALIASES.items():
            if field in mapping:
                continue
            if h in aliases or any(h == a or h.startswith(a) for a in aliases):
                mapping[field] = idx
                break
    return mapping


def _col(row: List, cols: Dict[str, int], key: str, default=None):
    idx = cols.get(key)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


@dataclass(frozen=True)
class DateAllocation:
    """How to assign one register date to an entire daily upload."""

    primary: Optional[date]
    clear_majority: bool
    counts: Dict[date, int] = field(default_factory=dict)
    candidates: tuple[date, ...] = ()


def analyze_date_allocation(
    row_dates: List[date],
    *,
    filename: str = "",
    sheet_name: str = "",
) -> DateAllocation:
    """Majority row date wins; ties / empty rows fall back or need a prompt.

    - Clear majority: one date has strictly more rows than any other.
    - Unclear: two or more dates share the top count → caller should ask.
    - No row dates: filename/sheet date is treated as a clear choice when present.
    """
    if row_dates:
        counts = Counter(row_dates)
        max_n = max(counts.values())
        tied = tuple(sorted(d for d, n in counts.items() if n == max_n))
        clear = len(tied) == 1
        return DateAllocation(
            primary=tied[0] if clear else None,
            clear_majority=clear,
            counts=dict(counts),
            candidates=tied,
        )
    fallback = detect_date_from_name(filename) or detect_date_from_name(sheet_name)
    if fallback is None:
        return DateAllocation(primary=None, clear_majority=False, counts={}, candidates=())
    return DateAllocation(
        primary=fallback,
        clear_majority=True,
        counts={},
        candidates=(fallback,),
    )


def pick_primary_date(
    row_dates: List[date],
    *,
    filename: str = "",
    sheet_name: str = "",
) -> Optional[date]:
    """Prefer majority row date; fall back to filename/sheet detection.

    On a tie, returns the earliest tied date (stable) — prefer
    ``analyze_date_allocation`` when you need to know if the majority is clear.
    """
    alloc = analyze_date_allocation(
        row_dates, filename=filename, sheet_name=sheet_name
    )
    if alloc.primary is not None:
        return alloc.primary
    return alloc.candidates[0] if alloc.candidates else None


_CLASSIC_MATUMIZI_COLS: Dict[str, int] = {
    "serial": 0,
    "date": 1,
    "description": 3,
    "truck": 4,
    "lpo": 5,
    "do": 6,
    "memo": 9,
    "notes": 10,
    "tzs": 11,
    "usd": 12,
    "receipt": 13,
    "ownership": 14,
    "approver": 15,
}


def _looks_like_classic_matumizi(header_row: List) -> bool:
    """True when row 0 matches classic MATUMIZI headers (Date + Description)."""
    cells = [_norm_header(c) for c in (header_row or [])]
    if len(cells) < 4:
        return False
    date_h = cells[1] if len(cells) > 1 else ""
    desc_h = cells[3] if len(cells) > 3 else ""
    date_ok = date_h == "date" or date_h.startswith("date")
    desc_ok = desc_h in ("description", "desc") or "description" in desc_h
    return date_ok and desc_ok


def parse_daily_expenses_excel(
    path: str | Path,
    *,
    sheet_name: Optional[str] = None,
    should_cancel: Optional[Callable[[], bool]] = None,
) -> Tuple[List[DailyImportRow], int, str]:
    """Return (rows, skipped_blank_count, sheet_name_used).

    Raises ValueError when the workbook does not match Daily Register format.
    Raises DailyImportCancelled when ``should_cancel`` returns True.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if should_cancel is not None and should_cancel():
            raise DailyImportCancelled()
        used_sheet = sheet_name or wb.sheetnames[0]
        if used_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{used_sheet}' not found in workbook.")
        ws = wb[used_sheet]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows_iter))
        except StopIteration:
            raise ValueError("The Excel file is empty.")
        cols = _map_headers(header)
        if "description" not in cols or "date" not in cols:
            if _looks_like_classic_matumizi(header):
                cols = dict(_CLASSIC_MATUMIZI_COLS)
            else:
                raise ValueError(
                    "This file does not match the Daily Register format.\n\n"
                    "Expected a MATUMIZI-style sheet with Date and Description "
                    "columns. Wrong or unrelated Excel files are rejected so they "
                    "do not create faulty uploads."
                )

        parsed: List[DailyImportRow] = []
        skipped = 0
        for i, row in enumerate(rows_iter):
            if should_cancel is not None and i % 16 == 0 and should_cancel():
                raise DailyImportCancelled()
            if not row:
                skipped += 1
                continue
            row = list(row)
            desc = cell_str(_col(row, cols, "description"))
            if not desc:
                skipped += 1
                continue
            # Skip total/summary rows
            if desc.upper().startswith("TOTAL") or str(_col(row, cols, "serial", "")).upper() == "TOTAL":
                skipped += 1
                continue
            dt = parse_date_value(_col(row, cols, "date"))
            if dt is None:
                skipped += 1
                continue

            tzs = parse_amount(_col(row, cols, "tzs"))
            usd = parse_amount(_col(row, cols, "usd"))
            amount = 0.0
            currency = "TZS"
            if usd is not None and usd != 0:
                amount = usd
                currency = "USD"
            elif tzs is not None:
                amount = tzs
                currency = "TZS"
            elif usd is not None:
                amount = usd
                currency = "USD"

            serial = None
            raw_serial = _col(row, cols, "serial")
            try:
                if raw_serial is not None and cell_str(raw_serial):
                    serial = int(float(raw_serial))
            except (TypeError, ValueError):
                pass

            notes = cell_str(_col(row, cols, "notes"))
            parsed.append(
                DailyImportRow(
                    serial=serial,
                    date=dt,
                    description=desc.upper(),
                    truck_number=cell_str(_col(row, cols, "truck")).upper(),
                    lpo_do=cell_str(_col(row, cols, "lpo")).upper(),
                    do_number=cell_str(_col(row, cols, "do")).upper(),
                    memo=cell_str(_col(row, cols, "memo")).upper(),
                    notes=notes,
                    amount=amount,
                    currency=currency,
                    receipt_status=normalize_receipt(cell_str(_col(row, cols, "receipt"))),
                    ownership=cell_str(_col(row, cols, "ownership")).upper(),
                    approver=cell_str(_col(row, cols, "approver")).upper(),
                )
            )
    finally:
        wb.close()
    return parsed, skipped, used_sheet


async def preview_daily_import(
    path: str | Path,
    *,
    should_cancel: Optional[Callable[[], bool]] = None,
) -> DailyImportPreview:
    path = Path(path)

    def _parse() -> Tuple[List[DailyImportRow], int, str]:
        return parse_daily_expenses_excel(path, should_cancel=should_cancel)

    parsed, skipped, sheet = await asyncio.to_thread(_parse)
    if should_cancel is not None and should_cancel():
        raise DailyImportCancelled()
    unique_descs = list({r.description for r in parsed})
    mappings = await get_mappings_for_descriptions(unique_descs)
    if should_cancel is not None and should_cancel():
        raise DailyImportCancelled()

    unmapped: Dict[str, int] = {}
    for row in parsed:
        key = normalize_description(row.description)
        mapping = mappings.get(key)
        if mapping:
            row.category_id = mapping.category_id
            row.category_name = mapping.category_name
        else:
            unmapped[key] = unmapped.get(key, 0) + 1

    row_dates = [r.date.date() for r in parsed]
    alloc = analyze_date_allocation(
        row_dates, filename=path.name, sheet_name=sheet
    )
    primary = alloc.primary
    if primary is None and alloc.candidates:
        # Unclear tie — keep a provisional primary for display; caller asks.
        primary = alloc.candidates[0]
    outliers = 0
    if primary is not None and row_dates:
        outliers = sum(1 for d in row_dates if d != primary)

    return DailyImportPreview(
        source_filename=path.name,
        source_path=str(path),
        rows=parsed,
        unmapped=unmapped,
        skipped_blank=skipped,
        primary_date=primary,
        detected_dates=sorted(set(row_dates)),
        date_counts=dict(alloc.counts),
        date_majority_clear=alloc.clear_majority,
        outlier_count=outliers,
        sheet_name=sheet,
    )


async def apply_mapping_to_preview(
    preview: DailyImportPreview,
    description_key: str,
    category_id: ObjectId,
    category_name: str,
) -> None:
    display = description_key
    for row in preview.rows:
        if normalize_description(row.description) == description_key:
            display = row.description
            break
    await save_mapping(display, category_id, category_name)
    for row in preview.rows:
        if normalize_description(row.description) == description_key:
            row.category_id = category_id
            row.category_name = category_name
            row.skipped_item = False
    preview.unmapped.pop(description_key, None)


def skip_description_in_preview(preview: DailyImportPreview, description_key: str) -> None:
    for row in preview.rows:
        if normalize_description(row.description) == description_key:
            row.skipped_item = True
            row.category_id = None
            row.category_name = None
    preview.unmapped.pop(description_key, None)


def skip_all_unmapped(preview: DailyImportPreview) -> None:
    for key in list(preview.unmapped.keys()):
        skip_description_in_preview(preview, key)


def apply_date_policy(
    preview: DailyImportPreview,
    *,
    force_primary: bool = False,
    flag_discrepancy: bool = False,
) -> None:
    """Attach a register day to the upload; never rewrite Excel row dates.

    Mixed dates in one file are normal (pending items cleared on the upload
    day). ``force_primary`` / ``flag_discrepancy`` are kept for call-site
    compatibility but do not change row dates or raise discrepancy flags.
    """
    del force_primary, flag_discrepancy  # no longer rewrite / flag
    preview.force_primary_date = False
    preview.flag_date_discrepancy = False


def _register_dt(primary: Optional[date]) -> Optional[datetime]:
    if primary is None:
        return None
    return datetime(primary.year, primary.month, primary.day)


def _ref_float_from_notes(notes: str) -> str:
    low = (notes or "").strip().lower()
    if "refund" in low and "float" in low:
        return "REFUND TO FLOAT"
    return (notes or "").strip().upper()


def staged_row_payload(row: DailyImportRow, preview: DailyImportPreview) -> dict:
    """Dict used to populate the Daily Register grid before Save."""
    primary = preview.primary_date
    return {
        "date": row.date,
        "item": row.category_name or "",
        "description": row.description,
        "truck_number": row.truck_number,
        "memo": row.memo,
        "ref_float": _ref_float_from_notes(row.notes),
        "amount": row.amount,
        "currency": row.currency,
        "receipt_status": row.receipt_status,
        "ownership": row.ownership,
        "approver": row.approver,
        "category_id": row.category_id,
        "category_name": row.category_name,
        "lpo_do": row.lpo_do,
        "do_number": row.do_number,
        "serial": row.serial,
        "daily_import_id": preview.upload_id,
        "daily_import_source": preview.source_filename,
        "date_discrepancy": False,
        "import_primary_date": _register_dt(primary),
    }


_MONTH_LABELS = (
    "", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
)


def _month_label(dt: datetime) -> str:
    return f"{_MONTH_LABELS[dt.month]} {str(dt.year)[-2:]}"


def preview_rows_as_truck_dicts(
    preview: DailyImportPreview,
) -> List[dict]:
    """Row dicts suitable for ``run_import_truck_gate`` (truck_number field)."""
    return [staged_row_payload(row, preview) for row in preview.rows]


def _payload_to_verified_transaction(
    payload: dict,
    *,
    verified_by: Optional[ObjectId] = None,
    day_order: Optional[int] = None,
) -> Transaction:
    """Build a verified master transaction from a staged daily-import payload."""
    dt = payload["date"]
    if not isinstance(dt, datetime):
        dt = datetime(dt.year, dt.month, dt.day)
    item_name = (payload.get("category_name") or payload.get("item") or "").strip()
    ref_float = (payload.get("ref_float") or "").strip()
    primary = payload.get("import_primary_date")
    return Transaction(
        date=dt,
        description=(payload.get("description") or "").strip(),
        truck_number=(payload.get("truck_number") or "").strip().upper(),
        amount=float(payload.get("amount") or 0),
        currency=(payload.get("currency") or "TZS").strip().upper() or "TZS",
        category_id=payload.get("category_id"),
        category_name=item_name or None,
        category_confidence=1.0 if item_name else 0.0,
        item=item_name,
        lpo_do=(payload.get("lpo_do") or "").upper(),
        do_number=(payload.get("do_number") or "").upper(),
        memo=(payload.get("memo") or "").strip(),
        receipt_status=payload.get("receipt_status") or "pending",
        notes_flag=ref_float == "REFUND TO FLOAT",
        ref_float=ref_float,
        ownership=(payload.get("ownership") or "").upper(),
        approver=(payload.get("approver") or "").upper(),
        # Attribute the batch to the importing accountant so rows show in
        # Table / Merged register like normal system-created entries.
        cashier_id=verified_by,
        verified=True,
        verified_by=verified_by,
        verified_at=datetime.utcnow(),
        register_status="submitted",
        day_order=day_order,
        month=_month_label(dt),
        year=dt.year,
        created_at=datetime.utcnow(),
        daily_import_id=payload.get("daily_import_id"),
        daily_import_source=payload.get("daily_import_source"),
        date_discrepancy=bool(payload.get("date_discrepancy")),
        import_primary_date=primary,
    )


async def commit_daily_to_master(
    payloads: List[dict],
    *,
    verified_by: Optional[ObjectId] = None,
) -> dict:
    """Insert daily-import rows as verified Master Expenses.

    Rows without an item are skipped. Each inserted doc carries
    ``import_row_key`` so a retry of the same upload cannot duplicate rows
    (unique partial index on ``daily_import_id`` + ``import_row_key``).
    """
    from tahmeed.db.import_idempotency import (
        daily_import_row_key,
        ensure_import_indexes,
        insert_many_idempotent,
    )

    db = get_db()
    await ensure_import_indexes()
    inserted = 0
    duplicates = 0
    skipped_no_item = 0
    batch: List[dict] = []
    upload_id = ""
    source = ""
    min_date: Optional[datetime] = None
    max_date: Optional[datetime] = None
    row_index = 0

    for payload in payloads:
        item_name = (payload.get("category_name") or payload.get("item") or "").strip()
        if not item_name:
            skipped_no_item += 1
            continue
        if not upload_id:
            upload_id = str(payload.get("daily_import_id") or uuid.uuid4())
        if not source:
            source = str(payload.get("daily_import_source") or "")
        # Keep batch tags consistent even if upstream mutated ids.
        payload = dict(payload)
        payload["daily_import_id"] = upload_id
        if source:
            payload["daily_import_source"] = source
        tx = _payload_to_verified_transaction(
            payload, verified_by=verified_by, day_order=row_index
        )
        row_index += 1
        if isinstance(tx.date, datetime):
            if min_date is None or tx.date < min_date:
                min_date = tx.date
            if max_date is None or tx.date > max_date:
                max_date = tx.date
        doc = tx.to_doc()
        doc["import_row_key"] = daily_import_row_key(payload)
        batch.append(doc)
        if len(batch) >= 500:
            n, d = await insert_many_idempotent(db.transactions, batch)
            inserted += n
            duplicates += d
            batch.clear()

    if batch:
        n, d = await insert_many_idempotent(db.transactions, batch)
        inserted += n
        duplicates += d

    result = {
        "inserted": inserted,
        "duplicates_skipped": duplicates,
        "skipped_no_item": skipped_no_item,
        "upload_id": upload_id,
        "source": source,
        "min_date": min_date,
        "max_date": max_date,
    }
    try:
        from tahmeed.services.audit_service import record_event

        await record_event(
            "import.daily_master",
            actor_id=verified_by,
            entity_type="import_batch",
            upload_id=upload_id or None,
            details=result,
        )
    except Exception:
        pass
    return result


async def cleanup_null_daily_import_ids() -> int:
    """Strip null/empty daily_import_id tags left by older saves / bad uploads.

    These used to appear in the Uploads browser as one undeleteable phantom group.
    """
    db = get_db()
    result = await db.transactions.update_many(
        {
            "$or": [
                {"daily_import_id": None},
                {"daily_import_id": ""},
                {"daily_import_id": {"$type": "null"}},
            ]
        },
        {
            "$unset": {
                "daily_import_id": "",
                "daily_import_source": "",
                "import_primary_date": "",
            }
        },
    )
    return int(result.modified_count or 0)


async def list_daily_uploads(limit: int = 100) -> List[dict]:
    db = get_db()
    # Heal legacy null-tagged rows so they never group as a phantom upload.
    await cleanup_null_daily_import_ids()
    pipeline = [
        # Only real string batch ids — never null / missing / empty.
        {"$match": {"daily_import_id": {"$type": "string", "$ne": ""}}},
        {
            "$group": {
                "_id": "$daily_import_id",
                "source_filename": {"$first": "$daily_import_source"},
                "count": {"$sum": 1},
                "primary_date": {"$first": "$import_primary_date"},
                "min_date": {"$min": "$date"},
                "max_date": {"$max": "$date"},
                "created_at": {"$min": "$created_at"},
                "cashier_id": {"$first": "$cashier_id"},
                "outlier_count": {
                    "$sum": {
                        "$cond": [
                            {
                                "$and": [
                                    {"$ne": ["$import_primary_date", None]},
                                    {"$ne": [
                                        {"$dateToString": {
                                            "format": "%Y-%m-%d",
                                            "date": "$date",
                                        }},
                                        {"$dateToString": {
                                            "format": "%Y-%m-%d",
                                            "date": "$import_primary_date",
                                        }},
                                    ]},
                                ]
                            },
                            1,
                            0,
                        ]
                    }
                },
                "duplicate_count": {
                    "$sum": {"$cond": ["$possible_duplicate", 1, 0]}
                },
            }
        },
        {"$sort": {"created_at": -1}},
        {"$limit": limit},
    ]
    return await db.transactions.aggregate(pipeline).to_list(length=limit)


async def get_daily_upload_records(
    upload_id: str,
    *,
    limit: int = 10000,
    skip: int = 0,
) -> List[Transaction]:
    """All transaction rows belonging to one daily Excel upload batch."""
    uid = str(upload_id or "").strip()
    if not uid:
        return []
    db = get_db()
    cursor = (
        db.transactions.find({"daily_import_id": uid})
        .sort([("date", 1), ("day_order", 1), ("created_at", 1)])
        .skip(max(0, skip))
        .limit(max(1, limit))
    )
    docs = await cursor.to_list(length=limit)
    return [Transaction.from_doc(d) for d in docs]


async def delete_daily_upload(upload_id: str) -> int:
    if not upload_id or not str(upload_id).strip():
        return 0
    db = get_db()
    result = await db.transactions.delete_many(
        {"daily_import_id": str(upload_id).strip()}
    )
    return result.deleted_count


def _build_issue_query(
    *,
    search: str = "",
    truck: str = "",
    cashier_id=None,
    date_from=None,
    date_to=None,
    item="",
    description="",
) -> dict:
    from tahmeed.services.accountant_service import _append_text_filters

    issue_or = {"$or": [
        {"possible_duplicate": True},
        {"date_discrepancy": True},
    ]}
    and_clauses: list = [issue_or]
    _append_text_filters(
        and_clauses, search=search, item=item, description=description,
    )

    query: dict = {
        "verified": {"$ne": True},
        "rejected": {"$ne": True},
        "$and": and_clauses,
    }
    if truck.strip():
        query["truck_number"] = {"$regex": re.escape(truck.strip()), "$options": "i"}
    if cashier_id:
        query["cashier_id"] = cashier_id
    if date_from or date_to:
        df: dict = {}
        if date_from:
            df["$gte"] = date_from
        if date_to:
            df["$lte"] = date_to
        query["date"] = df
    return query


async def get_issue_transactions(
    *,
    skip: int = 0,
    limit: int = 50,
    search: str = "",
    truck: str = "",
    cashier_id=None,
    date_from=None,
    date_to=None,
    item="",
    description="",
) -> List[Transaction]:
    db = get_db()
    query = _build_issue_query(
        search=search,
        truck=truck,
        cashier_id=cashier_id,
        date_from=date_from,
        date_to=date_to,
        item=item,
        description=description,
    )
    cursor = (
        db.transactions.find(query)
        .sort([("date", -1), ("created_at", -1)])
        .skip(skip)
        .limit(limit)
    )
    docs = await cursor.to_list(length=limit)
    return [Transaction.from_doc(d) for d in docs]


async def count_issue_transactions(
    *,
    search: str = "",
    truck: str = "",
    cashier_id=None,
    date_from=None,
    date_to=None,
    item="",
    description="",
) -> int:
    db = get_db()
    query = _build_issue_query(
        search=search,
        truck=truck,
        cashier_id=cashier_id,
        date_from=date_from,
        date_to=date_to,
        item=item,
        description=description,
    )
    return await db.transactions.count_documents(query)


async def clear_issue_flags(tx_id: ObjectId) -> bool:
    db = get_db()
    result = await db.transactions.update_one(
        {"_id": tx_id},
        {"$set": {"possible_duplicate": False, "date_discrepancy": False}},
    )
    return result.modified_count == 1
