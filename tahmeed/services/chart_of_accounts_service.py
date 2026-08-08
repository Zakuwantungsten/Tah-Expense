"""Import QuickBooks Chart of Accounts into the Items (categories) collection."""

from pathlib import Path
from typing import List

import openpyxl

from tahmeed.db.connection import get_db
from tahmeed.models.category import Category
from tahmeed.services.description_mapping_service import delete_all_mappings

_PALETTE = [
    "#4A90D9", "#16A34A", "#D97706", "#7C3AED", "#DC2626",
    "#0891B2", "#DB2777", "#65A30D", "#EA580C", "#6366F1",
]


def _cell_str(val) -> str:
    """Strip and uppercase text cells (same as Toll Plaza import)."""
    if val is None:
        return ""
    return str(val).strip().upper()


def _parse_chart_of_accounts(path: str | Path) -> List[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = "Chart of Accounts"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = _cell_str(row[0])
            if not name:
                continue
            rows.append({
                "name": name,
                "ref_num": _cell_str(row[1]) if len(row) > 1 else "",
                "account_type": _cell_str(row[3]) if len(row) > 3 else "",
                "coa_description": _cell_str(row[5]) if len(row) > 5 else "",
                "account_number": _cell_str(row[6]) if len(row) > 6 else "",
                "currency": _cell_str(row[14]) if len(row) > 14 else "",
            })
        return rows
    finally:
        wb.close()


async def import_chart_of_accounts(
    path: str | Path,
    *,
    replace_existing: bool = True,
) -> dict:
    """Load all Chart of Accounts rows into ``categories`` (Items tab).

    When ``replace_existing`` is True (default), all existing items, sub-items,
    keyword rules, and description mappings are removed first.

    Category / subtable / keyword wipe + insert run in one Mongo transaction
    when the server supports it (replica set). Description mappings are cleared
    via the API after the Mongo replace succeeds.
    """
    from tahmeed.db.mongo_txn import run_in_transaction, session_kwargs

    parsed = _parse_chart_of_accounts(path)
    if not parsed:
        raise ValueError("No accounts found in the Chart of Accounts file.")

    db = get_db()
    removed = {"categories": 0, "subtables": 0, "keyword_rules": 0, "mappings": 0}

    docs = []
    for i, row in enumerate(parsed):
        cat = Category(
            name=row["name"],
            description=row["coa_description"],
            color=_PALETTE[i % len(_PALETTE)],
            sort_order=i,
            account_type=row["account_type"],
            ref_num=row["ref_num"],
            account_number=row["account_number"],
            currency=row["currency"],
            coa_description=row["coa_description"],
            show_in_sidebar=False,
            show_in_cashier_sidebar=False,
            requires_truck=False,
            requires_receipt=False,
        )
        docs.append(cat.to_doc())

    async def _replace(session):
        kw = session_kwargs(session)
        local_removed = {"categories": 0, "subtables": 0, "keyword_rules": 0}
        if replace_existing:
            r = await db.categories.delete_many({}, **kw)
            local_removed["categories"] = r.deleted_count
            r = await db.category_subtables.delete_many({}, **kw)
            local_removed["subtables"] = r.deleted_count
            r = await db.keyword_rules.delete_many({}, **kw)
            local_removed["keyword_rules"] = r.deleted_count
        if docs:
            await db.categories.insert_many(docs, **kw)
        return local_removed

    mongo_removed = await run_in_transaction(_replace)
    removed.update(mongo_removed)

    if replace_existing:
        # After Mongo catalog is replaced — clear mappings (API path).
        removed["mappings"] = await delete_all_mappings()

    result = {
        "imported": len(docs),
        "removed": removed,
        "source": str(path),
    }
    try:
        from tahmeed.services.audit_service import record_event

        await record_event(
            "coa.replace",
            entity_type="category",
            details=result,
        )
    except Exception:
        pass
    return result
