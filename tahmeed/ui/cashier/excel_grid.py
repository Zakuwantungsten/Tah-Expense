"""
DailyRegister — unified QuickBooks-style cashier register.

Layout (top to bottom):
  ┌─ Date nav bar ────────────────────────────────────────┐
  │  ← Prev Day  |  09 June 2026 ▼  |  Next Day →  Today │
  ├─ Column headers ──────────────────────────────────────┤
  │  # | Date | Description | Truck | Memo | TZS | … │
  ├─ Saved rows (read-only, light-blue) ──────────────────┤
  │  ...existing transactions for the selected date...     │
  ├─ New entry rows (editable, white) ────────────────────┤
  │  ...blank rows for new entry...                        │
  ├─ Footer ──────────────────────────────────────────────┤
  │  5 entries  ·  TZS 2,202,500                           │
  └───────────────────────────────────────────────────────┘

Keyboard:
  Arrow / Tab / Enter       navigate cells
  Ctrl+C / Ctrl+V / Ctrl+X  clipboard (TSV — Excel-compatible)
  Delete / Backspace        clear selected cells
  Right-click               context menu (delete saved / edit row ops)

Save:
  save_rows() saves all non-empty editable rows and reloads the register.
  Saved rows are marked read-only with a blue tint.
"""

import asyncio
import csv
import time
from datetime import datetime, date
from typing import List, Optional

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QToolButton, QLabel,
    QTableWidget, QTableWidgetItem, QApplication,
    QAbstractItemDelegate, QStyledItemDelegate, QStyleOptionViewItem, QMenu, QFileDialog,
    QMessageBox, QAbstractItemView, QHeaderView, QDateEdit, QLineEdit,
    QStyle, QComboBox, QDialog, QFrame, QListWidget, QListWidgetItem, QPushButton,
)
from PySide6.QtCore import (
    Qt, Signal, QDate, QEvent, QRect, QSize, QObject, QTimer, QEventLoop,
    QItemSelection, QItemSelectionModel,
)
from PySide6.QtGui import (
    QAction, QKeyEvent, QColor, QBrush, QFont, QPen, QPainter, QMouseEvent, QCursor,
)

from tahmeed.models.category import Category
from tahmeed.models.transaction import Transaction, pack_money
from tahmeed.models.user import User
from tahmeed.services.truck_service import get_fleet_numbers
from tahmeed.services.truck_format import (
    normalize_truck_number, try_match_fleet, normalize_place_label,
    is_allowed_place_label, DEFAULT_PLACE_LABELS, merge_allowed_labels,
)
from tahmeed.services.cashier_service import (
    get_transactions_by_date, save_transaction, request_or_delete_transaction,
    update_transaction, insert_pending_edit,
    check_for_duplicates, submit_day_for_verify, recount_day_order,
    next_day_order,
)
from tahmeed.services.category_service import get_payment_target_categories, item_key
from tahmeed.services.subtable_service import get_subtables
from tahmeed.services.settings_service import get_setting, set_setting
from tahmeed.services.register_draft_service import (
    build_draft_payload,
    cells_for_json,
    cells_from_json,
    clear_register_draft,
    draft_is_empty,
    hydrate_pending_meta,
    load_register_draft,
    save_register_draft,
    serialize_pending_meta,
)
from tahmeed.ui.widgets.loading_overlay import LoadingOverlay
from tahmeed.ui.widgets.truck_autocomplete import TruckLineEdit
from tahmeed.ui.widgets.completer_line_edit import CompleterLineEdit, accept_completion
from tahmeed.ui.dialogs.truck_correction_dialog import TruckCorrectionDialog, TruckIssue
from tahmeed.ui.dialogs.duplicate_review_dialog import DuplicateReviewDialog
from tahmeed.services.duplicate_review import DuplicateReviewItem, format_amount_label

# ---------------------------------------------------------------------------
# Column indices / colors / delegates (shared with RejectedView)
# ---------------------------------------------------------------------------
from tahmeed.ui.cashier.register_delegates import (  # noqa: E402
    COL_SNO, COL_DATE, COL_ITEM, COL_DESC, COL_TRUCK, COL_MEMO,
    COL_REF, COL_TZS, COL_USD, COL_RECEIPT, COL_OWN, COL_APR, COL_PAYEE, COL_CHEQUE,
    COL_CASHIER,
    HEADERS, CHECK_COLS, READONLY_COLS, _DATA_SKIP_COLS, _UPPER_SKIP_COLS,
    DEFAULT_EDITABLE_ROWS, _REF_FLOAT_OPTS, _COL_PREFERRED, _COL_FLEX, _COL_MIN,
    _is_refund_float, _ref_float_text, _parse_optional_date, format_register_date,
    SAVED_BG, NEW_BG, EMPTY_BG, NEG_COLOR, EDIT_BG, DIRTY_BG, DRAFT_BG, DUP_BG, MISMATCH_BG,
    TRUCK_REQUIRED_BG,
    _FOOTER_BTN_STYLE,
    _accept_editor_completion, _upper_text,
    _ExcelCellDelegate, _DescriptionDelegate, _TruckDelegate, _DateDelegate,
    _RefFloatDelegate, _norm_receipt_text, _receipt_paste_value, _parse_amount_text,
    _parse_optional_amount_text, _ReceiptDelegate,
    _ItemDelegate, _CurrencyLineEdit, _TZSDelegate,
    _RCPT_COLORS, _RCPT_LABEL, _RECEIPT_OPTS, _RCPT_OPT_KEY, _RCPT_NORM, _VALID_RCPT,
)

from tahmeed.ui.cashier.excel_cursors import excel_hover_cursor, excel_drag_cursor
from tahmeed.ui.cashier.excel_row_header import (
    ExcelRowHeaderView,
    sync_row_header_labels,
)


def _amount_item_from_raw(raw: str) -> QTableWidgetItem:
    amt = _parse_amount_text(raw)
    text = f"{amt:,.2f}" if (raw or "").strip() else ""
    it = QTableWidgetItem(text)
    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
    return it


def _display_money_cells(tx: Transaction) -> tuple[str, str]:
    tzs, usd = tx.money_parts()
    tzs_txt = f"{tzs:,.2f}" if tzs else ""
    usd_txt = f"{usd:,.2f}" if usd else ""
    return tzs_txt, usd_txt


def _money_item(text: str) -> QTableWidgetItem:
    it = QTableWidgetItem(text)
    it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
    return it



# ---------------------------------------------------------------------------
# Key event filter — captures Tab before Qt's focus-chain system can steal it
# ---------------------------------------------------------------------------

class _TableKeyFilter(QObject):
    def __init__(self, handler):
        super().__init__()
        self._handler = handler

    def eventFilter(self, obj, event) -> bool:
        if event.type() == QEvent.KeyPress:
            self._handler(event)
            return True
        return False


# ---------------------------------------------------------------------------
# Table with Excel-like row gutter + fill handle
# ---------------------------------------------------------------------------

_ROWS_CLIP_PREFIX = "TAHMEED_ROWS_V1\n"
_FILL_HANDLE_SIZE = 7
_FILL_HANDLE_HIT = 14   # corner target for hover / click / drag start
# Excel autofill preview — green tint + ghost values (distinct from normal selection)
_FILL_RANGE_BORDER = QPen(QColor("#217346"), 2, Qt.DashLine)


class _ExcelTableViewport(QWidget):
    """Table viewport that paints the Excel fill handle and drag preview."""

    def __init__(self, table: "_ExcelTableWidget") -> None:
        super().__init__(table)
        self._table = table

    def mouseMoveEvent(self, event) -> None:
        table = self._table
        if table._fill_dragging and (event.buttons() & Qt.LeftButton):
            table._update_fill_drag_end(table._viewport_pos(event))
            table.setCursor(table._drag_cursor)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event) -> None:
        table = self._table
        if table._fill_dragging and event.button() == Qt.LeftButton:
            pos = table._viewport_pos(event)
            table._update_fill_drag_end(pos)
            table._finish_fill_drag(pos)
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def paintEvent(self, event) -> None:
        super().paintEvent(event)
        painter = QPainter(self)
        try:
            self._table._paint_excel_overlay(painter)
        finally:
            painter.end()


class _ExcelTableWidget(QTableWidget):
    """Excel-like grid: row gutter, cell cross cursor, fill handle + drag fill."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._grid_owner = None
        self._fill_dragging = False
        self._fill_anchor: tuple[int, int, int, int] | None = None
        self._fill_end: tuple[int, int] | None = None
        self._cursor_zone = "arrow"
        self.setViewport(_ExcelTableViewport(self))
        self.viewport().setMouseTracking(True)
        self.setMouseTracking(True)
        self._hover_cursor = excel_hover_cursor()
        self._drag_cursor = excel_drag_cursor()

    def leaveEvent(self, event) -> None:
        self.unsetCursor()
        self._cursor_zone = "arrow"
        super().leaveEvent(event)

    def _owner(self):
        return getattr(self, "_grid_owner", None)

    def _selection_bounds(self) -> tuple[int, int, int, int] | None:
        indexes = self.selectedIndexes()
        if not indexes:
            cur = self.currentIndex()
            if not cur.isValid():
                return None
            r, c = cur.row(), cur.column()
            return r, c, r, c
        rows = [i.row() for i in indexes]
        cols = [i.column() for i in indexes]
        return min(rows), min(cols), max(rows), max(cols)

    def _selection_viewport_rect(self) -> QRect | None:
        bounds = self._selection_bounds()
        if bounds is None:
            return None
        r0, c0, r1, c1 = bounds
        tl = self.visualRect(self.model().index(r0, c0))
        br = self.visualRect(self.model().index(r1, c1))
        if not tl.isValid() or not br.isValid():
            return None
        return tl.united(br)

    def _fill_handle_rect(self) -> QRect | None:
        owner = self._owner()
        if owner is None or not self._fill_handle_enabled():
            return None
        sel = self._selection_viewport_rect()
        if sel is None or sel.width() <= 0 or sel.height() <= 0:
            return None
        size = _FILL_HANDLE_SIZE
        return QRect(
            sel.right() - size + 2,
            sel.bottom() - size + 2,
            size,
            size,
        )

    def _fill_handle_enabled(self) -> bool:
        owner = self._owner()
        bounds = self._selection_bounds()
        if owner is None or bounds is None:
            return False
        r0, c0, r1, c1 = bounds
        for row in range(r0, r1 + 1):
            for col in range(c0, c1 + 1):
                if col in READONLY_COLS or col == COL_SNO:
                    continue
                if owner._cell_editable(row, col):
                    return True
        return False

    def _viewport_pos(self, pos) -> QPoint:
        if hasattr(pos, "position"):
            return pos.position().toPoint()
        return pos

    def _cell_at_viewport_pos(self, pos) -> tuple[int, int] | None:
        """Row/col under *pos* (viewport coords), even for empty cells."""
        pt = self._viewport_pos(pos)
        row = self.rowAt(pt.y())
        col = self.columnAt(pt.x())
        if col < 0 and self.columnCount():
            col = self.columnAt(max(0, min(pt.x(), self.viewport().width() - 1)))
        if row < 0 and self.rowCount():
            if pt.y() >= self.viewport().height() - 2:
                row = self.rowCount() - 1
            elif pt.y() <= 0:
                row = 0
        if row < 0 or col < 0:
            return None
        return row, col

    def _fill_handle_hit_rect(self) -> QRect | None:
        """Bottom-right corner hit zone — cursor + click (wider than the painted square)."""
        sel = self._selection_viewport_rect()
        if sel is None or not self._fill_handle_enabled():
            return None
        hit = _FILL_HANDLE_HIT
        return QRect(
            sel.right() - hit + 1,
            sel.bottom() - hit + 1,
            hit,
            hit,
        )

    def _pos_in_fill_handle(self, pos) -> bool:
        hit = self._fill_handle_hit_rect()
        return hit is not None and hit.contains(self._viewport_pos(pos))

    def _wants_fill_drag(self, pos) -> bool:
        return self._pos_in_fill_handle(pos) or self._cursor_zone == "fill"

    def _begin_fill_drag(self, pos) -> bool:
        if not self._wants_fill_drag(pos) or not self._fill_handle_enabled():
            return False
        bounds = self._selection_bounds()
        if bounds is None:
            return False
        r0, c0, r1, c1 = bounds
        self._fill_dragging = True
        self._fill_anchor = bounds
        self._fill_end = (r1, c1)
        self._cursor_zone = "drag"
        self.setCursor(self._drag_cursor)
        self.viewport().grabMouse()
        self._refresh_fill_drag_visual()
        return True

    def _is_fill_extension_cell(self, row: int, col: int) -> bool:
        if not self._fill_dragging or self._fill_anchor is None or self._fill_end is None:
            return False
        r0, c0, r1, c1 = self._fill_anchor
        er, ec = self._fill_end
        er = max(r1, er)
        ec = max(c1, ec)
        if er <= r1 and ec <= c1:
            return False
        if row < r0 or row > er or col < c0 or col > ec:
            return False
        return not (r0 <= row <= r1 and c0 <= col <= c1)

    def _update_fill_drag_end(self, pos) -> None:
        if self._fill_anchor is None:
            return
        cell = self._cell_at_viewport_pos(pos)
        if cell is None:
            return
        _r0, _c0, r1, c1 = self._fill_anchor
        row, col = cell
        self._fill_end = (max(row, r1), max(col, c1))
        self._restore_fill_source_selection()
        self._auto_scroll_for_fill(self._viewport_pos(pos))
        self._refresh_fill_drag_visual()

    def _refresh_fill_drag_visual(self) -> None:
        """Force an immediate paint pass so fill preview shows while dragging."""
        self.viewport().repaint()

    def _fill_drag_target_bounds(self) -> tuple[int, int, int, int] | None:
        if not self._fill_dragging or self._fill_anchor is None or self._fill_end is None:
            return None
        r0, c0, r1, c1 = self._fill_anchor
        er, ec = self._fill_end
        er = max(r1, er)
        ec = max(c1, ec)
        return r0, c0, er, ec

    def _restore_fill_source_selection(self) -> None:
        """Keep the original source block highlighted while previewing the extension."""
        if self._fill_anchor is None:
            return
        r0, c0, r1, c1 = self._fill_anchor
        sel = QItemSelection(
            self.model().index(r0, c0),
            self.model().index(r1, c1),
        )
        self.selectionModel().select(sel, QItemSelectionModel.ClearAndSelect)

    def _select_fill_result_range(self, anchor, end) -> None:
        r0, c0, r1, c1 = anchor
        er, ec = end
        er = max(r1, er)
        ec = max(c1, ec)
        sel = QItemSelection(
            self.model().index(r0, c0),
            self.model().index(er, ec),
        )
        self.selectionModel().select(sel, QItemSelectionModel.ClearAndSelect)

    def _fill_preview_source_coords(
        self, row: int, col: int,
    ) -> tuple[int, int] | None:
        if self._fill_anchor is None:
            return None
        r0, c0, r1, c1 = self._fill_anchor
        sel_h = r1 - r0 + 1
        sel_w = c1 - c0 + 1
        return r0 + (row - r0) % sel_h, c0 + (col - c0) % sel_w

    def _fill_preview_display_text(self, row: int, col: int) -> str:
        owner = self._owner()
        if owner is not None and (
            col in READONLY_COLS
            or col == COL_SNO
            or not owner._cell_editable(row, col)
        ):
            return ""
        src = self._fill_preview_source_coords(row, col)
        if src is None:
            return ""
        src_row, src_col = src
        item = self.item(src_row, src_col)
        if item is None:
            return ""
        if src_col in CHECK_COLS:
            return "✓" if item.data(Qt.UserRole) else ""
        return item.text()

    def _fill_preview_text_align(self, col: int) -> int:
        if col in (COL_TZS, COL_USD, COL_REF):
            return int(Qt.AlignRight | Qt.AlignVCenter)
        return int(Qt.AlignLeft | Qt.AlignVCenter)

    def _auto_scroll_for_fill(self, pos) -> None:
        vp = self.viewport()
        margin = 20
        bounds = self._fill_drag_target_bounds()
        if bounds is None:
            return
        _r0, _c0, er, ec = bounds
        if pos.y() >= vp.height() - margin and er < self.rowCount() - 1:
            self.scrollTo(
                self.model().index(er + 1, ec),
                QAbstractItemView.EnsureVisible,
            )
        elif pos.y() <= margin and er > self._fill_anchor[0]:
            self.scrollTo(
                self.model().index(max(self._fill_anchor[0], er - 1), ec),
                QAbstractItemView.EnsureVisible,
            )

    def _finish_fill_drag(self, pos) -> None:
        owner = self._owner()
        anchor = self._fill_anchor
        end = self._fill_end
        if owner is not None and anchor is not None and end is not None:
            owner._apply_fill_drag(anchor, end)
            self._select_fill_result_range(anchor, end)
        self._fill_dragging = False
        self._fill_anchor = None
        self._fill_end = None
        self.viewport().releaseMouse()
        self.viewport().update()
        self._update_hover_cursor(pos)

    def _update_hover_cursor(self, pos, buttons=Qt.NoButton) -> None:
        pos = self._viewport_pos(pos)
        if self._fill_dragging:
            zone = "drag"
        elif buttons & Qt.LeftButton:
            zone = "drag"
        elif self._pos_in_fill_handle(pos):
            zone = "fill"
        else:
            index = self.indexAt(pos)
            if index.isValid():
                zone = "hover"
            else:
                zone = "arrow"
        if zone == self._cursor_zone:
            return
        old_zone = self._cursor_zone
        self._cursor_zone = zone
        if zone == "hover":
            self.setCursor(self._hover_cursor)
        elif zone in ("fill", "drag"):
            self.setCursor(self._drag_cursor)
        else:
            self.unsetCursor()
        if old_zone == "fill" or zone == "fill":
            self.viewport().update()

    def _paint_excel_overlay(self, painter: QPainter) -> None:
        handle = self._fill_handle_rect()
        show_handle = (
            handle is not None
            and not self._fill_dragging
            and self._cursor_zone == "fill"
        )
        if show_handle:
            painter.save()
            painter.setPen(QPen(QColor("#111827"), 1))
            painter.setBrush(QColor("#ffffff"))
            painter.drawRect(handle.adjusted(0, 0, -1, -1))
            painter.restore()

        if self._fill_dragging:
            self._paint_fill_drag_preview(painter)

    def _paint_fill_drag_preview(self, painter: QPainter) -> None:
        """Dashed outline around the full autofill range (cells painted by delegate)."""
        if self._fill_anchor is None or self._fill_end is None:
            return
        r0, c0, r1, c1 = self._fill_anchor
        er, ec = self._fill_end
        er = max(r1, er)
        ec = max(c1, ec)
        if er <= r1 and ec <= c1:
            return

        tl = self.visualRect(self.model().index(r0, c0))
        br = self.visualRect(self.model().index(er, ec))
        if tl.isNull() or br.isNull():
            return
        full = tl.united(br)

        painter.save()
        painter.setBrush(Qt.NoBrush)
        painter.setPen(_FILL_RANGE_BORDER)
        painter.drawRect(full.adjusted(0, 0, -1, -1))
        painter.restore()

    def _fill_preview_rect(self) -> QRect | None:
        bounds = self._fill_drag_target_bounds()
        if bounds is None:
            return None
        r0, c0, er, ec = bounds
        if er <= self._fill_anchor[2] and ec <= self._fill_anchor[3]:
            return None
        tl = self.visualRect(self.model().index(r0, c0))
        br = self.visualRect(self.model().index(er, ec))
        if tl.isNull() or br.isNull():
            return None
        return tl.united(br)

    def mousePressEvent(self, event: QMouseEvent) -> None:
        pos = self._viewport_pos(event)
        if event.button() == Qt.LeftButton and self._begin_fill_drag(pos):
            event.accept()
            return

        super().mousePressEvent(event)
        self._update_hover_cursor(pos)
        self.viewport().update()

    def mouseMoveEvent(self, event: QMouseEvent) -> None:
        pos = self._viewport_pos(event)
        if self._fill_dragging and (event.buttons() & Qt.LeftButton):
            self._update_fill_drag_end(pos)
            self.setCursor(self._drag_cursor)
            event.accept()
            return

        super().mouseMoveEvent(event)
        self._update_hover_cursor(pos, event.buttons())

    def mouseReleaseEvent(self, event: QMouseEvent) -> None:
        pos = self._viewport_pos(event)
        if self._fill_dragging and event.button() == Qt.LeftButton:
            self._update_fill_drag_end(pos)
            self._finish_fill_drag(pos)
            event.accept()
            return

        super().mouseReleaseEvent(event)
        self._update_hover_cursor(pos)
        self.viewport().update()


# ---------------------------------------------------------------------------
# Column filter header
# ---------------------------------------------------------------------------

_FILTER_COLS = set(range(len(HEADERS))) - {COL_SNO}


def cascade_column_values(
    rows: List[dict],
    *,
    target_col: int,
    active_filters: dict,
) -> set:
    """Distinct values for *target_col* from rows that pass every *other* filter.

    Re-exported from the shared Excel filter widget for Daily Register.
    """
    from tahmeed.ui.widgets.excel_column_filter import (
        cascade_column_values as _cascade,
    )

    return _cascade(rows, target_col=target_col, active_filters=active_filters)


class _ColumnFilterPopup(QFrame):
    """Excel-style checklist popup: values from the table only, with search + Apply."""

    applied = Signal(object)  # set[str] | empty set = Show All

    def __init__(self, values: set, current: set, parent=None):
        super().__init__(parent, Qt.Popup | Qt.FramelessWindowHint)
        self.setObjectName("colFilterPopup")
        self.setStyleSheet(
            "QFrame#colFilterPopup{"
            " background:#ffffff;border:1px solid #D1D5DB;border-radius:6px;}"
        )
        self._all_values = sorted(values, key=lambda v: v.lower())
        self._current = set(current or [])

        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        hint = QLabel(f"{len(self._all_values)} value(s) in this table")
        hint.setStyleSheet(
            "font-size:10px;color:#6B7280;background:transparent;border:none;"
        )
        root.addWidget(hint)

        self._search = QLineEdit()
        self._search.setPlaceholderText("Search…")
        self._search.setClearButtonEnabled(True)
        self._search.setStyleSheet(
            "QLineEdit{border:1px solid #D1D5DB;border-radius:4px;"
            "padding:4px 8px;font-size:12px;}"
        )
        self._search.textChanged.connect(self._refilter)
        root.addWidget(self._search)

        self._list = QListWidget()
        self._list.setMinimumWidth(220)
        self._list.setMaximumHeight(260)
        self._list.setStyleSheet(
            "QListWidget{border:1px solid #E5E7EB;border-radius:4px;font-size:12px;}"
            "QListWidget::item{padding:3px 6px;}"
        )
        root.addWidget(self._list, 1)

        btns = QHBoxLayout()
        btns.setSpacing(6)
        show_all = QPushButton("Show All")
        show_all.setCursor(Qt.PointingHandCursor)
        show_all.setEnabled(bool(self._current))
        show_all.clicked.connect(self._on_show_all)
        apply_btn = QPushButton("Apply")
        apply_btn.setCursor(Qt.PointingHandCursor)
        apply_btn.setStyleSheet(
            "QPushButton{background:#0077C5;color:#fff;border:none;"
            "border-radius:4px;padding:5px 12px;font-weight:600;}"
        )
        apply_btn.clicked.connect(self._on_apply)
        btns.addWidget(show_all)
        btns.addStretch()
        btns.addWidget(apply_btn)
        root.addLayout(btns)

        self._refilter("")
        self._search.setFocus()

    def _refilter(self, text: str = "") -> None:
        needle = (text or self._search.text() or "").strip().lower()
        self._list.clear()
        for val in self._all_values:
            if needle and needle not in val.lower():
                continue
            it = QListWidgetItem(val)
            it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
            it.setCheckState(
                Qt.Checked if val in self._current else Qt.Unchecked
            )
            self._list.addItem(it)

    def _checked(self) -> set:
        # Start from prior selection, then sync visible rows' check states
        # (so search doesn't wipe hidden checked values).
        result = set(self._current)
        for i in range(self._list.count()):
            it = self._list.item(i)
            if it.checkState() == Qt.Checked:
                result.add(it.text())
            else:
                result.discard(it.text())
        return result

    def _on_show_all(self) -> None:
        self.applied.emit(set())
        self.close()

    def _on_apply(self) -> None:
        self.applied.emit(self._checked())
        self.close()


class _FilterHeaderView(QHeaderView):
    """Horizontal header ▾ filters — options only from table rows, with chaining."""

    filter_changed = Signal(int, set)   # (col_index, accepted_values); empty = cleared

    def __init__(self, parent=None):
        super().__init__(Qt.Horizontal, parent)
        self._active: dict = {}   # col -> set of accepted values
        self._value_provider = None  # optional callable(col) -> set[str]
        self._popup = None

    def set_value_provider(self, provider) -> None:
        self._value_provider = provider

    def clear_filters(self) -> None:
        self._active.clear()
        self.viewport().update()

    def sync_active(self, filters: dict) -> None:
        """Mirror DailyRegister._col_filters onto the chevron paint state."""
        self._active = {c: set(v) for c, v in (filters or {}).items() if v}
        self.viewport().update()

    def paintSection(self, painter, rect, logical_index):
        super().paintSection(painter, rect, logical_index)
        if logical_index not in _FILTER_COLS or rect.width() < 28:
            return
        painter.save()
        is_active = bool(self._active.get(logical_index))
        painter.setPen(QColor("#EA580C") if is_active else QColor("#94A3B8"))
        f = painter.font()
        f.setPointSize(7)
        painter.setFont(f)
        painter.drawText(
            QRect(rect.right() - 15, rect.top(), 13, rect.height()),
            Qt.AlignVCenter | Qt.AlignHCenter,
            "▾",
        )
        painter.restore()

    def mousePressEvent(self, event):
        col = self.logicalIndexAt(event.pos())
        if col in _FILTER_COLS:
            x      = event.pos().x()
            col_x  = self.sectionViewportPosition(col)
            col_w  = self.sectionSize(col)
            if x >= col_x + col_w - 20:
                self._open_menu(col, event.globalPosition().toPoint())
                return
        super().mousePressEvent(event)

    def _open_menu(self, col: int, global_pos) -> None:
        if not callable(self._value_provider):
            return
        values = set(self._value_provider(col) or [])
        current = set(self._active.get(col, set()) or [])
        # Keep currently selected values visible so they can be unchecked.
        values |= current
        if not values and not current:
            return

        if self._popup is not None:
            self._popup.close()
            self._popup = None

        popup = _ColumnFilterPopup(values, current, parent=self)
        self._popup = popup

        def _on_applied(new_filter):
            new_filter = set(new_filter or [])
            if new_filter:
                self._active[col] = new_filter
            else:
                self._active.pop(col, None)
            self.filter_changed.emit(col, new_filter)
            self.viewport().update()

        popup.applied.connect(_on_applied)
        # Position under the chevron
        popup.adjustSize()
        popup.move(global_pos)
        popup.show()


# ---------------------------------------------------------------------------
# DailyRegister
# ---------------------------------------------------------------------------

class DailyRegister(QWidget):
    """Unified daily expense register (replaces ExcelGrid + TransactionsTable)."""

    rows_saved        = Signal(int)
    stats_updated     = Signal(int, float, float, float, object)  # n, tzs, usd, refund, date
    register_status_updated = Signal(int, int)  # draft_count, submitted_count on saved rows
    drafts_changed        = Signal()
    undo_redo_changed     = Signal(bool, bool)  # can_undo, can_redo
    edit_state_changed = Signal(bool, int)         # (edit_mode_active, dirty_row_count)
    mode_changed      = Signal(bool)               # merged mode on/off
    attachment_count_changed = Signal(int)         # selected row attachment count
    save_busy_changed = Signal(bool)               # True while Save/Submit is in flight
    # Day-level Cheque for the always-visible header field
    active_cheque_changed = Signal(str, bool)  # cheque, editable

    def __init__(self, user: User, categories: List[Category], parent=None):
        super().__init__(parent)
        self._user        = user
        self._categories  = categories
        self._cat_by_name: dict = {c.name.lower(): c for c in categories}
        self._locked_subitems: dict = {}   # item name (lower) -> [sub-item names]
        self._restrict_items: bool = False
        self._export_restrict_surfaces: set = set()
        self._defer_item_to_verify: bool = False
        self._restrict_trucks: bool = True  # always on — only registered fleet numbers
        self._fleet_numbers: set = set()   # uppercased valid fleet numbers
        self._fleet_kinds: dict = {}       # number → "truck" | "trailer" | "motor_vehicle"
        self._allowed_truck_labels: set = set(DEFAULT_PLACE_LABELS)
        self._people_names: list = []      # Ownership / APR BY suggestions (unrestricted)
        self._cashier_names: dict = {}     # ObjectId -> display name
        self._merged_mode: bool = False    # Shared/Merged day (all cashiers)
        self._current_date: Optional[date] = date.today()
        self._saved_count: int   = 0
        self._saved_ids: dict    = {}   # row_index -> ObjectId
        self._saved_txs: dict    = {}   # row_index -> original Transaction (saved rows)
        self._edit_mode: bool    = False
        self._dirty_rows: set    = set()  # saved row indices modified while editing
        self._col_filters: dict   = {}   # col -> set of accepted values
        self._search_text: str    = ""
        self._pending_highlight: str = ""  # set by navigate_to_date; consumed in _populate
        self._load_upload_id: str = ""     # one-shot: load this Excel batch instead of a day
        # When True, skip local draft restore (used while staging a daily import).
        self._skip_draft_restore: bool = False
        # Day-level Cheque stamp (header field → all data rows)
        self._header_cheque: str = ""
        # row_index -> import metadata stamped onto Transaction at save time
        self._pending_row_meta: dict = {}
        # When True, skip async side-effects from itemChanged (bulk paste/import).
        self._bulk_mutating: bool = False
        # When True, queue truck issues but do not open the correction dialog yet
        # (avoids nested asyncio during daily import modals on Python 3.14).
        self._suppress_truck_dialog: bool = False
        # Coalesce truck issues into one combined dialog (paste / import / edit).
        self._pending_truck_issues: dict = {}  # row -> TruckIssue
        # row -> partner/non-fleet truck accepted via "Allow anyway"
        self._truck_allow_anyway: dict[int, str] = {}
        self._truck_dialog_scheduled: bool = False
        self._open_truck_dialog: object = None
        # Excel cut marquee (cells stay until paste / Insert Cut Cells / Esc)
        self._cut_cells: set = set()          # {(row, col), ...}
        self._cut_payload: dict = {}          # serialized cut buffer
        self._cut_is_rows: bool = False
        # Undo / redo stacks (Excel-style cell + row operations)
        self._undo_stack: list = []
        self._redo_stack: list = []
        self._undo_limit: int = 100
        self._undo_redo_active: bool = False
        # Row-level edit transaction: snapshot whole row so side effects
        # (S/N, Date, Item auto-fill) undo together with the edited cell.
        self._pending_row_edit: dict | None = None
        self._closing_row_edit: dict | None = None
        self._auto_fill_row: int | None = None
        self._edit_finalize_timer = QTimer(self)
        self._edit_finalize_timer.setSingleShot(True)
        self._edit_finalize_timer.setInterval(50)
        self._edit_finalize_timer.timeout.connect(self._finalize_pending_row_edit)
        # Re-entrancy guards — prevent double-click duplicate inserts.
        self._save_in_flight = False
        self._submit_in_flight = False
        # Local draft autosave (crash / power-loss recovery).
        self._draft_timer = QTimer(self)
        self._draft_timer.setSingleShot(True)
        self._draft_timer.setInterval(1_500)
        self._draft_timer.timeout.connect(self._flush_local_draft)
        # Debounce DB writes for day-level Cheque (saved rows only).
        self._cheque_persist_timer = QTimer(self)
        self._cheque_persist_timer.setSingleShot(True)
        self._cheque_persist_timer.setInterval(400)
        self._cheque_persist_timer.timeout.connect(
            self._kick_persist_header_cheque
        )
        self._restoring_draft = False
        self._load_gen = 0
        self._build_ui()
        self._show_register_loading("Loading…")
        self._emit_undo_redo_state()
        asyncio.ensure_future(self._load_date(self._current_date))
        asyncio.ensure_future(self._load_categories())
        asyncio.ensure_future(self._load_cashier_settings())
        asyncio.ensure_future(self._load_locked_subitems())
        asyncio.ensure_future(self._load_fleet_numbers())
        asyncio.ensure_future(self._load_people_names())
        asyncio.ensure_future(self._load_description_cache())

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_ui(self) -> None:
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Table ──────────────────────────────────────────────────────
        self._table = _ExcelTableWidget(DEFAULT_EDITABLE_ROWS, len(HEADERS))
        self._table._grid_owner = self
        _fhv = _FilterHeaderView(self._table)
        _fhv.set_value_provider(self._filter_menu_values)
        _fhv.filter_changed.connect(self._on_col_filter_changed)
        self._table.setHorizontalHeader(_fhv)
        self._table.setHorizontalHeaderLabels(HEADERS)
        sno_hdr = self._table.horizontalHeaderItem(COL_SNO)
        if sno_hdr is not None:
            sno_hdr.setTextAlignment(Qt.AlignCenter)
        self._table.setStyleSheet("""
            QTableWidget {
                background: #ffffff;
                gridline-color: #e5e7eb;
                border: none;
                font-family: Calibri;
                font-size: 11pt;
                selection-background-color: #cde0f5;
                selection-color: #1B2B4B;
            }
            QHeaderView::section:horizontal {
                background: #253A5C;
                color: #F9FAFB;
                font-weight: 600;
                font-size: 11px;
                padding: 5px 4px;
                border: none;
                border-right: 1px solid #1B2B4B;
                border-bottom: 2px solid #0077C5;
            }
            QHeaderView::section:vertical {
                background: #F2F2F2;
                color: #333333;
                font-family: Calibri;
                font-size: 11pt;
                padding: 0 2px;
                border: none;
                border-right: 1px solid #D4D4D4;
                border-bottom: 1px solid #D4D4D4;
            }
            QHeaderView:vertical {
                background: #ffffff;
                border: none;
            }
            QTableCornerButton::section {
                background: #F2F2F2;
                border: none;
                border-right: 1px solid #D4D4D4;
                border-bottom: 2px solid #0077C5;
            }
            QTableWidget::item         { padding: 2px 3px; color: #111827; }
            QTableWidget::item:selected { color: #1B2B4B; font-weight: 500; }
            QLineEdit { color: #111827; background: #ffffff; font-family: Calibri; font-size: 11pt; }
        """)

        hh = self._table.horizontalHeader()
        hh.setSectionsMovable(False)
        hh.setStretchLastSection(False)
        hh.setMinimumSectionSize(50)
        # Interactive columns; S/NO fixed; Description stretches to fill the viewport.
        for col in range(len(HEADERS)):
            hh.setSectionResizeMode(col, QHeaderView.Interactive)
        hh.setSectionResizeMode(COL_SNO,  QHeaderView.Fixed)
        hh.setSectionResizeMode(COL_DESC, QHeaderView.Stretch)

        for col, width in _COL_PREFERRED.items():
            self._table.setColumnWidth(col, width)
        QTimer.singleShot(0, self._fit_table_columns)
        self._table.setColumnHidden(COL_CASHIER, True)

        self._table.setSelectionMode(QAbstractItemView.ContiguousSelection)
        self._table.setVerticalHeader(ExcelRowHeaderView(self._table, owner=self))
        self._table.setTabKeyNavigation(False)
        sm = self._table.selectionModel()
        sm.selectionChanged.connect(lambda *_: self._table.viewport().update())
        sm.currentChanged.connect(lambda *_: self._table.viewport().update())

        # Excel selection model on every column; per-column delegates override as needed
        self._table.setItemDelegate(_ExcelCellDelegate(self._table))
        self._table.setItemDelegateForColumn(COL_ITEM,     _ItemDelegate(lambda: [c.name for c in self._categories], self._table))
        self._table.setItemDelegateForColumn(COL_DESC,     _DescriptionDelegate(
            cat_getter=lambda name: self._cat_by_name.get(name.lower()),
            subs_getter=lambda name: self._locked_subitems.get(name.lower(), []),
            parent=self._table,
        ))
        self._table.setItemDelegateForColumn(COL_TRUCK,    _TruckDelegate(
            lambda: sorted(self._fleet_numbers), self._table
        ))
        date_del = _DateDelegate(lambda: self._current_date, self._table)
        self._table.setItemDelegateForColumn(COL_DATE,     date_del)
        self._table.setItemDelegateForColumn(COL_REF,      _RefFloatDelegate(self._table))
        self._table.setItemDelegateForColumn(COL_TZS,      _TZSDelegate(self._table))
        self._table.setItemDelegateForColumn(COL_USD,      _TZSDelegate(self._table))
        self._table.setItemDelegateForColumn(COL_RECEIPT,  _ReceiptDelegate(self._table))
        # Ownership + APR BY — same Item-style autocomplete/preview; free text always allowed.
        people_del = _ItemDelegate(lambda: list(self._people_names), self._table)
        self._table.setItemDelegateForColumn(COL_OWN, people_del)
        self._table.setItemDelegateForColumn(COL_APR, people_del)

        self._table.itemChanged.connect(self._on_item_changed)
        self._table.model().dataChanged.connect(self._on_model_data_changed)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._show_context_menu)
        self._table.itemSelectionChanged.connect(self._emit_attachment_badge)
        self._table.currentCellChanged.connect(
            lambda *_: self._emit_active_cheque()
        )
        self._table.currentCellChanged.connect(self._on_current_cell_changed)

        self._table_host = QWidget(self)
        host_lay = QVBoxLayout(self._table_host)
        host_lay.setContentsMargins(0, 0, 0, 0)
        host_lay.setSpacing(0)
        host_lay.addWidget(self._table)
        root.addWidget(self._table_host, 1)
        self._loading = LoadingOverlay(self._table_host, "Loading…")

        # ── Footer — totals only ───────────────────────────────────────
        footer = QWidget()
        footer.setFixedHeight(48)
        footer.setStyleSheet(
            "background: #f5f6f7;"
            "border-top: 2px solid #d1d5db;"
        )
        fl = QHBoxLayout(footer)
        fl.setContentsMargins(10, 0, 14, 0)
        fl.setSpacing(4)

        fl.addStretch()

        self._totals_label = QLabel("0 entries")
        self._totals_label.setStyleSheet(
            "color: #374151; font-size: 12px; font-weight: 500;"
        )
        fl.addWidget(self._totals_label)

        root.addWidget(footer)

        # Init blank rows
        self._init_editable_rows(0, DEFAULT_EDITABLE_ROWS)
        self._sync_row_header_labels()
        self._install_key_handler()

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._fit_table_columns()

    def showEvent(self, event) -> None:
        super().showEvent(event)
        QTimer.singleShot(0, self._fit_table_columns)

    def _fit_table_columns(self) -> None:
        """Scale columns to the viewport so the register fits without H-scroll."""
        if not hasattr(self, "_table") or self._table is None:
            return
        vp = self._table.viewport().width()
        if vp <= 0:
            return

        widths = dict(_COL_PREFERRED)
        total = sum(widths.values())
        if total > vp:
            deficit = total - vp
            for col in _COL_FLEX:
                if deficit <= 0:
                    break
                floor = _COL_MIN.get(col, 50)
                cut = min(max(0, widths[col] - floor), deficit)
                widths[col] -= cut
                deficit -= cut

        hh = self._table.horizontalHeader()
        # Apply concrete widths, then let Description absorb any leftover slack.
        hh.setSectionResizeMode(COL_DESC, QHeaderView.Interactive)
        for col, width in widths.items():
            self._table.setColumnWidth(col, width)
        hh.setSectionResizeMode(COL_SNO,  QHeaderView.Fixed)
        hh.setSectionResizeMode(COL_DESC, QHeaderView.Stretch)

    def navigate_to_date(
        self, d: date, highlight_term: str = "", *, merged: bool | None = None
    ) -> bool:
        """Called by dashboard when TransactionBrowser 'Go To' is used.

        highlight_term — if provided, the register scrolls to the first row
        containing this text after the date loads and briefly flashes it.
        merged=True shows every cashier's rows (Browse is a merged view).
        """
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date):
            return False

        # Blank New table → user picked a Reconciled Date.
        if self._current_date is None:
            self._pending_highlight = highlight_term
            self._commit_open_editor()
            if merged is True and not self._merged_mode:
                self._merged_mode = True
                self.mode_changed.emit(True)
            if self.has_unsaved_work():
                # Keep typed rows; stamp them onto the chosen day.
                self._current_date = d
                self._stamp_empty_row_dates()
                self._update_footer()
                self._schedule_draft_autosave()
                return True
            self._reset_edit_state()
            self._current_date = d
            self._show_register_loading(f"Loading {d.strftime('%d %b %Y')}…")
            asyncio.ensure_future(self._load_date(d))
            return True

        if (
            d == self._current_date
            and not highlight_term
            and not getattr(self, "_load_upload_id", "")
            and merged is not True
        ):
            return True
        self._pending_highlight = highlight_term
        self._commit_open_editor()
        if merged is True and not self._merged_mode:
            self._merged_mode = True
            self.mode_changed.emit(True)
        if self.has_unsaved_work():
            self._flush_local_draft()
            choice = self._prompt_unsaved_work(
                "You have unsaved changes on this date.\n"
                "Save as draft before leaving?"
            )
            if choice == "cancel":
                self._pending_highlight = ""
                self._load_upload_id = ""
                return False
            if choice == "save":
                asyncio.ensure_future(self._save_then_navigate(d))
                return True
            # Discard → clear local recovery draft for this date, then leave.
            self._clear_local_draft()
        self._reset_edit_state()
        self._current_date = d
        if self._load_upload_id:
            self._show_register_loading("Loading upload…")
        else:
            self._show_register_loading(f"Loading {d.strftime('%d %b %Y')}…")
        asyncio.ensure_future(self._load_date(d))
        return True

    def current_date(self) -> Optional[date]:
        """Reconciled / register day Simple is currently open on (or None if unset)."""
        return self._current_date

    def navigate_to_upload(self, upload_id: str, primary_date=None) -> None:
        """Open every row of one Excel upload on the register table."""
        uid = str(upload_id or "").strip()
        if not uid:
            return
        self._load_upload_id = uid
        d = primary_date
        if isinstance(d, datetime):
            d = d.date()
        if not isinstance(d, date):
            d = self._current_date
        if not isinstance(d, date):
            d = date.today()
        self.navigate_to_date(d, merged=True)

    async def _save_then_navigate(self, d: date) -> None:
        await self._do_save()
        self._current_date = d
        if self._load_upload_id:
            self._show_register_loading("Loading upload…")
        else:
            self._show_register_loading(f"Loading {d.strftime('%d %b %Y')}…")
        await self._load_date(d)

    # ------------------------------------------------------------------
    # Load data
    # ------------------------------------------------------------------

    def _show_register_loading(self, message: str = "Loading…") -> None:
        overlay = getattr(self, "_loading", None)
        if overlay is not None:
            overlay.show_loading(message)

    def _hide_register_loading(self) -> None:
        overlay = getattr(self, "_loading", None)
        if overlay is not None:
            overlay.hide_loading()

    async def _load_date(self, d: date) -> None:
        from tahmeed.ui.async_utils import pause_background_polls

        with pause_background_polls(self):
            await self._load_date_body(d)

    async def _load_date_body(self, d: date) -> None:
        self._load_gen += 1
        seq = self._load_gen
        upload_id = self._load_upload_id
        if upload_id:
            self._show_register_loading("Loading upload…")
        else:
            label = d.strftime("%d %b %Y") if isinstance(d, date) else "register"
            self._show_register_loading(f"Loading {label}…")
        try:
            self._load_upload_id = ""
            if upload_id:
                from tahmeed.services.daily_import_service import get_daily_upload_records
                txs = await get_daily_upload_records(upload_id)
            elif self._merged_mode:
                txs = await get_transactions_by_date(d, merged=True)
            else:
                txs = await get_transactions_by_date(d, cashier_id=self._user._id)
            if seq != self._load_gen:
                return
            ids = [tx.cashier_id for tx in txs if tx.cashier_id]
            cashier_names = {}
            if ids:
                from tahmeed.services.accountant_service import get_cashier_names
                cashier_names = await get_cashier_names(ids)
            if seq != self._load_gen:
                return
            self._cashier_names = cashier_names
            self._pending_row_meta.clear()
            self._populate(txs)
            show_cashier = self._merged_mode or bool(upload_id)
            self._table.setColumnHidden(COL_CASHIER, not show_cashier)
            if not self._skip_draft_restore:
                restored = self._restore_local_draft()
                if restored:
                    self._show_draft_restored_notice(*restored)
        except Exception as exc:
            if seq == self._load_gen:
                QMessageBox.critical(self, "Error", f"Failed to load:\n{exc}")
        finally:
            if seq == self._load_gen:
                self._hide_register_loading()

    def set_merged_mode(self, merged: bool) -> None:
        """Switch My entries ↔ Merged (all cashiers for the day)."""
        if bool(merged) == self._merged_mode:
            return
        self._commit_open_editor()
        if self.has_unsaved_work():
            self._flush_local_draft()
            choice = self._prompt_unsaved_work(
                "You have unsaved changes.\nSave as draft before switching mode?"
            )
            if choice == "cancel":
                self.mode_changed.emit(self._merged_mode)
                return
            if choice == "save":
                asyncio.ensure_future(self._save_then_switch_mode(bool(merged)))
                return
            self._clear_local_draft()
        self._merged_mode = bool(merged)
        self._reset_edit_state()
        self.mode_changed.emit(self._merged_mode)
        if self._current_date is None:
            self._populate([])
            return
        self._show_register_loading("Loading…")
        asyncio.ensure_future(self._load_date(self._current_date))

    async def _save_then_switch_mode(self, merged: bool) -> None:
        ok = await self._do_save()
        if not ok:
            self.mode_changed.emit(self._merged_mode)
            return
        self._merged_mode = merged
        self.mode_changed.emit(self._merged_mode)
        if self._current_date is None:
            self._populate([])
            return
        self._show_register_loading("Loading…")
        await self._load_date(self._current_date)

    def submit_for_verify(self) -> None:
        """Submit every draft row for the current calendar day to Verify."""
        if self._submit_in_flight or self._save_in_flight:
            return
        asyncio.ensure_future(self._do_submit_for_verify())

    async def _do_submit_for_verify(self) -> None:
        if self._submit_in_flight or self._save_in_flight:
            return
        self._submit_in_flight = True
        self.save_busy_changed.emit(True)
        try:
            self._commit_open_editor()
            if self._current_date is None:
                QMessageBox.warning(
                    self,
                    "Reconciled Date required",
                    "Set Reconciled Date before submitting for verify.",
                )
                return
            if self.has_unsaved_work():
                resp = QMessageBox.question(
                    self, "Unsaved changes",
                    "Save changes before submitting this day for verify?",
                    QMessageBox.Yes | QMessageBox.Cancel,
                    QMessageBox.Yes,
                )
                if resp != QMessageBox.Yes:
                    return
                if not await self._do_save():
                    return
            d = self._current_date
            label = d.strftime("%d %b %Y")
            resp = QMessageBox.question(
                self, "Submit for Verify",
                f"Submit all draft entries for {label} to the Verify inbox?\n\n"
                "This sends the whole day's transactions (all cashiers).",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes,
            )
            if resp != QMessageBox.Yes:
                return
            try:
                n = await submit_day_for_verify(d)
                QMessageBox.information(
                    self, "Submitted",
                    f"{n:,} entr{'y' if n == 1 else 'ies'} sent to Verify for {label}.",
                )
                self.drafts_changed.emit()
                await self._load_date(d)
            except Exception as exc:
                QMessageBox.critical(self, "Submit Failed", str(exc))
        finally:
            self._submit_in_flight = False
            if not self._save_in_flight:
                self.save_busy_changed.emit(False)

    def refresh(self) -> None:
        if self._current_date is None:
            self._populate([])
        else:
            asyncio.ensure_future(self._load_date(self._current_date))
        asyncio.ensure_future(self._load_cashier_settings())

    def reload_settings(self) -> None:
        """Re-read the restrict toggles, locked sub-items and fleet list without
        touching the grid rows (so unsaved entries survive). Called on entering
        the table tab."""
        asyncio.ensure_future(self._load_categories())
        asyncio.ensure_future(self._load_cashier_settings())
        asyncio.ensure_future(self._load_locked_subitems())
        asyncio.ensure_future(self._load_fleet_numbers())
        asyncio.ensure_future(self._load_people_names())
        asyncio.ensure_future(self._load_description_cache())

    def update_people(self, names: list) -> None:
        """Refresh Ownership / APR BY suggestion list (free text still allowed)."""
        self._people_names = [str(n).strip().upper() for n in (names or []) if str(n).strip()]

    async def _load_people_names(self) -> None:
        try:
            from tahmeed.services.people_service import get_people_names
            self.update_people(await get_people_names())
        except Exception:
            self._people_names = []

    def _populate(self, transactions: List[Transaction]) -> None:
        # A fresh load always returns the grid to read-only state.
        self._cheque_persist_timer.stop()
        self._edit_mode = False
        self._dirty_rows = set()

        self._table.blockSignals(True)
        self._table.clearContents()
        self._saved_count = len(transactions)
        self._saved_ids   = {}
        self._saved_txs   = {}

        total_rows = self._saved_count + DEFAULT_EDITABLE_ROWS
        self._table.setRowCount(total_rows)

        for i, tx in enumerate(transactions):
            self._fill_saved_row(i, tx)
            self._saved_ids[i] = tx._id
            self._saved_txs[i] = tx

        self._init_editable_rows(self._saved_count, total_rows)
        self._table.blockSignals(False)
        self._renumber()
        self._clear_column_filters()
        self._clear_cut_marquee()
        self._undo_stack.clear()
        self._redo_stack.clear()
        self._emit_undo_redo_state()
        self._update_footer()
        self._apply_filters()
        self.edit_state_changed.emit(False, 0)
        if self._table.currentRow() < 0 and self._table.rowCount() > self._saved_count:
            self._table.setCurrentCell(self._saved_count, COL_DESC)
        self._sync_header_cheque_from_grid()
        self._emit_active_cheque()

        if self._pending_highlight:
            term = self._pending_highlight
            self._pending_highlight = ""
            # Small delay so Qt finishes laying out the rows before we scroll.
            QTimer.singleShot(80, lambda: self.scroll_and_highlight(term))
        self._refresh_truck_required_highlights()

    # ------------------------------------------------------------------
    # Row initialisation helpers
    # ------------------------------------------------------------------

    def _fill_saved_row(self, row: int, tx: Transaction) -> None:
        status = getattr(tx, "register_status", "") or "submitted"
        if status == "draft":
            bg = QBrush(DRAFT_BG)
            status_tip = "Draft — saved but not submitted to Verify"
        else:
            bg = QBrush(SAVED_BG)
            status_tip = "Submitted — awaiting accountant Verify"
        ro = Qt.ItemIsEnabled | Qt.ItemIsSelectable

        def saved_item(text: str, align=Qt.AlignVCenter | Qt.AlignLeft) -> QTableWidgetItem:
            it = QTableWidgetItem(text)
            it.setFlags(ro)
            it.setBackground(bg)
            it.setTextAlignment(align)
            it.setToolTip(status_tip)
            return it

        # S/NO — same row background as siblings (Excel-style continuous row)
        sno = saved_item(str(row + 1), Qt.AlignCenter)
        self._table.setItem(row, COL_SNO, sno)

        date_str = format_register_date(tx.date) if tx.date else ""
        date_item = saved_item(date_str)
        if tx.date and tx.created_at and tx.date.date() != tx.created_at.date():
            date_item.setBackground(QBrush(MISMATCH_BG))
            date_item.setToolTip(
                f"Transaction dated {tx.date.strftime('%d %b %y')} but submitted on "
                f"{tx.created_at.strftime('%d %b %y')}"
            )
        self._table.setItem(row, COL_DATE, date_item)

        self._table.setItem(row, COL_ITEM, saved_item(tx.item or ""))

        desc_item = saved_item(tx.description)
        if tx.possible_duplicate:
            desc_item.setBackground(QBrush(DUP_BG))
            desc_item.setToolTip("Possible duplicate — similar entry found within the check window")
        self._table.setItem(row, COL_DESC, desc_item)
        self._table.setItem(row, COL_TRUCK, saved_item(tx.truck_number or ""))
        self._table.setItem(row, COL_MEMO,  saved_item(tx.memo or ""))
        self._table.setItem(row, COL_REF,   saved_item(_ref_float_text(tx)))

        tzs_txt, usd_txt = _display_money_cells(tx)
        self._table.setItem(row, COL_TZS, _money_item(tzs_txt))
        self._table.setItem(row, COL_USD, _money_item(usd_txt))

        # Receipt
        rcpt_it = saved_item(tx.receipt_status or "pending")
        self._table.setItem(row, COL_RECEIPT, rcpt_it)

        self._table.setItem(row, COL_OWN,    saved_item(tx.ownership or ""))
        self._table.setItem(row, COL_APR,    saved_item(tx.approver or ""))
        self._table.setItem(row, COL_PAYEE,  saved_item(getattr(tx, "payee", "") or ""))
        self._table.setItem(row, COL_CHEQUE, saved_item(getattr(tx, "cheque", "") or ""))
        cashier = self._cashier_names.get(tx.cashier_id, "—") if tx.cashier_id else "—"
        self._table.setItem(row, COL_CASHIER, saved_item(cashier))

    def _init_editable_rows(self, start: int, end: int) -> None:
        # Preserve caller's blockSignals state — never force-unblock mid bulk load.
        prev = self._table.blockSignals(True)
        for row in range(start, end):
            # S/NO — blank until row is activated by data entry or Tab wrap
            sno = QTableWidgetItem("")
            sno.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            sno.setTextAlignment(Qt.AlignCenter)
            self._table.setItem(row, COL_SNO, sno)
            # Checkbox items are created lazily in _activate_row
        self._table.blockSignals(prev)

    # ------------------------------------------------------------------
    # Footer
    # ------------------------------------------------------------------

    def _update_footer(self) -> None:
        """Recompute entries / TZS / USD / refund from the live grid."""
        n, tzs, usd, refund = 0, 0.0, 0.0, 0.0
        for row in range(self._table.rowCount()):
            tzs_it = self._table.item(row, COL_TZS)
            usd_it = self._table.item(row, COL_USD)
            raw_tzs = tzs_it.text().strip() if tzs_it else ""
            raw_usd = usd_it.text().strip() if usd_it else ""
            if not raw_tzs and not raw_usd:
                continue
            tzs_amt = _parse_amount_text(raw_tzs) if raw_tzs else 0.0
            usd_amt = _parse_amount_text(raw_usd) if raw_usd else 0.0
            # Skip non-numeric leftovers that parse as 0
            if (
                raw_tzs
                and tzs_amt == 0.0
                and not any(ch.isdigit() for ch in raw_tzs)
                and not raw_usd
            ):
                continue
            if (
                raw_usd
                and usd_amt == 0.0
                and not any(ch.isdigit() for ch in raw_usd)
                and not raw_tzs
            ):
                continue
            n += 1
            tzs += tzs_amt
            usd += usd_amt
            ref_it = self._table.item(row, COL_REF)
            if ref_it and _is_refund_float(ref_it.text()):
                refund += tzs_amt  # refund-to-float stays TZS-only

        amount_str = f"TZS {tzs:,.0f}" if tzs else "—"
        usd_str = f"USD {usd:,.2f}" if usd else "—"
        self._totals_label.setText(
            f"{n} entr{'y' if n == 1 else 'ies'}   ·   {usd_str}   ·   {amount_str}"
        )
        self.stats_updated.emit(n, tzs, usd, refund, self._current_date)
        draft_n = submitted_n = 0
        for row in range(self._saved_count):
            tx = self._saved_txs.get(row)
            if tx is None:
                continue
            if (getattr(tx, "register_status", "") or "submitted") == "draft":
                draft_n += 1
            else:
                submitted_n += 1
        self.register_status_updated.emit(draft_n, submitted_n)

    # ------------------------------------------------------------------
    # Row → Transaction
    # ------------------------------------------------------------------

    def _build_transaction_from_row(self, row: int) -> Optional[Transaction]:
        """Read cell values for a single row and return a Transaction, or None
        if the row has no description. Raises ValueError on validation errors
        (bad item / locked description / unregistered truck) so callers can
        distinguish logical from network failures and skip retries."""
        def txt(col: int) -> str:
            it = self._table.item(row, col)
            return it.text().strip() if it else ""

        description = txt(COL_DESC)
        if not description:
            return None

        date_str = txt(COL_DATE)
        day = self._current_date or date.today()
        tx_date = _parse_optional_date(date_str, default_year=day.year)
        if tx_date is None:
            tx_date = datetime(day.year, day.month, day.day)

        raw_tzs = txt(COL_TZS)
        raw_usd = txt(COL_USD)
        amount, amount_usd, currency = pack_money(
            _parse_optional_amount_text(raw_tzs),
            _parse_optional_amount_text(raw_usd),
        )

        rcpt_status = txt(COL_RECEIPT).strip()

        item_name = txt(COL_ITEM)
        meta = self._pending_row_meta.get(row) or {}
        allow_blank_item = self._defer_item_to_verify or bool(meta.get("daily_import_id"))
        if not item_name and not allow_blank_item:
            raise ValueError("Item is required. Enter an item or ask the accountant to enable description-only entries.")

        cat = self._cat_by_name.get(item_name.lower()) if item_name else None
        if cat is not None:
            item_name = cat.name.upper()
        elif item_name and self._restrict_items:
            raise ValueError(f'"{item_name}" is not a known item.')
        elif item_name:
            item_name = item_name.upper()

        if cat is not None and getattr(cat, "lock_description", False):
            allowed = self._locked_subitems.get(item_name.lower(), [])
            if allowed:
                match = next(
                    (a for a in allowed if a.lower() == description.lower()), None
                )
                if match is None:
                    raise ValueError(
                        f'"{description}" is not an allowed description for "{item_name}".'
                    )
                description = match.upper()
            else:
                description = description.upper()
        else:
            description = description.upper()

        truck_raw = txt(COL_TRUCK)
        if cat is not None and getattr(cat, "requires_truck", True) and not truck_raw:
            raise ValueError(
                f'Truck number is required for item "{item_name}".'
            )
        truck_number = ""
        if truck_raw:
            allowed_anyway = self._truck_allow_anyway.get(row)
            if (
                allowed_anyway
                and self._truck_key(truck_raw) == self._truck_key(allowed_anyway)
            ):
                truck_number = allowed_anyway
            elif is_allowed_place_label(truck_raw, self._allowed_truck_labels):
                truck_number = normalize_place_label(truck_raw)
            else:
                matched = try_match_fleet(truck_raw, self._fleet_numbers)
                if matched is None:
                    norm = normalize_truck_number(
                        truck_raw, allowed_labels=self._allowed_truck_labels
                    )
                    label = norm.value if norm.status != "empty" else truck_raw
                    if norm.status == "invalid":
                        raise ValueError(
                            f'"{label}" is not a valid truck number '
                            f"(expected T + number + space + suffix, e.g. T688 EAF)."
                        )
                    if norm.status == "place_label":
                        truck_number = norm.value
                    else:
                        raise ValueError(
                            f'"{norm.value}" is not a registered fleet vehicle.'
                        )
                else:
                    truck_number = matched
            it_truck = self._table.item(row, COL_TRUCK)
            if it_truck and truck_number and it_truck.text() != truck_number:
                self._table.blockSignals(True)
                it_truck.setText(truck_number)
                self._table.blockSignals(False)

        ref_text = txt(COL_REF).upper()
        orig = self._saved_txs.get(row)
        primary_dt = self._resolve_import_primary_date(row)
        primary_day = primary_dt.date() if primary_dt else self._current_date
        tx_day = tx_date.date() if hasattr(tx_date, "date") else tx_date
        date_discrep = bool(meta.get("date_discrepancy")) or (tx_day != primary_day)
        return Transaction(
            date=tx_date,
            description=description,
            item=item_name,
            category_name=item_name or None,
            category_id=meta.get("category_id"),
            truck_number=truck_number,
            amount=amount,
            currency=currency,
            amount_usd=amount_usd,
            memo=txt(COL_MEMO).upper(),
            receipt_status=rcpt_status,
            ref_float=ref_text,
            notes_flag=_is_refund_float(ref_text),
            ownership=txt(COL_OWN).upper(),
            approver=txt(COL_APR).upper(),
            payee=txt(COL_PAYEE).upper(),
            cheque=txt(COL_CHEQUE).upper(),
            cashier_id=self._user._id,
            day_order=row,
            register_status="draft",
            daily_import_id=meta.get("daily_import_id"),
            daily_import_source=meta.get("daily_import_source"),
            date_discrepancy=date_discrep,
            import_primary_date=primary_dt,
            lpo_do=(meta.get("lpo_do") or "").upper(),
            do_number=(meta.get("do_number") or "").upper(),
            reported_date=getattr(orig, "reported_date", None) if orig is not None else None,
        )

    # ------------------------------------------------------------------
    # Edit mode
    # ------------------------------------------------------------------

    def toggle_edit_mode(self) -> None:
        """Public entry point for the Edit button: enter edit mode, or exit and
        discard pending changes on a second press."""
        if self._edit_mode:
            if self._dirty_rows:
                resp = QMessageBox.question(
                    self, "Discard changes?",
                    "Exit edit mode and discard your unsaved changes?",
                    QMessageBox.Discard | QMessageBox.Cancel,
                    QMessageBox.Cancel,
                )
                if resp == QMessageBox.Cancel:
                    return
            self._exit_edit_mode(discard=True)
        else:
            self._enter_edit_mode()

    def _enter_edit_mode(self) -> None:
        """Unlock every saved row for editing and tint it warm yellow."""
        self._edit_mode = True
        self._dirty_rows = set()
        editable = Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable
        self._table.blockSignals(True)
        for row in range(self._saved_count):
            for col in range(self._table.columnCount()):
                it = self._table.item(row, col)
                if it is None:
                    continue
                if col not in READONLY_COLS:
                    it.setFlags(editable)
                it.setBackground(QBrush(EDIT_BG))
        self._table.blockSignals(False)
        self.edit_state_changed.emit(True, 0)
        self._emit_active_cheque()
        self._refresh_truck_required_highlights()

    def _exit_edit_mode(self, discard: bool) -> None:
        """Leave edit mode. When discard is True the date is reloaded so the grid
        reverts to the stored values; otherwise the caller reloads after saving."""
        if discard:
            # Keep typed new rows in the local draft; drop dirty saved-row edits.
            self._flush_local_draft(include_dirty=False)
        self._reset_edit_state()
        if discard:
            if self._current_date is None:
                return
            asyncio.ensure_future(self._load_date(self._current_date))

    def _reset_edit_state(self) -> None:
        self._edit_mode = False
        self._dirty_rows = set()
        self.edit_state_changed.emit(False, 0)
        self._emit_active_cheque()

    def _emit_active_cheque(self) -> None:
        """Push the day-level Cheque stamp into the Table header field."""
        self.active_cheque_changed.emit(self._header_cheque, True)

    def _sync_header_cheque_from_grid(self) -> None:
        """Refresh day-level header Cheque from rows that already have data."""
        rows = self._data_rows()
        if not rows:
            self._header_cheque = ""
            return
        cheques = {self._cell_text(r, COL_CHEQUE) for r in rows}
        # Shared value when unanimous; blank when the day has mixed stamps.
        self._header_cheque = next(iter(cheques)) if len(cheques) == 1 else ""

    def set_active_cheque(self, text: str) -> None:
        """Stamp Cheque from the header onto every saved + filled unsaved row."""
        self._header_cheque = (text or "").upper()
        self._stamp_cheque_on_data_rows()
        self._schedule_cheque_persist()
        self._schedule_draft_autosave()

    def _stamp_cheque_on_data_rows(self) -> None:
        """Write header Cheque onto all data rows without dirtying saved rows."""
        rows = self._data_rows()
        if not rows:
            return
        prev = self._table.blockSignals(True)
        try:
            for row in rows:
                self._write_cheque_cell(row, self._header_cheque)
        finally:
            self._table.blockSignals(prev)

    def _write_cheque_cell(self, row: int, value: str) -> None:
        """Set Cheque cell, preserving row chrome (saved / edit / new)."""
        col = COL_CHEQUE
        it = self._table.item(row, col)
        if it is None:
            it = QTableWidgetItem("")
            flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
            if row >= self._saved_count or self._edit_mode:
                flags |= Qt.ItemIsEditable
            it.setFlags(flags)
            if row >= self._saved_count:
                it.setBackground(QBrush(NEW_BG))
            elif self._edit_mode:
                bg = DIRTY_BG if row in self._dirty_rows else EDIT_BG
                it.setBackground(QBrush(bg))
            else:
                it.setBackground(QBrush(SAVED_BG))
            self._table.setItem(row, col, it)
        if it.text() != value:
            it.setText(value)

    def _schedule_cheque_persist(self) -> None:
        """Debounce soft DB updates for saved rows (no Verify → Edited)."""
        if self._saved_count <= 0:
            return
        self._cheque_persist_timer.start()

    def _kick_persist_header_cheque(self) -> None:
        asyncio.ensure_future(self._persist_header_cheque())

    async def _persist_header_cheque(self) -> None:
        """Persist day-level Cheque on saved rows without re-verify flags."""
        target_date = self._current_date
        cheque = self._header_cheque
        snapshot = [
            (row, self._saved_ids.get(row), self._saved_txs.get(row))
            for row in range(self._saved_count)
        ]
        for _row, tx_id, tx in snapshot:
            if self._current_date != target_date:
                return
            if tx_id is None:
                continue
            cur_cheque = (getattr(tx, "cheque", "") or "") if tx is not None else ""
            if cur_cheque == cheque:
                continue
            try:
                await update_transaction(tx_id, {"cheque": cheque})
            except Exception:
                continue
            if tx is not None:
                tx.cheque = cheque

    def _mark_dirty(self, row: int) -> None:
        """Flag a saved row as modified and give it a stronger amber background."""
        if row in self._dirty_rows:
            return
        self._dirty_rows.add(row)
        self._table.blockSignals(True)
        for col in range(self._table.columnCount()):
            it = self._table.item(row, col)
            if it is not None:
                it.setBackground(QBrush(DIRTY_BG))
        self._table.blockSignals(False)
        self.edit_state_changed.emit(True, len(self._dirty_rows))
        self._schedule_draft_autosave()
        self._update_truck_required_highlight(row)

    def _updates_from_row(self, row: int) -> Optional[dict]:
        """Build the $set payload for an edited saved row from its cell values.
        Returns None when the row has no description. Raises ValueError on
        validation errors (bad item / locked description / unregistered truck)."""
        tx = self._build_transaction_from_row(row)
        if tx is None:
            return None
        return {
            "date": tx.date,
            "description": tx.description,
            "item": tx.item,
            "category_name": tx.category_name,
            "truck_number": tx.truck_number,
            "amount": tx.amount,
            "currency": tx.currency,
            "amount_usd": tx.amount_usd,
            "memo": tx.memo,
            "receipt_status": tx.receipt_status,
            "notes_flag": tx.notes_flag,
            "ref_float": tx.ref_float,
            "ownership": tx.ownership,
            "approver": tx.approver,
            "payee": tx.payee,
            "cheque": tx.cheque,
            "import_primary_date": tx.import_primary_date,
        }

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    _EXPORT_COLS = [
        COL_DATE, COL_ITEM, COL_DESC, COL_TRUCK, COL_MEMO,
        COL_REF, COL_TZS, COL_USD, COL_RECEIPT, COL_OWN, COL_APR, COL_PAYEE, COL_CHEQUE,
    ]

    def export_as(self, fmt: str = "xlsx") -> None:
        """Export visible rows as Excel, CSV, or PDF.

        ``fmt`` is one of ``xlsx``, ``csv``, or ``pdf``. Format is chosen from
        the Export menu so the save dialog only asks for a filename/location.
        """
        fmt = (fmt or "xlsx").lower().strip()
        if fmt not in ("xlsx", "csv", "pdf"):
            fmt = "xlsx"

        filters = {
            "xlsx": ("Excel Workbook (*.xlsx)", ".xlsx"),
            "csv": ("CSV File (*.csv)", ".csv"),
            "pdf": ("PDF Report (*.pdf)", ".pdf"),
        }
        file_filter, ext = filters[fmt]
        date_slug = (
            self._current_date.isoformat()
            if isinstance(self._current_date, date)
            else "new"
        )
        default_name = f"register_{date_slug}{ext}"
        path, _ = QFileDialog.getSaveFileName(
            self, f"Export register as {ext.lstrip('.').upper()}",
            default_name, file_filter,
        )
        if not path:
            return
        if not path.lower().endswith(ext):
            path = f"{path}{ext}"

        rows = self._visible_export_rows()
        if not rows:
            QMessageBox.information(self, "Nothing to export",
                                    "There are no visible rows to export.")
            return
        rows = self._apply_export_restriction(rows, fmt)
        if not rows:
            QMessageBox.information(self, "Nothing to export",
                                    "All visible rows belong to items restricted "
                                    f"from {fmt.upper()} export.")
            return
        try:
            if fmt == "pdf":
                self._write_pdf(path, rows)
            elif fmt == "csv":
                self._write_csv(path, rows)
            else:
                self._write_xlsx(path, rows)
        except Exception as exc:
            QMessageBox.critical(self, "Export failed", str(exc))
            return
        QMessageBox.information(self, "Export complete",
                               f"{len(rows)} row(s) exported to:\n{path}")

    def export_xlsx(self) -> None:
        """Backward-compatible alias — opens Export as Excel."""
        self.export_as("xlsx")

    def export_csv(self) -> None:
        """Backward-compatible alias — opens Export as CSV."""
        self.export_as("csv")

    def _visible_export_rows(self) -> List[list]:
        out: List[list] = []
        for row in range(self._table.rowCount()):
            if self._table.isRowHidden(row):
                continue
            if not (row < self._saved_count or self._row_has_data(row)):
                continue
            rec: list = []
            for col in self._EXPORT_COLS:
                it = self._table.item(row, col)
                if col == COL_RECEIPT:
                    val = (it.text().strip() if it else "")
                    rec.append(_RCPT_LABEL.get(val.lower(), val))
                else:
                    rec.append(it.text().strip() if it else "")
            out.append(rec)
        return out

    def _apply_export_restriction(self, rows: List[list], fmt: str) -> List[list]:
        from tahmeed.services.export_restriction_service import (
            filter_register_rows,
            restricted_names_from_categories,
        )

        if "cashier_register" not in getattr(self, "_export_restrict_surfaces", set()):
            return rows
        export_fmt = "pdf" if fmt == "pdf" else "excel"
        restricted = restricted_names_from_categories(self._categories, export_fmt)
        if not restricted:
            return rows
        item_idx = self._EXPORT_COLS.index(COL_ITEM)
        return filter_register_rows(rows, item_col_index=item_idx, restricted=restricted)

    def _write_csv(self, path: str, rows: List[list]) -> None:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow([HEADERS[c] for c in self._EXPORT_COLS])
            w.writerows(rows)

    def _write_xlsx(self, path: str, rows: List[list]) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Register"
        ws.append([HEADERS[c] for c in self._EXPORT_COLS])
        for rec in rows:
            ws.append(rec)
        wb.save(path)

    def _write_pdf(self, path: str, rows: List[list]) -> None:
        from tahmeed.services.daily_register_pdf import export_daily_register_pdf

        export_daily_register_pdf(
            path,
            rows=rows,
            register_date=self._current_date,
        )

    # ------------------------------------------------------------------
    # Search & column filtering
    # ------------------------------------------------------------------

    def set_search(self, text: str) -> None:
        self._search_text = text.strip().lower()
        self._apply_filters()

    def scroll_and_highlight(self, term: str) -> None:
        """Scroll to the first saved row that contains term and flash-highlight it.

        Scans all visible columns for a case-insensitive substring match.
        Applies a 2-second amber highlight then restores the original row background.
        Does NOT change the active search filter.
        """
        if not term:
            return
        needle = term.strip().lower()

        first_match = -1
        for row in range(self._saved_count):
            if self._table.isRowHidden(row):
                continue
            for col in range(self._table.columnCount()):
                it = self._table.item(row, col)
                if it and needle in it.text().lower():
                    first_match = row
                    break
            if first_match >= 0:
                break

        if first_match < 0:
            return

        self._table.scrollTo(self._table.model().index(first_match, COL_DESC))
        self._table.setCurrentCell(first_match, COL_DESC)

        highlight = QBrush(QColor("#FDE68A"))   # amber-200
        saved_bgs: dict = {}
        for col in range(self._table.columnCount()):
            it = self._table.item(first_match, col)
            if it:
                saved_bgs[col] = QBrush(it.background())
                it.setBackground(highlight)

        def _restore() -> None:
            for col, bg in saved_bgs.items():
                it = self._table.item(first_match, col)
                if it:
                    it.setBackground(bg)

        QTimer.singleShot(2000, _restore)

    def _on_col_filter_changed(self, col: int, accepted: set) -> None:
        if accepted:
            self._col_filters[col] = set(accepted)
        else:
            self._col_filters.pop(col, None)
        self._prune_stale_column_filters(changed_col=col)
        self._sync_filter_header()
        self._apply_filters()

    def _clear_column_filters(self) -> None:
        self._col_filters.clear()
        hdr = self._table.horizontalHeader()
        if isinstance(hdr, _FilterHeaderView):
            hdr.clear_filters()

    def _sync_filter_header(self) -> None:
        hdr = self._table.horizontalHeader()
        if isinstance(hdr, _FilterHeaderView):
            hdr.sync_active(self._col_filters)

    def _prune_stale_column_filters(self, *, changed_col: int) -> None:
        """Drop selections on other columns that no longer appear after chaining."""
        if not self._col_filters:
            return
        # Iterate until stable — narrowing one column can invalidate another.
        for _ in range(len(self._col_filters) + 1):
            changed = False
            for col in list(self._col_filters.keys()):
                if col == changed_col:
                    continue
                available = self._filter_menu_values(col)
                kept = {v for v in self._col_filters.get(col, set()) if v in available}
                if not kept:
                    self._col_filters.pop(col, None)
                    changed = True
                elif kept != self._col_filters[col]:
                    self._col_filters[col] = kept
                    changed = True
            if not changed:
                break

    def _cell_filter_value(self, row: int, col: int) -> str:
        it = self._table.item(row, col)
        if col == COL_RECEIPT:
            raw = it.text().strip().lower() if it else ""
            return _RCPT_LABEL.get(raw, raw)
        return it.text().strip() if it else ""

    def _iter_filter_source_indices(self):
        for row in range(self._table.rowCount()):
            if row < self._saved_count or self._row_has_data(row):
                yield row

    def _row_matches_other_filters(self, row: int, *, exclude_col: int) -> bool:
        """True if *row* matches search + every active column filter except *exclude_col*."""
        search = self._search_text
        if search:
            matched = False
            for c in range(self._table.columnCount()):
                it = self._table.item(row, c)
                if not it:
                    continue
                if c == COL_RECEIPT:
                    label = _RCPT_LABEL.get(it.text().strip().lower(), it.text())
                    if search in label.lower() or search in it.text().lower():
                        matched = True
                        break
                elif search in it.text().lower():
                    matched = True
                    break
            if not matched:
                return False

        for c, accepted in self._col_filters.items():
            if c == exclude_col or not accepted:
                continue
            if self._cell_filter_value(row, c) not in accepted:
                return False
        return True

    def _filter_menu_values(self, col: int) -> set:
        """Distinct values present in the table for *col*, chained through other filters."""
        rows: List[dict] = []
        for row in self._iter_filter_source_indices():
            if not self._row_matches_other_filters(row, exclude_col=col):
                continue
            m: dict = {}
            for c in range(self._table.columnCount()):
                if c == COL_SNO:
                    continue
                v = self._cell_filter_value(row, c)
                if v:
                    m[c] = v
            if m:
                rows.append(m)
        # active_filters already applied via _row_matches_other_filters; pass empty
        # here so cascade_column_values just collects target_col values.
        return cascade_column_values(rows, target_col=col, active_filters={})

    def _apply_filters(self) -> None:
        search = self._search_text
        for row in range(self._table.rowCount()):
            # Always show editable rows so new entry is never hidden
            if row >= self._saved_count:
                self._table.setRowHidden(row, False)
                continue

            # ── Search ───────────────────────────────────────────────
            if search:
                matched = False
                for col in range(self._table.columnCount()):
                    it = self._table.item(row, col)
                    if not it:
                        continue
                    if col == COL_RECEIPT:
                        label = _RCPT_LABEL.get(it.text().strip().lower(), it.text())
                        if search in label.lower() or search in it.text().lower():
                            matched = True
                            break
                    elif search in it.text().lower():
                        matched = True
                        break
                if not matched:
                    self._table.setRowHidden(row, True)
                    continue

            # ── Column filters ───────────────────────────────────────
            visible = True
            for col, accepted in self._col_filters.items():
                if not accepted:
                    continue
                if self._cell_filter_value(row, col) not in accepted:
                    visible = False
                    break

            self._table.setRowHidden(row, not visible)

    def _go_to_first_empty(self) -> None:
        """Scroll to and focus the first empty editable row (New button)."""
        row = self._first_empty_editable_row()
        if row >= self._table.rowCount():
            self._append_editable_rows(10)
        self._table.setCurrentCell(row, COL_DESC)
        self._table.scrollTo(self._table.model().index(row, COL_DESC))
        self._table.setFocus()

    # ------------------------------------------------------------------
    # Row numbering
    # ------------------------------------------------------------------

    def _sync_row_header_labels(self) -> None:
        sync_row_header_labels(self._table)

    def _renumber(self) -> None:
        self._table.blockSignals(True)
        for row in range(self._table.rowCount()):
            it = self._table.item(row, COL_SNO)
            if not it:
                continue
            is_saved = row < self._saved_count
            is_active = it.text() != "" or self._row_has_data(row)
            if is_saved or is_active:
                it.setText(str(row + 1))
                if not is_saved:
                    if not self._table.item(row, COL_RECEIPT):
                        ri = QTableWidgetItem("")
                        ri.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                        self._table.setItem(row, COL_RECEIPT, ri)
        self._table.blockSignals(False)
        self._sync_row_header_labels()
        # Keep header KPIs (entries / refund / total) in sync after row ops
        # that block itemChanged (delete, clear, paste, import).
        self._update_footer()

    # ------------------------------------------------------------------
    # Dynamic row expansion
    # ------------------------------------------------------------------

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        row = item.row()
        col = item.column()

        # Saved rows: only mutable while in edit mode. Track them as dirty and
        # uppercase free-text, but skip the new-row activation / expansion logic.
        if row < self._saved_count:
            if not self._edit_mode:
                return
            if col not in _UPPER_SKIP_COLS:
                text = item.text()
                if text and text != text.upper():
                    self._table.blockSignals(True)
                    item.setText(text.upper())
                    self._table.blockSignals(False)
            if col == COL_DESC and item.text().strip():
                from tahmeed.services.cashier_service import remember_description
                remember_description(item.text())
                if not self._bulk_mutating:
                    desc = item.text().strip()
                    QTimer.singleShot(
                        0, lambda r=row, d=desc: self._kick_auto_fill_item(r, d)
                    )
            self._mark_dirty(row)
            if col == COL_ITEM:
                self._validate_item_cell(row, item)
            elif col == COL_DESC and item.text().strip():
                self._validate_locked_description(row, item)
            elif col == COL_TRUCK and item.text().strip():
                self._validate_truck_cell(row, item)
            if col in (COL_ITEM, COL_DESC, COL_TRUCK):
                self._update_truck_required_highlight(row)
            if col == COL_TZS or col == COL_USD:
                self._update_footer()
            return

        # Uppercase all free-text cells
        if col not in _UPPER_SKIP_COLS:
            text = item.text()
            if text and text != text.upper():
                self._table.blockSignals(True)
                item.setText(text.upper())
                self._table.blockSignals(False)

        # Activate the row (show S/NO + checkboxes) on first data entry
        if col not in READONLY_COLS and col not in CHECK_COLS and item.text().strip():
            self._activate_row(row)

        # Keep Date in sync with whether the row has real entry data.
        # Skip when the Date cell itself is edited so a manual date is not
        # immediately cleared on an otherwise empty row.
        if col not in READONLY_COLS and col not in CHECK_COLS and col not in (COL_DATE,):
            self._table.blockSignals(True)
            self._sync_row_date(row)
            self._table.blockSignals(False)

        # Item / Description / Truck validation (canonicalise, restrict, locked lists)
        if col == COL_ITEM:
            self._validate_item_cell(row, item)
        elif col == COL_DESC and item.text().strip():
            from tahmeed.services.cashier_service import remember_description
            remember_description(item.text())
            self._validate_locked_description(row, item)
            if not self._bulk_mutating:
                # Defer off the current asyncio task so qasync/Py3.14 does not
                # try to nest _auto_fill inside an active import coroutine.
                desc = item.text().strip()
                QTimer.singleShot(
                    0, lambda r=row, d=desc: self._kick_auto_fill_item(r, d)
                )
        elif col == COL_TRUCK and item.text().strip():
            self._validate_truck_cell(row, item)
        if col in (COL_ITEM, COL_DESC, COL_TRUCK):
            self._update_truck_required_highlight(row)

        # Dynamic row expansion near the bottom
        if row >= self._table.rowCount() - 5 and item.text().strip():
            self._append_editable_rows(10)

        if col in (COL_TZS, COL_USD, COL_REF):
            self._update_footer()

        self._schedule_draft_autosave()

    def _on_model_data_changed(self, top_left, bottom_right, roles=()) -> None:
        # Kept for receipt/other UserRole updates if added later.
        pass

    def _activate_row(self, row: int) -> None:
        """Make a blank editable row visible: set S/NO number and create input items."""
        if row < self._saved_count:
            return
        sno_it = self._table.item(row, COL_SNO)
        if sno_it and sno_it.text():
            return  # already active
        if sno_it:
            sno_it.setText(str(row + 1))
        prev = self._table.blockSignals(True)
        if not self._table.item(row, COL_RECEIPT):
            ri = QTableWidgetItem("")
            ri.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self._table.setItem(row, COL_RECEIPT, ri)
        # Inherit the day-level Cheque stamp onto newly started rows.
        if self._header_cheque:
            self._write_cheque_cell(row, self._header_cheque)
        self._table.blockSignals(prev)

    def _deactivate_row(self, row: int) -> None:
        """Clear S/NO on an emptied editable row so it looks blank again."""
        if row < self._saved_count:
            return
        sno_it = self._table.item(row, COL_SNO)
        if sno_it and sno_it.text():
            sno_it.setText("")

    def _register_date_str(self) -> str:
        return format_register_date(self._current_date)

    def _register_primary_dt(self) -> datetime:
        """Register calendar day as datetime (same shape as daily-import primary)."""
        d = self._current_date or date.today()
        return datetime(d.year, d.month, d.day)

    def _stamp_empty_row_dates(self) -> None:
        """Fill blank Date cells on typed rows after Reconciled Date is chosen."""
        if not isinstance(self._current_date, date):
            return
        stamp = self._register_date_str()
        if not stamp:
            return
        prev = self._table.blockSignals(True)
        try:
            for row in range(self._saved_count, self._table.rowCount()):
                if not self._row_has_data(row):
                    continue
                date_it = self._table.item(row, COL_DATE)
                if date_it is not None and date_it.text().strip():
                    continue
                new_it = QTableWidgetItem(stamp)
                new_it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                self._table.setItem(row, COL_DATE, new_it)
        finally:
            self._table.blockSignals(prev)

    def _resolve_import_primary_date(self, row: int):
        """Register day ownership for this row (upload meta, existing stamp, or open day).

        Prior Excel/system dates stay on the open register day — same as uploads.
        """
        meta = self._pending_row_meta.get(row) or {}
        stamped = meta.get("import_primary_date")
        if stamped is not None:
            return stamped
        orig = self._saved_txs.get(row)
        if orig is not None and getattr(orig, "import_primary_date", None) is not None:
            return orig.import_primary_date
        return self._register_primary_dt()

    def _sync_row_date(self, row: int) -> None:
        """Fill Date when the row gains entry data; clear it when the row is emptied.

        Caller should block itemChanged signals when batching writes. Does not
        overwrite a date that is already set.
        """
        if row < self._saved_count:
            return
        has_data = self._row_has_data(row)
        date_it = self._table.item(row, COL_DATE)
        date_text = date_it.text().strip() if date_it else ""

        if has_data:
            if not date_text:
                new_it = QTableWidgetItem(self._register_date_str())
                new_it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                self._table.setItem(row, COL_DATE, new_it)
            self._activate_row(row)
        else:
            if date_text:
                if date_it is not None:
                    date_it.setText("")
                else:
                    self._table.setItem(row, COL_DATE, QTableWidgetItem(""))
            self._deactivate_row(row)

    def _append_editable_rows(self, n: int = 10) -> None:
        start = self._table.rowCount()
        self._table.setRowCount(start + n)
        self._init_editable_rows(start, start + n)
        self._renumber()

    # ------------------------------------------------------------------
    # Keyboard navigation
    # ------------------------------------------------------------------

    def keyPressEvent(self, event: QKeyEvent) -> None:
        super().keyPressEvent(event)

    def _table_key_press(self, event: QKeyEvent) -> None:
        mod = event.modifiers()
        key = event.key()

        if mod == Qt.ControlModifier:
            if key == Qt.Key_C:    self._copy();                               return
            if key == Qt.Key_X:    self._cut();                                return
            if key == Qt.Key_V:    self._paste();                              return
            if key == Qt.Key_Z:
                if mod & Qt.ShiftModifier:
                    self._redo()
                else:
                    self._undo()
                return
            if key in (Qt.Key_Y,):
                self._redo()
                return
            if key == Qt.Key_A:    self._table.selectAll();                    return
            if key == Qt.Key_D:    self._fill_down();                          return
            if key == Qt.Key_R:    self._fill_right();                         return
            if key == Qt.Key_Home: self._table.setCurrentCell(0, 0);          return
            if key == Qt.Key_End:  self._go_to_last_cell();                   return

        if key == Qt.Key_Escape:
            if self._cut_cells:
                self._clear_cut_marquee()
                return

        if mod == Qt.ShiftModifier:
            if key in (Qt.Key_Return, Qt.Key_Enter): self._step(-1, 0);      return
            if key == Qt.Key_Space:                  self._select_row();      return

        if key == Qt.Key_F2:
            it = self._table.currentItem()
            if it:
                self._table.editItem(it)
            return

        if key in (Qt.Key_Delete, Qt.Key_Backspace):
            sel_rows = sorted({i.row() for i in self._table.selectedIndexes()})
            if sel_rows and self._selection_is_full_rows(sel_rows):
                saved = [r for r in sel_rows if r < self._saved_count]
                if saved:
                    self._delete_saved_rows(saved)
                # Unsaved full rows still use the existing row-remove path.
                if any(r >= self._saved_count for r in sel_rows):
                    self._delete_rows()
                return
            self._clear_selected()
            return

        if key == Qt.Key_Tab:
            self._commit_date_suggestion()
            self._tab_forward(); return

        if key == Qt.Key_Backtab:
            self._step(0, -1, skip=CHECK_COLS | READONLY_COLS); return

        if key in (Qt.Key_Return, Qt.Key_Enter):
            self._commit_date_suggestion()
            self._step(+1, 0); return

        QTableWidget.keyPressEvent(self._table, event)

    def _tab_forward(self) -> None:
        """Advance Tab: skip readonly cols, wrap to next row at last column."""
        row, col = self._table.currentRow(), self._table.currentColumn()
        skip = READONLY_COLS
        next_col = col + 1
        while next_col < self._table.columnCount() and next_col in skip:
            next_col += 1
        if next_col >= self._table.columnCount():
            next_row = row + 1
            if next_row >= self._table.rowCount():
                self._append_editable_rows(10)
            first_col = 0
            while first_col < self._table.columnCount() and first_col in skip:
                first_col += 1
            self._activate_row(next_row)
            self._table.setCurrentCell(next_row, first_col)
        else:
            self._table.setCurrentCell(row, next_col)
        self._table.setFocus()
        # Blank editable rows have no QTableWidgetItem; Qt returns ItemIsDropEnabled
        # only when there is no item, so edit() silently fails.  Create a placeholder
        # so the cell is treated as editable before we attempt to open the editor.
        idx = self._table.currentIndex()
        trow, tcol = idx.row(), idx.column()
        if trow >= self._saved_count and not self._table.item(trow, tcol):
            self._table.blockSignals(True)
            it = QTableWidgetItem("")
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self._table.setItem(trow, tcol, it)
            self._table.blockSignals(False)
        self._table.edit(idx)

    def _install_key_handler(self) -> None:
        self._key_filter = _TableKeyFilter(self._table_key_press)
        self._table.installEventFilter(self._key_filter)
        vh = self._table.verticalHeader()
        if vh is not None:
            vh.installEventFilter(self._key_filter)

    def _commit_date_suggestion(self) -> None:
        """If the focused cell is an empty Date cell, write the register date."""
        row, col = self._table.currentRow(), self._table.currentColumn()
        if col != COL_DATE or row < self._saved_count:
            return
        it = self._table.item(row, col)
        if it is not None and it.text().strip():
            return
        cur = self._current_date
        today_str = format_register_date(cur)
        new_it = QTableWidgetItem(today_str)
        new_it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
        self._table.blockSignals(True)
        self._table.setItem(row, col, new_it)
        self._table.blockSignals(False)
        self._activate_row(row)

    def _step(self, dr: int, dc: int, skip: set = None) -> None:
        row, col = self._table.currentRow(), self._table.currentColumn()
        new_col, new_row = col + dc, row + dr
        if skip and dc != 0:
            while 0 <= new_col < self._table.columnCount():
                if new_col not in skip:
                    break
                new_col += dc
        new_row = max(0, min(new_row, self._table.rowCount() - 1))
        new_col = max(0, min(new_col, self._table.columnCount() - 1))
        self._table.setCurrentCell(new_row, new_col)

    # ------------------------------------------------------------------
    # Clipboard
    # ------------------------------------------------------------------

    # ------------------------------------------------------------------
    # Undo / redo (Excel-style)
    # ------------------------------------------------------------------

    def _normalize_cell_val(self, val) -> tuple:
        if isinstance(val, tuple) and len(val) >= 2:
            return val
        return ("text", "" if val is None else str(val))

    def _serialize_cell(self, row: int, col: int) -> tuple:
        if row < 0 or row >= self._table.rowCount() or col == COL_SNO:
            return ("text", "")
        it = self._table.item(row, col)
        if col in CHECK_COLS:
            checked = bool(it.data(Qt.UserRole)) if it else False
            return ("check", checked)
        return ("text", it.text() if it else "")

    def _cell_editable(self, row: int, col: int) -> bool:
        if col in READONLY_COLS or col == COL_SNO:
            return False
        if row < self._saved_count and not self._edit_mode:
            return False
        return True

    def _write_cell_value(self, row: int, col: int, val: tuple) -> None:
        kind, payload = self._normalize_cell_val(val)
        if col in CHECK_COLS or kind == "check":
            it = self._table.item(row, col) or QTableWidgetItem()
            it.setData(Qt.UserRole, bool(payload))
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self._table.setItem(row, col, it)
            return
        text = str(payload)
        if col == COL_RECEIPT:
            it = QTableWidgetItem(_receipt_paste_value(text))
        elif col in (COL_TZS, COL_USD):
            it = _amount_item_from_raw(text)
        elif col == COL_CASHIER:
            it = QTableWidgetItem(text)
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        else:
            it = QTableWidgetItem(_upper_text(col, text))
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
        self._table.setItem(row, col, it)

    def _normalize_undo_entry(self, entry: dict) -> dict:
        if "cells" in entry or "row_ops" in entry:
            return entry
        cells = {
            (int(r), int(c)): self._normalize_cell_val(v)
            for (r, c), v in entry.items()
        }
        return {"cells": cells}

    def _record_undo(self, entry: dict) -> None:
        normalized = self._normalize_undo_entry(entry)
        if not normalized.get("cells") and not normalized.get("row_ops"):
            return
        self._undo_stack.append(normalized)
        self._redo_stack.clear()
        if len(self._undo_stack) > self._undo_limit:
            self._undo_stack = self._undo_stack[-self._undo_limit:]
        self._emit_undo_redo_state()

    def _push_undo_cells(self, cells: dict) -> None:
        """Snapshot cell values before a mutating edit."""
        if not cells:
            return
        mapped = {}
        for (row, col), val in cells.items():
            if isinstance(val, tuple):
                mapped[(row, col)] = val
            else:
                mapped[(row, col)] = self._normalize_cell_val(val)
        self._record_undo({"cells": mapped})

    def _snapshot_selection(self) -> dict:
        snap = {}
        for it in self._table.selectedItems():
            if it.column() in READONLY_COLS:
                continue
            snap[(it.row(), it.column())] = self._serialize_cell(it.row(), it.column())
        return snap

    def _snapshot_rows(self, rows: list) -> dict:
        snap = {}
        for row in rows:
            snap.update(self._snapshot_row_cells(row))
        return snap

    def _snapshot_row_cells(self, row: int) -> dict:
        """Full row snapshot including S/N for undo of row activation."""
        snap = {}
        for col in range(self._table.columnCount()):
            if col == COL_SNO:
                sno_it = self._table.item(row, COL_SNO)
                snap[(row, col)] = ("text", sno_it.text() if sno_it else "")
            else:
                snap[(row, col)] = self._serialize_cell(row, col)
        return snap

    def _snapshot_cells_for_keys(self, keys) -> dict:
        return {(r, c): self._serialize_cell(r, c) for (r, c) in keys}

    def _apply_cells(self, cells: dict) -> None:
        if not cells:
            return
        self._undo_redo_active = True
        self._table.blockSignals(True)
        try:
            touched_rows: set = set()
            for (row, col), val in cells.items():
                if row < 0 or row >= self._table.rowCount():
                    continue
                if col == COL_SNO:
                    if row >= self._saved_count:
                        sno_it = self._table.item(row, COL_SNO)
                        if sno_it is not None:
                            kind, payload = self._normalize_cell_val(val)
                            sno_it.setText(str(payload))
                            touched_rows.add(row)
                    continue
                if not self._cell_editable(row, col) and row < self._saved_count:
                    continue
                self._write_cell_value(row, col, val)
                touched_rows.add(row)
                if row < self._saved_count and self._edit_mode:
                    self._dirty_rows.add(row)
            for row in touched_rows:
                self._sync_row_date(row)
        finally:
            self._table.blockSignals(False)
            self._undo_redo_active = False
        self._renumber()
        self._update_footer()
        self._table.viewport().update()
        self._schedule_draft_autosave()

    def _apply_row_ops(self, ops: list, *, forward: bool) -> None:
        """Apply row insert/remove ops. ``forward=True`` runs as recorded; False inverts."""
        if not ops:
            return
        self._undo_redo_active = True
        self._bulk_mutating = True
        self._table.blockSignals(True)
        try:
            removes = [o for o in ops if o.get("op") == "remove"]
            inserts = [o for o in ops if o.get("op") == "insert"]
            if forward:
                for op in sorted(removes, key=lambda o: int(o.get("at", 0)), reverse=True):
                    at = int(op.get("at", 0))
                    if at < self._table.rowCount():
                        self._shift_row_maps_on_remove(at)
                        self._table.removeRow(at)
                for op in inserts:
                    at = int(op.get("at", 0))
                    count = int(op.get("count") or 1)
                    for _ in range(count):
                        self._shift_row_maps_on_insert(at)
                        self._table.insertRow(at)
                        self._init_editable_rows(at, at + 1)
            else:
                for op in sorted(removes, key=lambda o: int(o.get("at", 0))):
                    at = int(op.get("at", 0))
                    values = dict(op.get("values") or {})
                    meta = dict(op.get("meta") or {})
                    self._shift_row_maps_on_insert(at)
                    self._table.insertRow(at)
                    self._init_editable_rows(at, at + 1)
                    truck_cells = self._write_row_values(at, values)
                    self._restore_moved_row_meta(at, meta)
                    self._sync_row_date(at)
                    self._finalize_truck_cells(truck_cells)
                for op in reversed(inserts):
                    at = int(op.get("at", 0))
                    count = int(op.get("count") or 1)
                    for _ in range(count):
                        if at < self._table.rowCount():
                            self._shift_row_maps_on_remove(at)
                            self._table.removeRow(at)
            min_rows = self._saved_count + DEFAULT_EDITABLE_ROWS
            if self._table.rowCount() < min_rows:
                start = self._table.rowCount()
                self._table.setRowCount(min_rows)
                self._init_editable_rows(start, min_rows)
        finally:
            self._table.blockSignals(False)
            self._bulk_mutating = False
            self._undo_redo_active = False
        self._renumber()
        self._update_footer()
        self._schedule_draft_autosave()

    def _invert_undo_entry(self, entry: dict) -> dict:
        entry = self._normalize_undo_entry(entry)
        inverted: dict = {}
        cells = entry.get("cells") or {}
        if cells:
            inverted["cells"] = self._snapshot_cells_for_keys(cells.keys())
        ops = entry.get("row_ops") or []
        if ops:
            inv_ops = []
            for op in reversed(ops):
                action = op.get("op")
                if action == "insert":
                    inv_ops.append({
                        "op": "insert",
                        "at": int(op.get("at", 0)),
                        "count": int(op.get("count") or 1),
                    })
                elif action == "remove":
                    inv_ops.append({"op": "remove", "at": int(op.get("at", 0))})
            inverted["row_ops"] = inv_ops
        return inverted

    def _apply_undo_entry(self, entry: dict, *, forward: bool) -> None:
        entry = self._normalize_undo_entry(entry)
        if not forward:
            ops = entry.get("row_ops") or []
            if ops:
                self._apply_row_ops(ops, forward=False)
            cells = entry.get("cells") or {}
            if cells:
                self._apply_cells(cells)
            return
        cells = entry.get("cells") or {}
        if cells:
            self._apply_cells(cells)
        ops = entry.get("row_ops") or []
        if ops:
            self._apply_row_ops(ops, forward=True)

    def _undo(self) -> None:
        if not self._undo_stack:
            return
        self._flush_pending_row_edits()
        self._clear_cut_marquee()
        entry = self._undo_stack.pop()
        redo_entry = self._invert_undo_entry(entry)
        self._apply_undo_entry(entry, forward=False)
        self._redo_stack.append(redo_entry)
        if len(self._redo_stack) > self._undo_limit:
            self._redo_stack = self._redo_stack[-self._undo_limit:]
        self._emit_undo_redo_state()

    def _redo(self) -> None:
        if not self._redo_stack:
            return
        self._flush_pending_row_edits()
        self._clear_cut_marquee()
        entry = self._redo_stack.pop()
        undo_entry = self._invert_undo_entry(entry)
        self._apply_undo_entry(entry, forward=True)
        self._undo_stack.append(undo_entry)
        if len(self._undo_stack) > self._undo_limit:
            self._undo_stack = self._undo_stack[-self._undo_limit:]
        self._emit_undo_redo_state()

    def _on_current_cell_changed(
        self, row: int, col: int, prev_row: int, prev_col: int,
    ) -> None:
        if prev_row >= 0 and prev_col >= 0:
            self._begin_closing_row_edit()
        if row >= 0 and col >= 0 and self._cell_editable(row, col):
            self._pending_row_edit = {
                "row": row,
                "col": col,
                "before": self._snapshot_row_cells(row),
            }
        else:
            self._pending_row_edit = None

    def _begin_closing_row_edit(self) -> None:
        if self._pending_row_edit is None:
            return
        self._closing_row_edit = self._pending_row_edit
        self._pending_row_edit = None
        self._schedule_finalize_row_edit()

    def _schedule_finalize_row_edit(self) -> None:
        self._edit_finalize_timer.start()

    def _row_edit_cell_value(self, row: int, col: int) -> tuple:
        if col == COL_SNO:
            sno_it = self._table.item(row, COL_SNO)
            return ("text", sno_it.text() if sno_it else "")
        return self._serialize_cell(row, col)

    def _finalize_one_row_edit(self, pending: dict) -> None:
        if self._undo_redo_active or self._bulk_mutating:
            return
        row = pending["row"]
        before = pending["before"]
        changes = {}
        for (r, c), old_val in before.items():
            if r != row:
                continue
            new_val = self._row_edit_cell_value(r, c)
            if new_val != old_val:
                changes[(r, c)] = old_val
        if changes:
            self._record_undo({"cells": changes})

    def _finalize_pending_row_edit(self) -> None:
        pending = self._closing_row_edit
        if pending is None:
            return
        row = pending["row"]
        if self._auto_fill_row == row:
            self._edit_finalize_timer.start(100)
            return
        self._finalize_one_row_edit(pending)
        self._closing_row_edit = None

    def _flush_pending_row_edits(self) -> None:
        """Finalize any open edit, waiting briefly for async Item auto-fill."""
        self._edit_finalize_timer.stop()
        if self._pending_row_edit is not None:
            self._closing_row_edit = self._pending_row_edit
            self._pending_row_edit = None
        deadline = time.monotonic() + 2.0
        while self._closing_row_edit is not None:
            row = self._closing_row_edit["row"]
            if self._auto_fill_row == row:
                if time.monotonic() > deadline:
                    break
                QApplication.processEvents(QEventLoop.ProcessEventsFlag.AllEvents, 50)
                continue
            self._finalize_one_row_edit(self._closing_row_edit)
            self._closing_row_edit = None
            break

    def _record_row_remove_undo(self, rows: list) -> None:
        ops = []
        for row in sorted(rows):
            ops.append({
                "op": "remove",
                "at": row,
                "values": self._row_value_map(row),
                "meta": self._capture_row_meta(row),
            })
        self._record_undo({"row_ops": ops})

    def _record_row_insert_undo(self, at: int, count: int = 1) -> None:
        self._record_undo({"row_ops": [{"op": "insert", "at": at, "count": count}]})

    def _emit_undo_redo_state(self) -> None:
        self.undo_redo_changed.emit(bool(self._undo_stack), bool(self._redo_stack))

    def toolbar_undo(self) -> None:
        self._commit_open_editor()
        self._undo()

    def toolbar_redo(self) -> None:
        self._commit_open_editor()
        self._redo()

    def _clear_cut_marquee(self) -> None:
        self._cut_cells = set()
        self._cut_payload = {}
        self._cut_is_rows = False
        self._table.viewport().update()

    def _has_cut_buffer(self) -> bool:
        return bool(self._cut_cells and self._cut_payload)

    def _copy(self) -> None:
        # Use selectedIndexes so blank cells stay in the rectangle (Excel-aligned TSV).
        indexes = self._table.selectedIndexes()
        if not indexes:
            return
        rows = sorted({i.row() for i in indexes})
        if self._selection_is_full_rows(rows):
            cols = sorted(
                c for c in range(self._table.columnCount()) if c not in READONLY_COLS
            )
        else:
            cols = sorted({i.column() for i in indexes})
        lines = []
        for row in rows:
            row_cells = []
            for col in cols:
                it = self._table.item(row, col)
                if it is None:
                    row_cells.append("")
                elif col in CHECK_COLS:
                    row_cells.append("1" if it.data(Qt.UserRole) else "0")
                else:
                    row_cells.append(it.text())
            lines.append("\t".join(row_cells))
        QApplication.clipboard().setText("\n".join(lines))

    def _cut(self) -> None:
        """Excel-style cut: copy + dashed marquee; content stays until paste/insert."""
        sel_rows = sorted({i.row() for i in self._table.selectedIndexes()})
        if sel_rows and self._selection_is_full_rows(sel_rows):
            self._cut_rows(sel_rows)
            return

        items = self._table.selectedItems()
        if not items:
            return
        editable = []
        for it in items:
            row, col = it.row(), it.column()
            if col in READONLY_COLS:
                continue
            if row < self._saved_count and not self._edit_mode:
                continue
            editable.append(it)
        if not editable:
            return

        self._push_undo_cells(self._snapshot_selection())
        self._copy()

        rows = sorted({it.row() for it in editable})
        cols = sorted({it.column() for it in editable})
        cell_map = {(it.row(), it.column()): it for it in editable}
        grid = []
        cut_cells = set()
        for row in rows:
            line = []
            for col in cols:
                it = cell_map.get((row, col))
                text = it.text() if it else ""
                line.append(text)
                if it is not None:
                    cut_cells.add((row, col))
            grid.append(line)

        self._cut_cells = cut_cells
        self._cut_is_rows = False
        self._cut_payload = {
            "kind": "cells",
            "rows": rows,
            "cols": cols,
            "grid": grid,
        }
        self._table.viewport().update()

    def _selection_is_full_rows(self, rows: list) -> bool:
        """True when the selection covers every data column for each row."""
        if not rows:
            return False
        indexes = self._table.selectedIndexes()
        if not indexes:
            return False
        sel_rows = {i.row() for i in indexes}
        if sel_rows != set(rows):
            return False
        sel_cols = {i.column() for i in indexes}
        ncols = self._table.columnCount()
        data_cols = {c for c in range(ncols) if c not in READONLY_COLS}
        if not data_cols:
            return False
        if data_cols.issubset(sel_cols):
            return True
        # Row gutter select: empty cells may be omitted from selectedIndexes().
        if COL_SNO in sel_cols and max(sel_cols) >= max(data_cols):
            span = set(range(min(sel_cols), max(sel_cols) + 1))
            return data_cols.issubset(span)
        return False

    def _serialize_row(self, row: int) -> list:
        cells = []
        for col in range(self._table.columnCount()):
            if col == COL_SNO:
                cells.append("")
                continue
            it = self._table.item(row, col)
            if it is None:
                cells.append("")
            elif col in CHECK_COLS:
                cells.append("1" if it.data(Qt.UserRole) else "0")
            else:
                cells.append(it.text())
        return cells

    def _row_value_map(self, row: int) -> dict:
        """Column index → exact cell text for cut/insert (preserves Receipt/Cashier)."""
        values = {}
        for col in range(self._table.columnCount()):
            if col == COL_SNO:
                continue
            it = self._table.item(row, col)
            if it is None:
                values[col] = ""
            elif col in CHECK_COLS:
                values[col] = "1" if it.data(Qt.UserRole) else "0"
            else:
                values[col] = it.text()
        return values

    def _capture_row_meta(self, row: int) -> dict:
        """Identity/metadata needed to move a row without losing cashier or tx id."""
        pending = self._pending_row_meta.get(row)
        return {
            "was_saved": row < self._saved_count,
            "saved_id": self._saved_ids.get(row),
            "saved_tx": self._saved_txs.get(row),
            "pending": dict(pending) if pending else None,
            "dirty": row in self._dirty_rows,
        }

    def _cut_rows(self, rows: list) -> None:
        """Mark whole rows as cut (marquee) — do not delete until paste/insert."""
        movable = []
        for row in rows:
            if row >= self._saved_count:
                movable.append(row)
            elif self._merged_mode and self._edit_mode:
                tx = self._saved_txs.get(row)
                if tx is not None and (getattr(tx, "register_status", "") or "") == "draft":
                    movable.append(row)
        if not movable:
            self._copy()
            return

        self._push_undo_cells(self._snapshot_rows(movable))
        maps = [self._row_value_map(r) for r in movable]
        metas = [self._capture_row_meta(r) for r in movable]
        lines = ["\t".join(self._serialize_row(r)) for r in movable]
        QApplication.clipboard().setText(_ROWS_CLIP_PREFIX + "\n".join(lines))

        cut_cells = set()
        for row in movable:
            for col in range(self._table.columnCount()):
                if col == COL_SNO:
                    continue
                cut_cells.add((row, col))

        self._cut_cells = cut_cells
        self._cut_is_rows = True
        self._cut_payload = {
            "kind": "rows",
            "rows": list(movable),
            "lines": lines,
            "maps": maps,
            "row_metas": metas,
        }
        self._table.viewport().update()

    def _write_row_values(self, row: int, values: dict) -> list:
        """Write a column→text map onto *row*. Returns truck cells to finalize."""
        truck_cells: list = []
        for col, cell in values.items():
            if col >= self._table.columnCount() or col == COL_SNO:
                continue
            if col == COL_CASHIER:
                it = QTableWidgetItem(str(cell) if cell is not None else "")
                it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                self._table.setItem(row, col, it)
                continue
            if col in CHECK_COLS:
                it = QTableWidgetItem()
                it.setData(Qt.UserRole, str(cell).strip() in ("1", "true", "True", "YES"))
                it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                self._table.setItem(row, col, it)
            elif col == COL_RECEIPT:
                it = QTableWidgetItem(_receipt_paste_value(str(cell)))
                it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                self._table.setItem(row, col, it)
            elif col == COL_TZS or col == COL_USD:
                self._table.setItem(row, col, _amount_item_from_raw(str(cell)))
            elif col == COL_TRUCK:
                raw = str(cell).strip()
                self._table.setItem(
                    row, col, QTableWidgetItem(raw.upper() if raw else "")
                )
                if raw:
                    truck_cells.append((row, raw))
            else:
                self._table.setItem(
                    row, col, QTableWidgetItem(_upper_text(col, str(cell).strip()))
                )
        return truck_cells

    def _style_moved_row(self, row: int, was_saved: bool) -> None:
        """Restore saved/edit styling after a cut→insert move."""
        if not was_saved:
            return
        if self._edit_mode:
            bg = QBrush(DIRTY_BG if row in self._dirty_rows else EDIT_BG)
            editable = Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable
            ro = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        else:
            bg = QBrush(SAVED_BG)
            editable = ro = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        for col in range(self._table.columnCount()):
            it = self._table.item(row, col)
            if it is None:
                continue
            it.setBackground(bg)
            it.setFlags(ro if col in READONLY_COLS else editable)

    def _restore_moved_row_meta(self, row: int, meta: dict) -> None:
        """Re-attach tx id / import meta / dirty flag onto an inserted cut row."""
        if not meta:
            return
        pending = meta.get("pending")
        if pending:
            self._pending_row_meta[row] = dict(pending)
        saved_id = meta.get("saved_id")
        saved_tx = meta.get("saved_tx")
        was_saved = bool(meta.get("was_saved") or saved_id is not None)
        if saved_id is not None:
            self._saved_ids[row] = saved_id
        if saved_tx is not None:
            self._saved_txs[row] = saved_tx
        if was_saved and self._edit_mode:
            self._dirty_rows.add(row)
        elif meta.get("dirty"):
            self._dirty_rows.add(row)
        self._style_moved_row(row, was_saved)

    def _insert_cut_cells(self) -> None:
        """Insert Cut Cells — move the cut buffer to the current position."""
        if not self._has_cut_buffer():
            return
        if self._cut_payload.get("kind") == "rows":
            maps = self._cut_payload.get("maps")
            if maps:
                self._paste_row_maps(maps, clear_cut_after=True)
            else:
                lines = list(self._cut_payload.get("lines") or [])
                self._paste_rows("\n".join(lines), clear_cut_after=True)
            return
        self._paste()

    def _paste_row_maps(self, maps: list, clear_cut_after: bool = True) -> None:
        """Insert rows from exact column→value maps (preserves Receipt etc.)."""
        if not maps:
            return
        metas: list = []
        source_rows: list = []
        if (
            clear_cut_after
            and self._has_cut_buffer()
            and self._cut_is_rows
            and self._cut_payload.get("kind") == "rows"
        ):
            metas = list(self._cut_payload.get("row_metas") or [])
            source_rows = list(self._cut_payload.get("rows") or [])

        cur = self._table.currentRow()
        if self._merged_mode and self._edit_mode:
            insert_at = max(cur, 0)
        else:
            insert_at = max(cur, self._saved_count)

        # Remove cut sources first so row maps stay consistent, then insert.
        if source_rows:
            removed_before = sum(1 for r in source_rows if r < insert_at)
            self._table.blockSignals(True)
            try:
                for row in sorted(source_rows, reverse=True):
                    if row < 0 or row >= self._table.rowCount():
                        continue
                    self._shift_row_maps_on_remove(row)
                    self._table.removeRow(row)
                    if row < self._saved_count:
                        self._saved_count -= 1
            finally:
                self._table.blockSignals(False)
            insert_at = max(0, insert_at - removed_before)
            self._clear_cut_marquee()
            clear_cut_after = False
            min_rows = self._saved_count + DEFAULT_EDITABLE_ROWS
            if self._table.rowCount() < min_rows:
                start = self._table.rowCount()
                self._table.setRowCount(min_rows)
                self._init_editable_rows(start, min_rows)

        self._record_row_insert_undo(insert_at, len(maps))
        truck_cells: list = []
        self._bulk_mutating = True
        prev = self._table.blockSignals(True)
        try:
            for i, values in enumerate(maps):
                meta = metas[i] if i < len(metas) else {}
                was_saved = bool(
                    meta.get("was_saved") or meta.get("saved_id") is not None
                )
                # Keep saved drafts in the saved prefix; new rows stay below it.
                if was_saved:
                    insert_at = min(insert_at, self._saved_count)
                else:
                    insert_at = max(insert_at, self._saved_count)

                self._shift_row_maps_on_insert(insert_at)
                self._table.insertRow(insert_at)
                self._init_editable_rows(insert_at, insert_at + 1)
                truck_cells.extend(self._write_row_values(insert_at, values))
                self._restore_moved_row_meta(insert_at, meta)
                self._sync_row_date(insert_at)
                if was_saved:
                    self._saved_count += 1
                insert_at += 1
        finally:
            self._table.blockSignals(prev)
            self._bulk_mutating = False
        self._renumber()
        self._update_footer()
        self._finalize_truck_cells(truck_cells)
        if clear_cut_after and self._has_cut_buffer() and self._cut_is_rows:
            self._clear_cut_source_cells()
        self.edit_state_changed.emit(self._edit_mode, len(self._dirty_rows))
        self._schedule_draft_autosave()

    def _clear_cut_source_cells(self) -> None:
        """After a successful paste/insert, clear/remove the original cut source."""
        if not self._cut_cells:
            return
        rows_touched = set()
        self._table.blockSignals(True)
        try:
            if self._cut_is_rows and self._cut_payload.get("kind") == "rows":
                for row in sorted(self._cut_payload.get("rows") or [], reverse=True):
                    if row < 0 or row >= self._table.rowCount():
                        continue
                    self._shift_row_maps_on_remove(row)
                    self._table.removeRow(row)
                    if row < self._saved_count:
                        self._saved_count -= 1
                min_rows = self._saved_count + DEFAULT_EDITABLE_ROWS
                if self._table.rowCount() < min_rows:
                    start = self._table.rowCount()
                    self._table.setRowCount(min_rows)
                    self._init_editable_rows(start, min_rows)
            else:
                for row, col in list(self._cut_cells):
                    if row < 0 or row >= self._table.rowCount():
                        continue
                    if row < self._saved_count and not self._edit_mode:
                        continue
                    it = self._table.item(row, col)
                    if it is not None:
                        it.setText("")
                        rows_touched.add(row)
                        if row < self._saved_count:
                            self._dirty_rows.add(row)
                for row in rows_touched:
                    self._sync_row_date(row)
        finally:
            self._table.blockSignals(False)
        self._clear_cut_marquee()
        self._renumber()
        self._update_footer()

    def _shift_row_maps_on_insert(self, at_row: int) -> None:
        def _shift(mapping: dict) -> dict:
            return {
                (k + 1 if k >= at_row else k): v
                for k, v in mapping.items()
            }
        self._pending_row_meta = _shift(self._pending_row_meta)
        self._saved_ids = _shift(self._saved_ids)
        self._saved_txs = _shift(self._saved_txs)
        self._truck_allow_anyway = {
            (k + 1 if k >= at_row else k): v
            for k, v in self._truck_allow_anyway.items()
        }
        self._dirty_rows = {(r + 1 if r >= at_row else r) for r in self._dirty_rows}
        if self._cut_cells:
            self._cut_cells = {
                (r + 1 if r >= at_row else r, c) for r, c in self._cut_cells
            }
        if self._cut_is_rows and self._cut_payload.get("rows"):
            self._cut_payload["rows"] = [
                r + 1 if r >= at_row else r for r in self._cut_payload["rows"]
            ]

    def _shift_row_maps_on_remove(self, at_row: int) -> None:
        def _shift(mapping: dict) -> dict:
            out = {}
            for k, v in mapping.items():
                if k == at_row:
                    continue
                out[k - 1 if k > at_row else k] = v
            return out
        self._pending_row_meta = _shift(self._pending_row_meta)
        self._saved_ids = _shift(self._saved_ids)
        self._saved_txs = _shift(self._saved_txs)
        self._truck_allow_anyway = {
            (k - 1 if k > at_row else k): v
            for k, v in self._truck_allow_anyway.items()
            if k != at_row
        }
        self._dirty_rows = {
            (r - 1 if r > at_row else r)
            for r in self._dirty_rows
            if r != at_row
        }
        if self._cut_cells:
            self._cut_cells = {
                (r - 1 if r > at_row else r, c)
                for r, c in self._cut_cells
                if r != at_row
            }
        if self._cut_is_rows and self._cut_payload.get("rows"):
            self._cut_payload["rows"] = [
                r - 1 if r > at_row else r
                for r in self._cut_payload["rows"]
                if r != at_row
            ]

    def _paste_rows(self, body: str, clear_cut_after: bool = True) -> None:
        """Insert cut/copied rows at the current position."""
        lines = [ln for ln in body.splitlines() if ln.strip() != "" or "\t" in ln]
        if not lines:
            return
        # Prefer exact maps + identity when this paste is finishing a row cut.
        if (
            clear_cut_after
            and self._has_cut_buffer()
            and self._cut_is_rows
            and self._cut_payload.get("maps")
        ):
            self._paste_row_maps(
                self._cut_payload["maps"], clear_cut_after=True
            )
            return

        cur = self._table.currentRow()
        if self._merged_mode and self._edit_mode:
            insert_at = max(cur, 0)
        else:
            insert_at = max(cur, self._saved_count)

        self._record_row_insert_undo(insert_at, len(lines))
        truck_cells: list = []
        self._bulk_mutating = True
        prev = self._table.blockSignals(True)
        try:
            for line in lines:
                self._shift_row_maps_on_insert(insert_at)
                self._table.insertRow(insert_at)
                self._init_editable_rows(insert_at, insert_at + 1)
                cells = line.split("\t")
                values = {
                    col: cell
                    for col, cell in enumerate(cells)
                    if col < self._table.columnCount() and col != COL_SNO
                }
                truck_cells.extend(self._write_row_values(insert_at, values))
                self._sync_row_date(insert_at)
                if insert_at < self._saved_count:
                    self._saved_count += 1
                insert_at += 1
        finally:
            self._table.blockSignals(prev)
            self._bulk_mutating = False
        self._renumber()
        self._update_footer()
        self._finalize_truck_cells(truck_cells)
        if clear_cut_after and self._has_cut_buffer() and self._cut_is_rows:
            self._clear_cut_source_cells()

    def _paste(self) -> None:
        text = QApplication.clipboard().text()
        if not text:
            return

        if text.startswith(_ROWS_CLIP_PREFIX):
            if self._has_cut_buffer() and self._cut_payload.get("maps"):
                self._paste_row_maps(
                    self._cut_payload["maps"], clear_cut_after=True
                )
            else:
                self._paste_rows(text[len(_ROWS_CLIP_PREFIX):])
            return

        self._push_undo_cells(self._snapshot_selection())

        lines = text.splitlines()

        # selectedIndexes() covers blank rows (which have no QTableWidgetItem and
        # therefore never appear in selectedItems()).
        # In edit mode, paste may land on saved rows; otherwise stay below them.
        min_row = 0 if self._edit_mode else self._saved_count
        sel_indexes = self._table.selectedIndexes()
        if sel_indexes:
            start_row = max(min(i.row() for i in sel_indexes), min_row)
            start_col = min(i.column() for i in sel_indexes)
            sel_rows = sorted({i.row() for i in sel_indexes if i.row() >= min_row})
            sel_cols = sorted({i.column() for i in sel_indexes})
        else:
            start_row = max(self._table.currentRow(), min_row)
            start_col = self._table.currentColumn()
            sel_rows = []
            sel_cols = []

        truck_cells: list = []  # (row, raw_text)
        self._bulk_mutating = True
        try:
            # Single clipboard value pasted onto a multi-cell selection: fill every
            # selected editable cell with that value (Excel behaviour).
            if len(lines) == 1 and "\t" not in lines[0] and sel_rows and (
                len(sel_rows) > 1 or len(sel_cols) > 1
            ):
                cell_value = lines[0].strip()
                prev = self._table.blockSignals(True)
                for row in sel_rows:
                    for col in sel_cols:
                        if col in READONLY_COLS:
                            continue
                        if col in CHECK_COLS:
                            it = self._table.item(row, col) or QTableWidgetItem()
                            it.setData(Qt.UserRole, cell_value in ("1", "true", "True", "YES"))
                            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                            self._table.setItem(row, col, it)
                        elif col == COL_RECEIPT:
                            norm = _receipt_paste_value(cell_value)
                            it = self._table.item(row, col) or QTableWidgetItem()
                            it.setText(norm)
                            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                            self._table.setItem(row, col, it)
                        elif col in (COL_TZS, COL_USD):
                            self._table.setItem(row, col, _amount_item_from_raw(cell_value))
                        elif col == COL_TRUCK:
                            self._table.setItem(row, col, QTableWidgetItem(cell_value.upper()))
                            if cell_value:
                                truck_cells.append((row, cell_value))
                        else:
                            self._table.setItem(row, col, QTableWidgetItem(_upper_text(col, cell_value)))
                    if row < self._saved_count and self._edit_mode:
                        self._mark_dirty(row)
                    self._sync_row_date(row)
                self._table.blockSignals(prev)
                self._renumber()
                self._finalize_truck_cells(truck_cells)
            else:
                # Multi-row / multi-column clipboard: paste starting at anchor (TSV layout).
                touched_rows: set = set()
                prev = self._table.blockSignals(True)
                for r, line in enumerate(lines):
                    for c, cell in enumerate(line.split("\t")):
                        row = start_row + r
                        col = start_col + c
                        if row >= self._table.rowCount():
                            self._append_editable_rows(20)
                        if col >= self._table.columnCount() or col in READONLY_COLS:
                            continue
                        if row < self._saved_count and not self._edit_mode:
                            continue
                        touched_rows.add(row)
                        if col in CHECK_COLS:
                            it = self._table.item(row, col) or QTableWidgetItem()
                            it.setData(Qt.UserRole, cell.strip() in ("1", "true", "True", "YES"))
                            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                            self._table.setItem(row, col, it)
                        elif col == COL_RECEIPT:
                            norm = _receipt_paste_value(cell)
                            it = self._table.item(row, col) or QTableWidgetItem()
                            it.setText(norm)
                            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                            self._table.setItem(row, col, it)
                        elif col in (COL_TZS, COL_USD):
                            self._table.setItem(row, col, _amount_item_from_raw(cell))
                        elif col == COL_TRUCK:
                            raw = cell.strip()
                            self._table.setItem(row, col, QTableWidgetItem(raw.upper() if raw else ""))
                            if raw:
                                truck_cells.append((row, raw))
                        else:
                            self._table.setItem(row, col, QTableWidgetItem(_upper_text(col, cell.strip())))
                for row in touched_rows:
                    if row < self._saved_count and self._edit_mode:
                        self._mark_dirty(row)
                    self._sync_row_date(row)
                self._table.blockSignals(prev)
                self._renumber()
                self._finalize_truck_cells(truck_cells)
        finally:
            self._bulk_mutating = False

        if self._has_cut_buffer() and not self._cut_is_rows:
            self._clear_cut_source_cells()
        self._update_footer()
        self._schedule_draft_autosave()

    def _clear_selected(self) -> None:
        snap = self._snapshot_selection()
        cleared_rows: set = set()
        self._table.blockSignals(True)
        for item in self._table.selectedItems():
            row = item.row()
            col = item.column()
            if row < self._saved_count or col in READONLY_COLS:
                continue
            cleared_rows.add(row)
            if col in CHECK_COLS:
                item.setData(Qt.UserRole, False)
            else:
                item.setText("")
        for row in cleared_rows:
            self._sync_row_date(row)
        self._table.blockSignals(False)
        if snap:
            self._push_undo_cells(snap)
        self._renumber()
        self._update_footer()

    def _fill_down(self) -> None:
        """Ctrl+D: copy the top row of the selection into all rows below it."""
        items = self._table.selectedItems()
        if not items:
            return
        self._push_undo_cells(self._snapshot_selection())
        rows = sorted(set(it.row() for it in items))
        cols = sorted(set(it.column() for it in items))
        if len(rows) < 2:
            return
        source_row = rows[0]
        cell_map = {(it.row(), it.column()): it for it in items}
        truck_cells: list = []
        self._table.blockSignals(True)
        for col in cols:
            if col in READONLY_COLS:
                continue
            src = cell_map.get((source_row, col))
            if src is None:
                continue
            for row in rows[1:]:
                if row < self._saved_count:
                    continue
                if col in CHECK_COLS:
                    it = self._table.item(row, col) or QTableWidgetItem()
                    it.setData(Qt.UserRole, src.data(Qt.UserRole))
                    it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                    self._table.setItem(row, col, it)
                elif col == COL_TRUCK:
                    raw = src.text().strip()
                    self._table.setItem(row, col, QTableWidgetItem(raw.upper() if raw else ""))
                    if raw:
                        truck_cells.append((row, raw))
                else:
                    self._table.setItem(row, col, QTableWidgetItem(_upper_text(col, src.text())))
        for row in rows[1:]:
            if row >= self._saved_count:
                self._sync_row_date(row)
        self._table.blockSignals(False)
        self._renumber()
        self._finalize_truck_cells(truck_cells)

    def _copy_cell_from_to(
        self, src_row: int, src_col: int, dst_row: int, dst_col: int,
    ) -> list:
        """Copy one cell's value onto another editable cell. Returns truck finalize pairs."""
        if src_col in READONLY_COLS or dst_col in READONLY_COLS:
            return []
        if not self._cell_editable(dst_row, dst_col):
            return []
        src = self._table.item(src_row, src_col)
        truck_cells: list = []
        if dst_col in CHECK_COLS:
            it = self._table.item(dst_row, dst_col) or QTableWidgetItem()
            checked = bool(src.data(Qt.UserRole)) if src else False
            it.setData(Qt.UserRole, checked)
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self._table.setItem(dst_row, dst_col, it)
        elif dst_col == COL_TRUCK:
            raw = src.text().strip() if src else ""
            self._table.setItem(
                dst_row, dst_col, QTableWidgetItem(raw.upper() if raw else "")
            )
            if raw:
                truck_cells.append((dst_row, raw))
        elif dst_col == COL_RECEIPT:
            text = src.text().strip() if src else ""
            it = QTableWidgetItem(_receipt_paste_value(text) if text else "")
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self._table.setItem(dst_row, dst_col, it)
        elif dst_col in (COL_TZS, COL_USD):
            self._table.setItem(
                dst_row, dst_col, _amount_item_from_raw(src.text() if src else "")
            )
        else:
            text = src.text() if src else ""
            self._table.setItem(
                dst_row, dst_col, QTableWidgetItem(_upper_text(dst_col, text))
            )
        if dst_row < self._saved_count and self._edit_mode:
            self._dirty_rows.add(dst_row)
        return truck_cells

    def _apply_fill_drag(
        self,
        anchor: tuple[int, int, int, int],
        end: tuple[int, int],
    ) -> None:
        """Fill down/right from the selection when the user drags the fill handle."""
        r0, c0, r1, c1 = anchor
        end_row, end_col = end
        er = max(r1, end_row)
        ec = max(c1, end_col)
        if er <= r1 and ec <= c1:
            return

        snap: dict = {}
        for row in range(r0, er + 1):
            for col in range(c0, ec + 1):
                if r0 <= row <= r1 and c0 <= col <= c1:
                    continue
                if not self._cell_editable(row, col) or col in READONLY_COLS:
                    continue
                snap[(row, col)] = self._serialize_cell(row, col)
        if not snap:
            return
        self._push_undo_cells(snap)

        sel_h = r1 - r0 + 1
        sel_w = c1 - c0 + 1
        truck_cells: list = []
        self._table.blockSignals(True)
        try:
            for row in range(r0, er + 1):
                for col in range(c0, ec + 1):
                    if r0 <= row <= r1 and c0 <= col <= c1:
                        continue
                    if not self._cell_editable(row, col) or col in READONLY_COLS:
                        continue
                    src_row = r0 + (row - r0) % sel_h
                    src_col = c0 + (col - c0) % sel_w
                    truck_cells.extend(
                        self._copy_cell_from_to(src_row, src_col, row, col)
                    )
                if row >= self._saved_count:
                    self._sync_row_date(row)
                elif row in self._dirty_rows:
                    for col_i in range(self._table.columnCount()):
                        it = self._table.item(row, col_i)
                        if it is not None:
                            it.setBackground(QBrush(DIRTY_BG))
        finally:
            self._table.blockSignals(False)

        self._renumber()
        self._finalize_truck_cells(truck_cells)
        self._update_footer()
        self._schedule_draft_autosave()
        self._table.viewport().update()

    def _fill_right(self) -> None:
        """Ctrl+R: copy the leftmost column of the selection into all cols to its right."""
        items = self._table.selectedItems()
        if not items:
            return
        self._push_undo_cells(self._snapshot_selection())
        rows = sorted(set(it.row() for it in items))
        cols = sorted(set(it.column() for it in items))
        if len(cols) < 2:
            return
        source_col = cols[0]
        cell_map = {(it.row(), it.column()): it for it in items}
        self._table.blockSignals(True)
        for row in rows:
            if row < self._saved_count:
                continue
            src = cell_map.get((row, source_col))
            if src is None:
                continue
            for col in cols[1:]:
                if col in READONLY_COLS or col in CHECK_COLS:
                    continue
                self._table.setItem(row, col, QTableWidgetItem(_upper_text(col, src.text())))
        for row in rows:
            if row >= self._saved_count:
                self._sync_row_date(row)
        self._table.blockSignals(False)
        self._renumber()

    def _go_to_last_cell(self) -> None:
        """Ctrl+End: jump to the last cell that contains data."""
        last_row, last_col = 0, 0
        for row in range(self._table.rowCount()):
            for col in range(1, self._table.columnCount()):
                it = self._table.item(row, col)
                if it and it.text().strip():
                    last_row = max(last_row, row)
                    last_col = max(last_col, col)
        self._table.setCurrentCell(last_row, last_col)
        self._table.scrollTo(self._table.model().index(last_row, last_col))

    def _select_row(self) -> None:
        """Shift+Space: select the entire current row."""
        self._table.selectRow(self._table.currentRow())

    # ------------------------------------------------------------------
    # Context menu
    # ------------------------------------------------------------------

    def _populate_context_menu(self, menu: QMenu) -> None:
        row = self._table.currentRow()
        sel_rows = sorted({i.row() for i in self._table.selectedIndexes()})
        saved_sel = [r for r in sel_rows if r < self._saved_count]
        if (
            saved_sel
            and self._selection_is_full_rows(saved_sel)
            and not self._edit_mode
        ):
            label = (
                "Delete Saved Entr"
                + ("ies" if len(saved_sel) > 1 else "y")
            )
            act = menu.addAction(label)
            act.triggered.connect(lambda: self._delete_saved_rows(saved_sel))
        elif 0 <= row < self._saved_count and not self._edit_mode:
            act = menu.addAction("Delete Saved Entry")
            act.triggered.connect(lambda: self._delete_saved_row(row))
        else:
            menu.addAction("Copy",  self._copy)
            menu.addAction("Cut",   self._cut)
            menu.addAction("Paste", self._paste)
            if self._has_cut_buffer():
                menu.addAction("Insert Cut Cells", self._insert_cut_cells)
            menu.addSeparator()
            menu.addAction("Insert Row Above",       self._insert_above)
            menu.addAction("Insert Row Below",       self._insert_below)
            menu.addAction("Delete Selected Row(s)", self._delete_rows)
            if self._undo_stack:
                menu.addSeparator()
                menu.addAction("Undo", self._undo)
            if self._redo_stack:
                menu.addAction("Redo", self._redo)

    def _show_context_menu(self, pos) -> None:
        menu = QMenu(self._table)
        self._populate_context_menu(menu)
        menu.exec(self._table.viewport().mapToGlobal(pos))

    def _show_row_header_context_menu(self, global_pos) -> None:
        menu = QMenu(self._table)
        self._populate_context_menu(menu)
        menu.exec(global_pos)

    def _insert_above(self) -> None:
        row = max(self._table.currentRow(), self._saved_count)
        self._record_row_insert_undo(row, 1)
        self._shift_row_maps_on_insert(row)
        self._table.insertRow(row)
        self._init_editable_rows(row, row + 1)
        self._renumber()

    def _insert_below(self) -> None:
        row = max(self._table.currentRow() + 1, self._saved_count)
        self._record_row_insert_undo(row, 1)
        self._shift_row_maps_on_insert(row)
        self._table.insertRow(row)
        self._init_editable_rows(row, row + 1)
        self._renumber()

    def _delete_rows(self) -> None:
        rows = sorted(
            {i.row() for i in self._table.selectedIndexes()
             if i.row() >= self._saved_count},
            reverse=True,
        )
        if not rows:
            return
        self._record_row_remove_undo(sorted(rows))
        for row in rows:
            self._shift_row_maps_on_remove(row)
            self._table.removeRow(row)
        min_rows = self._saved_count + DEFAULT_EDITABLE_ROWS
        if self._table.rowCount() < min_rows:
            start = self._table.rowCount()
            self._table.setRowCount(min_rows)
            self._init_editable_rows(start, min_rows)
        self._renumber()

    # ------------------------------------------------------------------
    # Delete saved row(s)
    # ------------------------------------------------------------------

    def _delete_saved_row(self, row: int) -> None:
        self._delete_saved_rows([row])

    def _delete_saved_rows(self, rows: list) -> None:
        rows = sorted({r for r in rows if r in self._saved_ids})
        n_rows = len(rows)
        if not n_rows:
            return

        verified_n = 0
        pending_n = 0
        draft_n = 0
        submitted_n = 0
        for row in rows:
            tx = self._saved_txs.get(row)
            is_pending_edit = bool(
                tx is not None
                and getattr(tx, "original_transaction_id", None)
                and not getattr(tx, "verified", False)
            )
            is_verified = bool(tx is not None and getattr(tx, "verified", False))
            if is_verified:
                verified_n += 1
            elif is_pending_edit:
                pending_n += 1
            elif (
                tx is not None
                and (getattr(tx, "register_status", "") or "submitted") == "draft"
            ):
                draft_n += 1
            else:
                submitted_n += 1

        if n_rows == 1:
            row = rows[0]
            it = self._table.item(row, COL_DESC)
            desc = it.text() if it else "?"
            if verified_n:
                msg = (
                    f'Request deletion of approved expense:\n"{desc}"?\n\n'
                    "It will leave Master Expenses immediately and appear in the "
                    "accountant's Verify → Deleted tab. Confirming moves it to Trash."
                )
            elif pending_n:
                msg = (
                    f'Delete pending edit:\n"{desc}"?\n\n'
                    "This undoes the edit. The original approved expense stays in Master."
                )
            elif submitted_n:
                msg = (
                    f'Delete submitted transaction:\n"{desc}"?\n\n'
                    "It will be removed from the Verify inbox."
                )
            else:
                msg = f'Delete saved transaction:\n"{desc}"?'
        else:
            parts = [f"Delete {n_rows} selected saved row(s)?"]
            if verified_n:
                parts.append(
                    f"{verified_n} approved row(s) will go to Verify → Deleted "
                    "(confirm moves them to Trash)."
                )
            if pending_n:
                parts.append(
                    f"{pending_n} pending edit(s) will be removed "
                    "(originals stay on Master)."
                )
            if submitted_n or draft_n:
                parts.append(
                    f"{submitted_n + draft_n} unverified row(s) will be deleted."
                )
            msg = "\n\n".join(parts)

        if (
            QMessageBox.question(
                self, "Delete Entry",
                msg,
                QMessageBox.Yes | QMessageBox.No,
            )
            == QMessageBox.Yes
        ):
            tx_ids = [self._saved_ids[r] for r in rows if r in self._saved_ids]
            asyncio.ensure_future(self._do_delete_saved_many(tx_ids))

    async def _do_delete_saved(self, tx_id) -> None:
        await self._do_delete_saved_many([tx_id])

    async def _do_delete_saved_many(self, tx_ids: list) -> None:
        if not tx_ids:
            return
        try:
            cashier_id = getattr(self._user, "_id", None)
            requested = 0
            deleted = 0
            missing = 0
            for tx_id in tx_ids:
                result = await request_or_delete_transaction(tx_id, cashier_id)
                if result == "not_found":
                    missing += 1
                elif result == "deletion_requested":
                    requested += 1
                elif result == "deleted":
                    deleted += 1
            if missing and not (requested or deleted):
                QMessageBox.warning(
                    self, "Not Found",
                    "Those entries were already removed.",
                )
            elif requested:
                QMessageBox.information(
                    self, "Deletion Requested",
                    (
                        f"{requested} approved expense(s) sent to Verify → Deleted.\n"
                        "An accountant must confirm (moves to Trash) or restore."
                        + (f"\n{deleted} other row(s) were removed." if deleted else "")
                    ),
                )
            await self._load_date(self._current_date)
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Failed to delete:\n{exc}")

    # ------------------------------------------------------------------
    # Import from Excel (daily MATUMIZI)
    # ------------------------------------------------------------------

    def import_from_file(self) -> None:
        asyncio.ensure_future(self._run_daily_import())

    async def _run_daily_import(self) -> None:
        from tahmeed.ui.cashier.daily_import_flow import run_daily_import_flow

        try:
            preview = await run_daily_import_flow(self)
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", f"Could not import this file:\n\n{exc}")
            return
        if preview is None:
            return
        await self.apply_daily_import_preview(preview)

    async def apply_daily_import_preview(self, preview) -> None:
        """Navigate to the Excel main date and stage rows for the user to Save."""
        from tahmeed.services.daily_import_service import staged_row_payload
        from tahmeed.ui.widgets.upload_busy import UploadBusy

        primary = preview.primary_date or self._current_date
        if not isinstance(primary, date):
            QMessageBox.warning(
                self,
                "Reconciled Date required",
                "This import has no reconciled date. Set Reconciled Date and try again.",
            )
            return

        # Always resolve unsaved work first — same-date imports used to append on
        # top of recovered local-draft / typed rows (extra rows vs the Excel file).
        if self.has_unsaved_work():
            ok = await self.confirm_leave()
            if not ok:
                return

        # Drop any crash-recovery draft for the target day so load cannot re-inject
        # old rows under the fresh import.
        try:
            clear_register_draft(
                self._user._id, primary, merged=self._merged_mode
            )
        except Exception:
            pass

        self._skip_draft_restore = True
        try:
            if primary != self._current_date:
                self._reset_edit_state()
                self._current_date = primary
                with UploadBusy(self, "Opening import date…", title="Import"):
                    await self._load_date(primary)
            else:
                # Same day: wipe leftover unsaved editable rows before staging.
                self._wipe_unsaved_editable_rows()
        finally:
            self._skip_draft_restore = False

        payloads = [staged_row_payload(row, preview) for row in preview.rows]
        # Queue truck issues now, but open the correction dialog only after this
        # async import task finishes — nested dialog.exec() + ensure_future crashes
        # under Python 3.14 / qasync.
        self._suppress_truck_dialog = True
        try:
            with UploadBusy(
                self,
                f"Loading {len(payloads):,} row(s) into table…",
                title="Import",
            ) as busy:
                busy.update(f"Loading {len(payloads):,} row(s) into table…")
                self._load_staged_import_rows(payloads)
            QMessageBox.information(
                self,
                "Import ready",
                f"Loaded {len(payloads):,} row(s) from \"{preview.source_filename}\" "
                f"under reconciled date {primary.strftime('%d/%m/%Y')}.\n\n"
                "Excel row dates are kept as written. Open this upload (or this "
                "reconciled day in Simple) anytime to see every row in the batch.\n\n"
                "Review the Table, make any edits, then click Save.\n"
                "Saved entries go to the accountant Verify inbox.",
            )
        finally:
            self._suppress_truck_dialog = False
        QTimer.singleShot(0, self._flush_truck_correction)

    def _wipe_unsaved_editable_rows(self) -> None:
        """Clear typed / staged unsaved rows without undo (import staging prep)."""
        self._commit_open_editor()
        self._draft_timer.stop()
        txs = [
            self._saved_txs[r]
            for r in range(self._saved_count)
            if self._saved_txs.get(r) is not None
        ]
        show_cashier = not self._table.isColumnHidden(COL_CASHIER)
        self._pending_row_meta.clear()
        self._truck_allow_anyway.clear()
        self._pending_truck_issues.clear()
        self._populate(txs)
        self._table.setColumnHidden(COL_CASHIER, not show_cashier)

    def _kick_auto_fill_item(self, row: int, description: str) -> None:
        if self._bulk_mutating or not description.strip():
            return
        self._auto_fill_row = row
        asyncio.ensure_future(self._auto_fill_item_from_mapping(row, description))

    def _load_staged_import_rows(self, payloads: list) -> None:
        if not payloads:
            return
        self._bulk_mutating = True
        start = self._first_empty_editable_row()
        prev = self._table.blockSignals(True)
        truck_cells: list = []
        loaded_rows: set = set()
        try:
            for r, data in enumerate(payloads):
                target = start + r
                if target >= self._table.rowCount():
                    self._append_editable_rows(max(20, len(payloads) - r + 5))
                if target < self._saved_count:
                    continue
                loaded_rows.add(target)
                self._pending_row_meta[target] = {
                    "daily_import_id": data.get("daily_import_id"),
                    "daily_import_source": data.get("daily_import_source"),
                    "date_discrepancy": data.get("date_discrepancy"),
                    "import_primary_date": data.get("import_primary_date"),
                    "category_id": data.get("category_id"),
                    "currency": data.get("currency") or "TZS",
                    "lpo_do": (data.get("lpo_do") or "").upper(),
                    "do_number": (data.get("do_number") or "").upper(),
                }

                dt = data.get("date")
                date_str = format_register_date(dt) if dt else ""
                self._table.setItem(target, COL_DATE, QTableWidgetItem(date_str))

                item_name = data.get("item") or data.get("category_name") or ""
                if item_name:
                    self._table.setItem(
                        target, COL_ITEM, QTableWidgetItem(_upper_text(COL_ITEM, item_name))
                    )

                desc = data.get("description") or ""
                self._table.setItem(
                    target, COL_DESC, QTableWidgetItem(_upper_text(COL_DESC, desc))
                )

                truck = data.get("truck_number") or ""
                if truck:
                    self._table.setItem(target, COL_TRUCK, QTableWidgetItem(truck.upper()))
                    truck_cells.append((target, truck))

                memo = data.get("memo") or ""
                if memo:
                    self._table.setItem(
                        target, COL_MEMO, QTableWidgetItem(_upper_text(COL_MEMO, memo))
                    )

                ref = data.get("ref_float") or ""
                if ref:
                    self._table.setItem(
                        target, COL_REF, QTableWidgetItem(_upper_text(COL_REF, ref))
                    )

                # Split amount into TZS / USD columns (legacy USD-only rows land in USD).
                staged = Transaction(
                    date=data.get("date") or datetime.utcnow(),
                    description=data.get("description") or "",
                    truck_number=data.get("truck_number") or "",
                    amount=float(data.get("amount") or 0),
                    currency=data.get("currency") or "TZS",
                    amount_usd=(
                        float(data["amount_usd"])
                        if data.get("amount_usd") is not None
                        else None
                    ),
                )
                tzs_txt, usd_txt = _display_money_cells(staged)
                self._table.setItem(target, COL_TZS, _money_item(tzs_txt))
                self._table.setItem(target, COL_USD, _money_item(usd_txt))

                rcpt = data.get("receipt_status") or ""
                rcpt_it = QTableWidgetItem(str(rcpt))
                rcpt_it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                self._table.setItem(target, COL_RECEIPT, rcpt_it)

                own = data.get("ownership") or ""
                if own:
                    self._table.setItem(
                        target, COL_OWN, QTableWidgetItem(_upper_text(COL_OWN, own))
                    )
                apr = data.get("approver") or ""
                if apr:
                    self._table.setItem(
                        target, COL_APR, QTableWidgetItem(_upper_text(COL_APR, apr))
                    )

            for row in loaded_rows:
                self._sync_row_date(row)
        finally:
            self._table.blockSignals(prev)
            self._bulk_mutating = False
        self._renumber()
        self._finalize_truck_cells(truck_cells)
        self._update_footer()
        self._schedule_draft_autosave()

    def _load_rows(self, file_rows: List[List]) -> None:
        """Legacy positional loader kept for CSV paste-compat helpers."""
        if not file_rows:
            return
        data = file_rows[1:] if _is_header(file_rows[0]) else file_rows

        FILE_MAP = {
            COL_DATE:     1,
            COL_DESC:     3,
            COL_TRUCK:    4,
            COL_MEMO:     9,
            COL_REF:      10,
            COL_TZS:      11,
            COL_USD:      12,
            COL_RECEIPT:  13,
            COL_OWN:      14,
            COL_APR:      15,
        }

        start = self._first_empty_editable_row()
        self._table.blockSignals(True)
        loaded_rows: set = set()
        truck_cells: list = []
        for r, row_data in enumerate(data):
            target = start + r
            if target >= self._table.rowCount():
                self._append_editable_rows(20)
            if target < self._saved_count:
                continue
            loaded_rows.add(target)

            for grid_col, file_col in FILE_MAP.items():
                if file_col >= len(row_data):
                    continue
                raw = str(row_data[file_col]).strip() if row_data[file_col] is not None else ""

                if grid_col == COL_RECEIPT:
                    it = self._table.item(target, grid_col) or QTableWidgetItem()
                    it.setText(raw)
                    it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                    self._table.setItem(target, grid_col, it)
                elif grid_col == COL_REF:
                    low = raw.lower()
                    if low in ("1", "true", "yes", "refund to float") or (
                        "refund" in low and "float" in low
                    ):
                        text = "REFUND TO FLOAT"
                    elif raw and raw != "None":
                        text = raw
                    else:
                        text = ""
                    self._table.setItem(target, grid_col, QTableWidgetItem(text))
                elif grid_col == COL_DATE:
                    try:
                        from datetime import datetime as _dt
                        if isinstance(row_data[file_col], _dt):
                            formatted = format_register_date(row_data[file_col])
                        else:
                            formatted = raw
                    except Exception:
                        formatted = raw
                    self._table.setItem(target, grid_col, QTableWidgetItem(formatted))
                elif grid_col in (COL_TZS, COL_USD):
                    if raw and raw != "None":
                        self._table.setItem(target, grid_col, _amount_item_from_raw(raw))
                elif grid_col == COL_TRUCK:
                    if raw and raw != "None":
                        self._table.setItem(target, grid_col, QTableWidgetItem(raw.upper()))
                        truck_cells.append((target, raw))
                else:
                    if raw and raw != "None":
                        self._table.setItem(
                            target, grid_col, QTableWidgetItem(_upper_text(grid_col, raw))
                        )

        for row in loaded_rows:
            self._sync_row_date(row)
        self._table.blockSignals(False)
        self._renumber()
        self._finalize_truck_cells(truck_cells)

    def _first_empty_editable_row(self) -> int:
        last = self._saved_count - 1
        for row in range(self._saved_count, self._table.rowCount()):
            if self._row_has_data(row):
                last = row
        return last + 1

    # ------------------------------------------------------------------
    # Save to MongoDB
    # ------------------------------------------------------------------

    def save_rows(self) -> None:
        if self._current_date is None:
            QMessageBox.warning(
                self,
                "Reconciled Date required",
                "Set Reconciled Date before saving.",
            )
            return
        if self._save_in_flight or self._submit_in_flight:
            return
        asyncio.ensure_future(self._do_save())

    # ------------------------------------------------------------------
    # QuickBooks-style toolbar actions
    # ------------------------------------------------------------------

    def _data_rows(self) -> List[int]:
        """Saved + non-empty editable rows (for Find navigation)."""
        rows: List[int] = []
        for row in range(self._table.rowCount()):
            if self._table.isRowHidden(row):
                continue
            if row < self._saved_count or self._row_has_data(row):
                rows.append(row)
        return rows

    def toolbar_find(self, direction: int) -> None:
        """Move selection to previous (-1) or next (+1) data row."""
        rows = self._data_rows()
        if not rows:
            return
        cur = self._table.currentRow()
        if cur not in rows:
            target = rows[0] if direction >= 0 else rows[-1]
        else:
            idx = rows.index(cur)
            target = rows[(idx + (1 if direction >= 0 else -1)) % len(rows)]
        self._table.selectRow(target)
        self._table.setCurrentCell(target, COL_DESC)
        self._table.scrollToItem(
            self._table.item(target, COL_DESC) or self._table.item(target, COL_SNO),
            QAbstractItemView.PositionAtCenter,
        )

    def toolbar_new_table(self) -> None:
        """Open a blank register for fresh entry (Reconciled Date shows '-')."""
        asyncio.ensure_future(self._toolbar_new_table())

    # Back-compat alias — New no longer inserts a row.
    toolbar_new_row = toolbar_new_table

    async def _toolbar_new_table(self) -> None:
        self._commit_open_editor()
        if self.has_unsaved_work():
            if self._current_date is None:
                box = QMessageBox(self)
                box.setIcon(QMessageBox.Warning)
                box.setWindowTitle("Unsaved changes")
                box.setText(
                    "This blank table has unsaved entries but no Reconciled Date.\n\n"
                    "Set Reconciled Date to save as draft, or discard this work."
                )
                discard_btn = box.addButton("Discard", QMessageBox.DestructiveRole)
                cancel_btn = box.addButton("Cancel", QMessageBox.RejectRole)
                box.setDefaultButton(cancel_btn)
                box.exec()
                if box.clickedButton() != discard_btn:
                    return
            else:
                choice = self._prompt_unsaved_work(
                    "You have unsaved entries in the Daily Register.\n"
                    "Save as draft under the current Reconciled Date, "
                    "discard, or cancel?"
                )
                if choice == "cancel":
                    return
                if choice == "save":
                    if not await self._do_save():
                        return
                else:
                    self._clear_local_draft()
        else:
            self._clear_local_draft()
        self._open_blank_register()

    def _open_blank_register(self) -> None:
        """Empty editable grid with unset reconciled date (New table)."""
        self._draft_timer.stop()
        if isinstance(self._current_date, date):
            try:
                clear_register_draft(
                    self._user._id, self._current_date, merged=self._merged_mode
                )
            except Exception:
                pass
        self._load_gen += 1  # cancel any in-flight date load
        self._load_upload_id = ""
        self._skip_draft_restore = False
        self._reset_edit_state()
        self._pending_row_meta.clear()
        self._truck_allow_anyway.clear()
        self._pending_truck_issues.clear()
        self._header_cheque = ""
        self._current_date = None
        self._table.setColumnHidden(COL_CASHIER, True)
        self._populate([])
        self._hide_register_loading()
        if self._table.rowCount() > 0:
            self._table.setCurrentCell(0, COL_DESC)

    def toolbar_delete(self) -> None:
        """Delete selected saved entries or clear/remove selected unsaved rows."""
        sel_rows = sorted({i.row() for i in self._table.selectedIndexes()})
        saved = [r for r in sel_rows if r < self._saved_count]
        if saved and self._selection_is_full_rows(saved):
            self._delete_saved_rows(saved)
            if any(r >= self._saved_count for r in sel_rows):
                self._delete_rows()
            return
        row = self._table.currentRow()
        if row < 0:
            return
        if row < self._saved_count:
            self._delete_saved_row(row)
            return
        self._delete_rows()

    def toolbar_clear_table(self) -> None:
        """Clear every unsaved new row and revert unsaved edits on saved rows."""
        if not self.has_unsaved_work():
            QMessageBox.information(
                self,
                "Clear Table",
                "There is no unsaved data to clear.",
            )
            return
        if QMessageBox.question(
            self,
            "Clear Table",
            "Clear all unsaved rows and typed data?\n\n"
            "Saved entries on this register will not be removed.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        ) != QMessageBox.Yes:
            return
        self._commit_open_editor()
        self._clear_unsaved_with_undo()

    def _clear_unsaved_with_undo(self) -> None:
        """Wipe unsaved work while recording one undo step."""
        undo_cells: dict = {}
        row_ops: list = []

        for row in sorted(self._dirty_rows):
            undo_cells.update(self._snapshot_row_cells(row))

        min_rows = self._saved_count + DEFAULT_EDITABLE_ROWS
        extra_rows = list(range(min_rows, self._table.rowCount()))
        for row in range(self._saved_count, self._table.rowCount()):
            if self._row_has_data(row):
                undo_cells.update(self._snapshot_row_cells(row))
        for row in extra_rows:
            row_ops.append({
                "op": "remove",
                "at": row,
                "values": self._row_value_map(row),
                "meta": self._capture_row_meta(row),
            })

        entry: dict = {}
        if undo_cells:
            entry["cells"] = undo_cells
        if row_ops:
            entry["row_ops"] = row_ops
        if entry:
            self._record_undo(entry)

        self._bulk_mutating = True
        self._table.blockSignals(True)
        truck_cells: list = []
        try:
            for row in sorted(self._dirty_rows, reverse=True):
                tx = self._saved_txs.get(row)
                if tx is not None:
                    truck_cells.extend(
                        self._write_row_values(row, self._tx_to_row_values(tx))
                    )
                self._dirty_rows.discard(row)
                for col in range(self._table.columnCount()):
                    it = self._table.item(row, col)
                    if it is not None and self._edit_mode:
                        it.setBackground(QBrush(EDIT_BG))

            for row in range(self._saved_count, min(self._table.rowCount(), min_rows)):
                self._clear_editable_row(row)

            for row in sorted(extra_rows, reverse=True):
                self._shift_row_maps_on_remove(row)
                self._table.removeRow(row)
                self._pending_row_meta.pop(row, None)

            if self._table.rowCount() < min_rows:
                start = self._table.rowCount()
                self._table.setRowCount(min_rows)
                self._init_editable_rows(start, min_rows)
        finally:
            self._table.blockSignals(False)
            self._bulk_mutating = False

        self._renumber()
        self._finalize_truck_cells(truck_cells)
        self._clear_cut_marquee()
        self._clear_local_draft()
        self._update_footer()
        self._refresh_truck_required_highlights()
        self.edit_state_changed.emit(self._edit_mode, len(self._dirty_rows))
        self._table.setCurrentCell(
            max(self._saved_count, 0),
            COL_DESC,
        )

    def _tx_to_row_values(self, tx: Transaction) -> dict:
        tzs_txt, usd_txt = _display_money_cells(tx)
        cashier = (
            self._cashier_names.get(tx.cashier_id, "—") if tx.cashier_id else "—"
        )
        return {
            COL_DATE: format_register_date(tx.date) if tx.date else "",
            COL_ITEM: tx.item or "",
            COL_DESC: tx.description or "",
            COL_TRUCK: tx.truck_number or "",
            COL_MEMO: tx.memo or "",
            COL_REF: _ref_float_text(tx),
            COL_TZS: tzs_txt,
            COL_USD: usd_txt,
            COL_RECEIPT: tx.receipt_status or "pending",
            COL_OWN: tx.ownership or "",
            COL_APR: tx.approver or "",
            COL_PAYEE: getattr(tx, "payee", "") or "",
            COL_CHEQUE: getattr(tx, "cheque", "") or "",
            COL_CASHIER: cashier,
        }

    def _clear_editable_row(self, row: int) -> None:
        """Blank one unsaved editable row back to the default empty grid state."""
        for col in range(self._table.columnCount()):
            if col in (COL_SNO, COL_CASHIER) or col in READONLY_COLS:
                continue
            self._table.takeItem(row, col)
        self._pending_row_meta.pop(row, None)
        self._sync_row_date(row)
        self._deactivate_row(row)

    def toolbar_copy_row(self) -> None:
        """Duplicate the current row into a new unsaved editable row."""
        row = self._table.currentRow()
        if row < 0 or not (row < self._saved_count or self._row_has_data(row)):
            QMessageBox.information(
                self, "Create a Copy",
                "Select an entry to copy first.",
            )
            return
        values: dict = {}
        for col in range(self._table.columnCount()):
            if col in (COL_SNO, COL_CASHIER):
                continue
            it = self._table.item(row, col)
            if col == COL_RECEIPT and it is not None:
                text = it.text().strip()
                values[col] = text or (it.data(Qt.UserRole) or "pending")
            else:
                values[col] = it.text() if it else ""
        insert_at = max(self._saved_count, row + 1)
        self._shift_row_maps_on_insert(insert_at)
        self._table.insertRow(insert_at)
        self._init_editable_rows(insert_at, insert_at + 1)
        self._write_row_values(insert_at, values)
        self._renumber()
        self._table.selectRow(insert_at)
        self._table.setCurrentCell(insert_at, COL_DESC)

    def toolbar_print(self) -> None:
        self.export_as("pdf")

    def toolbar_attach(self) -> None:
        """Open attachment manager for the selected saved transaction."""
        row = self._table.currentRow()
        if row < 0 or row >= self._saved_count:
            QMessageBox.information(
                self, "Attach File",
                "Save the entry first, then select it to attach a file.",
            )
            return
        tx_id = self._saved_ids.get(row)
        if not tx_id:
            QMessageBox.information(
                self, "Attach File",
                "Save the entry first, then select it to attach a file.",
            )
            return
        tx = self._saved_txs.get(row)
        desc = ""
        if tx is not None:
            desc = getattr(tx, "description", "") or ""
        else:
            it = self._table.item(row, COL_DESC)
            desc = it.text() if it else ""
        from tahmeed.ui.dialogs.attachment_dialog import AttachmentDialog
        dlg = AttachmentDialog(
            tx_id,
            description=desc,
            actor_id=getattr(self._user, "_id", None),
            parent=self,
        )
        dlg.exec()
        asyncio.ensure_future(self._refresh_attachment_meta(row, tx_id))

    async def _refresh_attachment_meta(self, row: int, tx_id) -> None:
        try:
            from tahmeed.services.attachment_service import get_attachments
            atts = await get_attachments(tx_id)
            tx = self._saved_txs.get(row)
            if tx is not None:
                tx.attachments = atts
            self.attachment_count_changed.emit(len(atts))
        except Exception:
            self.attachment_count_changed.emit(0)

    def selected_attachment_count(self) -> int:
        row = self._table.currentRow()
        if row < 0 or row >= self._saved_count:
            return 0
        tx = self._saved_txs.get(row)
        if tx is None:
            return 0
        return len(getattr(tx, "attachments", None) or [])

    def _emit_attachment_badge(self) -> None:
        self.attachment_count_changed.emit(self.selected_attachment_count())

    def has_unsaved_work(self) -> bool:
        """True when edit-mode dirty rows or typed-but-unsaved new rows exist."""
        if self._dirty_rows:
            return True
        for row in range(self._saved_count, self._table.rowCount()):
            if self._row_has_data(row):
                return True
        return False

    # ------------------------------------------------------------------
    # Local draft autosave (crash / power-loss recovery)
    # ------------------------------------------------------------------

    def _schedule_draft_autosave(self) -> None:
        if (
            self._restoring_draft
            or self._save_in_flight
            or self._bulk_mutating
            or self._current_date is None
        ):
            return
        self._draft_timer.start()

    def _flush_local_draft(self, *, include_dirty: bool = True) -> None:
        """Persist current unsaved grid state to disk (or clear if empty)."""
        if self._restoring_draft or self._current_date is None:
            return
        try:
            self._commit_open_editor()
            payload = self._capture_local_draft(include_dirty=include_dirty)
            save_register_draft(payload)
        except Exception:
            # Local draft must never break typing / save.
            pass

    def _clear_local_draft(self) -> None:
        self._draft_timer.stop()
        if self._current_date is None:
            return
        try:
            clear_register_draft(
                self._user._id, self._current_date, merged=self._merged_mode
            )
        except Exception:
            pass

    def _capture_local_draft(self, *, include_dirty: bool = True) -> dict:
        dirty_saved: list = []
        if include_dirty:
            for row in sorted(self._dirty_rows):
                tx_id = self._saved_ids.get(row)
                if tx_id is None:
                    continue
                tx = self._saved_txs.get(row)
                dirty_saved.append({
                    "tx_id": str(tx_id),
                    "cashier_id": str(tx.cashier_id) if tx and tx.cashier_id else None,
                    "cells": cells_for_json(self._row_value_map(row)),
                })

        new_rows: list = []
        for row in range(self._saved_count, self._table.rowCount()):
            if not self._row_has_data(row):
                continue
            pending = self._pending_row_meta.get(row)
            new_rows.append({
                "cells": cells_for_json(self._row_value_map(row)),
                "pending_meta": serialize_pending_meta(pending),
            })

        return build_draft_payload(
            user_id=self._user._id,
            username=getattr(self._user, "username", "") or "",
            register_date=self._current_date,
            merged=self._merged_mode,
            edit_mode=self._edit_mode and include_dirty and bool(dirty_saved),
            dirty_saved=dirty_saved,
            new_rows=new_rows,
        )

    def _restore_local_draft(self) -> Optional[tuple]:
        """Apply a saved draft onto the freshly populated grid.

        Returns ``(dirty_count, new_count)`` when anything was restored, else None.
        """
        if self._current_date is None:
            return None
        draft = load_register_draft(
            self._user._id, self._current_date, merged=self._merged_mode
        )
        if draft is None or draft_is_empty(draft):
            return None

        dirty_entries = list(draft.get("dirty_saved") or [])
        new_entries = list(draft.get("new_rows") or [])
        if not dirty_entries and not new_entries:
            return None

        self._restoring_draft = True
        self._bulk_mutating = True
        dirty_applied = 0
        new_applied = 0
        truck_cells: list = []
        prev = self._table.blockSignals(True)
        try:
            id_to_row = {str(tx_id): row for row, tx_id in self._saved_ids.items()}
            need_edit = bool(dirty_entries) and bool(draft.get("edit_mode", True))
            if need_edit and not self._edit_mode:
                # Unlock saved rows without clearing dirty set prematurely.
                self._edit_mode = True
                editable = Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable
                for row in range(self._saved_count):
                    for col in range(self._table.columnCount()):
                        it = self._table.item(row, col)
                        if it is None:
                            continue
                        if col not in READONLY_COLS:
                            it.setFlags(editable)
                        it.setBackground(QBrush(EDIT_BG))

            for entry in dirty_entries:
                tx_id = str(entry.get("tx_id") or "")
                row = id_to_row.get(tx_id)
                if row is None:
                    continue
                values = cells_from_json(entry.get("cells"))
                truck_cells.extend(self._write_row_values(row, values))
                self._dirty_rows.add(row)
                for col in range(self._table.columnCount()):
                    it = self._table.item(row, col)
                    if it is not None:
                        it.setBackground(QBrush(DIRTY_BG))
                dirty_applied += 1

            if new_entries:
                start = self._first_empty_editable_row()
                needed = start + len(new_entries) - self._table.rowCount()
                if needed > 0:
                    self._append_editable_rows(needed + 5)
                for offset, entry in enumerate(new_entries):
                    row = start + offset
                    if row < self._saved_count:
                        continue
                    values = cells_from_json(entry.get("cells"))
                    truck_cells.extend(self._write_row_values(row, values))
                    meta = hydrate_pending_meta(entry.get("pending_meta"))
                    if meta:
                        self._pending_row_meta[row] = meta
                    self._activate_row(row)
                    self._sync_row_date(row)
                    new_applied += 1
        finally:
            self._table.blockSignals(prev)
            self._bulk_mutating = False
            self._restoring_draft = False

        self._renumber()
        self._finalize_truck_cells(truck_cells)
        self._update_footer()
        if dirty_applied:
            self.edit_state_changed.emit(True, len(self._dirty_rows))
        elif self._edit_mode:
            self.edit_state_changed.emit(True, 0)

        if dirty_applied or new_applied:
            return dirty_applied, new_applied
        return None

    def _show_draft_restored_notice(self, dirty_count: int, new_count: int) -> None:
        parts = []
        if dirty_count:
            parts.append(
                f"{dirty_count} edited row{'s' if dirty_count != 1 else ''}"
            )
        if new_count:
            parts.append(
                f"{new_count} new entr{'ies' if new_count != 1 else 'y'}"
            )
        detail = " and ".join(parts) if parts else "unsaved work"
        QMessageBox.information(
            self,
            "Draft restored",
            f"Recovered {detail} from before the app closed.\n\n"
            "Click Save to store them on the server.",
        )

    def _prompt_unsaved_work(self, message: str) -> str:
        """Return ``save``, ``discard``, or ``cancel`` for unsaved register work."""
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Question)
        box.setWindowTitle("Unsaved changes")
        box.setText(message)
        save_btn = box.addButton("Save as Draft", QMessageBox.AcceptRole)
        discard_btn = box.addButton("Discard", QMessageBox.DestructiveRole)
        cancel_btn = box.addButton("Cancel", QMessageBox.RejectRole)
        box.setDefaultButton(save_btn)
        box.exec()
        clicked = box.clickedButton()
        if clicked == save_btn:
            return "save"
        if clicked == discard_btn:
            return "discard"
        return "cancel"

    def _commit_open_editor(self) -> None:
        """Flush the active cell editor into the model before save/leave checks."""
        row, col = self._table.currentRow(), self._table.currentColumn()
        w = QApplication.focusWidget()
        if w is not None and self._table.isAncestorOf(w):
            self._table.commitData(w)
            self._table.closeEditor(w, QAbstractItemDelegate.NoHint)
        self._begin_closing_row_edit()
        self._flush_pending_row_edits()

    async def confirm_leave(self) -> bool:
        """Ask to save/discard before logout or app exit. False = stay put."""
        self._commit_open_editor()
        if not self.has_unsaved_work():
            self._clear_local_draft()
            return True
        self._flush_local_draft()
        if self._current_date is None:
            box = QMessageBox(self)
            box.setIcon(QMessageBox.Warning)
            box.setWindowTitle("Unsaved changes")
            box.setText(
                "You have unsaved entries but no Reconciled Date.\n\n"
                "Set Reconciled Date to save as draft, or discard before leaving."
            )
            discard_btn = box.addButton("Discard", QMessageBox.DestructiveRole)
            cancel_btn = box.addButton("Cancel", QMessageBox.RejectRole)
            box.setDefaultButton(cancel_btn)
            box.exec()
            if box.clickedButton() != discard_btn:
                return False
            self._clear_local_draft()
            return True
        choice = self._prompt_unsaved_work(
            "You have unsaved entries in the Daily Register.\n"
            "Save as draft before leaving?"
        )
        if choice == "cancel":
            return False
        if choice == "discard":
            self._clear_local_draft()
            return True
        # Save as draft; if the user cancels mid-save (duplicates / off-date), stay.
        return await self._do_save()

    async def _do_save(self) -> bool:
        """Persist dirty + new rows. Returns False if the user cancelled mid-save."""
        if self._save_in_flight:
            return False
        self._save_in_flight = True
        # Submit already holds the busy UI lock; avoid flickering it off early.
        nested_under_submit = self._submit_in_flight
        if not nested_under_submit:
            self.save_busy_changed.emit(True)
        try:
            return await self._do_save_body()
        finally:
            self._save_in_flight = False
            if not nested_under_submit and not self._submit_in_flight:
                self.save_busy_changed.emit(False)

    async def _do_save_body(self) -> bool:
        """Inner save implementation (caller holds `_save_in_flight`)."""
        saved, updated, errors = 0, 0, []
        self._commit_open_editor()

        if self._current_date is None:
            QMessageBox.warning(
                self,
                "Reconciled Date required",
                "Set Reconciled Date before saving as draft.",
            )
            return False

        missing_truck_rows = self._rows_missing_required_truck()
        if missing_truck_rows:
            detail_lines = []
            for row in missing_truck_rows[:12]:
                item_label = self._cell_text(row, COL_ITEM) or "—"
                detail_lines.append(f"  Row {row + 1}: {item_label}")
            if len(missing_truck_rows) > 12:
                detail_lines.append(
                    f"  …and {len(missing_truck_rows) - 12} more"
                )
            plural = "s" if len(missing_truck_rows) != 1 else ""
            QMessageBox.warning(
                self,
                "Truck Required",
                f"{len(missing_truck_rows)} row{plural} require a truck number "
                f"before saving:\n\n"
                + "\n".join(detail_lines)
                + "\n\nEnter the truck number for each row, then save again.",
                QMessageBox.Ok,
            )
            return False

        # ── Pass 1: commit edits to already-saved rows (UPDATE) ──────────
        for row in sorted(self._dirty_rows):
            tx_id = self._saved_ids.get(row)
            if tx_id is None:
                continue
            try:
                updates = self._updates_from_row(row)
            except ValueError as exc:
                errors.append(f"Row {row + 1}: {exc}")
                continue
            if updates is None:
                continue
            updates["last_edited_at"] = datetime.utcnow()
            updates["last_edited_by"] = self._user._id
            orig = self._saved_txs.get(row)
            try:
                if orig is not None and getattr(orig, "original_transaction_id", None):
                    # Already a pending-edit clone — refresh it in place.
                    updates["edited_after_verification"] = True
                    await update_transaction(tx_id, updates)
                elif orig is not None and orig.verified:
                    # Leave the original in Master Expenses intact; insert a
                    # pending-edit document that the accountant reviews in the
                    # Edited tab. On re-approval the new values cascade to the
                    # original in-place.
                    await insert_pending_edit(tx_id, updates, self._user._id)
                elif orig is not None and (getattr(orig, "register_status", "") or "") == "draft":
                    # Still in Merged draft — update in place; do not send to Edited yet.
                    await update_transaction(tx_id, updates)
                elif orig is not None:
                    # Option B: any edit of a saved (unverified / rejected) row
                    # moves it to Verify → Edited for accountant re-approval.
                    updates["edited_after_verification"] = True
                    if getattr(orig, "rejected", False):
                        updates["rejected"] = False
                        updates["rejection_reason"] = None
                        updates["discarded"] = False
                    await update_transaction(tx_id, updates)
                else:
                    await update_transaction(tx_id, updates)
                updated += 1
            except Exception as exc:
                errors.append(f"Row {row + 1}: {exc}")

        # ── Pass 2: insert brand-new rows (INSERT) ───────────────────────
        try:
            dup_days = int(await get_setting("duplicate_check_days") or 5)
        except Exception:
            dup_days = 5

        # ── Pre-scan: warn once if any new rows carry a date other than the open register day
        register_label = self._current_date.strftime("%d %b %Y")
        _off_date = 0
        for _s in range(self._saved_count, self._table.rowCount()):
            if not self._row_has_data(_s):
                continue
            _it = self._table.item(_s, COL_DATE)
            _ds = _it.text().strip() if _it else ""
            _parsed = _parse_optional_date(_ds, default_year=self._current_date.year)
            if _parsed is not None:
                _td = _parsed.date()
            else:
                _td = self._current_date
            if _td != self._current_date:
                _off_date += 1
        if _off_date:
            _plural = "s" if _off_date != 1 else ""
            _are    = "are" if _off_date != 1 else "is"
            if QMessageBox.warning(
                self, "Off-date Entries",
                f"{_off_date} row{_plural} {_are} not dated {register_label}.\n\n"
                "They will stay on this register day (visible after Save) but "
                "will be flagged in the accountant's Verify inbox.\n\n"
                "Proceed with save?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            ) == QMessageBox.No:
                return False

        pending_inserts: list[tuple[int, Transaction]] = []
        duplicate_items: list[DuplicateReviewItem] = []

        append_order = None
        if not self._merged_mode:
            try:
                append_order = await next_day_order(self._current_date)
            except Exception:
                append_order = None

        for row in range(self._saved_count, self._table.rowCount()):
            if not self._row_has_data(row):
                continue
            try:
                tx = self._build_transaction_from_row(row)
            except ValueError as exc:
                errors.append(f"Row {row + 1}: {exc}")
                continue
            if tx is None:
                continue

            if self._merged_mode:
                tx.day_order = row
            elif append_order is not None:
                tx.day_order = append_order
                append_order += 1
            else:
                tx.day_order = row

            dupes = []
            try:
                dupes = await check_for_duplicates(
                    truck_number=tx.truck_number or "",
                    amount=tx.amount,
                    item=tx.item or "",
                    description=tx.description or "",
                    days=dup_days,
                )
            except Exception:
                dupes = []

            if dupes:
                duplicate_items.append(DuplicateReviewItem(
                    row=row,
                    row_display=row + 1,
                    description=tx.description or "—",
                    truck_number=tx.truck_number or "",
                    item=tx.item or "",
                    amount=tx.amount,
                    amount_label=format_amount_label(tx),
                    existing=dupes[0],
                ))
            pending_inserts.append((row, tx))

        save_anyway: set[int] = set()
        if duplicate_items:
            dlg = DuplicateReviewDialog(
                duplicate_items, dup_days=dup_days, parent=self,
            )
            if dlg.exec() != QDialog.Accepted:
                self._flush_local_draft()
                return False
            save_anyway = dlg.save_anyway_rows()

        duplicate_rows = {item.row for item in duplicate_items}
        for row, tx in pending_inserts:
            if row in duplicate_rows and row not in save_anyway:
                continue
            tx.possible_duplicate = row in save_anyway
            try:
                await save_transaction(tx)
                saved += 1
                self._pending_row_meta.pop(row, None)
            except Exception as exc:
                errors.append(f"Row {row + 1}: {exc}")

        if self._merged_mode and (saved or updated):
            try:
                await self._persist_visual_day_order()
            except Exception as exc:
                errors.append(f"Could not save row order: {exc}")

        if errors:
            QMessageBox.warning(
                self, "Save — partial errors",
                f"{saved} added, {updated} updated.\n\nErrors:\n" + "\n".join(errors),
            )
            # Keep the grid as-is and refresh the local draft so a crash still
            # recovers remaining unsaved / failed rows.
            self._flush_local_draft()
            return True
        elif saved == 0 and updated == 0:
            QMessageBox.information(self, "Nothing to save", "No changes to save.")
            self._clear_local_draft()
            return True
        # else: clean save — reload silently, no popup

        self._clear_local_draft()
        self._reset_edit_state()
        self.rows_saved.emit(saved)
        self.drafts_changed.emit()
        await self._load_date(self._current_date)
        return True

    def _ordered_saved_ids(self) -> list:
        """Transaction ids in current on-screen order (saved prefix)."""
        return [
            self._saved_ids[r]
            for r in range(self._saved_count)
            if self._saved_ids.get(r)
        ]

    async def _persist_visual_day_order(self) -> None:
        """Write Merged-table sequence to ``day_order`` before a reload can scramble it."""
        if not self._merged_mode:
            return
        ordered_ids = self._ordered_saved_ids()
        if ordered_ids:
            await recount_day_order(self._current_date, ordered_ids)

    def _row_has_data(self, row: int) -> bool:
        for col in range(self._table.columnCount()):
            if col in _DATA_SKIP_COLS:
                continue
            it = self._table.item(row, col)
            if it and it.text().strip():
                return True
        return False

    # ------------------------------------------------------------------
    # Category update
    # ------------------------------------------------------------------

    def update_categories(self, categories: List[Category]) -> None:
        self._categories = categories
        self._cat_by_name = {c.name.lower(): c for c in categories}
        asyncio.ensure_future(self._load_locked_subitems())
        self._refresh_truck_required_highlights()

    async def _load_categories(self) -> None:
        """Ensure the Item column lists expense items and suppliers."""
        try:
            cats = await get_payment_target_categories()
        except Exception:
            return
        if cats:
            self.update_categories(cats)
            self._revalidate_visible_item_cells()

    def _revalidate_visible_item_cells(self) -> None:
        """Re-run Item validation after the catalog loads (clears false flags)."""
        for row in range(self._table.rowCount()):
            it = self._table.item(row, COL_ITEM)
            if it is None or not it.text().strip():
                continue
            if row < self._saved_count and not self._edit_mode:
                continue
            self._validate_item_cell(row, it)

    # ------------------------------------------------------------------
    # Settings / locked sub-item cache
    # ------------------------------------------------------------------

    async def _load_cashier_settings(self) -> None:
        try:
            self._restrict_items = bool(await get_setting("restrict_items"))
        except Exception:
            self._restrict_items = False
        try:
            self._defer_item_to_verify = bool(await get_setting("defer_item_to_verify"))
        except Exception:
            self._defer_item_to_verify = False
        try:
            self._restrict_trucks = True
            # Persist intended default so accountant UI / other clients stay in sync.
            current = await get_setting("restrict_trucks")
            if current is not True:
                await set_setting("restrict_trucks", True)
        except Exception:
            self._restrict_trucks = True
        try:
            from tahmeed.services.export_restriction_service import (
                get_enabled_export_surfaces,
            )

            self._export_restrict_surfaces = await get_enabled_export_surfaces()
        except Exception:
            self._export_restrict_surfaces = set()

    async def _auto_fill_item_from_mapping(self, row: int, description: str) -> None:
        """Pre-fill Item from a saved description map or prior entries."""
        try:
            if not description.strip():
                return
            item_it = self._table.item(row, COL_ITEM)
            if item_it and item_it.text().strip():
                return
            from tahmeed.services.cashier_service import resolve_item_name_for_description

            try:
                cat_name = await resolve_item_name_for_description(description)
            except Exception:
                return
            if not cat_name:
                return
            prev = self._table.blockSignals(True)
            if item_it is None:
                item_it = QTableWidgetItem(cat_name)
                item_it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                self._table.setItem(row, COL_ITEM, item_it)
            else:
                item_it.setText(cat_name)
            self._table.blockSignals(prev)
            self._validate_item_cell(row, item_it)
            self._update_truck_required_highlight(row)
        finally:
            if self._auto_fill_row == row:
                self._auto_fill_row = None
            self._schedule_finalize_row_edit()

    async def _load_fleet_numbers(self) -> None:
        from tahmeed.services.truck_service import get_fleet_kinds, get_fleet_numbers
        try:
            self._fleet_numbers = await get_fleet_numbers()
        except Exception:
            self._fleet_numbers = set()
        try:
            self._fleet_kinds = await get_fleet_kinds()
        except Exception:
            self._fleet_kinds = {}
        try:
            raw = await get_setting("allowed_truck_labels")
            if isinstance(raw, list) and raw:
                self._allowed_truck_labels = merge_allowed_labels(raw, DEFAULT_PLACE_LABELS)
            else:
                self._allowed_truck_labels = set(DEFAULT_PLACE_LABELS)
        except Exception:
            self._allowed_truck_labels = set(DEFAULT_PLACE_LABELS)

    async def _load_description_cache(self) -> None:
        """Warm system-wide description history for Excel-style autocomplete."""
        from tahmeed.services.cashier_service import ensure_description_cache
        try:
            await ensure_description_cache()
        except Exception:
            pass

    async def _remember_truck_labels(self, labels: list) -> None:
        if not labels:
            return
        try:
            merged = merge_allowed_labels(
                self._allowed_truck_labels, labels, DEFAULT_PLACE_LABELS
            )
            self._allowed_truck_labels = merged
            await set_setting("allowed_truck_labels", sorted(merged))
        except Exception:
            pass

    async def _load_locked_subitems(self) -> None:
        """Cache the allowed sub-item names for every lock-description item."""
        cache: dict = {}
        for c in self._categories:
            if getattr(c, "lock_description", False):
                try:
                    subs = await get_subtables(item_key(c.name))
                    cache[c.name.lower()] = [s.name for s in subs]
                except Exception:
                    cache[c.name.lower()] = []
        self._locked_subitems = cache

    # ------------------------------------------------------------------
    # Item-column validation (canonicalise / restrict / flag unknown)
    # ------------------------------------------------------------------

    def _item_row_background(self, row: int) -> QBrush:
        if row < self._saved_count:
            if row in self._dirty_rows:
                return QBrush(DIRTY_BG)
            if self._edit_mode:
                return QBrush(EDIT_BG)
            return QBrush(SAVED_BG)
        return QBrush(NEW_BG)

    def _flag_unknown_item(self, item: QTableWidgetItem, text: str) -> None:
        """Mark an unknown Item cell (restrict on) — keep text, no add dialog."""
        item.setForeground(QBrush(NEG_COLOR))
        item.setToolTip(
            f'"{text}" is not a known item. Pick an existing item from the list, '
            "or ask the accountant to add it in Manage Items."
        )

    def _clear_item_flag(self, row: int, item: QTableWidgetItem) -> None:
        tip = item.toolTip() or ""
        if "is not a known item" not in tip:
            return
        item.setToolTip("")
        item.setForeground(QBrush())
        # Restore row background in case an older build painted DUP_BG.
        if item.background().color() == DUP_BG:
            item.setBackground(self._item_row_background(row))

    def _validate_item_cell(self, row: int, item: QTableWidgetItem) -> None:
        text = item.text().strip()
        if not text:
            self._clear_item_flag(row, item)
            self._update_truck_required_highlight(row)
            return
        cat = self._cat_by_name.get(text.lower())
        if cat is not None:
            # Known item — snap to uppercase (table-view convention).
            canonical = cat.name.upper()
            if item.text() != canonical:
                self._table.blockSignals(True)
                item.setText(canonical)
                self._table.blockSignals(False)
            self._clear_item_flag(row, item)
            self._update_truck_required_highlight(row)
            return
        if not self._restrict_items:
            self._clear_item_flag(row, item)
            self._update_truck_required_highlight(row)
            return
        # Do not flag every cell when the catalog failed to load — that paints
        # known names red and makes Restrict look broken.
        if not self._categories:
            return
        # Unknown item with restriction on — keep the typed text and flag the cell.
        # Do not prompt to add; save still rejects unknown items.
        self._flag_unknown_item(item, text)
        self._update_truck_required_highlight(row)
    # ------------------------------------------------------------------
    # Truck-column validation (format + restrict to fleet registry)
    # ------------------------------------------------------------------

    def _cell_text(self, row: int, col: int) -> str:
        it = self._table.item(row, col)
        return it.text().strip() if it else ""

    def _item_requires_truck(self, item_name: str) -> bool:
        if not item_name:
            return False
        cat = self._cat_by_name.get(item_name.lower())
        if cat is None:
            return False
        return bool(getattr(cat, "requires_truck", True))

    def _row_missing_required_truck(self, row: int) -> bool:
        if not self._cell_text(row, COL_DESC):
            return False
        item_name = self._cell_text(row, COL_ITEM)
        if not self._item_requires_truck(item_name):
            return False
        return not self._cell_text(row, COL_TRUCK)

    def _rows_missing_required_truck(self) -> list[int]:
        missing: list[int] = []
        seen: set[int] = set()
        for row in self._dirty_rows:
            if row in seen:
                continue
            if self._row_missing_required_truck(row):
                missing.append(row)
                seen.add(row)
        for row in range(self._saved_count, self._table.rowCount()):
            if row in seen or not self._row_has_data(row):
                continue
            if self._row_missing_required_truck(row):
                missing.append(row)
                seen.add(row)
        return sorted(missing)

    _TRUCK_REQUIRED_TIP = "Truck number is required for this item."

    def _truck_cell_background(self, row: int) -> QBrush:
        if row >= self._saved_count:
            return QBrush(NEW_BG)
        if row in self._dirty_rows:
            return QBrush(DIRTY_BG)
        if self._edit_mode:
            return QBrush(EDIT_BG)
        tx = self._saved_txs.get(row)
        if tx is not None and (getattr(tx, "register_status", "") or "submitted") == "draft":
            return QBrush(DRAFT_BG)
        return QBrush(SAVED_BG)

    def _ensure_truck_cell(self, row: int) -> QTableWidgetItem:
        it = self._table.item(row, COL_TRUCK)
        if it is not None:
            return it
        it = QTableWidgetItem("")
        flags = Qt.ItemIsEnabled | Qt.ItemIsSelectable
        if row >= self._saved_count or self._edit_mode:
            flags |= Qt.ItemIsEditable
        it.setFlags(flags)
        it.setBackground(self._truck_cell_background(row))
        self._table.setItem(row, COL_TRUCK, it)
        return it

    def _update_truck_required_highlight(self, row: int) -> None:
        if row < 0 or row >= self._table.rowCount():
            return
        if row < self._saved_count and not self._edit_mode and not self._cell_text(row, COL_DESC):
            return
        truck_it = self._table.item(row, COL_TRUCK)
        if self._row_missing_required_truck(row):
            truck_it = self._ensure_truck_cell(row)
            prev = self._table.blockSignals(True)
            truck_it.setBackground(QBrush(TRUCK_REQUIRED_BG))
            truck_it.setToolTip(self._TRUCK_REQUIRED_TIP)
            self._table.blockSignals(prev)
            return
        if truck_it is None:
            return
        if truck_it.background().color() != TRUCK_REQUIRED_BG:
            return
        prev = self._table.blockSignals(True)
        if (truck_it.toolTip() or "") == self._TRUCK_REQUIRED_TIP:
            truck_it.setToolTip("")
        truck_it.setBackground(self._truck_cell_background(row))
        self._table.blockSignals(prev)

    def _refresh_truck_required_highlights(self) -> None:
        prev = self._table.blockSignals(True)
        try:
            for row in range(self._table.rowCount()):
                if not self._cell_text(row, COL_DESC):
                    continue
                self._update_truck_required_highlight(row)
        finally:
            self._table.blockSignals(prev)

    def _can_add_fleet(self) -> bool:
        return getattr(self._user, "role", "") in ("admin", "accountant")

    def _set_truck_cell(self, row: int, value: str) -> None:
        it = self._table.item(row, COL_TRUCK)
        prev = self._table.blockSignals(True)
        if it is None:
            it = QTableWidgetItem(value)
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self._table.setItem(row, COL_TRUCK, it)
        else:
            it.setText(value)
        self._table.blockSignals(prev)
        self._update_truck_required_highlight(row)

    def _truck_key(self, text: str) -> str:
        return " ".join((text or "").upper().split())

    def _resolve_truck_text(
        self, raw: str, row: Optional[int] = None
    ) -> tuple[str, Optional[str]]:
        """
        Return (status, value) where status is:
          'empty' | 'ok' | 'invalid_format' | 'not_in_registry'
        and value is the canonical / display string.
        """
        norm = normalize_truck_number(raw, allowed_labels=self._allowed_truck_labels)
        if norm.status == "empty":
            return "empty", ""
        if norm.status == "place_label":
            return "ok", norm.value
        if norm.status == "invalid":
            allowed = self._truck_allow_anyway.get(row) if row is not None else None
            if allowed and self._truck_key(raw) == self._truck_key(allowed):
                return "ok", allowed
            return "invalid_format", norm.value or raw.strip().upper()
        matched = try_match_fleet(norm.value, self._fleet_numbers)
        if matched is None:
            allowed = self._truck_allow_anyway.get(row) if row is not None else None
            if allowed and self._truck_key(raw) == self._truck_key(allowed):
                return "ok", allowed
            return "not_in_registry", norm.value
        return "ok", matched

    def _finalize_truck_cells(self, cells: list) -> None:
        """Normalize / validate truck cells; queue one combined correction dialog.

        ``cells`` is a list of (row, raw_text) for truck values just written.
        """
        if not cells:
            return

        for row, raw in cells:
            status, value = self._resolve_truck_text(raw, row=row)
            if status == "empty":
                self._set_truck_cell(row, "")
                self._pending_truck_issues.pop(row, None)
            elif status == "ok":
                self._set_truck_cell(row, value)
                self._pending_truck_issues.pop(row, None)
            elif status == "invalid_format":
                self._set_truck_cell(row, value)
                self._pending_truck_issues[row] = TruckIssue(
                    row=row, original=raw, kind="invalid_format"
                )
            else:
                self._set_truck_cell(row, value)
                self._pending_truck_issues[row] = TruckIssue(
                    row=row, original=raw, kind="not_in_registry"
                )

        for row, _raw in cells:
            self._update_truck_required_highlight(row)
        self._schedule_truck_correction()

    def _schedule_truck_correction(self) -> None:
        """Debounce so paste/import opens one combined dialog, not one per truck."""
        if not self._pending_truck_issues:
            return
        if self._suppress_truck_dialog:
            # Issues stay queued; caller flushes after the enclosing async work.
            return
        # If a dialog is already open, push new issues into it immediately.
        dlg = self._open_truck_dialog
        if dlg is not None and getattr(dlg, "isVisible", lambda: False)():
            batch = list(self._pending_truck_issues.values())
            self._pending_truck_issues.clear()
            try:
                dlg.add_issues(batch)
            except Exception:
                for issue in batch:
                    self._pending_truck_issues[issue.row] = issue
            return
        if self._truck_dialog_scheduled:
            return
        self._truck_dialog_scheduled = True
        QTimer.singleShot(0, self._flush_truck_correction)

    def _flush_truck_correction(self) -> None:
        self._truck_dialog_scheduled = False
        if not self._pending_truck_issues:
            return
        batch = list(self._pending_truck_issues.values())
        self._pending_truck_issues.clear()
        self._show_truck_correction(batch)

    def _show_truck_correction(self, issues: list) -> None:
        # Drop issues whose cell was cleared/changed already
        live: list[TruckIssue] = []
        for issue in issues:
            it = self._table.item(issue.row, COL_TRUCK)
            current = (it.text().strip() if it else "")
            if not current:
                continue
            status, value = self._resolve_truck_text(current, row=issue.row)
            if status == "ok":
                self._set_truck_cell(issue.row, value)
                continue
            issue.kind = "invalid_format" if status == "invalid_format" else "not_in_registry"
            issue.original = current
            live.append(issue)
        if not live:
            return

        # Merge with any dialog already open (should be rare after schedule coalesce)
        dlg = self._open_truck_dialog
        if dlg is not None and getattr(dlg, "isVisible", lambda: False)():
            dlg.add_issues(live)
            return

        dlg = TruckCorrectionDialog(
            live,
            self._fleet_numbers,
            can_add=self._can_add_fleet(),
            allowed_labels=self._allowed_truck_labels,
            on_resolved=self._on_truck_issue_resolved_live,
            fleet_kinds=getattr(self, "_fleet_kinds", None) or {},
            parent=self,
        )
        self._open_truck_dialog = dlg
        result = dlg.exec()
        self._open_truck_dialog = None

        pending_adds = list(getattr(dlg, "pending_registry_adds", None) or [])
        if pending_adds:
            asyncio.ensure_future(self._persist_truck_registry_adds(pending_adds))
        # Live callback already wrote resolved rows to the grid. On cancel,
        # remaining unresolved issues still need clearing.
        if getattr(dlg, "new_labels", None):
            asyncio.ensure_future(self._remember_truck_labels(dlg.new_labels))
        asyncio.ensure_future(self._load_fleet_numbers())
        if result != QDialog.Accepted:
            resolved_rows = {i.row for i in dlg.issues}
            for issue in live:
                if issue.row not in resolved_rows:
                    self._set_truck_cell(issue.row, "")

    async def _persist_truck_registry_adds(self, adds: list) -> None:
        from tahmeed.services.truck_service import add_fleet_by_collection

        for kind, number in adds:
            try:
                label = await add_fleet_by_collection(kind, number)
                self._fleet_kinds[number] = label
                self._fleet_numbers.add(number)
            except Exception as exc:
                QMessageBox.critical(
                    self, "Error", f"Failed to add {number} to registry:\n{exc}"
                )
        await self._load_fleet_numbers()

    def _on_truck_issue_resolved_live(self, issue: TruckIssue) -> None:
        """Apply one resolved truck to the grid as soon as it leaves the dialog list."""
        if issue.skip or not issue.corrected:
            self._truck_allow_anyway.pop(issue.row, None)
            self._set_truck_cell(issue.row, "")
            return
        if getattr(issue, "is_place_label", False):
            self._truck_allow_anyway.pop(issue.row, None)
            self._allowed_truck_labels.add(normalize_place_label(issue.corrected))
        elif getattr(issue, "allow_anyway", False):
            self._truck_allow_anyway[issue.row] = issue.corrected
        else:
            self._truck_allow_anyway.pop(issue.row, None)
            self._fleet_numbers.add(issue.corrected)
        self._set_truck_cell(issue.row, issue.corrected)

    def _validate_truck_cell(self, row: int, item: QTableWidgetItem) -> None:
        raw = item.text().strip()
        if not raw:
            self._truck_allow_anyway.pop(row, None)
            self._update_truck_required_highlight(row)
            return
        allowed = self._truck_allow_anyway.get(row)
        if allowed and self._truck_key(raw) != self._truck_key(allowed):
            self._truck_allow_anyway.pop(row, None)
        status, value = self._resolve_truck_text(raw, row=row)
        if status == "ok":
            if item.text() != value:
                prev = self._table.blockSignals(True)
                item.setText(value)
                self._table.blockSignals(prev)
            self._pending_truck_issues.pop(row, None)
            self._update_truck_required_highlight(row)
            return
        if status == "empty":
            self._update_truck_required_highlight(row)
            return
        kind = "invalid_format" if status == "invalid_format" else "not_in_registry"
        if item.text() != value:
            prev = self._table.blockSignals(True)
            item.setText(value)
            self._table.blockSignals(prev)
        self._pending_truck_issues[row] = TruckIssue(
            row=row, original=value or raw, kind=kind
        )
        self._schedule_truck_correction()

    def _reject_truck(self, row: int, number: str) -> None:
        # Kept for compatibility; correction dialog handles messaging.
        it = self._table.item(row, COL_TRUCK)
        if it is None or it.text().strip().upper() != number.upper():
            return
        self._pending_truck_issues[row] = TruckIssue(
            row=row, original=number, kind="not_in_registry"
        )
        self._schedule_truck_correction()

    # ------------------------------------------------------------------
    # Description-lock validation
    # ------------------------------------------------------------------

    def _validate_locked_description(self, row: int, item: QTableWidgetItem) -> None:
        item_name = self._cell_text(row, COL_ITEM)
        if not item_name:
            return
        cat = self._cat_by_name.get(item_name.lower())
        if cat is None or not getattr(cat, "lock_description", False):
            return
        allowed = self._locked_subitems.get(item_name.lower(), [])
        if not allowed:
            return  # locked but no sub-items defined → don't block
        text = item.text().strip()
        if not text:
            return
        match = next((a for a in allowed if a.lower() == text.lower()), None)
        if match is not None:
            if item.text() != match:
                self._table.blockSignals(True)
                item.setText(match)
                self._table.blockSignals(False)
            return
        QTimer.singleShot(
            0, lambda: self._reject_locked_description(row, text, cat.name, allowed)
        )

    def _reject_locked_description(
        self, row: int, text: str, item_name: str, allowed: List[str]
    ) -> None:
        it = self._table.item(row, COL_DESC)
        if it is None or it.text().strip().lower() != text.lower():
            return
        QMessageBox.information(
            self, "Description locked",
            f'"{item_name}" only allows these descriptions:\n\n• '
            + "\n• ".join(allowed)
            + "\n\nPlease pick one of the above.",
        )
        self._table.blockSignals(True)
        it.setText("")
        self._table.blockSignals(False)

    def _cell_text(self, row: int, col: int) -> str:
        it = self._table.item(row, col)
        return it.text().strip() if it else ""


# ---------------------------------------------------------------------------
# File helpers
# ---------------------------------------------------------------------------

def _read_spreadsheet(path: str) -> List[List]:
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8-sig") as f:
            return [row for row in csv.reader(f)]
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    result = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return result


def _is_header(row: List) -> bool:
    numeric = sum(
        1 for c in row
        if c is not None
        and str(c).replace(".", "").replace(",", "").replace("-", "").strip().isdigit()
    )
    return numeric == 0


# Backward-compat alias
ExcelGrid = DailyRegister
