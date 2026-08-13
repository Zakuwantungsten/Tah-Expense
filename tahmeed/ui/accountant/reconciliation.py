"""AccountantDashboard — Reconciliation › SM Burhani.

SM Burhani is an expandable sidebar parent with two sub-tables, each matching
the Excel files the accountant receives from the bonding company:

  RPAScheduleWidget — the "SM BURHANI RPA SCHEDULE" sheet (12 columns, adds
                      ASYCUDA AMOUNT / EXPORTER / DESCRIPTION OF SHIPMENT).
  BondsWidget       — the "SM BURHANI - BONDS" workbook, one tab per border
                      station (Nakonde / Kasumbalesa / Sakania, + Add Station),
                      with a reconciliation summary card.

Both use the upload-browse pattern (list of imports → click → full records),
allow Import (upload the schedule .xlsx, header-row auto-detected, dedup by
PRN + ENTRY REG) and Export (branded .xlsx), and use the same QuickBooks-style
striped table as the rest of the dashboard.
"""

from __future__ import annotations

import asyncio
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import qtawesome as qta

from PySide6.QtCore import Qt, QSize, Signal
from PySide6.QtWidgets import (
    QAbstractItemView, QComboBox, QDialog, QFileDialog, QFormLayout,
    QFrame, QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMenu, QMessageBox,
    QPushButton, QStackedWidget, QVBoxLayout, QWidget,
)

# Reuse the shared primitives / design tokens from the Separate Expenses module
# so SM Burhani looks identical to the rest of the dashboard.
from tahmeed.ui.accountant.separate_expenses import (
    _WHITE, _BG, _BORDER, _BLUE, _BLUE_L, _GREEN, _GREEN_L, _AMBER, _AMBER_L,
    _RED, _RED_L, _T1, _T2, _TM, _HDR_BG,
    _lbl, _btn, _input_ss, _hsep, _make_table, _cell, _fmt_num,
    _DropZone, _PageHeader, _TotalsBar, _SegmentTabBar,
    _table_style, _finish_table_row, _populate_year_combo, _TOLL_MONTHS, _SCROLL_CHUNK,
    _write_xlsx_template,
)
from tahmeed.models.reconciliation import ReconciliationEntry
from tahmeed.services import reconciliation_service as recon_svc
from tahmeed.ui.dialog_theme import show_question

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ═══════════════════════════════════════════════════════════════════════════════
#  Column specs   (label, field, width, align, mono)
# ═══════════════════════════════════════════════════════════════════════════════

_BONDS_COLS: List[Tuple[str, str, int, str, bool]] = [
    ("SR",              "sr_no",             44,  "center", False),
    ("SM REF NO",       "sm_ref_no",         100, "left",   False),
    ("PRN NUMBER",      "prn_number",        130, "left",   True),
    ("ENTRY REG NO",    "entry_reg_no",      95,  "left",   False),
    ("T1 NO",           "t1_no",             70,  "left",   False),
    ("T1 DATE",         "t1_date",           95,  "left",   False),
    ("IMPORTER",        "importer",          180, "left",   False),
    ("CONSIGNMENT",     "consignment",       140, "left",   False),
    ("TRUCK & TRAILER", "truck_and_trailer", 165, "left",   False),
    ("CHARGE",          "charge",            80,  "right",  True),
]

_RPA_COLS: List[Tuple[str, str, int, str, bool]] = [
    ("SR",              "sr_no",             44,  "center", False),
    ("SM REF NO",       "sm_ref_no",         100, "left",   False),
    ("PRN",             "prn_number",        130, "left",   True),
    ("REG NO",          "entry_reg_no",      80,  "left",   False),
    ("ASYCUDA",         "asycuda_amount",    95,  "right",  True),
    ("T1 NO",           "t1_no",             70,  "left",   False),
    ("T1 DATE",         "t1_date",           95,  "left",   False),
    ("IMPORTER",        "importer",          165, "left",   False),
    ("EXPORTER",        "exporter",          150, "left",   False),
    ("TRUCK & TRAILER", "truck_and_trailer", 160, "left",   False),
    ("SHIPMENT",        "consignment",       120, "left",   False),
    ("CHARGE",          "charge",            80,  "right",  True),
]

# Header-name candidates per field (case-insensitive substring match).
_BONDS_FIELDS = [
    ("sr_no",             ["sr"]),
    ("sm_ref_no",         ["sm ref"]),
    ("prn_number",        ["prn"]),
    ("entry_reg_no",      ["entry reg", "reg no"]),
    ("t1_no",             ["t1 no", "t1 number"]),
    ("t1_date",           ["t1 date"]),
    ("importer",          ["importer"]),
    ("consignment",       ["consignment", "description"]),
    ("truck_and_trailer", ["truck"]),
    ("charge",            ["charge"]),
]
_RPA_FIELDS = [
    ("sr_no",             ["sr"]),
    ("sm_ref_no",         ["sm ref"]),
    ("prn_number",        ["prn"]),
    ("entry_reg_no",      ["reg no", "entry reg"]),
    ("asycuda_amount",    ["asycuda"]),
    ("t1_no",             ["t1 number", "t1 no"]),
    ("t1_date",           ["t1 date"]),
    ("importer",          ["importer"]),
    ("exporter",          ["exporter"]),
    ("truck_and_trailer", ["truck"]),
    ("consignment",       ["description", "shipment"]),
    ("charge",            ["charge"]),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  Excel parsing
# ═══════════════════════════════════════════════════════════════════════════════

def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _is_header_row(row: tuple) -> bool:
    """True for a column-header row (has both an SR cell and a PRN cell)."""
    cells = [_norm(c) for c in row]
    has_sr = any(c == "sr" or c.startswith("sr.") or c.startswith("sr ") for c in cells)
    has_prn = any("prn" in c for c in cells)
    return has_sr and has_prn


def _find_header_row(rows: List[tuple]) -> Optional[int]:
    """First header row index, or None."""
    for i, row in enumerate(rows):
        if _is_header_row(row):
            return i
    return None


def _find_header_indices(rows: List[tuple]) -> List[int]:
    """Every stacked table header on a sheet (Nakonde, then Kasumbalesa, …)."""
    return [i for i, row in enumerate(rows) if _is_header_row(row)]


def _looks_like_section_row(row: tuple) -> bool:
    """Title / schedule banner / totals — not a data row."""
    text = " ".join(str(c).strip() for c in row if c is not None and str(c).strip())
    if not text:
        return True
    low = text.lower()
    if "schedule from" in low:
        return True
    if "tahmeed rpa" in low or "tahmeed bonds" in low or "tahmeed rits" in low:
        return True
    return False


def _map_columns(headers: List[str], fields) -> Dict[str, int]:
    used: set = set()
    colmap: Dict[str, int] = {}
    for field, candidates in fields:
        for cand in candidates:
            idx = next(
                (i for i, h in enumerate(headers)
                 if i not in used and cand in h),
                None,
            )
            if idx is not None:
                colmap[field] = idx
                used.add(idx)
                break
    return colmap


def _find_period(rows: List[tuple], start: int = 0, end: Optional[int] = None) -> str:
    stop = len(rows) if end is None else end
    for row in rows[start:stop]:
        for c in row:
            if c is not None and "schedule from" in str(c).lower():
                return str(c).strip()
    return ""


def _station_from_title_block(rows: List[tuple], header_idx: int, start: int) -> Tuple[str, str]:
    """Station slug/name from the banner above a table header."""
    for i in range(header_idx - 1, start - 1, -1):
        for c in rows[i]:
            if c is None or not str(c).strip():
                continue
            text = str(c).strip()
            if "schedule from" in text.lower():
                break
            word = text.split()[0].split("-")[0].strip()
            if word:
                return recon_svc._slug(word), word.title()
            break
    return "", ""


def _station_from_title(rows: List[tuple]) -> Tuple[str, str]:
    """First non-empty title cell → ('nakonde', 'Nakonde')."""
    return _station_from_title_block(rows, len(rows), 0)


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def _to_int(v) -> Optional[int]:
    f = _to_float(v)
    return int(f) if f is not None else None


def _to_dt(v) -> Optional[datetime]:
    if isinstance(v, datetime):
        return v
    if v in (None, ""):
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(str(v).strip()[:19], fmt)
        except ValueError:
            continue
    return None


def _cell_at(row: tuple, idx: Optional[int]) -> object:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_recon_sheet_rows(
    rows: List[tuple],
    table: str,
    sheet_title: str = "",
) -> List[ReconciliationEntry]:
    """Parse one worksheet, including stacked RPA tables with their own headers."""
    fields = _RPA_FIELDS if table == "rpa_schedule" else _BONDS_FIELDS
    header_indices = _find_header_indices(rows)
    if not header_indices:
        return []

    entries: List[ReconciliationEntry] = []
    for ti, hidx in enumerate(header_indices):
        next_h = header_indices[ti + 1] if ti + 1 < len(header_indices) else len(rows)
        block_start = 0 if ti == 0 else header_indices[ti - 1] + 1
        headers = [_norm(c) for c in rows[hidx]]
        colmap = _map_columns(headers, fields)
        period = _find_period(rows, block_start, next_h) or _find_period(rows)

        banner = _station_from_title_block(rows, hidx, block_start)[0]
        sheet_slug = recon_svc._slug(sheet_title) if sheet_title else ""
        if table == "bonds":
            station = banner or sheet_slug
        else:
            station = banner or sheet_slug or "nakonde"

        for raw in rows[hidx + 1 : next_h]:
            if _is_header_row(raw) or _looks_like_section_row(raw):
                continue
            prn = str(_cell_at(raw, colmap.get("prn_number")) or "").strip()
            sm_ref = str(_cell_at(raw, colmap.get("sm_ref_no")) or "").strip()
            if not prn and not sm_ref:
                continue
            if _norm(prn) in ("prn", "prn number") or _norm(sm_ref) in ("sm ref no", "sm ref"):
                continue
            truck = str(_cell_at(raw, colmap.get("truck_and_trailer")) or "").strip()
            if "truck" in _norm(truck) and "trailer" in _norm(truck):
                continue
            entries.append(ReconciliationEntry(
                table=table,
                station=station,
                schedule_period=period,
                sr_no=_to_int(_cell_at(raw, colmap.get("sr_no"))),
                sm_ref_no=sm_ref,
                prn_number=prn,
                entry_reg_no=str(_cell_at(raw, colmap.get("entry_reg_no")) or "").strip(),
                t1_no=str(_cell_at(raw, colmap.get("t1_no")) or "").strip(),
                t1_date=_to_dt(_cell_at(raw, colmap.get("t1_date"))),
                importer=str(_cell_at(raw, colmap.get("importer")) or "").strip(),
                consignment=str(_cell_at(raw, colmap.get("consignment")) or "").strip(),
                truck_and_trailer=truck,
                charge=_to_float(_cell_at(raw, colmap.get("charge"))) or 0.0,
                asycuda_amount=_to_float(_cell_at(raw, colmap.get("asycuda_amount"))),
                exporter=str(_cell_at(raw, colmap.get("exporter")) or "").strip(),
            ))
    return entries


def parse_recon_workbook(path: str, table: str) -> List[ReconciliationEntry]:
    """Parse every sheet of an SM Burhani workbook into reconciliation entries.

    For Bonds, each worksheet is usually a border station (tab title), and the
    banner may read ``NAKONDE - TAHMEED RITS``. Both feeds also accept stacked
    tables on one sheet (Nakonde, then Kasumbalesa), each with its own title
    banner and header row.
    """
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to read .xlsx files.")

    wb = openpyxl.load_workbook(path, data_only=True)
    entries: List[ReconciliationEntry] = []
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            entries.extend(_parse_recon_sheet_rows(rows, table, ws.title))
    finally:
        wb.close()
    return entries


def _fmt_recon_value(entry: ReconciliationEntry, field: str) -> str:
    val = getattr(entry, field, None)
    if field == "sr_no":
        return str(val) if val is not None else "—"
    if field == "t1_date":
        return val.strftime("%d %b %y") if isinstance(val, datetime) else "—"
    if field == "charge":
        return _fmt_num(val, "$ ", 0) if val else "—"
    if field == "asycuda_amount":
        return _fmt_num(val, "", 0) if val else "—"
    return str(val) if val not in (None, "") else "—"


# ═══════════════════════════════════════════════════════════════════════════════
#  Import dialog
# ═══════════════════════════════════════════════════════════════════════════════

class ReconImportDialog(QDialog):
    """Upload an SM Burhani schedule, dedup by PRN+ENTRY REG, preview, import."""

    imported = Signal(int)

    def __init__(self, table: str, title: str, cols, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._table = table
        self._cols = cols
        self._entries: List[ReconciliationEntry] = []
        self._new: List[ReconciliationEntry] = []
        self._source_path = ""
        self._last_skipped = 0

        self.setWindowTitle(f"Import — {title}")
        self.setMinimumWidth(720)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        browse_row = QWidget()
        browse_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(browse_row)
        brl.setContentsMargins(0, 0, 0, 0)
        brl.addStretch()
        browse_btn = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse_btn.clicked.connect(self._browse)
        brl.addWidget(browse_btn)
        vl.addWidget(browse_row)

        self._stats = _lbl("No file loaded.", size=12, color=_T2)
        vl.addWidget(self._stats)
        vl.addWidget(_hsep())

        vl.addWidget(_lbl("Preview (first 10 new rows)", size=12, weight=600))
        self._preview = _make_table([c[0] for c in self._cols])
        self._preview.setMinimumHeight(220)
        vl.addWidget(self._preview)

        vl.addWidget(_hsep())
        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.addStretch()
        cancel = _btn("Cancel", primary=False)
        cancel.clicked.connect(self.reject)
        bbl.addWidget(cancel)
        self._import_btn = _btn("Import Records", "mdi.check-circle-outline")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)
        vl.addWidget(btn_row)

    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open SM Burhani Schedule", "", "Excel Files (*.xlsx *.xls)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _on_file(self, path: str) -> None:
        from tahmeed.ui.widgets.upload_busy import UploadBusy

        self._source_path = path
        self._stats.setText("Reading file…")
        try:
            with UploadBusy(self, f"Reading {Path(path).name}…", title="Import"):
                self._entries = parse_recon_workbook(path, self._table)
        except Exception as exc:
            self._stats.setText(f"Error reading file: {exc}")
            return
        if not self._entries:
            self._stats.setText("No recognizable schedule rows found in this file.")
            self._import_btn.setEnabled(False)
            return
        asyncio.ensure_future(self._check_dupes())

    async def _check_dupes(self) -> None:
        keys = [e.dedup_key for e in self._entries]
        try:
            existing = await recon_svc.get_existing_recon_keys(keys)
        except Exception:
            existing = set()

        seen: set = set()
        new: List[ReconciliationEntry] = []
        for e in self._entries:
            k = e.dedup_key
            if k in existing or k in seen:
                continue
            seen.add(k)
            new.append(e)
        self._new = new

        stations = sorted({e.station for e in new})
        dupe = len(self._entries) - len(new)
        station_txt = f"  ·  Stations: {', '.join(s.title() for s in stations)}" if stations else ""
        self._stats.setText(
            f"New records: {len(new):,}     Duplicates (skipped): {dupe:,}{station_txt}"
        )
        self._import_btn.setEnabled(bool(new))
        self._import_btn.setText(f"Import {len(new):,} Records")
        self._fill_preview(new[:10])

    def _fill_preview(self, rows: List[ReconciliationEntry]) -> None:
        t = self._preview
        t.setRowCount(0)
        for e in rows:
            r = t.rowCount()
            t.insertRow(r)
            for c, (_, field, _w, align, mono) in enumerate(self._cols):
                flag = {"left": Qt.AlignLeft, "right": Qt.AlignRight,
                        "center": Qt.AlignHCenter}[align] | Qt.AlignVCenter
                t.setItem(r, c, _cell(_fmt_recon_value(e, field), flag, mono=mono))

    def _do_import(self) -> None:
        self._import_btn.setEnabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        from tahmeed.ui.accountant.import_truck_gate import run_import_truck_gate

        try:
            upload_id = str(uuid.uuid4())
            filename = Path(self._source_path).name if self._source_path else "Unknown"
            docs = []
            for e in self._new:
                e.upload_id = upload_id
                e.source_filename = filename
                docs.append(e.to_doc())
            gate = await run_import_truck_gate(
                self,
                docs,
                feed_key=self._table,
                upload_id=upload_id,
                source_filename=filename,
                can_add=True,
            )
            if gate.aborted:
                self._import_btn.setEnabled(True)
                self._import_btn.setText(f"Import {len(self._new):,} Records")
                return
            self._last_skipped = gate.skipped_count
            if not gate.rows:
                if gate.skipped_count:
                    QMessageBox.information(
                        self,
                        "Import",
                        f"No rows imported. {gate.skipped_count:,} parked in Skipped.",
                    )
                    self.imported.emit(0)
                    self.accept()
                else:
                    self._import_btn.setEnabled(True)
                    self._import_btn.setText(f"Import {len(self._new):,} Records")
                return
            entries = [ReconciliationEntry.from_doc(d) for d in gate.rows]
            saved = await recon_svc.save_reconciliation_rows(entries)
            self.imported.emit(saved)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(self._new):,} Records")


# ═══════════════════════════════════════════════════════════════════════════════
#  Upload browse / detail widgets
# ═══════════════════════════════════════════════════════════════════════════════

_RECON_BROWSE_HEADERS = [
    "UPLOAD DATE", "FILE NAME", "RECORDS", "TOTAL CHARGE", "SCHEDULE PERIOD",
]


class _ReconAllEntries(QWidget):
    """Flat, filterable list of every SM Burhani record for one table."""

    def __init__(self, table: str, cols, bonds: bool = False, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._table = table
        self._cols = cols
        self._bonds = bonds
        self._station = ""
        self._search = ""
        self._year = 0
        self._month = 0
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        if self._bonds:
            self._tabs = _StationTabBar()
            self._tabs.station_changed.connect(self._on_station)
            self._tabs.add_requested.connect(self._add_station)
            vl.addWidget(self._tabs)

            self._summary = QFrame()
            self._summary.setStyleSheet(
                f"QFrame{{background:{_BLUE_L};border:1px solid #BFDBFE;border-radius:6px;}}"
            )
            sl = QHBoxLayout(self._summary)
            sl.setContentsMargins(16, 10, 16, 10)
            sl.setSpacing(28)
            self._inv_lbl = _lbl("Invoiced: —", size=12, weight=600)
            self._conf_lbl = _lbl("Confirmed: —", size=12, weight=600, color=_GREEN)
            self._var_lbl = _lbl("Variance: —", size=12, weight=600, color=_AMBER)
            self._disp_lbl = _lbl("Disputed: 0", size=12, color=_T2)
            sl.addWidget(self._inv_lbl)
            sl.addWidget(self._conf_lbl)
            sl.addWidget(self._var_lbl)
            sl.addWidget(self._disp_lbl)
            sl.addStretch()
            vl.addWidget(self._summary)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search PRN, SM ref, importer, truck…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)
        tbl.addStretch()
        vl.addWidget(tb)

        self._totals = _TotalsBar([("charge", "$ "), ("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._tbl = _make_table([c[0] for c in self._cols])
        hdr = self._tbl.horizontalHeader()
        for i, (_, _f, width, _a, _m) in enumerate(self._cols):
            self._tbl.setColumnWidth(i, width)
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        hdr.setStretchLastSection(True)
        self._tbl.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._tbl, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        if self._bonds:
            asyncio.ensure_future(self._load_stations())
        else:
            asyncio.ensure_future(self._reload_years_and_data())

    async def _load_stations(self) -> None:
        stations = await recon_svc.get_recon_stations("bonds")
        self._tabs.set_stations(stations, active=self._station or self._tabs.active)
        self._station = self._tabs.active
        await self._reload_years_and_data()

    async def _reload_years_and_data(self) -> None:
        try:
            years = await recon_svc.get_recon_available_years(self._table, self._station)
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._tbl.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        month = self._effective_month()
        try:
            rows, total, totals = await asyncio.gather(
                recon_svc.get_recon_all_records(
                    self._table, self._station, self._search, self._year, month, _SCROLL_CHUNK, 0,
                ),
                recon_svc.count_recon_all_records(
                    self._table, self._station, self._search, self._year, month,
                ),
                recon_svc.get_recon_all_totals(
                    self._table, self._station, self._search, self._year, month,
                ),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(rows)
        self._loaded = len(rows)
        self._totals.set_total("charge", totals["invoiced"], "$ ")
        self._totals.set_total("count", totals["count"], "")
        if self._bonds:
            self._inv_lbl.setText(f"Invoiced: $ {totals['invoiced']:,.0f}")
            self._conf_lbl.setText(f"Confirmed: $ {totals['confirmed']:,.0f}")
            self._var_lbl.setText(f"Variance: $ {totals['variance']:,.0f}")
            self._disp_lbl.setText(f"Disputed: {totals['disputed']}")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        month = self._effective_month()
        try:
            rows = await recon_svc.get_recon_all_records(
                self._table, self._station, self._search, self._year, month, _SCROLL_CHUNK, self._loaded,
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if rows:
            self._append_rows(rows)
            self._loaded += len(rows)
        self._loading = False
        self._update_status()

    def _append_rows(self, rows: List[ReconciliationEntry]) -> None:
        t = self._tbl
        for e in rows:
            r = t.rowCount()
            t.insertRow(r)
            for c, (_, field, _w, align, mono) in enumerate(self._cols):
                flag = {"left": Qt.AlignLeft, "right": Qt.AlignRight, "center": Qt.AlignHCenter}[align] | Qt.AlignVCenter
                color = _RED if (field == "charge" and (e.charge or 0) < 0) else ""
                t.setItem(r, c, _cell(_fmt_recon_value(e, field), flag, mono=mono, color=color))
            _finish_table_row(t, r)

    def _on_scroll(self, value: int) -> None:
        bar = self._tbl.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        self._reset_and_load()

    def _on_station(self, slug: str) -> None:
        self._station = slug
        self._year = 0
        self._month = 0
        asyncio.ensure_future(self._reload_years_and_data())

    def _add_station(self) -> None:
        dlg = _AddStationDialog(parent=self)
        dlg.submitted.connect(lambda name, border: asyncio.ensure_future(self._do_add_station(name, border)))
        dlg.exec()

    async def _do_add_station(self, name: str, border: str) -> None:
        await recon_svc.add_recon_station(name, border, "bonds")
        await self._load_stations()


class _StationTabBar(QWidget):
    station_changed = Signal(str)   # slug
    add_requested = Signal()

    def __init__(self, parent: QWidget | None = None, *, allow_add: bool = True) -> None:
        super().__init__(parent)
        self.setStyleSheet("background:transparent;")
        self._allow_add = allow_add
        self._hl = QHBoxLayout(self)
        self._hl.setContentsMargins(0, 0, 0, 0)
        self._hl.setSpacing(6)
        self._buttons: Dict[str, QPushButton] = {}
        self._active = ""

    def set_stations(self, stations: List[dict], active: str = "") -> None:
        while self._hl.count():
            item = self._hl.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._buttons.clear()

        for st in stations:
            slug, name = st["slug"], st["name"]
            chip = QPushButton(name)
            chip.setCursor(Qt.PointingHandCursor)
            chip.setFixedHeight(30)
            chip.clicked.connect(lambda _=False, s=slug: self._select(s))
            self._buttons[slug] = chip
            self._hl.addWidget(chip)

        if self._allow_add:
            add = QPushButton("+ Add Station")
            add.setCursor(Qt.PointingHandCursor)
            add.setFixedHeight(30)
            add.setStyleSheet(
                f"QPushButton{{background:transparent;color:{_BLUE};border:1px dashed {_BLUE};"
                "border-radius:15px;font-size:12px;font-weight:600;"
                "font-family:'Segoe UI';padding:0 14px;}}"
                f"QPushButton:hover{{background:{_BLUE_L};}}"
            )
            add.clicked.connect(self.add_requested.emit)
            self._hl.addWidget(add)
        self._hl.addStretch()

        if stations:
            self._select(active if active in self._buttons else stations[0]["slug"],
                         emit=False)

    def _select(self, slug: str, emit: bool = True) -> None:
        self._active = slug
        for s, chip in self._buttons.items():
            if s == slug:
                chip.setStyleSheet(
                    f"QPushButton{{background:{_BLUE};color:#FFF;border:none;"
                    "border-radius:15px;font-size:12px;font-weight:600;"
                    "font-family:'Segoe UI';padding:0 16px;}}"
                )
            else:
                chip.setStyleSheet(
                    f"QPushButton{{background:{_WHITE};color:{_T2};border:1px solid {_BORDER};"
                    "border-radius:15px;font-size:12px;"
                    "font-family:'Segoe UI';padding:0 16px;}}"
                    f"QPushButton:hover{{background:{_BG};color:{_T1};}}"
                )
        if emit:
            self.station_changed.emit(slug)

    @property
    def active(self) -> str:
        return self._active


class _AddStationDialog(QDialog):
    submitted = Signal(str, str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("Add Station")
        self.setMinimumWidth(380)
        self.setStyleSheet(f"background:{_WHITE};")
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        form = QFormLayout()
        form.setSpacing(10)
        self._name = QLineEdit()
        self._name.setPlaceholderText("e.g. Mwami")
        self._name.setStyleSheet(_input_ss())
        self._border = QLineEdit()
        self._border.setPlaceholderText("e.g. Zambia – Malawi border")
        self._border.setStyleSheet(_input_ss())
        form.addRow("Station Name *", self._name)
        form.addRow("Border Post", self._border)
        vl.addLayout(form)
        vl.addWidget(_hsep())

        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(btn_row)
        brl.setContentsMargins(0, 0, 0, 0)
        brl.addStretch()
        cancel = _btn("Cancel", primary=False)
        cancel.clicked.connect(self.reject)
        brl.addWidget(cancel)
        add = _btn("Add Station", "mdi.plus-circle-outline")
        add.clicked.connect(self._submit)
        brl.addWidget(add)
        vl.addWidget(btn_row)

    def _submit(self) -> None:
        name = self._name.text().strip()
        if not name:
            QMessageBox.warning(self, "Validation", "Station name is required.")
            return
        self.submitted.emit(name, self._border.text().strip())
        self.accept()


class _ReconUploadBrowse(QWidget):
    """Table of every import batch for one SM Burhani table type."""

    upload_clicked = Signal(object)
    delete_clicked = Signal(object)

    def __init__(self, table: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._table = table
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        self._totals = _TotalsBar([("charge", "$ "), ("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table_w = _make_table(_RECON_BROWSE_HEADERS)
        self._table_w.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table_w.horizontalHeader().setStretchLastSection(True)
        self._table_w.setColumnWidth(0, 160)
        self._table_w.setColumnWidth(1, 260)
        self._table_w.setColumnWidth(2, 80)
        self._table_w.setColumnWidth(3, 120)
        self._table_w.setStyleSheet(_table_style())
        self._table_w.setCursor(Qt.PointingHandCursor)
        self._table_w.cellClicked.connect(self._on_row_clicked)
        self._table_w.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table_w.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table_w, 1)

        hint = _lbl("Click any row to view its records · right-click to delete an upload.", size=11, color=_TM)
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        uploads = await recon_svc.get_recon_uploads(self._table)
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table_w
        t.setRowCount(0)
        total_charge = 0.0
        total_recs = 0
        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)
            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count = int(up.get("record_count", 0))
            charge = float(up.get("total_charge", 0))
            period = str(up.get("schedule_period") or "—").strip() or "—"

            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 2, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 3, _cell(
                _fmt_num(charge, "$ ", 0),
                mono=True, align=Qt.AlignRight | Qt.AlignVCenter,
            ))
            t.setItem(r, 4, _cell(period))
            _finish_table_row(t, r)
            total_charge += charge
            total_recs += count

        self._totals.set_total("charge", total_charge, "$ ")
        self._totals.set_total("count", total_recs, "")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table_w.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table_w.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


class _ReconUploadDetail(QWidget):
    """Full record table for a single SM Burhani upload batch."""

    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(
        self,
        table: str,
        title: str,
        cols,
        bonds: bool = False,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._table = table
        self._title = title
        self._cols = cols
        self._bonds = bonds
        self._upload_id = ""
        self._station = ""
        self._search = ""
        self._upload_doc: dict = {}
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)
        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)
        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()
        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        export_btn = _btn("Export", "mdi.download-outline", primary=False, height=30)
        export_btn.clicked.connect(self._export)
        navl.addWidget(export_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        self._tabs = _StationTabBar(allow_add=False)
        self._tabs.station_changed.connect(self._on_station)
        vl.addWidget(self._tabs)

        if self._bonds:
            self._summary = QFrame()
            self._summary.setStyleSheet(
                f"QFrame{{background:{_BLUE_L};border:1px solid #BFDBFE;border-radius:6px;}}"
            )
            sl = QHBoxLayout(self._summary)
            sl.setContentsMargins(16, 10, 16, 10)
            sl.setSpacing(28)
            self._inv_lbl = _lbl("Invoiced: —", size=12, weight=600)
            self._conf_lbl = _lbl("Confirmed: —", size=12, weight=600, color=_GREEN)
            self._var_lbl = _lbl("Variance: —", size=12, weight=600, color=_AMBER)
            self._disp_lbl = _lbl("Disputed: 0", size=12, color=_T2)
            sl.addWidget(self._inv_lbl)
            sl.addWidget(self._conf_lbl)
            sl.addWidget(self._var_lbl)
            sl.addWidget(self._disp_lbl)
            sl.addStretch()
            vl.addWidget(self._summary)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search PRN, SM ref, importer, truck…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        tbl.addStretch()
        self._schedule_lbl = _lbl("", size=11, color=_TM)
        tbl.addWidget(self._schedule_lbl)
        vl.addWidget(tb)

        self._tbl = _make_table([c[0] for c in self._cols])
        hdr = self._tbl.horizontalHeader()
        for i, (_, _f, width, _a, _m) in enumerate(self._cols):
            self._tbl.setColumnWidth(i, width)
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        hdr.setStretchLastSection(True)
        vl.addWidget(self._tbl, 1)

        self._totals = _TotalsBar([("charge", "$ "), ("count", "Records: ")])
        vl.addWidget(self._totals)
        self._tbl.verticalScrollBar().valueChanged.connect(self._on_scroll)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc or {}
        self._upload_id = str(upload_doc.get("_id") or "")
        self._station = ""
        filename = upload_doc.get("source_filename") or "Unknown file"
        count = int(upload_doc.get("record_count", 0))
        charge = float(upload_doc.get("total_charge", 0))
        period = str(upload_doc.get("schedule_period") or "").strip()
        import_dt = upload_doc.get("import_date")
        date_str = (
            import_dt.strftime("%d %b %Y")
            if isinstance(import_dt, datetime) else ""
        )
        self._crumb_lbl.setText(f"Uploads  ›  {filename}")
        self._info_lbl.setText(
            f"{filename}   •   {count:,} records   •   $ {charge:,.0f}   •   {date_str}"
        )
        self._schedule_lbl.setText(period)
        self._totals.set_total("charge", charge, "$ ")
        self._totals.set_total("count", count, "")

        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)

        asyncio.ensure_future(self._load_stations())

    def _request_delete(self) -> None:
        doc = self._upload_doc or ({"_id": self._upload_id} if self._upload_id else {})
        if doc.get("_id"):
            self.delete_requested.emit(doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._tbl.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_stations(self) -> None:
        stations = await recon_svc.get_recon_upload_stations(
            self._upload_id, self._table,
        )
        if not stations:
            self._tabs.setVisible(False)
            self._station = ""
            self._reset_and_load()
            return
        self._tabs.setVisible(True)
        self._tabs.set_stations(stations, active=self._station or self._tabs.active)
        self._station = self._tabs.active
        self._reset_and_load()

    def _on_station(self, slug: str) -> None:
        self._station = slug
        self._reset_and_load()

    async def _load_initial(self) -> None:
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            rows, total, totals = await asyncio.gather(
                recon_svc.get_recon_upload_records(
                    self._upload_id, self._table, self._station, self._search, _SCROLL_CHUNK, 0,
                ),
                recon_svc.count_recon_upload_records(
                    self._upload_id, self._table, self._station, self._search,
                ),
                recon_svc.get_recon_upload_totals(
                    self._upload_id, self._table, self._station,
                ),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(rows)
        self._loaded = len(rows)
        self._totals.set_total("charge", totals["invoiced"], "$ ")
        self._totals.set_total("count", totals["count"], "")
        if self._bonds:
            self._inv_lbl.setText(f"Invoiced: $ {totals['invoiced']:,.0f}")
            self._conf_lbl.setText(f"Confirmed: $ {totals['confirmed']:,.0f}")
            self._var_lbl.setText(f"Variance: $ {totals['variance']:,.0f}")
            self._disp_lbl.setText(f"Disputed: {totals['disputed']}")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            rows = await recon_svc.get_recon_upload_records(
                self._upload_id, self._table, self._station, self._search, _SCROLL_CHUNK, self._loaded,
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if rows:
            self._append_rows(rows)
            self._loaded += len(rows)
        self._loading = False
        self._update_status()

    def _append_rows(self, rows: List[ReconciliationEntry]) -> None:
        t = self._tbl
        for e in rows:
            r = t.rowCount()
            t.insertRow(r)
            for c, (_, field, _w, align, mono) in enumerate(self._cols):
                flag = {"left": Qt.AlignLeft, "right": Qt.AlignRight,
                        "center": Qt.AlignHCenter}[align] | Qt.AlignVCenter
                color = _RED if (field == "charge" and (e.charge or 0) < 0) else ""
                t.setItem(r, c, _cell(_fmt_recon_value(e, field), flag, mono=mono, color=color))
            _finish_table_row(t, r)

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _on_scroll(self, value: int) -> None:
        bar = self._tbl.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _export(self) -> None:
        asyncio.ensure_future(self._do_export())

    async def _do_export(self) -> None:
        if not _HAS_OPENPYXL:
            QMessageBox.critical(self, "Missing Dependency",
                                 "openpyxl is required for Excel export.")
            return
        from openpyxl.styles import Font, PatternFill, Alignment

        rows = await recon_svc.get_recon_upload_records(
            self._upload_id, self._table, self._station,
            self._search, limit=100_000, skip=0,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self._title[:28]
        last_col = len(self._cols)
        col_end = ws.cell(1, last_col).column_letter

        ws.merge_cells(f"A1:{col_end}1")
        ws["A1"] = "TAHMEED COACH TZ LTD"
        ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1B2B4B")
        ws["A1"].alignment = Alignment(horizontal="center")
        sub = self._title + (f" — {self._station.title()}" if self._station else "")
        ws.merge_cells(f"A2:{col_end}2")
        ws["A2"] = sub
        ws["A2"].font = Font(name="Segoe UI", bold=True, size=11, color="374151")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.append([])

        ws.append([c[0] for c in self._cols])
        hdr_row = ws.max_row
        grey = PatternFill("solid", fgColor="F1F5F9")
        for cell in ws[hdr_row]:
            cell.font = Font(name="Segoe UI", bold=True, size=10, color="6B7280")
            cell.fill = grey
            cell.alignment = Alignment(horizontal="center", vertical="center")

        alt = PatternFill("solid", fgColor="F9FAFB")
        charge_total = 0.0
        for i, e in enumerate(rows):
            charge_total += e.charge or 0.0
            ws.append([_fmt_recon_value(e, c[1]).replace("$ ", "") for c in self._cols])
            if i % 2:
                for cell in ws[ws.max_row]:
                    cell.fill = alt

        ws.append([])
        total_row = ["" for _ in self._cols]
        total_row[0] = "TOTAL"
        total_row[-1] = f"$ {charge_total:,.0f}"
        ws.append(total_row)
        ws.cell(ws.max_row, 1).font = Font(name="Segoe UI", bold=True, size=11)
        ws.cell(ws.max_row, last_col).font = Font(name="Cascadia Code", bold=True, size=11)

        for idx, (_, _f, width, _a, _m) in enumerate(self._cols, 1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = max(10, width // 7)
        ws.freeze_panes = ws.cell(hdr_row + 1, 1)

        tag = (self._station or "all").title()
        default = f"SM_Burhani_{self._title.replace(' ', '_')}_{tag}.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Export", default, "Excel Files (*.xlsx)")
        if path:
            try:
                wb.save(path)
                QMessageBox.information(self, "Export Complete",
                                        f"Exported {len(rows):,} records to:\n{path}")
            except Exception as exc:
                QMessageBox.critical(self, "Save Error", f"Could not save file:\n{exc}")


class _ReconShellWidget(QWidget):
    """All Entries + Uploads + Skipped shell for one SM Burhani table."""

    def __init__(
        self,
        table: str,
        title: str,
        icon: str,
        cols,
        import_label: str,
        bonds: bool = False,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._table = table
        self._title = title
        self._icon = icon
        self._cols = cols
        self._import_label = import_label
        self._bonds = bonds
        self._build()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader(self._title, self._icon)
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        self._import_btn = _btn(self._import_label, "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Uploads", "Skipped"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _ReconAllEntries(self._table, self._cols, bonds=self._bonds)
        self._main_stack.addWidget(self._all_entries)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _ReconUploadBrowse(self._table)
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _ReconUploadDetail(
            self._table, self._title, self._cols, bonds=self._bonds,
        )
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        from tahmeed.ui.accountant.skipped_trucks_tab import SkippedTrucksTab
        self._skipped = SkippedTrucksTab(self._table)
        self._main_stack.addWidget(self._skipped)

        self._tabs.tab_changed.connect(self._on_main_tab)
        vl.addWidget(self._main_stack, 1)

    def _on_main_tab(self, idx: int) -> None:
        self._main_stack.setCurrentIndex(idx)
        if idx == 2:
            self._skipped.refresh()

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._show_browse()
        if hasattr(self, "_skipped"):
            self._skipped.refresh()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(1, emit=False)
        self._main_stack.setCurrentIndex(1)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _template_filename(self) -> str:
        if self._table == "rpa_schedule":
            return "SM_Burhani_RPA_Import_Template.xlsx"
        return "SM_Burhani_Bonds_Import_Template.xlsx"

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            self._template_filename(),
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        headers = [c[0] for c in self._cols]
        try:
            _write_xlsx_template(path, self._title, headers)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        dlg = ReconImportDialog(self._table, self._title, self._cols, parent=self)
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        skipped = int(getattr(self.sender(), "_last_skipped", 0) or 0)
        msg = f"Imported {n:,} new records."
        if skipped:
            msg += f"\n{skipped:,} parked in Skipped for follow-up."
        QMessageBox.information(self, "Import Complete", msg)
        self._all_entries.refresh()
        self._show_browse()
        if skipped:
            self._tabs.set_index(2)
            self._skipped.refresh()
        else:
            self._tabs.set_index(1)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        upload_id = str(upload_doc.get("_id") or "")
        if not upload_id:
            return
        filename = upload_doc.get("source_filename") or "Unknown file"
        count = int(upload_doc.get("record_count", 0))
        if show_question(
            self,
            "Delete Upload",
            f"Delete upload from \"{filename}\" and all {count:,} records?\n\nThis cannot be undone.",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(upload_id))

    async def _delete_upload(self, upload_id: str) -> None:
        try:
            deleted = await recon_svc.delete_recon_upload(upload_id, self._table)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        if deleted <= 0:
            QMessageBox.warning(self, "Delete Upload", "No records were deleted.")
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._show_browse()
        if hasattr(self, "_skipped"):
            self._skipped.refresh()


# ═══════════════════════════════════════════════════════════════════════════════
#  RPA Schedule
# ═══════════════════════════════════════════════════════════════════════════════

class RPAScheduleWidget(_ReconShellWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(
            table="rpa_schedule",
            title="SM Burhani — RPA Schedule",
            icon="mdi.file-table-outline",
            cols=_RPA_COLS,
            import_label="Import RPA Schedule",
            parent=parent,
        )
        self.refresh()


class BondsWidget(_ReconShellWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(
            table="bonds",
            title="SM Burhani — Bonds",
            icon="mdi.bank-outline",
            cols=_BONDS_COLS,
            import_label="Import Bonds Schedule",
            bonds=True,
            parent=parent,
        )
        self.refresh()
