"""Fuel Consumption widgets — Infinity, Lake Zambia, Lake Tunduma, GBP Diesel.

Each station owns a fixed column structure (mirroring its sheet in the
reconciliation workbook).  Uploading works like Separate Expenses:

    Import  →  pick the sheet  →  headers are validated against this
    station's structure  →  preview  →  import  →  land on the Uploads
    list (one row per import batch)  →  click a batch to drill into its
    records.

If the chosen sheet's headers don't match the station's structure the
import is blocked and the user is asked to pick the correct sheet.
"""

from __future__ import annotations

import asyncio
import csv
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import qtawesome as qta

from PySide6.QtCore import Qt, QSize, Signal
from PySide6.QtGui import QColor, QDragEnterEvent, QDropEvent, QFont
from PySide6.QtWidgets import (
    QAbstractItemView, QComboBox, QDialog, QFileDialog, QFrame,
    QHBoxLayout, QHeaderView, QLabel, QLineEdit, QMenu, QMessageBox,
    QPushButton, QStackedWidget, QTableWidget, QTableWidgetItem,
    QVBoxLayout, QWidget, QDateEdit,
)

from tahmeed.services.diesel_amounts import apply_diesel_computed_fields, diesel_line_total
from tahmeed.services.excel_dates import format_excel_date, parse_excel_date

try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from tahmeed.ui.accountant.date_filters import (
    add_from_to_editors, read_from_to, sync_from_to, clear_list_filters,
)
from tahmeed.ui.accountant.separate_expenses import (
    _SegmentTabBar, _populate_year_combo, _TOLL_MONTHS, _SCROLL_CHUNK,
    _write_xlsx_template,
)
from tahmeed.ui.widgets.checkable_multi_combo import CheckableMultiCombo
from tahmeed.ui.accountant.fuel_sort_helpers import diesel_columns_sort
from tahmeed.ui.accountant.feed_sort_helpers import (
    wire_feed_table_sort, sort_kw, reset_feed_sort, clear_upload_detail_filters,
)

# ── Design tokens ──────────────────────────────────────────────────────────────
_WHITE   = "#FFFFFF"
_BG      = "#F4F6F8"
_BORDER  = "#E5E7EB"
_BLUE    = "#0077C5"
_BLUE_L  = "#E8F4FD"
_T1      = "#111827"
_T2      = "#6B7280"
_TM      = "#9CA3AF"
_RED     = "#DC2626"
_GREEN   = "#059669"
_HDR_BG  = "#F1F5F9"

# Row styling — mirrors SM Burhani / Bonds: compact rows, manual slate stripe.
_ROW_EVEN = "#FFFFFF"
_ROW_ODD  = "#F1F5F9"   # slate-100

_PAGE_SIZES = [25, 50, 100]
_ROW_H      = 28
_HDR_H      = 26


# ═══════════════════════════════════════════════════════════════════════════════
#  Column schema
#
#  Each station defines an ordered list of columns:
#     (display_label, field_key, kind)      kind ∈ {text, date, num, money}
#  plus the set of field_keys that MUST resolve for a sheet to be accepted.
# ═══════════════════════════════════════════════════════════════════════════════

# Candidate header names per field (case-insensitive, exact match after strip).
_FIELD_CANDIDATES: Dict[str, List[str]] = {
    "sn":            ["s/no.", "s/n", "sn", "s/no", "no.", "#"],
    "date":          ["date"],
    "lpo_no":        ["lpo no.", "lpo no", "lpo"],
    "do_sdo_no":     ["do/sdo no.", "do/sdo no", "do/sdo", "do no.", "do no", "do", "sdo no."],
    "diesel_at":     ["diesel @", "diesel at", "supplier", "station"],
    "ownership":     ["ownership", "owner"],
    "clients_name":  ["clients name", "client name", "client"],
    "destinations":  ["destinations", "destination", "destijnation", "dest"],
    "truck_no":      ["truck no.", "truck no", "truck"],
    "ltrs":          ["ltrs", "litres", "liters", "qty", "quantity"],
    "price_per_ltr": ["price per ltr", "price/ltr", "unit price", "rate"],
    "total_amount":  ["total amount", "total", "amount"],
}

# Column layout + required fields per station (matches the real workbook sheets).
# currency: prefix for money columns / totals (None when the station has no amount).
_FUEL_SCHEMAS: Dict[str, dict] = {
    "diesel_infinity": {
        "title":      "Infinity Diesel",
        "icon":       "mdi.gas-station",
        "sheet_hint": "INFINITY",
        "currency":   "TZS",
        "columns": [
            ("S/NO",        "sn",            "text"),
            ("DATE",        "date",          "date"),
            ("LPO",         "lpo_no",        "text"),
            ("DO",          "do_sdo_no",     "text"),
            ("SUPPLIER",    "diesel_at",     "text"),
            ("OWNERSHIP",   "ownership",     "text"),
            ("DESTINATION", "destinations",  "text"),
            ("TRUCK NO",    "truck_no",      "text"),
            ("LTRS",        "ltrs",          "num"),
            ("RATE",        "price_per_ltr", "num"),
            ("AMOUNT",      "total_amount",  "money"),
        ],
        "required": ["date", "lpo_no", "truck_no", "ltrs", "price_per_ltr"],
    },
    "diesel_lake_zambia": {
        "title":      "Lake Zambia Diesel",
        "icon":       "mdi.water-pump",
        "sheet_hint": "LAKE OIL",
        "currency":   "USD",
        "columns": [
            ("S/NO",         "sn",            "text"),
            ("DATE",         "date",          "date"),
            ("LPO NO.",      "lpo_no",        "text"),
            ("DIESEL @",     "diesel_at",     "text"),
            ("DO/SDO",       "do_sdo_no",     "text"),
            ("TRUCK NO.",    "truck_no",      "text"),
            ("LTRS",         "ltrs",          "num"),
            ("PRICE/LTR",    "price_per_ltr", "num"),
            ("AMOUNT",       "total_amount",  "money"),
            ("DESTINATIONS", "destinations",  "text"),
        ],
        "required": ["date", "lpo_no", "diesel_at", "truck_no", "ltrs",
                     "price_per_ltr"],
    },
    "diesel_gbp": {
        "title":      "GBP Diesel",
        "icon":       "mdi.fuel",
        "sheet_hint": "GBP",
        "currency":   None,
        "columns": [
            ("S/NO",         "sn",            "text"),
            ("DATE",         "date",          "date"),
            ("LPO NO.",      "lpo_no",        "text"),
            ("DO/SDO NO.",   "do_sdo_no",     "text"),
            ("DIESEL @",     "diesel_at",     "text"),
            ("CLIENTS NAME", "clients_name",  "text"),
            ("DESTINATIONS", "destinations",  "text"),
            ("TRUCK NO.",    "truck_no",      "text"),
            ("LTRS",         "ltrs",          "num"),
            ("PRICE/LTR",    "price_per_ltr", "num"),
            ("AMOUNT",       "total_amount",  "money"),
        ],
        "required": ["date", "lpo_no", "truck_no", "ltrs", "price_per_ltr", "clients_name"],
    },
    "diesel_lake_tunduma": {
        "title":      "Lake Tunduma Diesel",
        "icon":       "mdi.water-pump",
        "sheet_hint": "TUNDUMA",
        "currency":   None,
        "columns": [
            ("S/NO",        "sn",            "text"),
            ("DATE",        "date",         "date"),
            ("LPO NO.",     "lpo_no",       "text"),
            ("DO NO.",      "do_sdo_no",    "text"),
            ("STATION",     "diesel_at",    "text"),
            ("DESTINATION", "destinations", "text"),
            ("TRUCK NO.",   "truck_no",     "text"),
            ("LTRS",        "ltrs",         "num"),
            ("RATE",        "price_per_ltr", "num"),
            ("AMOUNT",      "total_amount",  "money"),
        ],
        "required": ["date", "lpo_no", "truck_no", "ltrs", "do_sdo_no"],
    },
}


def _pretty_field(key: str) -> str:
    return {
        "sn": "S/No", "date": "Date", "lpo_no": "LPO", "do_sdo_no": "DO/SDO",
        "diesel_at": "Station/Supplier", "ownership": "Ownership",
        "clients_name": "Client", "destinations": "Destination",
        "truck_no": "Truck No.", "ltrs": "Litres",
        "price_per_ltr": "Price/Ltr", "total_amount": "Amount",
        "upload_label": "File Name",
    }.get(key, key)


_FILE_COL = ("FILE NAME", "upload_label", "text")


def _display_columns(schema: dict) -> List[Tuple[str, str, str]]:
    """Station columns plus the last File Name (upload description) column."""
    return list(schema["columns"]) + [_FILE_COL]


def _row_label(rec: dict) -> str:
    from tahmeed.services.accountant_service import diesel_display_label
    return diesel_display_label(rec)


# ═══════════════════════════════════════════════════════════════════════════════
#  UI helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _lbl(text: str = "", size: int = 13, weight: int = 400,
         color: str = _T1) -> QLabel:
    w = QLabel(text)
    w.setStyleSheet(
        f"color:{color};font-size:{size}px;font-weight:{weight};"
        "font-family:'Segoe UI';background:transparent;"
    )
    return w


def _btn(text: str, icon: str = "", primary: bool = True,
         danger: bool = False, height: int = 34) -> QPushButton:
    b = QPushButton(f"  {text}" if icon else text)
    b.setCursor(Qt.PointingHandCursor)
    b.setFixedHeight(height)
    if icon:
        try:
            b.setIcon(qta.icon(icon, color="#FFFFFF" if (primary or danger) else _T1))
            b.setIconSize(QSize(16, 16))
        except Exception:
            pass
    if danger:
        ss = (f"QPushButton{{background:{_RED};color:#FFF;border:none;border-radius:5px;"
              f"font-size:12px;font-weight:600;font-family:'Segoe UI';padding:0 14px;}}"
              f"QPushButton:hover{{background:#B91C1C;}}"
              f"QPushButton:disabled{{background:#FCA5A5;}}")
    elif primary:
        ss = (f"QPushButton{{background:{_BLUE};color:#FFF;border:none;border-radius:5px;"
              f"font-size:12px;font-weight:600;font-family:'Segoe UI';padding:0 14px;}}"
              f"QPushButton:hover{{background:#005EA3;}}"
              f"QPushButton:disabled{{background:#93C5FD;}}")
    else:
        ss = (f"QPushButton{{background:{_WHITE};color:{_T1};border:1px solid {_BORDER};"
              f"border-radius:5px;font-size:12px;font-family:'Segoe UI';padding:0 14px;}}"
              f"QPushButton:hover{{background:{_BG};}}"
              f"QPushButton:disabled{{color:{_TM};}}")
    b.setStyleSheet(ss)
    return b


def _input_ss() -> str:
    return (
        f"QLineEdit,QComboBox,QDateEdit{{border:1px solid {_BORDER};border-radius:5px;"
        f"background:{_WHITE};color:{_T1};font-size:12px;"
        f"font-family:'Segoe UI';padding:0 8px;"
        f"min-height:32px;max-height:32px;}}"
        f"QLineEdit:focus,QComboBox:focus,QDateEdit:focus{{border-color:{_BLUE};}}"
        "QComboBox::drop-down{border:none;width:20px;}"
    )


def _hsep() -> QFrame:
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setFixedHeight(1)
    f.setStyleSheet(f"background:{_BORDER};")
    return f


def _table_style() -> str:
    return (
        f"QTableWidget{{background:{_WHITE};gridline-color:{_BORDER};"
        "border:none;font-size:11px;font-family:'Segoe UI';}}"
        f"QTableWidget::item{{padding:2px 8px;color:{_T1};}}"
        f"QTableWidget::item:selected{{background:{_BLUE_L};color:{_T1};}}"
        f"QHeaderView::section{{background:{_HDR_BG};color:{_T2};"
        "font-size:10px;font-weight:600;font-family:'Segoe UI';"
        f"border:none;border-bottom:1px solid {_BORDER};"
        f"padding:0 8px;height:{_HDR_H}px;}}"
        "QScrollBar:vertical{width:8px;background:transparent;}"
        f"QScrollBar::handle:vertical{{background:#D1D5DB;border-radius:4px;}}"
    )


def _make_table(headers: List[str]) -> QTableWidget:
    t = QTableWidget(0, len(headers))
    t.setHorizontalHeaderLabels(headers)
    t.setEditTriggers(QAbstractItemView.NoEditTriggers)
    t.setSelectionBehavior(QAbstractItemView.SelectRows)
    t.setAlternatingRowColors(False)
    t.verticalHeader().setVisible(False)
    t.verticalHeader().setDefaultSectionSize(_ROW_H)
    t.horizontalHeader().setStretchLastSection(True)
    t.setStyleSheet(_table_style())
    return t


def _finish_table_row(t: QTableWidget, row: int, bg: str | None = None) -> None:
    """Apply the Bonds-style manual stripe + compact row height."""
    color = bg if bg else (_ROW_ODD if row % 2 else _ROW_EVEN)
    for col in range(t.columnCount()):
        item = t.item(row, col)
        if item is not None:
            item.setBackground(QColor(color))
    t.setRowHeight(row, _ROW_H)


def _cell(text: str, align: Qt.AlignmentFlag = Qt.AlignLeft | Qt.AlignVCenter,
          mono: bool = False, color: str = "") -> QTableWidgetItem:
    item = QTableWidgetItem(str(text) if text is not None else "—")
    item.setTextAlignment(align)
    if mono:
        item.setFont(QFont("Cascadia Code", 11))
    if color:
        item.setForeground(QColor(color))
    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
    return item


def _fmt_num(v: Any, prefix: str = "", decimals: int = 2) -> str:
    if v is None or v == "" or v == "None":
        return "—"
    try:
        return f"{prefix}{float(v):,.{decimals}f}"
    except Exception:
        return str(v)


def _fmt_date_str(val: Any) -> str:
    text = format_excel_date(val, "%d %b %y", fallback="")
    if text:
        return text
    if val is None or str(val) in ("None", ""):
        return "—"
    return str(val).strip() or "—"


def _looks_like_date(val: Any) -> bool:
    if val in (None, ""):
        return False
    return parse_excel_date(val) is not None


def _fmt_cell(
    kind: str,
    value: Any,
    currency: str | None = "TZS",
) -> Tuple[str, Qt.AlignmentFlag, bool]:
    """Return (text, alignment, mono) for a value given its column kind.

    Amounts use the same Segoe UI table font as Separate Expenses (no mono).
    """
    if kind == "date":
        return _fmt_date_str(value), Qt.AlignLeft | Qt.AlignVCenter, False
    if kind == "num":
        return _fmt_num(value, decimals=2), Qt.AlignRight | Qt.AlignVCenter, False
    if kind == "money":
        prefix = f"{currency} " if currency else ""
        return _fmt_num(value, prefix, 0), Qt.AlignRight | Qt.AlignVCenter, False
    return (str(value) if value not in (None, "") else "—",
            Qt.AlignLeft | Qt.AlignVCenter, False)


def _fill_diesel_row(
    table: QTableWidget,
    row: int,
    rec: dict,
    columns: List[Tuple[str, str, str]],
    sn_offset: int = 0,
    currency: str | None = "TZS",
) -> None:
    for c, (_, key, kind) in enumerate(columns):
        if key == "sn":
            val = sn_offset + 1
        elif key == "total_amount":
            val = diesel_line_total(rec.get("ltrs"), rec.get("price_per_ltr"))
        elif key == "upload_label":
            val = _row_label(rec)
        else:
            val = rec.get(key, "")
        text, align, mono = _fmt_cell(kind, val, currency=currency)
        table.setItem(row, c, _cell(text, align, mono=mono))
    _finish_table_row(table, row)


# ═══════════════════════════════════════════════════════════════════════════════
#  Page header
# ═══════════════════════════════════════════════════════════════════════════════

class _PageHeader(QWidget):
    def __init__(self, title: str, icon_name: str = "mdi.gas-station",
                 parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setStyleSheet("background:transparent;")
        hl = QHBoxLayout(self)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(10)
        try:
            icon_lbl = QLabel()
            icon_lbl.setPixmap(qta.icon(icon_name, color=_BLUE).pixmap(22, 22))
            icon_lbl.setFixedSize(22, 22)
            icon_lbl.setStyleSheet("background:transparent;")
            hl.addWidget(icon_lbl)
        except Exception:
            pass
        hl.addWidget(_lbl(title, size=18, weight=700))
        hl.addStretch()
        self._right = hl

    def add_right(self, widget: QWidget) -> None:
        self._right.addWidget(widget)


# ═══════════════════════════════════════════════════════════════════════════════
#  Pagination bar
# ═══════════════════════════════════════════════════════════════════════════════

class _PaginationBar(QWidget):
    page_changed = Signal(int)
    size_changed = Signal(int)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setStyleSheet("background:transparent;")
        self._page = 1
        self._total = 0
        self._page_size = 25

        hl = QHBoxLayout(self)
        hl.setContentsMargins(0, 4, 0, 4)
        hl.setSpacing(8)

        self._size_cb = QComboBox()
        self._size_cb.setFixedWidth(80)
        for s in _PAGE_SIZES:
            self._size_cb.addItem(f"Show {s}", s)
        self._size_cb.setStyleSheet(_input_ss())
        self._size_cb.currentIndexChanged.connect(self._on_size_change)
        hl.addWidget(self._size_cb)

        self._info_lbl = _lbl("", size=11, color=_T2)
        hl.addWidget(self._info_lbl)
        hl.addStretch()

        self._prev_btn = _btn("← Prev", primary=False, height=30)
        self._prev_btn.clicked.connect(self._prev)
        hl.addWidget(self._prev_btn)

        self._page_lbl = _lbl("Page 1 of 1", size=11, color=_T2)
        hl.addWidget(self._page_lbl)

        self._next_btn = _btn("Next →", primary=False, height=30)
        self._next_btn.clicked.connect(self._next)
        hl.addWidget(self._next_btn)

    def set_total(self, total: int, page_size: int, current_page: int) -> None:
        self._total = total
        self._page_size = page_size
        self._page = current_page
        pages = max(1, (total + page_size - 1) // page_size)
        self._page_lbl.setText(f"Page {current_page} of {pages}")
        self._info_lbl.setText(f"{total:,} records")
        self._prev_btn.setEnabled(current_page > 1)
        self._next_btn.setEnabled(current_page < pages)

    def _prev(self) -> None:
        if self._page > 1:
            self._page -= 1
            self.page_changed.emit(self._page)

    def _next(self) -> None:
        pages = max(1, (self._total + self._page_size - 1) // self._page_size)
        if self._page < pages:
            self._page += 1
            self.page_changed.emit(self._page)

    def _on_size_change(self) -> None:
        self._page = 1
        self._page_size = self._size_cb.currentData()
        self.size_changed.emit(self._page_size)


# ═══════════════════════════════════════════════════════════════════════════════
#  Totals bar
# ═══════════════════════════════════════════════════════════════════════════════

class _TotalsBar(QFrame):
    def __init__(self, labels: List[Tuple[str, str]],
                 parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setFixedHeight(38)
        self.setStyleSheet(
            f"QFrame{{background:{_HDR_BG};border-top:1px solid {_BORDER};"
            "border-bottom-left-radius:6px;border-bottom-right-radius:6px;}}"
        )
        hl = QHBoxLayout(self)
        hl.setContentsMargins(16, 0, 16, 0)
        hl.setSpacing(32)

        self._lbl_map: Dict[str, QLabel] = {}
        for key, prefix in labels:
            sub = QWidget()
            sub.setStyleSheet("background:transparent;")
            sub_hl = QHBoxLayout(sub)
            sub_hl.setContentsMargins(0, 0, 0, 0)
            sub_hl.setSpacing(4)
            sub_hl.addWidget(_lbl(prefix, size=11, weight=600, color=_T2))
            val = _lbl("—", size=12, weight=700, color=_T1)
            sub_hl.addWidget(val)
            self._lbl_map[key] = val
            hl.addWidget(sub)
        hl.addStretch()

    def set_total(self, key: str, value: float, prefix: str = "") -> None:
        if key in self._lbl_map:
            self._lbl_map[key].setText(_fmt_num(value, prefix=prefix, decimals=0))


# ═══════════════════════════════════════════════════════════════════════════════
#  Drop zone
# ═══════════════════════════════════════════════════════════════════════════════

class _DropZone(QFrame):
    file_dropped = Signal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setMinimumHeight(90)
        self._normal_ss = (
            f"QFrame{{border:2px dashed {_BORDER};border-radius:8px;background:{_BG};}}"
            f"QFrame:hover{{border-color:{_BLUE};background:{_BLUE_L};}}"
        )
        self._hover_ss = (
            f"QFrame{{border:2px dashed {_BLUE};border-radius:8px;background:{_BLUE_L};}}"
        )
        self.setStyleSheet(self._normal_ss)

        vl = QVBoxLayout(self)
        vl.setAlignment(Qt.AlignCenter)
        vl.setSpacing(4)
        try:
            ic = QLabel()
            ic.setPixmap(qta.icon("mdi.upload-outline", color=_TM).pixmap(28, 28))
            ic.setAlignment(Qt.AlignCenter)
            ic.setStyleSheet("background:transparent;border:none;")
            vl.addWidget(ic)
        except Exception:
            pass
        vl.addWidget(_lbl("Drag & drop .xlsx or .csv here", size=12, color=_T2))
        self._path_lbl = _lbl("", size=11, color=_TM)
        self._path_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._path_lbl)

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(self._hover_ss)

    def dragLeaveEvent(self, event) -> None:
        self.setStyleSheet(self._normal_ss)

    def dropEvent(self, event: QDropEvent) -> None:
        urls = event.mimeData().urls()
        if urls:
            path = urls[0].toLocalFile()
            self._path_lbl.setText(Path(path).name)
            self.file_dropped.emit(path)
        self.setStyleSheet(self._normal_ss)

    def set_path(self, path: str) -> None:
        self._path_lbl.setText(Path(path).name)


# ═══════════════════════════════════════════════════════════════════════════════
#  Fuel Import Dialog  (sheet picker + structure validation + preview)
# ═══════════════════════════════════════════════════════════════════════════════

class _FuelImportDialog(QDialog):
    """Import diesel feed data.

    Flow: choose file → pick the sheet → the sheet's headers are validated
    against this station's required columns.  A mismatched sheet blocks the
    import and prompts the user to pick the correct one.  Matching sheets
    are previewed and imported as a single tagged upload batch.
    """

    imported = Signal(int)

    _PREVIEW_ROWS = 12

    def __init__(self, feed_type: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._feed_type = feed_type
        self._schema    = _FUEL_SCHEMAS[feed_type]
        self._columns   = self._schema["columns"]
        self._required  = self._schema["required"]
        self._currency  = self._schema.get("currency")
        self._headers   = [c[0] for c in self._columns]

        self._wb: Any = None
        self._source_filename = ""
        self._sheet_label = ""
        self._rows: List[dict] = []          # rows that will be imported
        self._already_exists = False
        self._check_gen = 0

        self.setWindowTitle(f"Import — {self._schema['title']}")
        self.setMinimumWidth(760)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()

    # ── Build ────────────────────────────────────────────────────────────────
    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        # Expected structure hint
        expect = ", ".join(_pretty_field(k) for k in self._required)
        hint = _lbl(
            f"This sheet must contain: {expect}. "
            "S/No is assigned by the system. Amount is litres × rate "
            "(Excel totals are ignored).",
            size=11, color=_T2,
        )
        hint.setWordWrap(True)
        vl.addWidget(hint)

        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        # Browse + sheet selector
        ctrl = QWidget()
        ctrl.setStyleSheet("background:transparent;")
        ctrl_hl = QHBoxLayout(ctrl)
        ctrl_hl.setContentsMargins(0, 0, 0, 0)
        ctrl_hl.setSpacing(8)

        browse = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse.clicked.connect(self._browse)
        ctrl_hl.addWidget(browse)

        ctrl_hl.addSpacing(12)
        ctrl_hl.addWidget(_lbl("Sheet:", size=12, color=_T2))

        self._sheet_cb = QComboBox()
        self._sheet_cb.setFixedWidth(240)
        self._sheet_cb.setStyleSheet(_input_ss())
        self._sheet_cb.setEnabled(False)
        self._sheet_cb.currentIndexChanged.connect(self._on_sheet_changed)
        ctrl_hl.addWidget(self._sheet_cb)
        ctrl_hl.addStretch()
        vl.addWidget(ctrl)

        desc_row = QWidget()
        desc_row.setStyleSheet("background:transparent;")
        desc_hl = QHBoxLayout(desc_row)
        desc_hl.setContentsMargins(0, 0, 0, 0)
        desc_hl.setSpacing(8)
        desc_hl.addWidget(_lbl("Description:", size=12, color=_T2))
        self._desc_edit = QLineEdit()
        self._desc_edit.setPlaceholderText("e.g. 16th - 31st Mar 2026")
        self._desc_edit.setStyleSheet(_input_ss())
        desc_hl.addWidget(self._desc_edit, 1)
        vl.addWidget(desc_row)

        self._stats_lbl = _lbl("No file loaded.", size=12, color=_T2)
        self._stats_lbl.setWordWrap(True)
        vl.addWidget(self._stats_lbl)

        vl.addWidget(_hsep())
        vl.addWidget(_lbl("Preview", size=12, weight=600))

        self._preview_tbl = _make_table(self._headers)
        self._preview_tbl.setMinimumHeight(200)
        vl.addWidget(self._preview_tbl)

        vl.addWidget(_hsep())

        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.setSpacing(8)
        bbl.addStretch()

        cancel = _btn("Cancel", primary=False)
        cancel.clicked.connect(self.reject)
        bbl.addWidget(cancel)

        self._import_btn = _btn("Import Records", "mdi.check-circle-outline")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)
        vl.addWidget(btn_row)

    # ── File handling ──────────────────────────────────────────────────────────
    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open File", "", "Excel / CSV (*.xlsx *.xls *.csv)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _on_file(self, path: str) -> None:
        from tahmeed.ui.widgets.upload_busy import UploadBusy

        self._bump_check_gen()
        self._already_exists = False
        self._source_filename = Path(path).name
        self._sheet_label = ""
        if hasattr(self, "_desc_edit"):
            self._desc_edit.setText("")
            self._desc_edit.setPlaceholderText(
                self._source_filename or "e.g. 16th - 31st Mar 2026"
            )
        self._stats_lbl.setText("Reading file…")
        self._sheet_cb.blockSignals(True)
        self._sheet_cb.clear()
        self._sheet_cb.setEnabled(False)
        self._sheet_cb.blockSignals(False)
        self._set_import_enabled(False)
        self._preview_tbl.setRowCount(0)

        p = Path(path)
        if p.suffix.lower() in (".xlsx", ".xls") and _HAS_OPENPYXL:
            try:
                with UploadBusy(self, f"Reading {p.name}…", title="Import"):
                    self._wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            except Exception as exc:
                self._stats_lbl.setText(f"Error opening file: {exc}")
                return

            names = self._wb.sheetnames
            self._sheet_cb.blockSignals(True)
            # No name-based auto-detection: sheets may be named "Sheet1",
            # "Sheet8", etc. The user must actively choose which one to import.
            self._sheet_cb.addItem("— Select a sheet —", None)
            for name in names:
                self._sheet_cb.addItem(name, name)
            self._sheet_cb.setCurrentIndex(0)
            self._sheet_cb.setEnabled(True)
            self._sheet_cb.blockSignals(False)
            self._reset_preview(
                f"{len(names)} sheet(s) found. Select a sheet to preview it."
            )
        else:
            self._wb = None
            try:
                with UploadBusy(self, f"Reading {p.name}…", title="Import"):
                    with open(path, newline="", encoding="utf-8-sig") as f:
                        rows = list(csv.reader(f))
            except Exception as exc:
                self._stats_lbl.setText(f"Error reading CSV: {exc}")
                return
            if not rows:
                self._stats_lbl.setText("File is empty.")
                return
            headers = [str(c) if c is not None else "" for c in rows[0]]
            data = [
                [str(c) if c is not None else "" for c in r]
                for r in rows[1:] if any(str(c).strip() for c in r)
            ]
            self._process(headers, data, sheet_label="")

    def _on_sheet_changed(self, _idx: int = 0) -> None:
        if self._wb is None:
            return
        name = self._sheet_cb.currentData()
        if not name:   # placeholder "— Select a sheet —"
            self._reset_preview("Select a sheet to preview it.")
            return
        try:
            ws = self._wb[name]
        except Exception:
            return
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self._show_mismatch(f"Sheet “{name}” is empty.")
            return
        headers = [str(c) if c is not None else "" for c in rows[0]]
        data = [
            [str(c) if c is not None else "" for c in r]
            for r in rows[1:] if any(c is not None and str(c).strip() for c in r)
        ]
        self._process(headers, data, sheet_label=name)

    # ── Mapping + validation ─────────────────────────────────────────────────
    def _resolve_indices(self, headers: List[str]) -> Dict[str, Optional[int]]:
        hdr_lower = {h.strip().lower(): i for i, h in enumerate(headers)}

        def _find(cands: List[str]) -> Optional[int]:
            for c in cands:
                idx = hdr_lower.get(c.lower())
                if idx is not None:
                    return idx
            return None

        return {
            key: _find(_FIELD_CANDIDATES.get(key, [key]))
            for _, key, _ in self._columns
        }

    def _process(self, headers: List[str], rows: List[List[str]],
                 sheet_label: str) -> None:
        idxs = self._resolve_indices(headers)

        missing = [k for k in self._required if idxs.get(k) is None]
        if missing:
            names = ", ".join(_pretty_field(k) for k in missing)
            where = f" in “{sheet_label}”" if sheet_label else ""
            self._show_mismatch(
                f"This sheet{where} doesn't match {self._schema['title']}. "
                f"Missing column(s): {names}. Please choose the correct sheet."
            )
            return

        records: List[dict] = []
        for row in rows:
            rec: dict = {"feed_type": self._feed_type}
            for _, key, _kind in self._columns:
                idx = idxs.get(key)
                rec[key] = (
                    str(row[idx]).strip()
                    if idx is not None and idx < len(row) and row[idx] is not None
                    else ""
                )
            self._normalize_row(rec)
            apply_diesel_computed_fields(rec)
            records.append(rec)

        self._sheet_label = sheet_label
        if not records:
            self._reset_preview("This sheet has no data rows to import.")
            return

        gen = self._bump_check_gen()
        self._already_exists = False
        self._rows = records
        self._fill_preview(records[: self._PREVIEW_ROWS])
        self._stats_lbl.setStyleSheet(
            f"color:{_T2};font-size:12px;font-family:'Segoe UI';background:transparent;"
        )
        where = f"“{sheet_label}” · " if sheet_label else ""
        self._stats_lbl.setText(
            f"{where}{len(records):,} rows. Checking for duplicates…"
        )
        self._set_import_enabled(False)
        self._import_btn.setText("Checking…")
        asyncio.ensure_future(self._check_already_uploaded(gen, records, sheet_label))

    def _bump_check_gen(self) -> int:
        self._check_gen += 1
        return self._check_gen

    def _reset_preview(self, message: str) -> None:
        """Neutral state: clear preview + disable import, show an info message."""
        self._bump_check_gen()
        self._already_exists = False
        self._rows = []
        self._preview_tbl.setRowCount(0)
        self._stats_lbl.setStyleSheet(
            f"color:{_T2};font-size:12px;font-family:'Segoe UI';background:transparent;"
        )
        self._stats_lbl.setText(message)
        self._set_import_enabled(False)
        self._import_btn.setText("Import Records")

    @staticmethod
    def _normalize_row(rec: dict) -> None:
        """Recover a leading S/No column that the header row doesn't declare.

        Some sheets stack rows that carry an extra leading S/No (and drop the
        DO column) beneath the same header, shifting DATE/LPO/DO one column to
        the right.  Detect it per row: when the DATE cell isn't a date but the
        LPO cell is, treat the first value as S/No and shift back.
        """
        if _looks_like_date(rec.get("date")) or not _looks_like_date(rec.get("lpo_no")):
            return
        rec["sn"] = rec.get("date", "")
        rec["date"] = rec.get("lpo_no", "")
        rec["lpo_no"] = rec.get("do_sdo_no", "")
        rec["do_sdo_no"] = ""

    def _show_mismatch(self, message: str) -> None:
        self._bump_check_gen()
        self._already_exists = False
        self._rows = []
        self._preview_tbl.setRowCount(0)
        self._stats_lbl.setStyleSheet(
            f"color:{_RED};font-size:12px;font-weight:600;"
            "font-family:'Segoe UI';background:transparent;"
        )
        self._stats_lbl.setText(message)
        self._set_import_enabled(False)
        self._import_btn.setText("Import Records")

    def _show_already_uploaded(self, message: str) -> None:
        self._already_exists = True
        self._stats_lbl.setStyleSheet(
            f"color:{_RED};font-size:12px;font-weight:600;"
            "font-family:'Segoe UI';background:transparent;"
        )
        self._stats_lbl.setText(message)
        self._set_import_enabled(False)
        self._import_btn.setText("Already Uploaded")

    async def _check_already_uploaded(
        self, gen: int, records: List[dict], sheet_label: str,
    ) -> None:
        from tahmeed.services import accountant_service as svc

        matching: List[dict] = []
        try:
            matching = await svc.diesel_already_uploaded(
                self._feed_type,
                records,
                source_filename=self._source_filename,
                sheet_label=sheet_label,
            )
        except Exception:
            matching = []
        if gen != self._check_gen:
            return
        if matching:
            names = ", ".join(
                (u.get("source_filename") or u.get("sheet_label") or "prior upload")
                for u in matching[:3]
            )
            extra = "…" if len(matching) > 3 else ""
            where = f"“{sheet_label}” · " if sheet_label else ""
            self._show_already_uploaded(
                f"{where}This upload was already imported ({names}{extra}) "
                "— import blocked."
            )
            return
        self._already_exists = False
        self._stats_lbl.setStyleSheet(
            f"color:{_GREEN};font-size:12px;font-family:'Segoe UI';background:transparent;"
        )
        where = f"“{sheet_label}” · " if sheet_label else ""
        self._stats_lbl.setText(
            f"{where}{len(records):,} rows ready to import."
        )
        self._set_import_enabled(True)
        self._import_btn.setText(f"Import {len(records):,} Records")

    def _set_import_enabled(self, ok: bool) -> None:
        self._import_btn.setEnabled(ok)

    def _resolved_upload_label(self) -> str:
        typed = ""
        if hasattr(self, "_desc_edit"):
            typed = self._desc_edit.text().strip()
        return typed or self._source_filename

    def _fill_preview(self, rows: List[dict]) -> None:
        t = self._preview_tbl
        t.setRowCount(0)
        for i, row in enumerate(rows):
            r = t.rowCount()
            t.insertRow(r)
            _fill_diesel_row(
                t, r, row, self._columns, sn_offset=i, currency=self._currency,
            )

    # ── Import ─────────────────────────────────────────────────────────────────
    def _do_import(self) -> None:
        if self._already_exists or not self._rows:
            return
        self._set_import_enabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        from tahmeed.ui.accountant.import_truck_gate import run_import_truck_gate

        upload_id = str(uuid.uuid4())
        sheet_label = self._sheet_label or (
            (self._sheet_cb.currentData() or "") if self._wb is not None else ""
        )
        upload_label = self._resolved_upload_label()
        docs = []
        for rec in self._rows:
            doc = dict(rec)
            doc["upload_id"] = upload_id
            doc["source_filename"] = self._source_filename
            doc["upload_label"] = upload_label
            doc["sheet_label"] = sheet_label
            docs.append(doc)
        try:
            matching = await svc.diesel_already_uploaded(
                self._feed_type,
                self._rows,
                source_filename=self._source_filename,
                sheet_label=sheet_label,
            )
            if matching:
                names = ", ".join(
                    (u.get("source_filename") or u.get("sheet_label") or "prior upload")
                    for u in matching[:3]
                )
                extra = "…" if len(matching) > 3 else ""
                QMessageBox.warning(
                    self,
                    "Already Uploaded",
                    f"This upload was already imported ({names}{extra}) — import blocked.",
                )
                self._show_already_uploaded(
                    f"This upload was already imported ({names}{extra}) — import blocked."
                )
                return

            file_hash = svc.diesel_batch_content_hash(self._rows)
            for row in docs:
                row["content_hash"] = file_hash

            gate = await run_import_truck_gate(
                self,
                docs,
                feed_key=self._feed_type,
                upload_id=upload_id,
                source_filename=self._source_filename,
                sheet_label=sheet_label,
                can_add=True,
            )
            if gate.aborted:
                self._set_import_enabled(True)
                self._import_btn.setText(f"Import {len(self._rows):,} Records")
                return
            self._last_skipped = gate.skipped_count
            if not gate.rows:
                if gate.skipped_count:
                    QMessageBox.information(
                        self,
                        "Import",
                        f"No rows imported. {gate.skipped_count:,} parked in Skipped.",
                    )
                    self.imported.emit(0)
                    self.accept()
                else:
                    self._set_import_enabled(True)
                    self._import_btn.setText(f"Import {len(self._rows):,} Records")
                return
            saved = await svc.save_imported_feed(gate.rows)
            self.imported.emit(saved)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            self._set_import_enabled(True)
            self._import_btn.setText(f"Import {len(self._rows):,} Records")


# ═══════════════════════════════════════════════════════════════════════════════
#  All entries — cross-upload flat list
# ═══════════════════════════════════════════════════════════════════════════════

class _DieselAllEntries(QWidget):
    """Flat, filterable list of every diesel record with infinite scroll."""

    def __init__(self, feed_type: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._feed_type = feed_type
        self._schema = _FUEL_SCHEMAS[feed_type]
        self._columns = _display_columns(self._schema)
        self._currency = self._schema.get("currency")
        self._has_amount = any(k == "total_amount" for _, k, _ in self._columns)
        self._search = ""
        self._year = 0
        self._month = 0
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText(
            "Search truck, LPO, client, destination, station, file name…"
        )
        self._search_edit.setFixedWidth(260)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._file_cb = CheckableMultiCombo(
            "All File Names", noun_plural="file names", parent=self,
        )
        self._file_cb.setFixedWidth(180)
        self._file_cb.setStyleSheet(_input_ss())
        self._file_cb.selectionChanged.connect(self._reset_and_load)
        tbl.addWidget(self._file_cb)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        clear_btn = _btn("Clear", "mdi.filter-remove-outline", primary=False)
        clear_btn.setToolTip("Clear search, year, date, file filters, and column sort.")
        clear_btn.clicked.connect(self._clear_filters)
        tbl.addWidget(clear_btn)

        tbl.addStretch()
        vl.addWidget(tb)

        totals_defs = [("ltrs", "Ltrs: ")]
        if self._has_amount:
            amt_prefix = f"{self._currency}: " if self._currency else "Amount: "
            totals_defs.append(("amount", amt_prefix))
        totals_defs.append(("count", "Records: "))
        self._totals = _TotalsBar(totals_defs)
        vl.addWidget(self._totals)

        self._table = _make_table([c[0] for c in self._columns])
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        self._sort_state = wire_feed_table_sort(
            self._table,
            diesel_columns_sort(self._columns),
            default_field="transaction_date",
            default_asc=False,
            on_sort_changed=self._on_sort_changed,
        )
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        years_r, labels_r = await asyncio.gather(
            svc.get_diesel_available_years(self._feed_type),
            svc.get_diesel_file_labels(self._feed_type),
            return_exceptions=True,
        )
        years = years_r if isinstance(years_r, list) else []
        labels = labels_r if isinstance(labels_r, list) else []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        self._file_cb.set_options(labels, keep_selected=True, emit=False)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _on_sort_changed(self, field: str, asc: bool) -> None:
        self._reset_and_load()

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_diesel_all_totals(
                    self._feed_type, self._search, self._year, month, **self._list_kw(),
                ),
                svc.get_diesel_all_records(
                    self._feed_type, self._search, self._year, month,
                    limit=_SCROLL_CHUNK, skip=0, **self._list_kw(), **sort_kw(self._sort_state),
                ),
                svc.count_diesel_all_records(
                    self._feed_type, self._search, self._year, month, **self._list_kw(),
                ),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return

        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("count", int(totals.get("count", 0) or 0))
        self._totals.set_total("ltrs", float(totals.get("ltrs", 0) or 0))
        if self._has_amount:
            self._totals.set_total("amount", float(totals.get("total_amount", 0) or 0))
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_diesel_all_records(
                self._feed_type, self._search, self._year, month,
                limit=_SCROLL_CHUNK, skip=self._loaded, **self._list_kw(), **sort_kw(self._sort_state),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for i, rec in enumerate(recs):
            r = self._table.rowCount()
            self._table.insertRow(r)
            _fill_diesel_row(
                self._table, r, rec, self._columns, self._loaded + i,
                currency=self._currency,
            )

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _list_kw(self) -> dict:
        kw = self._date_kw()
        labels = self._file_cb.selected_values() if hasattr(self, "_file_cb") else []
        if labels:
            kw["file_labels"] = labels
        return kw

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _clear_filters(self) -> None:
        self._search = ""
        if hasattr(self, "_file_cb"):
            self._file_cb.reset_to_all(emit=False)
        self._year, self._month = clear_list_filters(
            search_edit=self._search_edit,
            year_cb=self._year_cb,
            month_cb=self._month_cb,
            from_edit=self._from_date,
            to_edit=self._to_date,
        )
        reset_feed_sort(self._sort_state)
        self._reset_and_load()


# ═══════════════════════════════════════════════════════════════════════════════
#  Upload browse — one row per import batch
# ═══════════════════════════════════════════════════════════════════════════════

def _browse_headers(currency: str | None) -> List[str]:
    total_hdr = f"TOTAL ({currency})" if currency else "TOTAL"
    return ["UPLOAD DATE", "SHEET", "FILE NAME", "RECORDS", "LTRS", total_hdr]


class _DieselUploadBrowse(QWidget):
    upload_clicked  = Signal(object)
    delete_clicked  = Signal(object)

    def __init__(self, feed_type: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._feed_type = feed_type
        self._currency = _FUEL_SCHEMAS[feed_type].get("currency")
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        self._table = _make_table(_browse_headers(self._currency))
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 160)
        self._table.setColumnWidth(1, 160)
        self._table.setColumnWidth(2, 220)
        self._table.setColumnWidth(3, 90)
        self._table.setColumnWidth(4, 110)
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        amt_prefix = f"{self._currency}: " if self._currency else "Total: "
        self._totals = _TotalsBar([("count", "Total records: "),
                                   ("ltrs", "Ltrs: "), ("amount", amt_prefix)])
        vl.addWidget(self._totals)

        hint = _lbl("Click a row to view its records · right-click to delete an upload.",
                    size=11, color=_TM)
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            uploads = await svc.get_diesel_uploads(self._feed_type)
        except Exception:
            self._uploads = []
            self._table.setRowCount(0)
            self._totals.set_total("count", 0)
            return
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        tot_recs = 0
        tot_ltrs = 0.0
        tot_amt = 0.0
        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)
            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count = int(up.get("record_count", 0))
            ltrs  = float(up.get("ltrs", 0) or 0)
            amt   = float(up.get("total_amount", 0) or 0)
            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("sheet_label") or "—"))
            t.setItem(r, 2, _cell(
                up.get("upload_label") or up.get("source_filename") or "Unknown"
            ))
            t.setItem(r, 3, _cell(f"{count:,}", Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 4, _cell(_fmt_num(ltrs, decimals=0),
                                  Qt.AlignRight | Qt.AlignVCenter))
            t.setItem(r, 5, _cell(_fmt_num(amt, decimals=0),
                                  Qt.AlignRight | Qt.AlignVCenter))
            _finish_table_row(t, r)
            tot_recs += count
            tot_ltrs += ltrs
            tot_amt  += amt
        self._totals.set_total("count", tot_recs)
        self._totals.set_total("ltrs", tot_ltrs)
        self._totals.set_total("amount", tot_amt)

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if 0 <= row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


# ═══════════════════════════════════════════════════════════════════════════════
#  Upload detail — records within one batch
# ═══════════════════════════════════════════════════════════════════════════════

class _DieselUploadDetail(QWidget):
    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, feed_type: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._feed_type = feed_type
        self._schema    = _FUEL_SCHEMAS[feed_type]
        self._columns   = _display_columns(self._schema)
        self._currency  = self._schema.get("currency")
        self._has_amount = any(k == "total_amount" for _, k, _ in self._columns)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._search    = ""
        self._loaded    = 0
        self._total     = 0
        self._loading   = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)
        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)
        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()
        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search truck, LPO, client, destination…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        clear_btn = _btn("Clear", "mdi.filter-remove-outline", primary=False)
        clear_btn.clicked.connect(lambda: clear_upload_detail_filters(self))
        tbl.addWidget(clear_btn)
        tbl.addStretch()
        vl.addWidget(tb)

        self._table = _make_table([c[0] for c in self._columns])
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        self._sort_state = wire_feed_table_sort(
            self._table,
            diesel_columns_sort(self._columns),
            default_field="transaction_date",
            default_asc=False,
            on_sort_changed=self._on_sort_changed,
        )
        vl.addWidget(self._table, 1)

        totals_defs = [("ltrs", "Ltrs: ")]
        if self._has_amount:
            amt_prefix = f"{self._currency}: " if self._currency else "Amount: "
            totals_defs.append(("amount", amt_prefix))
        totals_defs.append(("count", "Records: "))
        self._totals = _TotalsBar(totals_defs)
        vl.addWidget(self._totals)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        label     = (
            upload_doc.get("upload_label")
            or upload_doc.get("source_filename")
            or "Unknown file"
        )
        sheet     = upload_doc.get("sheet_label") or "—"
        count     = int(upload_doc.get("record_count", 0))
        import_dt = upload_doc.get("import_date")
        date_str  = import_dt.strftime("%d %b %Y") if isinstance(import_dt, datetime) else ""
        self._crumb_lbl.setText(f"Uploads  ›  {sheet}")
        self._info_lbl.setText(
            f"{label}   •   sheet: {sheet}   •   {count:,} records   •   {date_str}"
        )
        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._reset_and_load()

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _on_sort_changed(self, field: str, asc: bool) -> None:
        self._reset_and_load()

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_initial(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            recs, total, totals = await asyncio.gather(
                svc.get_diesel_upload_records(
                    self._feed_type, self._upload_id, self._search, _SCROLL_CHUNK, 0,
                    **sort_kw(self._sort_state),
                ),
                svc.count_diesel_upload_records(
                    self._feed_type, self._upload_id, self._search,
                ),
                svc.get_diesel_upload_totals(
                    self._feed_type, self._upload_id, self._search,
                ),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("count", total)
        self._totals.set_total("ltrs", float(totals.get("ltrs", 0) or 0))
        if self._has_amount:
            self._totals.set_total("amount", float(totals.get("total_amount", 0) or 0))
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            recs = await svc.get_diesel_upload_records(
                self._feed_type, self._upload_id, self._search,
                _SCROLL_CHUNK, self._loaded, **sort_kw(self._sort_state),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for i, rec in enumerate(recs):
            r = self._table.rowCount()
            self._table.insertRow(r)
            _fill_diesel_row(
                self._table, r, rec, self._columns, self._loaded + i,
                currency=self._currency,
            )

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())


# ═══════════════════════════════════════════════════════════════════════════════
#  Base station widget — browse ↔ detail shell
# ═══════════════════════════════════════════════════════════════════════════════

class _BaseDieselWidget(QWidget):
    """All Entries + Uploads shell shared by all four fuel stations."""

    _FEED_TYPE: str = ""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._schema = _FUEL_SCHEMAS[self._FEED_TYPE]
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader(self._schema["title"], self._schema["icon"])
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        self._import_btn = _btn("Import from Excel", "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Uploads", "Skipped"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _DieselAllEntries(self._FEED_TYPE)
        self._main_stack.addWidget(self._all_entries)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _DieselUploadBrowse(self._FEED_TYPE)
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _DieselUploadDetail(self._FEED_TYPE)
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        from tahmeed.ui.accountant.skipped_trucks_tab import SkippedTrucksTab
        self._skipped = SkippedTrucksTab(self._FEED_TYPE)
        self._main_stack.addWidget(self._skipped)

        self._tabs.tab_changed.connect(self._on_main_tab)
        vl.addWidget(self._main_stack, 1)

    def _on_main_tab(self, idx: int) -> None:
        self._main_stack.setCurrentIndex(idx)
        if idx == 2:
            self._skipped.refresh()

    # ── Public API ───────────────────────────────────────────────────────────
    def refresh(self) -> None:
        self._all_entries.refresh()
        self._show_browse()
        if hasattr(self, "_skipped"):
            self._skipped.refresh()

    # ── Internal ───────────────────────────────────────────────────────────────
    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(1, emit=False)
        self._main_stack.setCurrentIndex(1)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _template_filename(self) -> str:
        safe = self._schema["title"].replace(" ", "_")
        return f"{safe}_Import_Template.xlsx"

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            self._template_filename(),
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        headers = [
            c[0] for c in self._schema["columns"]
            if c[1] not in ("sn", "total_amount")
        ]
        try:
            _write_xlsx_template(path, self._schema["title"], headers)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        dlg = _FuelImportDialog(self._FEED_TYPE, parent=self)
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        skipped = int(getattr(self.sender(), "_last_skipped", 0) or 0)
        msg = f"Imported {n:,} new records."
        if skipped:
            msg += f"\n{skipped:,} parked in Skipped for follow-up."
        QMessageBox.information(self, "Import Complete", msg)
        self._all_entries.refresh()
        self._show_browse()
        if skipped:
            self._tabs.set_index(2)
            self._skipped.refresh()
        else:
            self._tabs.set_index(1)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        count = int(upload_doc.get("record_count", 0))
        sheet = upload_doc.get("sheet_label") or upload_doc.get("source_filename") or ""
        if QMessageBox.question(
            self, "Delete upload",
            f"Delete this upload ({sheet}) and its {count:,} records?",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(str(upload_doc.get("_id") or "")))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            await svc.delete_diesel_upload(self._FEED_TYPE, upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        self._all_entries.refresh()
        self._show_browse()


# ═══════════════════════════════════════════════════════════════════════════════
#  Concrete stations
# ═══════════════════════════════════════════════════════════════════════════════

class InfinityWidget(_BaseDieselWidget):
    _FEED_TYPE = "diesel_infinity"


class LakeZambiaWidget(_BaseDieselWidget):
    _FEED_TYPE = "diesel_lake_zambia"


class LakeTundumaWidget(_BaseDieselWidget):
    _FEED_TYPE = "diesel_lake_tunduma"


class GBPDieselWidget(_BaseDieselWidget):
    _FEED_TYPE = "diesel_gbp"
