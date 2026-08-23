"""Truck Overview — cross-source truck-centric expense view.

This page lets the accountant select a truck and review related rows pulled
from verified cashier transactions, diesel imports, USD sheet-ledgers, and
selected imported feeds such as Toll Plaza and Zambia Parking.
"""

from __future__ import annotations

import asyncio
from datetime import datetime
from time import monotonic
from typing import Dict, List, Optional, Sequence, Set

import qtawesome as qta

from PySide6.QtCore import Qt, QTimer, QSize, QDate, QEvent, Signal
from PySide6.QtGui import QColor, QStandardItem, QStandardItemModel
from PySide6.QtWidgets import (
    QButtonGroup,
    QComboBox,
    QDateEdit,
    QFileDialog,
    QFrame,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from tahmeed.app_state import app_state
from tahmeed.services.truck_service import get_fleet_numbers, search_fleet, search_fleet_sync
from tahmeed.ui.widgets.truck_autocomplete import TruckLineEdit
from tahmeed.ui.widgets.loading_overlay import LoadingOverlay
from tahmeed.ui.accountant.date_filters import (
    style_calendar_popup,
    sync_from_to,
)
from tahmeed.ui.accountant.separate_expenses import _make_table, _cell, _finish_table_row
from tahmeed.ui.widgets.excel_column_filter import (
    ExcelFilterHeaderView,
    SORT_ASC,
    cascade_column_values,
)

_WHITE = "#FFFFFF"
_BG = "#F4F6F8"
_BORDER = "#E5E7EB"
_BLUE = "#0077C5"
_BLUE_L = "#E8F4FD"
_GREEN = "#16A34A"
_AMBER = "#D97706"
_RED = "#DC2626"
_T1 = "#111827"
_T2 = "#6B7280"
_TM = "#9CA3AF"
_HDR_BG = "#F1F5F9"
_NAVY = "#1B2B4B"

_SCROLL_CHUNK = 50
_ROW_H = 32
_MIN_FILTER_DATE = QDate(2000, 1, 1)
_CURRENCY_FILTERS = ("All", "TZS", "USD", "ZMW")
_MONTHS = [
    ("All Months", 0),
    ("January", 1), ("February", 2), ("March", 3), ("April", 4),
    ("May", 5), ("June", 6), ("July", 7), ("August", 8),
    ("September", 9), ("October", 10), ("November", 11), ("December", 12),
]

_SOURCE_OPTIONS = [
    ("All Sources", "all"),
    ("Master Expenses", "master"),
    ("Diesel Cash", "diesel_cash"),
    ("Diesel Imports", "diesel_imports"),
    ("Afritrack", "afritrack"),
    ("Toll Plaza", "toll_plaza"),
    ("Parking Congo", "parking_congo"),
    ("Zambia Parking", "zambia_parking"),
    ("Congo Expenses", "congo_expenses"),
    ("Ahmed Kimvi", "ahmed_kimvi"),
    ("RahnTech", "rahntech"),
    ("COMESA", "comesa"),
    ("Third Party Covers", "third_party"),
    ("SM Burhani", "sm_burhani"),
]

# Widths are preferred defaults; DESCRIPTION and STATION stretch to fill.
_COLS = [
    ("DATE", 90, Qt.AlignLeft),
    ("SOURCE", 120, Qt.AlignLeft),
    ("DESCRIPTION", 200, Qt.AlignLeft),
    ("REFERENCE", 120, Qt.AlignLeft),
    ("TRUCK FIELD", 90, Qt.AlignLeft),
    ("TZS", 110, Qt.AlignRight),
    ("USD", 100, Qt.AlignRight),
    ("ZMW", 100, Qt.AlignRight),
    ("LTRS", 70, Qt.AlignRight),
    ("RATE", 80, Qt.AlignRight),
    ("STATION / OWNER", 130, Qt.AlignLeft),
    ("RECEIPT", 85, Qt.AlignCenter),
]
_STRETCH_COLS = {2, 10}  # DESCRIPTION, STATION / OWNER
_UPLOAD_DESCRIPTION_COL = ("UPLOAD DESCRIPTION", 160, Qt.AlignLeft)
_FUEL_OVERVIEW_COLS = [
    ("DATE", 82, Qt.AlignLeft),
    ("SOURCE", 105, Qt.AlignLeft),
    ("DESCRIPTION", 120, Qt.AlignLeft),
    ("REFERENCE", 72, Qt.AlignLeft),
    ("TRUCK FIELD", 78, Qt.AlignLeft),
    ("TZS", 92, Qt.AlignRight),
    ("USD", 82, Qt.AlignRight),
    ("ZMW", 82, Qt.AlignRight),
    ("LTRS", 52, Qt.AlignRight),
    ("RATE", 62, Qt.AlignRight),
    ("STATION / OWNER", 100, Qt.AlignLeft),
    _UPLOAD_DESCRIPTION_COL,
    ("RECEIPT", 72, Qt.AlignCenter),
]
_FUEL_OVERVIEW_STRETCH_COLS = {2, 10, 11}  # DESCRIPTION, STATION, UPLOAD DESCRIPTION
_CTRL_H = 32


def _lbl(text: str = "", size: int = 13, weight: int = 400, color: str = _T1) -> QLabel:
    w = QLabel(text)
    w.setStyleSheet(
        f"QLabel {{"
        f"  color: {color}; font-size: {size}px; font-weight: {weight};"
        f"  font-family:'Segoe UI'; background: transparent; border: none;"
        f"}}"
    )
    return w


def _input_ss() -> str:
    return (
        f"QLineEdit, QComboBox, QDateEdit {{"
        f" border: 1px solid {_BORDER}; border-radius: 5px;"
        f" background: {_WHITE}; color: {_T1}; font-size: 12px;"
        " font-family:'Segoe UI'; padding: 0 8px;"
        f" min-height: {_CTRL_H}px; max-height: {_CTRL_H}px; }}"
        f"QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{ border-color: {_BLUE}; }}"
        "QComboBox::drop-down { border: none; width: 20px; }"
        "QDateEdit::drop-down { border: none; width: 20px; }"
        "QComboBox QAbstractItemView {"
        f"  border: 1px solid {_BORDER}; background: {_WHITE}; selection-background-color: {_BLUE_L};"
        "  outline: none; padding: 2px;"
        "}"
    )


class _SourceMultiCombo(QComboBox):
    """Checkable source filter — pick one or many sources, or All Sources."""

    selectionChanged = Signal()

    def __init__(self, options: Sequence[tuple[str, str]], parent=None) -> None:
        super().__init__(parent)
        self.setEditable(True)
        self.lineEdit().setReadOnly(True)
        self.lineEdit().setPlaceholderText("All Sources")
        self.lineEdit().setCursor(Qt.PointingHandCursor)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.setMaxVisibleItems(14)
        self.setFocusPolicy(Qt.StrongFocus)

        self._model = QStandardItemModel(self)
        self.setModel(self._model)
        self._updating = False

        all_item = QStandardItem("All Sources")
        all_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
        all_item.setData("all", Qt.UserRole)
        all_item.setCheckState(Qt.Checked)
        self._model.appendRow(all_item)

        for label, key in options:
            if key == "all":
                continue
            item = QStandardItem(label)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            item.setData(key, Qt.UserRole)
            item.setCheckState(Qt.Unchecked)
            self._model.appendRow(item)

        self.lineEdit().installEventFilter(self)
        self.view().viewport().installEventFilter(self)
        self._model.itemChanged.connect(self._on_item_changed)
        self._refresh_label()

    def showPopup(self) -> None:
        # Keep the popup at least as wide as the combo for readable checkboxes.
        self.view().setMinimumWidth(max(self.width(), 200))
        self._popup_shown_at = monotonic()
        super().showPopup()

    def hidePopup(self) -> None:
        # QScrollArea toolbars often fire an immediate hide right after show.
        if monotonic() - getattr(self, "_popup_shown_at", 0) < 0.2:
            return
        super().hidePopup()
        # Editable combo can snap the line edit back to a row label on close.
        self._refresh_label()

    def mousePressEvent(self, event) -> None:
        # Whole control opens the list (not only the tiny arrow).
        self.showPopup()

    def eventFilter(self, obj, event) -> bool:
        # Line-edit is read-only; without this, clicks on the text do nothing.
        if obj is self.lineEdit() and event.type() == QEvent.MouseButtonPress:
            self.showPopup()
            return True
        if obj is self.view().viewport():
            # Swallow press so Qt does not select-and-close before we toggle.
            if event.type() == QEvent.MouseButtonPress:
                index = self.view().indexAt(event.position().toPoint())
                if index.isValid():
                    return True
            if event.type() == QEvent.MouseButtonRelease:
                index = self.view().indexAt(event.position().toPoint())
                if index.isValid():
                    item = self._model.itemFromIndex(index)
                    if item is not None and item.flags() & Qt.ItemIsUserCheckable:
                        item.setCheckState(
                            Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked
                        )
                        # Keep popup open for multi-select / bulk toggling.
                        return True
        return super().eventFilter(obj, event)

    def _on_item_changed(self, item: QStandardItem) -> None:
        if self._updating:
            return
        self._updating = True
        try:
            key = item.data(Qt.UserRole)
            if key == "all" and item.checkState() == Qt.Checked:
                for row in range(1, self._model.rowCount()):
                    self._model.item(row).setCheckState(Qt.Unchecked)
            elif key != "all" and item.checkState() == Qt.Checked:
                self._model.item(0).setCheckState(Qt.Unchecked)
            # If nothing is checked, fall back to All Sources.
            if not any(
                self._model.item(row).checkState() == Qt.Checked
                for row in range(self._model.rowCount())
            ):
                self._model.item(0).setCheckState(Qt.Checked)
        finally:
            self._updating = False
        self._refresh_label()
        self.selectionChanged.emit()

    def selected_keys(self) -> List[str]:
        """Return selected source keys, or ``[]`` meaning all sources."""
        if self._model.item(0).checkState() == Qt.Checked:
            return []
        keys: List[str] = []
        for row in range(1, self._model.rowCount()):
            item = self._model.item(row)
            if item.checkState() == Qt.Checked:
                keys.append(str(item.data(Qt.UserRole)))
        return keys

    def summary_text(self) -> str:
        if self._model.item(0).checkState() == Qt.Checked:
            return "All Sources"
        labels = [
            self._model.item(row).text()
            for row in range(1, self._model.rowCount())
            if self._model.item(row).checkState() == Qt.Checked
        ]
        if not labels:
            return "All Sources"
        if len(labels) == 1:
            return labels[0]
        if len(labels) == 2:
            return f"{labels[0]}, {labels[1]}"
        return f"{len(labels)} sources"

    def _refresh_label(self) -> None:
        text = self.summary_text()
        # Avoid setCurrentText — it can change currentIndex and fight the popup.
        self.lineEdit().setText(text)
        self.lineEdit().setToolTip(text)

    def reset_to_all(self) -> None:
        self._updating = True
        try:
            self._model.item(0).setCheckState(Qt.Checked)
            for row in range(1, self._model.rowCount()):
                self._model.item(row).setCheckState(Qt.Unchecked)
        finally:
            self._updating = False
        self._refresh_label()
        self.selectionChanged.emit()


_CARD_SS = (
    "QFrame#truckCard {"
    f"  background: {_WHITE};"
    f"  border: 1px solid {_BORDER};"
    "  border-radius: 12px;"
    "}"
)


def _normalize_currency(currency: str) -> str:
    cur = (currency or "").strip().upper()
    if cur in ("TZS", "TSH", "TZ"):
        return "TZS"
    if cur == "USD":
        return "USD"
    if cur in ("ZMW", "ZMB", "ZK"):
        return "ZMW"
    return cur


def _currency_toggle_style(checked: bool) -> str:
    if checked:
        return (
            f"QPushButton {{"
            f"  background: {_NAVY}; color: {_WHITE}; border: 1px solid {_NAVY};"
            f"  border-radius: 6px; font-size: 11px; font-weight: 600;"
            f"  font-family: 'Segoe UI'; padding: 0 10px;"
            f"}}"
        )
    return (
        f"QPushButton {{"
        f"  background: {_WHITE}; color: {_T2}; border: 1px solid {_BORDER};"
        f"  border-radius: 6px; font-size: 11px; font-weight: 600;"
        f"  font-family: 'Segoe UI'; padding: 0 10px;"
        f"}}"
        f"QPushButton:hover {{ background: #F8FAFC; color: {_T1}; }}"
    )


class _CurrencyFilterToggle(QWidget):
    """Exclusive All / TZS / USD / ZMW filter for the truck overview table."""

    changed = Signal(str)

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._currency = "All"
        self.setStyleSheet("background: transparent; border: none;")
        row = QHBoxLayout(self)
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(4)
        self._group = QButtonGroup(self)
        self._group.setExclusive(True)
        self._buttons: dict[str, QPushButton] = {}
        for cur in _CURRENCY_FILTERS:
            btn = QPushButton(cur)
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setFixedHeight(26)
            btn.setChecked(cur == "All")
            btn.setStyleSheet(_currency_toggle_style(cur == "All"))
            btn.clicked.connect(lambda _=False, c=cur: self._on_clicked(c))
            self._group.addButton(btn)
            self._buttons[cur] = btn
            row.addWidget(btn)

    def currency(self) -> str:
        return self._currency

    def filter_value(self) -> str:
        """Empty string means all currencies."""
        return "" if self._currency == "All" else self._currency

    def set_currency(self, currency: str, *, emit: bool = True) -> None:
        cur = currency if currency in _CURRENCY_FILTERS else "All"
        if cur == self._currency:
            return
        self._currency = cur
        for code, btn in self._buttons.items():
            btn.setChecked(code == cur)
            btn.setStyleSheet(_currency_toggle_style(code == cur))
        if emit:
            self.changed.emit(cur)

    def _on_clicked(self, currency: str) -> None:
        self.set_currency(currency)


def _amount_columns(currency: str, amount) -> tuple[Optional[float], Optional[float], Optional[float]]:
    """Return (tzs, usd, zmw) with the amount only in its currency column."""
    if amount is None or amount == "":
        return None, None, None
    try:
        val = float(amount)
    except (TypeError, ValueError):
        return None, None, None
    cur = _normalize_currency(currency)
    if cur == "TZS":
        return val, None, None
    if cur == "USD":
        return None, val, None
    if cur == "ZMW":
        return None, None, val
    return None, None, None


def _fmt_currency_cell(currency: str, value) -> str:
    if value is None:
        return "—"
    try:
        val = float(value)
    except (TypeError, ValueError):
        return "—"
    decimals = 2 if currency == "USD" else 0
    return f"{val:,.{decimals}f}"


def _btn(text: str, icon_name: str = "", primary: bool = True) -> QPushButton:
    b = QPushButton(f"  {text}" if icon_name else text)
    b.setCursor(Qt.PointingHandCursor)
    b.setFixedHeight(32)
    if icon_name:
        try:
            b.setIcon(qta.icon(icon_name, color="#FFFFFF" if primary else _T2))
            b.setIconSize(QSize(15, 15))
        except Exception:
            pass
    if primary:
        b.setStyleSheet(
            f"QPushButton {{ background: {_BLUE}; color: #FFF; border: none;"
            f" border-radius: 5px; font-size: 12px; font-weight: 600;"
            f" font-family:'Segoe UI'; padding: 0 12px; }}"
            f"QPushButton:hover {{ background: #005EA3; }}"
        )
    else:
        b.setStyleSheet(
            f"QPushButton {{ background: {_WHITE}; color: {_T1};"
            f" border: 1px solid {_BORDER}; border-radius: 5px;"
            f" font-size: 12px; font-family:'Segoe UI'; padding: 0 12px; }}"
            f"QPushButton:hover {{ background: {_BG}; }}"
        )
    return b


def _fmt_amount(currency: str, value) -> str:
    if value is None:
        return "—"
    try:
        val = float(value)
    except (TypeError, ValueError):
        return "—"
    prefix = f"{currency} " if currency else ""
    decimals = 2 if currency == "USD" else 0
    return f"{prefix}{val:,.{decimals}f}"


def _fmt_num(value, decimals: int = 0) -> str:
    if value in (None, "", 0, 0.0):
        return "—"
    try:
        return f"{float(value):,.{decimals}f}"
    except (TypeError, ValueError):
        return "—"


class _SummaryCard(QFrame):
    def __init__(self, label: str, value: str = "—", parent=None) -> None:
        super().__init__(parent)
        self.setObjectName("truckCard")
        self.setMinimumHeight(88)
        self.setStyleSheet(_CARD_SS)
        vl = QVBoxLayout(self)
        vl.setContentsMargins(16, 14, 16, 14)
        vl.setSpacing(6)
        self._label = _lbl(label, size=12, weight=400, color=_T2)
        self._value = _lbl(value, size=22, weight=700, color=_T1)
        vl.addWidget(self._label)
        vl.addWidget(self._value)
        vl.addStretch()

    def set_value(self, value: str, color: Optional[str] = None) -> None:
        self._value.setText(value)
        self._value.setStyleSheet(
            f"QLabel {{"
            f"  color: {color or _T1}; font-size: 22px; font-weight: 700;"
            f"  font-family:'Segoe UI'; background: transparent; border: none;"
            f"}}"
        )


class _StatusFooter(QFrame):
    """Footer that shows infinite-scroll progress (Master / Verify style)."""

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setObjectName("truckStatusFooter")
        self.setFixedHeight(36)
        self.setStyleSheet(
            f"QFrame#truckStatusFooter {{"
            f"  background: {_WHITE}; border: none; border-top: 1px solid {_BORDER};"
            f"}}"
        )
        hl = QHBoxLayout(self)
        hl.setContentsMargins(16, 0, 16, 0)
        self._info = _lbl("Select a truck to load records", size=12, color=_T2)
        hl.addWidget(self._info)
        hl.addStretch()

    def set_text(self, text: str) -> None:
        self._info.setText(text)


def _overview_export_values(row: dict, cols: Sequence) -> list:
    """Map an overview row to Excel cell values for the given column layout."""
    tzs_amt, usd_amt, zmw_amt = _amount_columns(row.get("currency", ""), row.get("amount"))
    date_value = row.get("date")
    date_txt = (
        date_value.strftime("%d/%m/%Y")
        if hasattr(date_value, "strftime") and date_value.year > 1
        else ""
    )
    receipt = (row.get("receipt_status") or "").strip().lower()
    receipt_txt = receipt.title() if receipt else ""
    values = []
    for header, _, _align in cols:
        if header == "DATE":
            values.append(date_txt)
        elif header == "SOURCE":
            values.append(row.get("source", ""))
        elif header == "DESCRIPTION":
            values.append(row.get("description", ""))
        elif header == "REFERENCE":
            values.append(row.get("reference", ""))
        elif header == "TRUCK FIELD":
            values.append(row.get("truck_value", ""))
        elif header == "TZS":
            values.append(tzs_amt)
        elif header == "USD":
            values.append(usd_amt)
        elif header == "ZMW":
            values.append(zmw_amt)
        elif header == "LTRS":
            values.append(row.get("liters"))
        elif header == "RATE":
            values.append(row.get("rate"))
        elif header == "STATION / OWNER":
            values.append(row.get("station", ""))
        elif header == "UPLOAD DESCRIPTION":
            label = row.get("upload_description") or ""
            values.append("" if label in ("", "—") else label)
        elif header == "RECEIPT":
            values.append(receipt_txt)
        else:
            values.append("")
    return values


def _write_truck_overview_excel(
    path: str,
    truck: str,
    source_label: str,
    date_range_label: str,
    rows: list,
    summary: dict,
    *,
    sheet_title: str = "Truck Overview",
    banner_label: str = "TRUCK OVERVIEW",
    cols: Sequence = _COLS,
) -> None:
    """Build and save the truck overview workbook (runs off the UI event loop)."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    # Match results.xlsx style: green section banner, navy headers,
    # thin #CCCCCC borders, alternating #EEF2FF rows, Calibri.
    n_cols = len(cols)
    last_col = openpyxl.utils.get_column_letter(n_cols)

    title_fill = PatternFill("solid", fgColor="1F6B2E")
    header_fill = PatternFill("solid", fgColor="2C5282")
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    zmw_fill = PatternFill("solid", fgColor="E8F4FD")
    summary_label_fill = PatternFill("solid", fgColor="2C5282")
    summary_value_fill = PatternFill("solid", fgColor="FFFFFF")
    receipt_ok_fill = PatternFill("solid", fgColor="C6EFCE")
    receipt_pending_fill = PatternFill("solid", fgColor="FFEB9C")
    receipt_missing_fill = PatternFill("solid", fgColor="FFCCCC")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    meta_font = Font(name="Calibri", italic=True, size=10, color="475569")
    hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11, color="000000")
    summary_label_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    summary_value_font = Font(name="Calibri", bold=True, size=11, color="000000")
    amount_font = Font(name="Calibri", size=11, color="000000")
    red_font = Font(name="Calibri", bold=True, size=11, color="9C0006")
    green_font = Font(name="Calibri", bold=True, size=11, color="276221")
    receipt_ok_font = Font(name="Calibri", bold=True, size=11, color="276221")
    receipt_pending_font = Font(name="Calibri", bold=True, size=11, color="9C6500")
    receipt_missing_font = Font(name="Calibri", bold=True, size=11, color="9C0006")

    def _style_merged_row(row: int, fill=None, font=None, align=None, apply_border: bool = False) -> None:
        for col in range(1, n_cols + 1):
            cell = ws.cell(row, col)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if align is not None:
                cell.alignment = align
            if apply_border:
                cell.border = border

    record_count = summary.get("record_count", len(rows))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{banner_label}  -  {truck}   ({record_count:,} records)"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    _style_merged_row(1, fill=title_fill, font=title_font,
                      align=Alignment(horizontal="left", vertical="center"))
    ws.row_dimensions[1].height = 18

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        f"Source: {source_label}  |  "
        f"Date Range: {date_range_label}  |  "
        f"Exported: {datetime.now().strftime('%d %b %Y %H:%M')}"
    )
    ws["A2"].font = meta_font
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    summary_pairs = [
        ("Records", summary["record_count"]),
        ("Sources", summary["source_count"]),
        ("TZS Total", summary["tzs_total"]),
        ("USD Total", summary["usd_total"]),
        ("ZMW Total", summary["zmw_total"]),
        ("Fuel Liters", summary["liters_total"]),
    ]
    summary_row = 4
    for idx, (label, value) in enumerate(summary_pairs):
        col = 1 + idx
        if col > n_cols:
            break
        label_cell = ws.cell(summary_row, col, label)
        label_cell.fill = summary_label_fill
        label_cell.font = summary_label_font
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal="center", vertical="center")

        value_cell = ws.cell(summary_row + 1, col, value)
        value_cell.fill = summary_value_fill
        value_cell.font = summary_value_font
        value_cell.border = border
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(value, float):
            value_cell.number_format = '#,##0.00'
        elif isinstance(value, int):
            value_cell.number_format = '#,##0'

    headers = [c[0] for c in cols]
    table_row = 7
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(table_row, col, header)
        cell.fill = header_fill
        cell.border = border
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[table_row].height = 18

    header_index = {header: idx + 1 for idx, header in enumerate(headers)}
    amount_cols = (
        header_index.get("TZS"),
        header_index.get("USD"),
        header_index.get("ZMW"),
        header_index.get("LTRS"),
        header_index.get("RATE"),
    )
    wrap_cols = {
        header_index[h]
        for h in ("SOURCE", "DESCRIPTION", "REFERENCE", "STATION / OWNER", "UPLOAD DESCRIPTION")
        if h in header_index
    }
    receipt_col = header_index.get("RECEIPT")

    for i, row in enumerate(rows, start=1):
        values = _overview_export_values(row, cols)
        ws.append(values)
        excel_row = ws.max_row
        is_zmw = _normalize_currency(row.get("currency") or "") == "ZMW"
        fill = zmw_fill if is_zmw else (alt_fill if i % 2 == 0 else white_fill)
        for col in range(1, len(values) + 1):
            cell = ws.cell(excel_row, col)
            cell.fill = fill
            cell.border = border
            cell.font = cell_font
            cell.alignment = Alignment(vertical="center", wrap_text=col in wrap_cols)

        for col in amount_cols:
            if not col:
                continue
            ws.cell(excel_row, col).font = amount_font
            ws.cell(excel_row, col).alignment = Alignment(horizontal="right", vertical="center")
        if header_index.get("TZS"):
            ws.cell(excel_row, header_index["TZS"]).number_format = '#,##0'
        if header_index.get("USD"):
            ws.cell(excel_row, header_index["USD"]).number_format = '#,##0.00'
        if header_index.get("ZMW"):
            ws.cell(excel_row, header_index["ZMW"]).number_format = '#,##0'
        if header_index.get("LTRS"):
            ws.cell(excel_row, header_index["LTRS"]).number_format = '#,##0.00'
        if header_index.get("RATE"):
            ws.cell(excel_row, header_index["RATE"]).number_format = '#,##0.00'

        amount = row.get("amount")
        cur = _normalize_currency(row.get("currency") or "")
        if isinstance(amount, (int, float)):
            amount_col = {
                "TZS": header_index.get("TZS"),
                "USD": header_index.get("USD"),
                "ZMW": header_index.get("ZMW"),
            }.get(cur)
            if amount_col:
                if amount < 0:
                    ws.cell(excel_row, amount_col).font = red_font
                elif cur == "USD":
                    ws.cell(excel_row, amount_col).font = green_font

        if receipt_col:
            receipt = (row.get("receipt_status") or "").strip().lower()
            receipt_cell = ws.cell(excel_row, receipt_col)
            receipt_cell.alignment = Alignment(horizontal="center", vertical="center")
            if receipt == "received":
                receipt_cell.fill = receipt_ok_fill
                receipt_cell.font = receipt_ok_font
            elif receipt == "pending":
                receipt_cell.fill = receipt_pending_fill
                receipt_cell.font = receipt_pending_font
            elif receipt == "missing":
                receipt_cell.fill = receipt_missing_fill
                receipt_cell.font = receipt_missing_font

    default_widths = {
        "DATE": 12,
        "SOURCE": 18,
        "DESCRIPTION": 36,
        "REFERENCE": 22,
        "TRUCK FIELD": 14,
        "TZS": 14,
        "USD": 12,
        "ZMW": 12,
        "LTRS": 10,
        "RATE": 10,
        "STATION / OWNER": 20,
        "UPLOAD DESCRIPTION": 24,
        "RECEIPT": 12,
    }
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = default_widths.get(header, 14)
    ws.freeze_panes = f"A{table_row + 1}"
    ws.auto_filter.ref = f"A{table_row}:{last_col}{ws.max_row}"
    wb.save(path)


class TruckOverviewWidget(QWidget):
    PAGE_TITLE = "Truck Overview"
    PAGE_ICON = "mdi.truck-fast-outline"
    PAGE_OBJECT_NAME = "truckOverview"
    SOURCE_OPTIONS = _SOURCE_OPTIONS
    TABLE_COLS = _COLS
    TABLE_STRETCH_COLS = _STRETCH_COLS
    SUBTITLE_EMPTY = "Select a truck to gather cross-source expenses and fuel."
    LOADING_OVERLAY_TEXT = "Loading truck overview…"
    STATUS_EMPTY = "No truck selected yet."
    ROW_NOUN = "cross-source row"
    STATUS_HINT = " Zambia entries are summarized under ZMW."
    LOADED_SUBTITLE = "Cross-source view for {truck}  ·  {period}"
    LOAD_ERROR_LABEL = "truck overview"
    HIGHLIGHT_SOURCE_GROUPS = ("master", "diesel_cash")
    EXPORT_PREFIX = "Truck_Overview"
    EXCEL_DIALOG_TITLE = "Export Truck Overview"
    PDF_DIALOG_TITLE = "Export Truck Overview PDF"
    EXCEL_SHEET_TITLE = "Truck Overview"
    EXCEL_BANNER = "TRUCK OVERVIEW"
    PDF_EYEBROW = "FLEET EXPENSE REPORT"
    PDF_SUBTITLE = "Consolidated expense record across all logged sources"
    USE_EXCEL_COLUMN_FILTERS = False
    DEFAULT_SORT_FIELD = "date"
    DEFAULT_SORT_ASC = True

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._loaded = 0
        self._total = 0
        self._active_truck = ""
        self._loading = False
        self._scroll_loading = False
        self._reload_generation = 0
        self._fleet_numbers: List[str] = []
        self._year = app_state.fiscal_year
        self._month = 0
        self._sort_field = self.DEFAULT_SORT_FIELD
        self._sort_asc = self.DEFAULT_SORT_ASC
        self._col_filters: Dict[int, Set[str]] = {}
        self._col_value_cache: Dict[int, Set[str]] = {}

        self._debounce = QTimer(self)
        self._debounce.setSingleShot(True)
        self._debounce.setInterval(300)
        self._debounce.timeout.connect(self._reset_and_load)

        self._build()
        # Defer off any current asyncio Task — Python 3.14 + qasync forbids
        # entering a new Task while another is executing (sidebar nav can nest
        # via processEvents and previously tore down the whole app).
        from tahmeed.ui.async_utils import schedule_coro

        schedule_coro(self._preload_fleet())

    def _build(self) -> None:
        self.setObjectName(self.PAGE_OBJECT_NAME)
        self.setStyleSheet(
            f"QWidget#{self.PAGE_OBJECT_NAME} {{ background: {_BG}; border: none; }}"
            f"QWidget#{self.PAGE_OBJECT_NAME} QLabel {{ border: none; }}"
        )
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 16)
        root.setSpacing(12)

        title_bar = QFrame()
        title_bar.setObjectName("truckTitleBar")
        title_bar.setFixedHeight(52)
        title_bar.setStyleSheet(
            f"QFrame#truckTitleBar {{"
            f"  background: {_WHITE}; border: none; border-bottom: 1px solid {_BORDER};"
            f"}}"
        )
        tb = QHBoxLayout(title_bar)
        tb.setContentsMargins(16, 0, 16, 0)
        tb.setSpacing(10)
        try:
            icon_lbl = QLabel()
            icon_lbl.setFixedSize(22, 22)
            icon_lbl.setStyleSheet("QLabel { border: none; background: transparent; }")
            icon_lbl.setPixmap(qta.icon(self.PAGE_ICON, color=_BLUE).pixmap(22, 22))
            tb.addWidget(icon_lbl)
        except Exception:
            pass
        tb.addWidget(_lbl(self.PAGE_TITLE, size=16, weight=700))
        self._subtitle = _lbl(self.SUBTITLE_EMPTY, size=12, color=_T2)
        tb.addWidget(self._subtitle)
        tb.addStretch()
        export_excel_btn = _btn("Excel", "mdi.microsoft-excel", primary=False)
        export_excel_btn.clicked.connect(self._export_excel)
        tb.addWidget(export_excel_btn)
        export_pdf_btn = _btn("PDF", "mdi.file-pdf-box", primary=False)
        export_pdf_btn.clicked.connect(self._export_pdf)
        tb.addWidget(export_pdf_btn)
        root.addWidget(title_bar)

        root.addWidget(self._build_toolbar())
        root.addLayout(self._build_summary_cards())

        status_row = QHBoxLayout()
        status_row.setContentsMargins(0, 0, 0, 0)
        status_row.setSpacing(12)
        self._status = _lbl(self.STATUS_EMPTY, size=11, color=_TM)
        status_row.addWidget(self._status, 1)
        self._currency_toggle = _CurrencyFilterToggle()
        self._currency_toggle.changed.connect(self._on_currency_changed)
        status_row.addWidget(self._currency_toggle, 0, Qt.AlignRight | Qt.AlignVCenter)
        root.addLayout(status_row)

        self._table = _make_table([c[0] for c in self.TABLE_COLS])
        self._table.setShowGrid(True)
        self._table.verticalHeader().setDefaultSectionSize(_ROW_H)
        hdr = self._table.horizontalHeader()
        hdr.setSectionsMovable(False)
        hdr.setStretchLastSection(False)
        for idx, (_, width, _align) in enumerate(self.TABLE_COLS):
            self._table.setColumnWidth(idx, width)
            if idx in self.TABLE_STRETCH_COLS:
                hdr.setSectionResizeMode(idx, QHeaderView.Stretch)
            else:
                hdr.setSectionResizeMode(idx, QHeaderView.Interactive)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        if type(self).USE_EXCEL_COLUMN_FILTERS:
            self._attach_excel_column_filters()
        root.addWidget(self._table, 1)

        self._footer = _StatusFooter()
        root.addWidget(self._footer)

        self._loading_overlay = LoadingOverlay(self, self.LOADING_OVERLAY_TEXT)

    async def _preload_fleet(self) -> None:
        """Warm fleet cache and give TruckLineEdit a sync suggestion list."""
        try:
            numbers = await get_fleet_numbers()
            self._fleet_numbers = sorted(numbers)
            self._truck_edit.set_local_numbers(lambda: list(self._fleet_numbers))
        except Exception:
            # Keep async search_fleet as fallback when API fleet load fails.
            pass

    def _build_toolbar(self) -> QFrame:
        toolbar = QFrame()
        toolbar.setObjectName("truckToolbar")
        toolbar.setStyleSheet(
            f"QFrame#truckToolbar {{"
            f"  background: {_WHITE}; border: 1px solid {_BORDER}; border-radius: 12px;"
            f"}}"
        )
        toolbar_v = QVBoxLayout(toolbar)
        toolbar_v.setContentsMargins(12, 10, 12, 10)
        toolbar_v.setSpacing(0)

        # Single row: filters + actions; scroll horizontally when narrow
        filter_scroll = QScrollArea()
        filter_scroll.setWidgetResizable(True)
        filter_scroll.setFrameShape(QFrame.NoFrame)
        filter_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        filter_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        filter_scroll.setFixedHeight(_CTRL_H + 14)
        filter_scroll.setStyleSheet(
            "QScrollArea { background: transparent; border: none; }"
            f"QScrollBar:horizontal {{ height: 8px; background: {_BG}; }}"
            f"QScrollBar::handle:horizontal {{ background: {_BORDER}; border-radius: 4px; min-width: 24px; }}"
        )

        filter_inner = QWidget()
        filter_inner.setStyleSheet("background: transparent; border: none;")
        filter_row = QHBoxLayout(filter_inner)
        filter_row.setContentsMargins(0, 0, 0, 0)
        filter_row.setSpacing(10)

        self._truck_edit = TruckLineEdit(search_fleet, sync_fn=search_fleet_sync)
        self._truck_edit.setPlaceholderText("Search truck, trailer, or bike/car…")
        self._truck_edit.setFixedWidth(180)
        self._truck_edit.setFixedHeight(_CTRL_H)
        self._truck_edit.setStyleSheet(_input_ss())
        self._truck_edit.returnPressed.connect(self._on_load_clicked)
        filter_row.addWidget(self._truck_edit)

        self._source_cb = _SourceMultiCombo(self.SOURCE_OPTIONS)
        self._source_cb.setFixedWidth(160)
        self._source_cb.setFixedHeight(_CTRL_H)
        self._source_cb.setStyleSheet(_input_ss())
        self._source_cb.selectionChanged.connect(self._on_filter_changed)
        filter_row.addWidget(self._source_cb)
        QWidget.setTabOrder(self._truck_edit, self._source_cb)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        current_yr = datetime.now().year
        for yr in range(current_yr - 5, current_yr + 2):
            self._year_cb.addItem(str(yr), yr)
        yr_idx = self._year_cb.findData(self._year)
        self._year_cb.setCurrentIndex(yr_idx if yr_idx >= 0 else 0)
        self._year_cb.setFixedWidth(100)
        self._year_cb.setFixedHeight(_CTRL_H)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.setToolTip("Limit truck history to one year (recommended for large data).")
        self._year_cb.currentIndexChanged.connect(self._on_year)
        filter_row.addWidget(self._year_cb)
        QWidget.setTabOrder(self._source_cb, self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(120)
        self._month_cb.setFixedHeight(_CTRL_H)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(self._year > 0)
        self._month_cb.setToolTip("Narrow further to a single month within the selected year.")
        self._month_cb.currentIndexChanged.connect(self._on_month)
        filter_row.addWidget(self._month_cb)
        QWidget.setTabOrder(self._year_cb, self._month_cb)

        self._from_date = QDateEdit()
        self._from_date.setCalendarPopup(True)
        self._from_date.setDisplayFormat("dd MMM yyyy")
        self._from_date.setMinimumDate(_MIN_FILTER_DATE)
        self._from_date.setSpecialValueText("From")
        self._from_date.setDate(_MIN_FILTER_DATE)
        self._from_date.setFixedWidth(130)
        self._from_date.setFixedHeight(_CTRL_H)
        self._from_date.setStyleSheet(_input_ss())
        style_calendar_popup(self._from_date)
        self._from_date.dateChanged.connect(lambda _d: self._on_filter_changed())
        filter_row.addWidget(self._from_date)
        QWidget.setTabOrder(self._month_cb, self._from_date)

        self._to_date = QDateEdit()
        self._to_date.setCalendarPopup(True)
        self._to_date.setDisplayFormat("dd MMM yyyy")
        self._to_date.setMinimumDate(_MIN_FILTER_DATE)
        self._to_date.setSpecialValueText("To")
        self._to_date.setDate(_MIN_FILTER_DATE)
        self._to_date.setFixedWidth(130)
        self._to_date.setFixedHeight(_CTRL_H)
        self._to_date.setStyleSheet(_input_ss())
        style_calendar_popup(self._to_date)
        self._to_date.dateChanged.connect(lambda _d: self._on_filter_changed())
        filter_row.addWidget(self._to_date)
        QWidget.setTabOrder(self._from_date, self._to_date)

        # Default to fiscal year so loads stay bounded on large histories.
        if self._year > 0:
            sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)

        load_btn = _btn("Load", "mdi.magnify")
        load_btn.clicked.connect(self._on_load_clicked)
        filter_row.addWidget(load_btn)

        refresh_btn = _btn("Refresh", "mdi.refresh", primary=False)
        refresh_btn.clicked.connect(self.refresh)
        filter_row.addWidget(refresh_btn)

        clear_btn = _btn("Clear", "mdi.filter-remove-outline", primary=False)
        clear_btn.clicked.connect(self._clear_results)
        filter_row.addWidget(clear_btn)

        filter_row.addStretch()

        filter_inner.setMinimumWidth(1200)
        filter_inner.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Fixed)
        filter_scroll.setWidget(filter_inner)
        toolbar_v.addWidget(filter_scroll)
        return toolbar

    def _build_summary_cards(self) -> QHBoxLayout:
        summary = QHBoxLayout()
        summary.setSpacing(12)
        self._records_card = _SummaryCard("Records")
        self._sources_card = _SummaryCard("Sources")
        self._tzs_card = _SummaryCard("TZS Total")
        self._usd_card = _SummaryCard("USD Total")
        self._zmw_card = _SummaryCard("ZMW Total")
        self._liters_card = _SummaryCard("Fuel Liters")
        for card in (
            self._records_card,
            self._sources_card,
            self._tzs_card,
            self._usd_card,
            self._zmw_card,
            self._liters_card,
        ):
            summary.addWidget(card)
        return summary

    def refresh(self) -> None:
        if self._active_truck:
            self._reset_and_load()

    def _selected_truck(self) -> str:
        return self._truck_edit.text().strip().upper()

    def _selected_sources(self) -> List[str]:
        """Selected source keys; empty list means all sources."""
        return self._source_cb.selected_keys()

    def _source_filter_param(self):
        keys = self._selected_sources()
        return keys if keys else "all"

    def _source_filter_label(self) -> str:
        return self._source_cb.summary_text()

    def _source_filter_tag(self) -> str:
        keys = self._selected_sources()
        if not keys:
            return "all"
        if len(keys) == 1:
            return keys[0]
        return "multi"

    def _date_filters(self) -> tuple[Optional[datetime], Optional[datetime]]:
        date_from = None
        date_to = None
        if self._from_date.date() > _MIN_FILTER_DATE:
            d = self._from_date.date()
            date_from = datetime(d.year(), d.month(), d.day(), 0, 0, 0)
        if self._to_date.date() > _MIN_FILTER_DATE:
            d = self._to_date.date()
            date_to = datetime(d.year(), d.month(), d.day(), 23, 59, 59)
        return date_from, date_to

    def _has_valid_date_range(self, *, warn: bool = False) -> bool:
        date_from, date_to = self._date_filters()
        valid = not (date_from and date_to and date_from > date_to)
        if not valid and warn:
            QMessageBox.warning(self, "Invalid Date Range", "'From' date cannot be later than 'To' date.")
        return valid

    def _column_filter_specs(self) -> List[tuple[str, Optional[str], str]]:
        """Return (header label, sort/filter field, kind) per table column."""
        return []

    def _attach_excel_column_filters(self) -> None:
        specs = self._column_filter_specs()
        filterable: Set[int] = set()
        sort_kinds: Dict[int, str] = {}
        for idx, (_label, field, kind) in enumerate(specs):
            if field:
                filterable.add(idx)
                sort_kinds[idx] = kind

        filter_hdr = ExcelFilterHeaderView(
            self._table,
            filterable_columns=filterable,
            sort_kinds=sort_kinds,
        )
        filter_hdr.set_value_provider(self._filter_menu_values)
        filter_hdr.set_label_provider(
            lambda c: specs[c][0] if 0 <= c < len(specs) else "",
        )
        filter_hdr.filter_changed.connect(self._on_col_filter_changed)
        filter_hdr.sort_requested.connect(self._on_excel_sort)
        self._table.setHorizontalHeader(filter_hdr)

        hdr = self._table.horizontalHeader()
        hdr.setSortIndicatorShown(True)
        hdr.sectionClicked.connect(self._on_header_click)
        self._update_sort_indicator()

    def _field_for_col(self, col: int) -> Optional[str]:
        specs = self._column_filter_specs()
        if 0 <= col < len(specs):
            return specs[col][1]
        return None

    def _column_filters_for_query(self) -> Optional[Dict[str, List[str]]]:
        out: Dict[str, List[str]] = {}
        for col, accepted in self._col_filters.items():
            field = self._field_for_col(col)
            if not field or not accepted:
                continue
            vals = sorted(v for v in accepted if v)
            if vals:
                out[field] = vals
        return out or None

    def _filter_menu_values(self, col: int) -> set:
        if col in self._col_value_cache:
            return set(self._col_value_cache.get(col) or set())
        rows: List[dict] = []
        for r in range(self._table.rowCount()):
            row: dict = {}
            field = self._field_for_col(col)
            if not field:
                continue
            item = self._table.item(r, col)
            txt = (item.text() if item else "").strip()
            if txt:
                row[col] = txt
            if row:
                rows.append(row)
        return cascade_column_values(
            rows,
            target_col=col,
            active_filters=self._col_filters,
        )

    def _clear_column_filters(self) -> None:
        self._col_filters.clear()
        self._col_value_cache.clear()
        hdr = self._table.horizontalHeader()
        if isinstance(hdr, ExcelFilterHeaderView):
            hdr.clear_filters()
        self._sort_field = self.DEFAULT_SORT_FIELD
        self._sort_asc = self.DEFAULT_SORT_ASC
        self._update_sort_indicator()

    def _on_col_filter_changed(self, col: int, accepted) -> None:
        accepted = set(accepted or [])
        if accepted:
            self._col_filters[col] = accepted
        else:
            self._col_filters.pop(col, None)
        hdr = self._table.horizontalHeader()
        if isinstance(hdr, ExcelFilterHeaderView):
            hdr.sync_active(self._col_filters)
        self._reset_and_load()

    def _on_header_click(self, col: int) -> None:
        field = self._field_for_col(col)
        if not field:
            return
        if self._sort_field == field:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_field = field
            self._sort_asc = False
        self._update_sort_indicator()
        self._reset_and_load()

    def _on_excel_sort(self, col: int, mode: str) -> None:
        field = self._field_for_col(col)
        if not field:
            return
        self._sort_field = field
        self._sort_asc = mode == SORT_ASC
        self._update_sort_indicator()
        self._reset_and_load()

    def _update_sort_indicator(self) -> None:
        if not type(self).USE_EXCEL_COLUMN_FILTERS:
            return
        order = Qt.AscendingOrder if self._sort_asc else Qt.DescendingOrder
        hdr = self._table.horizontalHeader()
        for idx, (_label, field, _kind) in enumerate(self._column_filter_specs()):
            if field == self._sort_field:
                hdr.setSortIndicator(idx, order)
                return

    async def _refresh_column_filter_cache(self) -> None:
        if not type(self).USE_EXCEL_COLUMN_FILTERS:
            return
        from tahmeed.services.accountant_service import get_fuel_overview_column_filter_cache

        kw = self._filter_kw()
        by_field = await get_fuel_overview_column_filter_cache(
            truck=kw["truck"],
            source=kw["source"],
            date_from=kw["date_from"],
            date_to=kw["date_to"],
            currency=kw["currency"],
            column_filters=self._column_filters_for_query(),
        )
        self._col_value_cache = {}
        for idx, (_label, field, _kind) in enumerate(self._column_filter_specs()):
            if field:
                self._col_value_cache[idx] = set(by_field.get(field) or set())
        hdr = self._table.horizontalHeader()
        if isinstance(hdr, ExcelFilterHeaderView):
            hdr.sync_active(self._col_filters)

    def _filter_kw(self) -> dict:
        date_from, date_to = self._date_filters()
        kw = dict(
            truck=self._active_truck or self._selected_truck(),
            source=self._source_filter_param(),
            date_from=date_from,
            date_to=date_to,
            currency=self._currency_toggle.filter_value(),
        )
        if type(self).USE_EXCEL_COLUMN_FILTERS:
            col_filters = self._column_filters_for_query()
            if col_filters:
                kw["column_filters"] = col_filters
            kw["sort_field"] = self._sort_field
            kw["sort_asc"] = self._sort_asc
        return kw

    def _on_currency_changed(self, _currency: str) -> None:
        if not self._active_truck:
            return
        self._reset_and_load()

    def _overview_apis(self):
        from tahmeed.services.accountant_service import (
            get_truck_overview_records,
            count_truck_overview_records,
            get_truck_overview_summary,
        )
        return (
            get_truck_overview_records,
            count_truck_overview_records,
            get_truck_overview_summary,
        )

    def _pdf_report_title(self, truck: str) -> str:
        return f"{self.PAGE_TITLE} — {truck}"

    def _status_loaded_text(self, truck: str, total: int) -> str:
        currency = self._currency_toggle.filter_value()
        noun = self.ROW_NOUN if total == 1 else f"{self.ROW_NOUN}s"
        suffix = ""
        if type(self).USE_EXCEL_COLUMN_FILTERS and self._column_filters_for_query():
            suffix = "  ·  column filters applied"
        if currency:
            return f"Loaded {total:,} {currency} {noun} for {truck}.{suffix}"
        return f"Loaded {total:,} {noun} for {truck}.{self.STATUS_HINT}{suffix}"

    def _update_footer(self) -> None:
        if not self._active_truck:
            self._footer.set_text("Select a truck to load records")
        elif self._loading and self._loaded == 0:
            self._footer.set_text("Loading…")
        elif self._total == 0:
            self._footer.set_text("No matching records")
        elif self._loaded >= self._total:
            self._footer.set_text(f"Showing all {self._total:,} records")
        else:
            self._footer.set_text(
                f"Showing {self._loaded:,} of {self._total:,}  •  Scroll down for more"
            )

    def _reset_and_load(self) -> None:
        if not self._has_valid_date_range(warn=True):
            return
        truck = self._active_truck or self._selected_truck()
        if not truck:
            self._status.setText("Enter a truck number to load the overview.")
            self._table.setRowCount(0)
            self._footer.set_text("Select a truck to load records")
            return
        self._active_truck = truck
        self._reload_generation += 1
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        self._update_footer()
        from tahmeed.ui.async_utils import schedule_coro

        schedule_coro(self._load_initial(self._reload_generation))

    def _on_load_clicked(self) -> None:
        self._active_truck = self._selected_truck()
        self._reset_and_load()

    def _on_year(self, _idx: int = 0) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        else:
            app_state.fiscal_year = self._year
        sync_from_to(
            self._from_date, self._to_date, self._year, self._month, optional=True,
        )
        self._on_filter_changed()

    def _on_month(self, _idx: int = 0) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        sync_from_to(
            self._from_date, self._to_date, self._year, self._month, optional=True,
        )
        self._on_filter_changed()

    def _on_filter_changed(self) -> None:
        if not self._active_truck:
            return
        self._debounce.start()

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            from tahmeed.ui.async_utils import schedule_coro

            schedule_coro(self._load_more())

    async def _load_initial(self, generation: int) -> None:
        if self._loading:
            return
        truck = self._active_truck or self._selected_truck()
        if not truck:
            return

        self._loading = True
        self._loading_overlay.show_loading(f"Loading data for {truck}…")
        self._status.setText(f"Loading data for {truck}…")
        self._update_footer()
        try:
            get_records, count_records, get_summary = self._overview_apis()

            kw = self._filter_kw()
            records, total, summary = await asyncio.gather(
                get_records(**kw, limit=_SCROLL_CHUNK, skip=0),
                count_records(**kw),
                get_summary(**kw),
            )
            if generation != self._reload_generation:
                return

            self._active_truck = truck
            self._total = total
            period = self._format_date_range_label()
            self._subtitle.setText(
                self.LOADED_SUBTITLE.format(truck=truck, period=period)
            )
            self._records_card.set_value(f"{summary['record_count']:,}")
            self._sources_card.set_value(f"{summary['source_count']:,}")
            self._tzs_card.set_value(_fmt_amount("TZS", summary["tzs_total"]))
            self._usd_card.set_value(_fmt_amount("USD", summary["usd_total"]))
            self._zmw_card.set_value(_fmt_amount("ZMW", summary["zmw_total"]))
            self._liters_card.set_value(_fmt_num(summary["liters_total"], 0))
            self._fill_table(records, append=False)
            self._loaded = len(records)
            self._status.setText(self._status_loaded_text(truck, total))
            if type(self).USE_EXCEL_COLUMN_FILTERS:
                await self._refresh_column_filter_cache()
        except Exception as exc:
            if generation == self._reload_generation:
                self._table.setRowCount(0)
                self._status.setText(f"Failed to load {self.LOAD_ERROR_LABEL}: {exc}")
        finally:
            self._loading = False
            self._loading_overlay.hide_loading()
            if generation == self._reload_generation:
                self._update_footer()

    async def _load_more(self) -> None:
        if self._scroll_loading or self._loading:
            return
        if self._loaded >= self._total:
            return
        self._scroll_loading = True
        self._update_footer()
        try:
            get_records, _count_records, _get_summary = self._overview_apis()

            kw = self._filter_kw()
            gen = self._reload_generation
            records = await get_records(
                **kw, limit=_SCROLL_CHUNK, skip=self._loaded,
            )
            if gen != self._reload_generation:
                return
            if records:
                self._fill_table(records, append=True)
                self._loaded += len(records)
        except Exception:
            pass
        finally:
            self._scroll_loading = False
            self._update_footer()

    def _overview_cell(
        self,
        header: str,
        row: dict,
        *,
        tzs_amt,
        usd_amt,
        zmw_amt,
        amount_color: str,
        source_color: str,
        receipt: str,
        receipt_color: str,
    ) -> tuple[str, str, Optional[int]]:
        """Return (text, color, optional alignment override) for one table column."""
        date_value = row.get("date")
        if header == "DATE":
            text = (
                date_value.strftime("%d %b %Y")
                if hasattr(date_value, "strftime") and date_value.year > 1
                else "—"
            )
            return text, _T1, None
        if header == "SOURCE":
            return row.get("source", "—"), source_color, None
        if header == "DESCRIPTION":
            return row.get("description", "—"), _T1, None
        if header == "REFERENCE":
            return row.get("reference", "—"), _T1, None
        if header == "TRUCK FIELD":
            return row.get("truck_value", "—"), _T1, None
        if header == "TZS":
            return (
                _fmt_currency_cell("TZS", tzs_amt),
                amount_color if tzs_amt is not None else _TM,
                Qt.AlignRight | Qt.AlignVCenter,
            )
        if header == "USD":
            return (
                _fmt_currency_cell("USD", usd_amt),
                (_GREEN if isinstance(usd_amt, float) and usd_amt >= 0 else amount_color)
                if usd_amt is not None else _TM,
                Qt.AlignRight | Qt.AlignVCenter,
            )
        if header == "ZMW":
            return (
                _fmt_currency_cell("ZMW", zmw_amt),
                amount_color if zmw_amt is not None else _TM,
                Qt.AlignRight | Qt.AlignVCenter,
            )
        if header == "LTRS":
            return _fmt_num(row.get("liters"), 0), _T1, Qt.AlignRight | Qt.AlignVCenter
        if header == "RATE":
            return _fmt_num(row.get("rate"), 2), _T1, Qt.AlignRight | Qt.AlignVCenter
        if header == "STATION / OWNER":
            return row.get("station", "—"), _T1, None
        if header == "UPLOAD DESCRIPTION":
            return row.get("upload_description", "—"), _T2, None
        if header == "RECEIPT":
            return receipt, receipt_color, Qt.AlignCenter | Qt.AlignVCenter
        return "—", _T1, None

    def _fill_table(self, rows: list, *, append: bool = False) -> None:
        if not append:
            self._table.setRowCount(0)
        for idx, row in enumerate(rows):
            r = self._table.rowCount()
            self._table.insertRow(r)

            amount = row.get("amount")
            tzs_amt, usd_amt, zmw_amt = _amount_columns(row.get("currency", ""), amount)
            amount_color = _RED if isinstance(amount, (int, float)) and amount < 0 else _T1
            source_color = _BLUE if row.get("source_group") in self.HIGHLIGHT_SOURCE_GROUPS else _T2
            receipt = row.get("receipt_status") or "—"
            receipt_color = (
                _GREEN if receipt == "received" else
                _AMBER if receipt == "pending" else
                _RED if receipt == "missing" else
                _TM
            )

            for col_idx, (header, _width, align) in enumerate(self.TABLE_COLS):
                text, color, align_override = self._overview_cell(
                    header,
                    row,
                    tzs_amt=tzs_amt,
                    usd_amt=usd_amt,
                    zmw_amt=zmw_amt,
                    amount_color=amount_color,
                    source_color=source_color,
                    receipt=receipt,
                    receipt_color=receipt_color,
                )
                self._table.setItem(
                    r,
                    col_idx,
                    _cell(text, align=align_override or align, color=color),
                )
            _finish_table_row(self._table, r)

            # Slightly tint ZMW rows so Zambia-related entries stand out.
            if _normalize_currency(row.get("currency") or "") == "ZMW":
                for c in range(self._table.columnCount()):
                    item = self._table.item(r, c)
                    if item is not None:
                        item.setBackground(QColor(_BLUE_L))

    def _show_export_busy(self, message: str) -> None:
        self._loading_overlay.show_loading(message)
        self._status.setText(message)

    def _hide_export_busy(self) -> None:
        self._loading_overlay.hide_loading()

    def _export_excel(self) -> None:
        self._start_background_task(self._do_export_excel(), "Excel export")

    async def _do_export_excel(self) -> None:
        truck = self._active_truck or self._selected_truck()
        if not truck:
            QMessageBox.warning(self, "Export", "Select a truck first.")
            return
        if not self._has_valid_date_range(warn=True):
            return
        try:
            import openpyxl  # noqa: F401 — validate dependency before fetching
        except ImportError:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n\nRun: pip install openpyxl",
            )
            return

        get_records, _count_records, get_summary = self._overview_apis()

        self._show_export_busy(f"Preparing Excel export for {truck}…")
        rows = None
        summary = None
        prepare_error = None
        try:
            kw = self._filter_kw()
            rows = await get_records(**kw, limit=100000, skip=0)
            summary = await get_summary(**kw)
        except Exception as exc:
            prepare_error = exc
        finally:
            self._hide_export_busy()

        if prepare_error is not None:
            QMessageBox.critical(
                self, "Export Error", f"Failed to prepare export data:\n{prepare_error}"
            )
            return
        if not rows:
            QMessageBox.information(
                self,
                "Export",
                "No records match the current truck, toolbar, and column filters.",
            )
            return

        source_tag = self._source_filter_tag()
        currency = self._currency_toggle.filter_value() or "ALL"
        default_name = (
            f"{self.EXPORT_PREFIX}_{truck}_{source_tag}_{currency}.xlsx".replace(" ", "_")
        )
        path, _ = QFileDialog.getSaveFileName(
            self, self.EXCEL_DIALOG_TITLE, default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"

        self._show_export_busy(f"Writing Excel export for {truck}…")
        write_error = None
        try:
            await asyncio.to_thread(
                _write_truck_overview_excel,
                path,
                truck,
                self._source_filter_label(),
                self._format_date_range_label(),
                rows,
                summary,
                sheet_title=self.EXCEL_SHEET_TITLE,
                banner_label=self.EXCEL_BANNER,
                cols=self.TABLE_COLS,
            )
        except Exception as exc:
            write_error = exc
        finally:
            self._hide_export_busy()

        if write_error is not None:
            QMessageBox.critical(self, "Export Error", f"Could not save file:\n{write_error}")
            self._status.setText(f"Excel export failed for {truck}.")
            return
        QMessageBox.information(self, "Export Complete", f"Excel report saved to:\n{path}")
        self._status.setText(f"Excel export saved for {truck}.")

    def _export_pdf(self) -> None:
        self._start_background_task(self._do_export_pdf(), "PDF export")

    async def _do_export_pdf(self) -> None:
        truck = self._active_truck or self._selected_truck()
        if not truck:
            QMessageBox.warning(self, "Export", "Select a truck first.")
            return
        if not self._has_valid_date_range(warn=True):
            return

        get_records, _count_records, get_summary = self._overview_apis()

        self._show_export_busy(f"Preparing PDF export for {truck}…")
        rows = None
        summary = None
        prepare_error = None
        try:
            kw = self._filter_kw()
            rows = await get_records(**kw, limit=100000, skip=0)
            summary = await get_summary(**kw)
        except Exception as exc:
            prepare_error = exc
        finally:
            self._hide_export_busy()

        if prepare_error is not None:
            QMessageBox.critical(
                self, "Export Error", f"Failed to prepare export data:\n{prepare_error}"
            )
            return
        if not rows:
            QMessageBox.information(
                self,
                "Export",
                "No records match the current truck, toolbar, and column filters.",
            )
            return

        currency = self._currency_toggle.filter_value() or "ALL"
        default_name = f"{self.EXPORT_PREFIX}_{truck}_{currency}_Report.pdf".replace(" ", "_")
        path, _ = QFileDialog.getSaveFileName(
            self, self.PDF_DIALOG_TITLE, default_name, "PDF Files (*.pdf)"
        )
        if not path:
            return
        if not path.endswith(".pdf"):
            path += ".pdf"

        self._show_export_busy(f"Writing PDF export for {truck}…")
        write_error = None
        try:
            from tahmeed.services.truck_overview_pdf import export_truck_overview_pdf

            date_from, date_to = self._date_filters()
            await asyncio.to_thread(
                export_truck_overview_pdf,
                path,
                truck=truck,
                rows=rows,
                summary=summary,
                date_from=date_from,
                date_to=date_to,
                source_label=self._source_filter_label(),
                eyebrow=self.PDF_EYEBROW,
                report_title=self._pdf_report_title(truck),
                subtitle=self.PDF_SUBTITLE,
            )
        except Exception as exc:
            write_error = exc
        finally:
            self._hide_export_busy()

        if write_error is not None:
            QMessageBox.critical(self, "Export Error", f"Could not create PDF:\n{write_error}")
            self._status.setText(f"PDF export failed for {truck}.")
            return
        QMessageBox.information(self, "Export Complete", f"PDF report saved to:\n{path}")
        self._status.setText(f"PDF export saved for {truck}.")

    async def _run_with_polls_paused(self, coro):
        """Run export work with connectivity / badge polls paused (Py3.14 + qasync)."""
        from tahmeed.ui.async_utils import pause_background_polls

        with pause_background_polls(self):
            return await coro

    def _start_background_task(self, coro, action: str) -> None:
        from tahmeed.ui.async_utils import create_task, in_running_task

        wrapped = self._run_with_polls_paused(coro)
        if in_running_task():
            def _kick(c=wrapped, a=action) -> None:
                task = create_task(c)
                task.add_done_callback(lambda t: self._handle_task_result(t, a))

            from PySide6.QtCore import QTimer

            QTimer.singleShot(0, _kick)
            return
        task = create_task(wrapped)
        task.add_done_callback(lambda t: self._handle_task_result(t, action))

    def _handle_task_result(self, task: asyncio.Task, action: str) -> None:
        if task.cancelled():
            return
        try:
            exc = task.exception()
        except Exception as err:
            QMessageBox.critical(self, action, f"{action} failed:\n{err}")
            return
        if exc is not None:
            QMessageBox.critical(self, action, f"{action} failed:\n{exc}")

    def _format_date_range_label(self) -> str:
        date_from, date_to = self._date_filters()
        if date_from and date_to:
            return f"{date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}"
        if date_from:
            return f"From {date_from.strftime('%d %b %Y')}"
        if date_to:
            return f"Up to {date_to.strftime('%d %b %Y')}"
        return "All Dates"

    def _clear_results(self) -> None:
        self._debounce.stop()
        self._active_truck = ""
        self._loaded = 0
        self._total = 0
        self._reload_generation += 1
        self._truck_edit.clear()
        self._source_cb.reset_to_all()
        # Keep fiscal-year scope after clear so the next load stays bounded.
        self._year = app_state.fiscal_year
        self._month = 0
        self._year_cb.blockSignals(True)
        self._month_cb.blockSignals(True)
        try:
            yr_idx = self._year_cb.findData(self._year)
            self._year_cb.setCurrentIndex(yr_idx if yr_idx >= 0 else 0)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(self._year > 0)
        finally:
            self._year_cb.blockSignals(False)
            self._month_cb.blockSignals(False)
        sync_from_to(
            self._from_date, self._to_date, self._year, self._month, optional=True,
        )
        if type(self).USE_EXCEL_COLUMN_FILTERS:
            self._clear_column_filters()
        self._subtitle.setText(self.SUBTITLE_EMPTY)
        self._status.setText(self.STATUS_EMPTY)
        self._table.setRowCount(0)
        self._footer.set_text("Select a truck to load records")
        self._records_card.set_value("—")
        self._sources_card.set_value("—")
        self._tzs_card.set_value("—")
        self._usd_card.set_value("—")
        self._zmw_card.set_value("—")
        self._liters_card.set_value("—")
