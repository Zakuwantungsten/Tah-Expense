"""AccountantDashboard — Master Expenses Table (Task 5).

QuickBooks-style full verified ledger with:
  - FY year selector (persists in app_state)
  - Month tab bar (All · Jan–Dec) with per-month TZS totals
  - Single unified table (Bonds / Diesel Cash row styling), ITEM column sourced
    from the approved cashier entry's category
  - Excel-style ▾ header filters (sort / search / multi-select / Apply)
  - ReceiptBadge color pill
  - Split TZS + USD amount columns + footer totals
  - Ref_Float column (cashier free-text; empty when unset)
  - Export to Excel — Export Filtered (current view) or Export All (FY/month tab)
"""

from __future__ import annotations

import asyncio
from datetime import datetime
import calendar
from typing import Dict, List, Optional, Set

import qtawesome as qta

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame, QLabel,
    QTableWidget, QTableWidgetItem,
    QLineEdit, QComboBox, QPushButton, QToolButton, QMenu, QDateEdit,
    QMessageBox, QFileDialog, QAbstractItemView, QDialog,
    QFormLayout, QDialogButtonBox,
)
from PySide6.QtCore import Qt, Signal, QTimer, QSize, QDate
from PySide6.QtGui import QColor, QAction

from tahmeed.models.transaction import Transaction
from tahmeed.app_state import app_state
from tahmeed.ui.widgets.loading_overlay import LoadingOverlay
from tahmeed.ui.widgets.split_export_button import make_export_menu_btn
from tahmeed.ui.widgets.export_runner import (
    attach_export_overlay,
    export_file_ready,
    fetch_records_with_progress,
    hide_export_busy,
    normalize_xlsx_path,
    notify_export_error,
    notify_export_info,
    pick_export_path,
    run_export_write,
    show_export_busy,
    FAST_STYLE_ROW_LIMIT,
    PROGRESS_EVERY,
)
from tahmeed.ui.accountant.date_filters import style_calendar_popup
from tahmeed.ui.accountant.separate_expenses import _fmt_num, _SCROLL_CHUNK
from tahmeed.ui.accountant.master_ledger_table import MasterLedgerTable

# ── Design tokens ──────────────────────────────────────────────────────────────
_WHITE   = "#FFFFFF"
_BG      = "#F4F6F8"
_BORDER  = "#E5E7EB"
_BLUE    = "#0077C5"
_BLUE_L  = "#EFF6FF"
_GREEN   = "#16A34A"
_GREEN_L = "#DCFCE7"
_AMBER   = "#D97706"
_AMBER_L = "#FEF3C7"
_RED     = "#DC2626"
_RED_L   = "#FEE2E2"
_T1      = "#111827"
_T2      = "#6B7280"
_TM      = "#9CA3AF"
_HDR_BG  = "#F1F5F9"
_ALT_ROW = "#F9FAFB"
_NAVY    = "#1B2B4B"

_ROW_H = 28
_MIN_FILTER_DATE = QDate(2000, 1, 1)

# (label, align, mongo sort field or None)
_COLS = [
    ("S/NO",           "center", None),
    ("DATE",           "left",   "date"),
    ("ITEM",           "left",   "category_name"),
    ("DESCRIPTION",    "left",   "description"),
    ("TRUCK NO",       "left",   "truck_number"),
    ("MEMO",           "left",   "memo"),
    ("REF_FLOAT",      "left",   "ref_float"),
    ("TZS",            "right",  "amount"),
    ("USD",            "right",  "amount"),
    ("RECEIPT",        "center", "receipt_status"),
    ("OWNERSHIP",      "left",   "ownership"),
    ("APPROVED BY",    "left",   "approver"),
    ("CASHIER",        "left",   None),
]

# Default pixel widths; 0 = stretch (DESCRIPTION fills remaining space).
_COL_DEFAULTS = [52, 72, 110, 0, 95, 120, 110, 110, 100, 100, 90, 100, 100]
_DESC_COL = 3
_COL_DATE = 1
_COL_ITEM = 2
_COL_DESC = 3
_COL_TRUCK = 4
_COL_MEMO = 5
_COL_REF = 6
_COL_TZS = 7
_COL_USD = 8
_COL_RCPT = 9
_COL_OWN = 10
_COL_APP = 11
_COL_CASH = 12

# Columns accountants may edit (S/NO + CASHIER stay read-only).
_EDITABLE_COLS: Set[int] = {
    _COL_DATE, _COL_ITEM, _COL_DESC, _COL_TRUCK, _COL_MEMO, _COL_REF,
    _COL_TZS, _COL_USD, _COL_RCPT, _COL_OWN, _COL_APP,
}

_TX_ID_ROLE = Qt.UserRole

_RECEIPT_EDIT_OPTIONS = [
    ("received", "Received"),
    ("pending", "Pending"),
    ("no_receipt", "No Receipt"),
]

_FILTERABLE_COLS: Set[int] = set(range(len(_COLS))) - {0}
_SORT_KINDS = {
    1: "date",
    _COL_TZS: "number",
    _COL_USD: "number",
}

_MONTHS = [
    (0,  "All"),
    (1,  "Jan"), (2,  "Feb"), (3,  "Mar"),
    (4,  "Apr"), (5,  "May"), (6,  "Jun"),
    (7,  "Jul"), (8,  "Aug"), (9,  "Sep"),
    (10, "Oct"), (11, "Nov"), (12, "Dec"),
]

_RECEIPT_MAP = {
    "received":   ("Received",   _GREEN),
    "pending":    ("Pending",    _AMBER),
    "missing":    ("No Receipt", _RED),
    "no_receipt": ("No Receipt", _RED),
}

_TABLE_SS = (
    f"QTableWidget {{"
    f"  background: {_WHITE};"
    f"  gridline-color: {_BORDER};"
    f"  font-size: 11px;"
    f"  font-family:'Segoe UI';"
    f"  color: {_T1};"
    f"  border: none;"
    f"  selection-background-color: #DBEAFE;"
    f"  selection-color: {_T1};"
    f"}}"
    f"QTableWidget::item {{"
    f"  padding: 2px 8px;"
    f"  border: none;"
    f"}}"
    f"QHeaderView::section {{"
    f"  background: {_HDR_BG};"
    f"  color: {_T2};"
    f"  font-size: 10px;"
    f"  font-weight: 600;"
    f"  font-family:'Segoe UI';"
    f"  border: none;"
    f"  border-bottom: 1px solid {_BORDER};"
    f"  padding: 0 18px 0 8px;"
    f"  min-height: 28px;"
    f"}}"
    f"QHeaderView::section:hover {{ background: #E2E8F0; }}"
    f"QScrollBar:vertical {{"
    f"  background: transparent; width: 8px; margin: 0;"
    f"}}"
    f"QScrollBar::handle:vertical {{"
    f"  background: #D1D5DB; border-radius: 4px; min-height: 24px;"
    f"}}"
    f"QScrollBar::add-line, QScrollBar::sub-line {{ width: 0; height: 0; }}"
)


# ── Helpers ────────────────────────────────────────────────────────────────────

def _lbl(text: str = "", size: int = 13, weight: int = 400,
         color: str = _T1) -> QLabel:
    w = QLabel(text)
    w.setStyleSheet(
        f"color: {color}; font-size: {size}px; font-weight: {weight};"
        " font-family:'Segoe UI'; background: transparent;"
    )
    return w


def _input_ss() -> str:
    return (
        f"QLineEdit, QComboBox, QDateEdit {{"
        f"  border: 1px solid {_BORDER}; border-radius: 5px;"
        f"  background: {_WHITE}; color: {_T1}; font-size: 12px;"
        "  font-family:'Segoe UI'; padding: 0 8px;"
        "  min-height: 32px; max-height: 32px; }}"
        f"QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{ border-color: {_BLUE}; }}"
        "QComboBox::drop-down { border: none; width: 20px; }"
    )


def _qdate_to_dt_start(qd: QDate) -> datetime:
    return datetime(qd.year(), qd.month(), qd.day())


def _qdate_to_dt_end(qd: QDate) -> datetime:
    return datetime(qd.year(), qd.month(), qd.day(), 23, 59, 59)


def _normalize_receipt(status: str) -> str:
    """Map stored receipt_status values to canonical keys."""
    s = (status or "pending").strip().lower().replace(" ", "_")
    if s in ("no_receipt", "no", "n/a", "none"):
        return "no_receipt"
    if s in _RECEIPT_MAP:
        return s
    if "received" in s or s in ("yes", "receipt", "rcvd"):
        return "received"
    if "missing" in s or "no receipt" in s:
        return "no_receipt"
    return "pending"


def _fmt_short(amount: float) -> str:
    """Format large numbers compactly for the tab bar: 847,341,200 → 847M."""
    a = abs(amount)
    if a >= 1_000_000:
        return f"{amount / 1_000_000:.1f}M"
    if a >= 1_000:
        return f"{amount / 1_000:.0f}K"
    return f"{amount:,.0f}"


def _action_btn(text: str, icon_name: str, primary: bool = True) -> QPushButton:
    b = QPushButton(f"  {text}")
    b.setCursor(Qt.PointingHandCursor)
    b.setFixedHeight(32)
    try:
        b.setIcon(qta.icon(icon_name, color="#FFFFFF" if primary else _T2))
        b.setIconSize(QSize(15, 15))
    except Exception:
        pass
    if primary:
        ss = (
            f"QPushButton {{ background: {_BLUE}; color: #FFF; border: none;"
            " border-radius: 5px; font-size: 12px; font-weight: 600;"
            " font-family:'Segoe UI'; padding: 0 12px; }}"
            "QPushButton:hover { background: #005EA3; }"
            "QPushButton:disabled { background: #93C5FD; }"
        )
    else:
        ss = (
            f"QPushButton {{ background: {_WHITE}; color: {_T1};"
            f" border: 1px solid {_BORDER};"
            " border-radius: 5px; font-size: 12px;"
            " font-family:'Segoe UI'; padding: 0 12px; }}"
            f"QPushButton:hover {{ background: {_BG}; }}"
        )
    b.setStyleSheet(ss)
    return b


def _export_menu_btn(
    on_filtered,
    on_all,
    *,
    parent=None,
) -> QToolButton:
    """Export split button: main click = filtered, menu = filtered + all."""
    return make_export_menu_btn(
        on_filtered,
        on_all,
        parent=parent,
        height=32,
        btn_tip=(
            "Export Filtered — current search, filters, column filters, and sort.\n"
            "Use the ▾ menu for Export All (full FY/month tab, no extra filters)."
        ),
        filtered_tip="Export rows matching the current filters and sort order on screen.",
        all_tip=(
            "Export every record for the selected fiscal year and month tab "
            "(ignores search, toolbar filters, and column filters)."
        ),
    )


def _receipt_text(status: str) -> tuple:
    """Return (display text, text color) for a receipt status."""
    key = _normalize_receipt(status)
    text, fg = _RECEIPT_MAP.get(key, ("Pending", _AMBER))
    return text, fg


def _fmt_tx_date(dt: Optional[datetime]) -> str:
    return dt.strftime("%d %b") if dt else "—"


def _currency_key(tx: Transaction) -> str:
    return (tx.currency or "TZS").upper()


def _is_tzs(currency: str) -> bool:
    return currency in {"TZS", "TSH", "TZ"}


def _amount_cells(tx: Transaction) -> tuple[str, str]:
    """Return (tzs_text, usd_text) — both may be filled on the same row."""
    tzs, usd = tx.money_parts()
    tzs_txt = _fmt_num(tzs, "", 0) if tzs else "—"
    usd_txt = _fmt_num(usd, "", 2) if usd else "—"
    # Other non-TZS/USD currencies stay visible in the TZS column with a prefix.
    cur = _currency_key(tx)
    if cur not in {"TZS", "TSH", "TZ", "USD"} and tx.amount_usd is None:
        return _fmt_num(tx.amount, f"{cur} ", 2), "—"
    return tzs_txt, usd_txt


def _ref_float_display(tx: Transaction) -> str:
    """Show cashier Ref_Float text; fall back for older notes_flag-only rows."""
    text = (getattr(tx, "ref_float", None) or "").strip()
    if text:
        return text.upper()
    if tx.notes_flag:
        return "REFUND TO FLOAT"
    return ""


def _short_name(name: str) -> str:
    parts = (name or "").strip().split()
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1][0]}."
    return name or "—"



class _MasterBulkItemDialog(QDialog):
    """Bulk assign Item (category) to selected Master rows — picker only."""

    def __init__(self, row_count: int, item_names: list, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle(f"Set Item ({row_count} rows)")
        self.setMinimumWidth(380)
        self.setStyleSheet(
            f"QDialog {{ background: {_WHITE}; }}"
            f"QLabel {{ color: {_T1}; font-size: 12px; font-family:'Segoe UI'; }}"
            f"QComboBox {{ border: 1px solid {_BORDER}; border-radius: 5px;"
            f"  background: {_WHITE}; color: {_T1}; font-size: 12px;"
            "  font-family:'Segoe UI'; padding: 0 8px; min-height: 30px; }"
        )
        form = QFormLayout()
        form.setContentsMargins(16, 16, 16, 8)
        form.setSpacing(10)
        self._item = QComboBox()
        self._item.setEditable(True)
        self._item.setInsertPolicy(QComboBox.NoInsert)
        names = sorted({(n or "").strip() for n in item_names if (n or "").strip()}, key=str.lower)
        self._item.addItems(names)
        form.addRow("Item", self._item)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        root = QVBoxLayout(self)
        root.addLayout(form)
        root.addWidget(buttons)

    def item_name(self) -> str:
        return self._item.currentText().strip()


class _MonthTabBar(QFrame):
    month_selected = Signal(int)  # 0=All, 1-12

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setFixedHeight(44)
        self.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-bottom: 1px solid {_BORDER}; }}"
        )
        self._active_month = 0
        self._buttons: Dict[int, QPushButton] = {}
        self._build()

    def _build(self) -> None:
        hl = QHBoxLayout(self)
        hl.setContentsMargins(16, 6, 16, 6)
        hl.setSpacing(6)

        for idx, label in _MONTHS:
            btn = QPushButton(label)
            btn.setCheckable(False)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setFixedHeight(28)
            btn.setMinimumWidth(44)
            self._apply_btn_style(btn, active=(idx == 0))
            btn.clicked.connect(lambda _=False, i=idx: self._on_click(i))
            self._buttons[idx] = btn
            hl.addWidget(btn)

        hl.addStretch()

    def _apply_btn_style(self, btn: QPushButton, active: bool) -> None:
        if active:
            btn.setStyleSheet(
                f"QPushButton {{ background: {_BLUE}; color: #fff;"
                " border: none; border-radius: 14px;"
                " font-size: 11px; font-weight: 700;"
                " font-family:'Segoe UI'; padding: 0 10px; }}"
            )
        else:
            btn.setStyleSheet(
                f"QPushButton {{ background: transparent; color: {_T2};"
                f" border: 1px solid {_BORDER}; border-radius: 14px;"
                " font-size: 11px;"
                " font-family:'Segoe UI'; padding: 0 10px; }}"
                f"QPushButton:hover {{ background: {_BG}; color: {_T1}; }}"
            )

    def _on_click(self, month: int) -> None:
        if month == self._active_month:
            return
        old_btn = self._buttons.get(self._active_month)
        if old_btn:
            self._apply_btn_style(old_btn, active=False)
        self._active_month = month
        self._apply_btn_style(self._buttons[month], active=True)
        self.month_selected.emit(month)

    def update_totals(self, month_data: Dict[int, dict]) -> None:
        """Update button labels with TZS totals once loaded."""
        for idx, label in _MONTHS:
            if idx == 0:
                continue
            btn = self._buttons.get(idx)
            if btn is None:
                continue
            data = month_data.get(idx)
            if data and data.get("tzs"):
                btn.setText(f"{label}  {_fmt_short(data['tzs'])}")
                btn.setMinimumWidth(64)
            else:
                btn.setText(label)
                btn.setMinimumWidth(44)

    def active_month(self) -> int:
        return self._active_month

    def reset(self) -> None:
        self._on_click(0)


# ── Filter Bar ─────────────────────────────────────────────────────────────────

class _FilterBar(QFrame):
    filter_changed = Signal()
    filters_cleared = Signal()
    import_requested = Signal()

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._sync_year = 0
        self._sync_month = 0
        self.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-bottom: 1px solid {_BORDER}; }}"
        )
        self._build()

    def _build(self) -> None:
        hl = QHBoxLayout(self)
        hl.setContentsMargins(16, 8, 16, 8)
        hl.setSpacing(10)

        try:
            si = QLabel()
            si.setFixedSize(16, 16)
            si.setPixmap(qta.icon("mdi.magnify", color=_TM).pixmap(16, 16))
            si.setStyleSheet("background: transparent;")
            hl.addWidget(si)
        except Exception:
            pass

        self._search = QLineEdit()
        self._search.setPlaceholderText("Search description or truck…")
        self._search.setMinimumWidth(180)
        self._search.setStyleSheet(_input_ss())
        # Drop signal args — filter_changed / import/export are zero-arg Signals.
        self._search.textChanged.connect(lambda *_: self.filter_changed.emit())
        hl.addWidget(self._search, 1)

        self._truck_cb = QComboBox()
        self._truck_cb.addItem("All Trucks")
        self._truck_cb.setMinimumWidth(110)
        self._truck_cb.setStyleSheet(_input_ss())
        self._truck_cb.currentTextChanged.connect(lambda *_: self.filter_changed.emit())
        hl.addWidget(self._truck_cb)

        self._cat_cb = QComboBox()
        self._cat_cb.addItem("All Categories")
        self._cat_cb.setMinimumWidth(130)
        self._cat_cb.setStyleSheet(_input_ss())
        self._cat_cb.currentTextChanged.connect(lambda *_: self.filter_changed.emit())
        hl.addWidget(self._cat_cb)

        self._rcpt_cb = QComboBox()
        for item in [("All Receipts", "all"), ("Received", "received"),
                     ("Pending", "pending"), ("No Receipt", "missing")]:
            self._rcpt_cb.addItem(item[0], item[1])
        self._rcpt_cb.setMinimumWidth(110)
        self._rcpt_cb.setStyleSheet(_input_ss())
        self._rcpt_cb.currentIndexChanged.connect(lambda *_: self.filter_changed.emit())
        hl.addWidget(self._rcpt_cb)

        hl.addWidget(_lbl("From", size=12, color=_T2))
        self._from_date = QDateEdit()
        self._from_date.setCalendarPopup(True)
        self._from_date.setDisplayFormat("dd MMM yyyy")
        self._from_date.setMinimumDate(_MIN_FILTER_DATE)
        self._from_date.setSpecialValueText("From")
        self._from_date.setDate(_MIN_FILTER_DATE)
        self._from_date.setFixedWidth(120)
        self._from_date.setStyleSheet(_input_ss())
        style_calendar_popup(self._from_date)
        self._from_date.dateChanged.connect(lambda *_: self.filter_changed.emit())
        hl.addWidget(self._from_date)

        hl.addWidget(_lbl("To", size=12, color=_T2))
        self._to_date = QDateEdit()
        self._to_date.setCalendarPopup(True)
        self._to_date.setDisplayFormat("dd MMM yyyy")
        self._to_date.setMinimumDate(_MIN_FILTER_DATE)
        self._to_date.setSpecialValueText("To")
        self._to_date.setDate(_MIN_FILTER_DATE)
        self._to_date.setFixedWidth(120)
        self._to_date.setStyleSheet(_input_ss())
        style_calendar_popup(self._to_date)
        self._to_date.dateChanged.connect(lambda *_: self.filter_changed.emit())
        hl.addWidget(self._to_date)

        clear_btn = _action_btn("Clear", "mdi.filter-remove-outline", primary=False)
        clear_btn.setToolTip("Clear search, filters, column filters, and column sort.")
        clear_btn.clicked.connect(lambda *_: self.clear_filters())
        hl.addWidget(clear_btn)

        import_btn = _action_btn("Import Excel", "mdi.file-upload-outline", primary=False)
        import_btn.clicked.connect(lambda *_: self.import_requested.emit())
        hl.addWidget(import_btn)

    def set_date_range(self, year: int, month: int = 0) -> None:
        """Sync date pickers to FY or month selection without firing filters."""
        self._sync_year = year
        self._sync_month = month
        if month and 1 <= month <= 12:
            last = calendar.monthrange(year, month)[1]
            start = QDate(year, month, 1)
            end = QDate(year, month, last)
        else:
            start = QDate(year, 1, 1)
            end = QDate(year, 12, 31)
        self._from_date.blockSignals(True)
        self._to_date.blockSignals(True)
        self._from_date.setDate(start)
        self._to_date.setDate(end)
        self._from_date.blockSignals(False)
        self._to_date.blockSignals(False)

    def clear_filters(self) -> None:
        """Reset search/combos and clear From/To back to unset placeholders."""
        self._search.blockSignals(True)
        self._truck_cb.blockSignals(True)
        self._cat_cb.blockSignals(True)
        self._rcpt_cb.blockSignals(True)
        self._from_date.blockSignals(True)
        self._to_date.blockSignals(True)
        try:
            self._search.clear()
            self._truck_cb.setCurrentIndex(0)
            self._cat_cb.setCurrentIndex(0)
            self._rcpt_cb.setCurrentIndex(0)
            self._from_date.setDate(_MIN_FILTER_DATE)
            self._to_date.setDate(_MIN_FILTER_DATE)
        finally:
            self._search.blockSignals(False)
            self._truck_cb.blockSignals(False)
            self._cat_cb.blockSignals(False)
            self._rcpt_cb.blockSignals(False)
            self._from_date.blockSignals(False)
            self._to_date.blockSignals(False)
        self.filters_cleared.emit()

    def populate_trucks(self, trucks: List[str]) -> None:
        cur = self._truck_cb.currentText()
        self._truck_cb.blockSignals(True)
        self._truck_cb.clear()
        self._truck_cb.addItem("All Trucks")
        for t in trucks:
            self._truck_cb.addItem(t)
        idx = self._truck_cb.findText(cur)
        self._truck_cb.setCurrentIndex(max(0, idx))
        self._truck_cb.blockSignals(False)

    def populate_categories(self, cats: List[str]) -> None:
        cur = self._cat_cb.currentText()
        self._cat_cb.blockSignals(True)
        self._cat_cb.clear()
        self._cat_cb.addItem("All Categories")
        for c in cats:
            self._cat_cb.addItem(c)
        idx = self._cat_cb.findText(cur)
        self._cat_cb.setCurrentIndex(max(0, idx))
        self._cat_cb.blockSignals(False)

    def search_text(self) -> str:
        return self._search.text().strip()

    def truck_filter(self) -> str:
        t = self._truck_cb.currentText()
        return "" if t == "All Trucks" else t

    def category_filter(self) -> str:
        c = self._cat_cb.currentText()
        return "" if c == "All Categories" else c

    def receipt_filter(self) -> str:
        return self._rcpt_cb.currentData() or "all"

    def date_from(self) -> Optional[datetime]:
        if self._from_date.date() <= _MIN_FILTER_DATE:
            return None
        return _qdate_to_dt_start(self._from_date.date())

    def date_to(self) -> Optional[datetime]:
        if self._to_date.date() <= _MIN_FILTER_DATE:
            return None
        return _qdate_to_dt_end(self._to_date.date())


# ── Unified ledger table (single QTableWidget, Bonds / Diesel Cash styling) ────

# Ledger grid lives in master_ledger_table (Excel-like edit/copy/paste).
_LedgerTable = MasterLedgerTable


class _FooterBar(QFrame):
    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setFixedHeight(40)
        self.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 2px solid {_BORDER}; }}"
        )
        hl = QHBoxLayout(self)
        hl.setContentsMargins(16, 0, 16, 0)
        hl.setSpacing(24)

        hl.addWidget(_lbl("TOTAL (filtered view)", size=11, weight=700, color=_T2))
        hl.addSpacing(8)

        self._tzs_lbl = _lbl("TZS  —", size=13, weight=700, color=_T1)
        self._tzs_lbl.setStyleSheet(
            f"color: {_T1}; font-size: 13px; font-weight: 700;"
            " font-family: 'Cascadia Code', 'Consolas', monospace;"
            " background: transparent;"
        )
        hl.addWidget(self._tzs_lbl)

        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setFixedWidth(1)
        sep.setStyleSheet(f"background: {_BORDER};")
        hl.addWidget(sep)

        self._usd_lbl = _lbl("USD  —", size=13, weight=700, color=_T1)
        self._usd_lbl.setStyleSheet(
            f"color: {_T1}; font-size: 13px; font-weight: 700;"
            " font-family: 'Cascadia Code', 'Consolas', monospace;"
            " background: transparent;"
        )
        hl.addWidget(self._usd_lbl)

        hl.addStretch()

        self._count_lbl = _lbl("", size=11, color=_T2)
        hl.addWidget(self._count_lbl)

    def update_totals(self, tzs: float, usd: float, record_count: int) -> None:
        self._tzs_lbl.setText(f"TZS  {tzs:,.0f}")
        self._tzs_lbl.setStyleSheet(
            f"color: {_RED if tzs < 0 else _T1}; font-size: 13px; font-weight: 700;"
            " font-family: 'Cascadia Code', 'Consolas', monospace;"
            " background: transparent;"
        )
        usd_prefix = "$" if usd != 0 else ""
        self._usd_lbl.setText(f"USD  {usd_prefix}{usd:,.2f}" if usd else "USD  —")
        self._count_lbl.setText(f"{record_count:,} records")


# ── MasterExpensesWidget (public) ─────────────────────────────────────────────

class MasterExpensesWidget(QWidget):
    """Full verified ledger — Task 5, Master Expenses Table."""

    def __init__(self, user=None, parent=None) -> None:
        super().__init__(parent)
        self._user = user
        self._year = app_state.fiscal_year
        self._month = 0
        self._loaded = 0
        self._total = 0
        self._sort_field = "date"
        self._sort_asc = False
        self._loading = False
        self._scroll_loading = False
        self._filters_loaded = False
        self._reload_generation = 0
        self._import_in_flight = False
        self._export_in_flight = False
        self._item_names: List[str] = []
        self._item_by_name: Dict = {}
        self._edit_in_flight = False

        self._debounce = QTimer(self)
        self._debounce.setSingleShot(True)
        self._debounce.setInterval(350)
        self._debounce.timeout.connect(self._reset_and_load)

        self._build()
        self._filter_bar.set_date_range(self._year, self._month)

    # ── Build ─────────────────────────────────────────────────────────────

    def _build(self) -> None:
        self.setStyleSheet(f"background: {_BG};")
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Title bar ────────────────────────────────────────────────────
        title_bar = QFrame()
        title_bar.setFixedHeight(52)
        title_bar.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-bottom: 1px solid {_BORDER}; }}"
        )
        tb = QHBoxLayout(title_bar)
        tb.setContentsMargins(20, 0, 20, 0)
        tb.setSpacing(12)

        try:
            icon_lbl = QLabel()
            icon_lbl.setFixedSize(22, 22)
            icon_lbl.setPixmap(qta.icon("mdi.table-large", color=_BLUE).pixmap(22, 22))
            icon_lbl.setStyleSheet("background: transparent;")
            tb.addWidget(icon_lbl)
        except Exception:
            pass

        tb.addWidget(_lbl("Master Expenses", size=16, weight=700))

        self._count_lbl = _lbl("", size=12, color=_T2)
        tb.addWidget(self._count_lbl)

        tb.addStretch()

        # FY selector — years filled dynamically from Master data on load
        tb.addWidget(_lbl("FY", size=12, color=_T2))
        self._fy_cb = QComboBox()
        current_yr = datetime.now().year
        for yr in range(current_yr + 1, current_yr - 4, -1):
            self._fy_cb.addItem(str(yr), yr)
        self._fy_cb.setCurrentIndex(
            self._fy_cb.findData(self._year) if self._fy_cb.findData(self._year) >= 0
            else self._fy_cb.findData(current_yr)
        )
        self._fy_cb.setFixedWidth(80)
        self._fy_cb.setStyleSheet(_input_ss())
        self._fy_cb.setToolTip(
            "Calendar year for Master (by Excel transaction date). "
            "Use Uploads / Open upload to see a full daily import batch together."
        )
        self._fy_cb.currentIndexChanged.connect(self._on_year_changed)
        tb.addWidget(self._fy_cb)

        self._edit_btn = _action_btn("Edit", "mdi.pencil", primary=False)
        self._edit_btn.setToolTip(
            "Unlock the grid for Excel-style editing (type in cells, copy/paste, Ctrl+D fill)."
        )
        self._edit_btn.clicked.connect(self._on_toggle_edit)
        tb.addWidget(self._edit_btn)

        self._save_btn = _action_btn("Save", "mdi.content-save", primary=True)
        self._save_btn.setEnabled(False)
        self._save_btn.setToolTip("Save dirty Master rows (Ctrl+S in the grid).")
        self._save_btn.clicked.connect(self._on_save_edits)
        tb.addWidget(self._save_btn)

        export_btn = _export_menu_btn(
            self._on_export_filtered, self._on_export_all, parent=self,
        )
        tb.addWidget(export_btn)

        # Refresh
        refresh_btn = QPushButton()
        refresh_btn.setFixedSize(32, 32)
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(
            "QPushButton { background: transparent; border: none; border-radius: 4px; }"
            f"QPushButton:hover {{ background: {_BG}; }}"
        )
        try:
            refresh_btn.setIcon(qta.icon("mdi.refresh", color=_T2))
            refresh_btn.setIconSize(QSize(18, 18))
        except Exception:
            refresh_btn.setText("↻")
        refresh_btn.clicked.connect(self.refresh)
        tb.addWidget(refresh_btn)

        root.addWidget(title_bar)

        # ── Month tab bar ────────────────────────────────────────────────
        self._month_bar = _MonthTabBar()
        self._month_bar.month_selected.connect(self._on_month_changed)
        root.addWidget(self._month_bar)

        # ── Filter bar ───────────────────────────────────────────────────
        self._filter_bar = _FilterBar()
        self._filter_bar.filter_changed.connect(self._on_filter_changed)
        self._filter_bar.filters_cleared.connect(self._on_filters_cleared)
        self._filter_bar.import_requested.connect(self._on_import)
        root.addWidget(self._filter_bar)

        # ── Table ────────────────────────────────────────────────────────
        self._table = _LedgerTable()
        self._table.sort_changed.connect(self._on_sort_changed)
        self._table.col_filter_changed.connect(self._on_col_filter_changed)
        self._table.edit_state_changed.connect(self._on_edit_state_changed)
        self._table.bulk_set_item_requested.connect(self._on_bulk_set_item)
        self._table.save_requested.connect(self._on_save_edits)
        self._table.table().verticalScrollBar().valueChanged.connect(self._on_scroll)
        root.addWidget(self._table, 1)

        # ── Footer totals ────────────────────────────────────────────────
        self._footer = _FooterBar()
        root.addWidget(self._footer)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        self._status_lbl.setFixedHeight(28)
        self._status_lbl.setStyleSheet(
            f"background: {_WHITE}; border-top: 1px solid {_BORDER};"
        )
        root.addWidget(self._status_lbl)

        self._loading_overlay = LoadingOverlay(self, "Loading master expenses…")

    # ── Public API ─────────────────────────────────────────────────────────

    def refresh(self) -> None:
        self._filters_loaded = False
        self._reset_and_load()

    async def _reload_year_options(self) -> None:
        """Populate FY combo from years that actually exist in Master."""
        try:
            from tahmeed.services.accountant_service import get_master_available_years
            years = await get_master_available_years()
        except Exception:
            return
        if not years:
            return
        cur = self._fy_cb.currentData() or self._year
        self._fy_cb.blockSignals(True)
        self._fy_cb.clear()
        for yr in years:
            self._fy_cb.addItem(str(yr), yr)
        idx = self._fy_cb.findData(cur)
        if idx < 0:
            idx = self._fy_cb.findData(self._year)
        if idx < 0:
            idx = 0
        self._fy_cb.setCurrentIndex(idx)
        self._fy_cb.blockSignals(False)
        chosen = self._fy_cb.currentData()
        if chosen and chosen != self._year:
            self._year = int(chosen)
            app_state.fiscal_year = self._year

    def _filter_kw(self) -> dict:
        col_filters = self._table.column_filters_for_query()
        kw = dict(
            year=self._year,
            month=self._month,
            search=self._filter_bar.search_text(),
            truck=self._filter_bar.truck_filter(),
            category=self._filter_bar.category_filter(),
            receipt=self._filter_bar.receipt_filter(),
            date_from=self._filter_bar.date_from(),
            date_to=self._filter_bar.date_to(),
        )
        if col_filters:
            kw["column_filters"] = col_filters
        return kw

    def _export_scope_kw(self, *, all_records: bool) -> dict:
        """Query kwargs for export — filtered view or full FY/month scope."""
        kw = dict(
            year=self._year,
            month=self._month,
            search="",
            truck="",
            category="",
            receipt="",
            date_from=None,
            date_to=None,
        )
        if not all_records:
            kw["search"] = self._filter_bar.search_text()
            kw["truck"] = self._filter_bar.truck_filter()
            kw["category"] = self._filter_bar.category_filter()
            kw["receipt"] = self._filter_bar.receipt_filter()
            kw["date_from"] = self._filter_bar.date_from()
            kw["date_to"] = self._filter_bar.date_to()
            col_filters = self._table.column_filters_for_query()
            if col_filters:
                kw["column_filters"] = col_filters
        return kw

    def _reset_and_load(self, *, keep_col_filters: bool = False) -> None:
        self._reload_generation += 1
        self._loaded = 0
        self._total = 0
        if not keep_col_filters:
            self._table.clear_column_filters()
        self._table.clear_rows()
        self._update_status()
        asyncio.ensure_future(self._load_initial(self._reload_generation))

    def _update_status(self) -> None:
        filtered = bool(self._table.column_filters_for_query())
        if self._total == 0:
            self._status_lbl.setText(
                "No records match the current filters." if not self._loading else "Loading…"
            )
        elif self._loaded >= self._total:
            suffix = " (column filters applied)" if filtered else ""
            self._status_lbl.setText(f"Showing all {self._total:,} records{suffix}")
        else:
            extra = "  •  Column filters on full range" if filtered else ""
            self._status_lbl.setText(
                f"Showing {self._loaded:,} of {self._total:,}  •  Scroll down for more{extra}"
            )

    # ── Event handlers ─────────────────────────────────────────────────────

    def _on_year_changed(self) -> None:
        yr = self._fy_cb.currentData()
        if yr and yr != self._year:
            self._year = yr
            app_state.fiscal_year = yr
            self._filters_loaded = False
            self._month = 0
            self._month_bar.reset()
            self._filter_bar.set_date_range(self._year, 0)
            self._reset_and_load()

    def _on_month_changed(self, month: int) -> None:
        self._month = month
        self._filter_bar.set_date_range(self._year, month)
        self._reset_and_load()

    def _on_filter_changed(self) -> None:
        self._debounce.start()

    def _on_filters_cleared(self) -> None:
        """Reset toolbar filters, Excel column filters, and sort to defaults."""
        self._sort_field = "date"
        self._sort_asc = False
        self._table.reset_default_sort()
        self._reset_and_load()

    def _on_col_filter_changed(self) -> None:
        """Excel ▾ filters query the whole selected year/month range, not one page."""
        self._reset_and_load(keep_col_filters=True)

    def _on_sort_changed(self, field: str, asc: bool) -> None:
        self._sort_field = field
        self._sort_asc = asc
        self._reset_and_load(keep_col_filters=True)

    def _on_scroll(self, value: int) -> None:
        if self._table.scroll_frozen():
            return
        bar = self._table.table().verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    # ── Edit / save (Excel-style grid) ─────────────────────────────────────

    def _accountant_id(self):
        return getattr(self._user, "_id", None) if self._user else None

    async def _ensure_item_names(self) -> List[str]:
        if self._item_names:
            return self._item_names
        names: List[str] = []
        try:
            from tahmeed.services.category_service import get_all_categories
            cats = await get_all_categories(include_inactive=True)
            names = [c.name for c in cats if getattr(c, "name", None)]
            self._item_by_name = {
                (c.name or "").strip().lower(): c
                for c in cats if getattr(c, "name", None)
            }
        except Exception:
            self._item_by_name = {}
            try:
                from tahmeed.services.accountant_service import get_master_categories
                names = list(await get_master_categories(self._year) or [])
            except Exception:
                names = []
        self._item_names = names
        self._table.set_lookups(
            item_names=names,
            default_year=self._year,
        )
        return names

    def _with_category_id(self, updates: dict) -> dict:
        out = dict(updates)
        name = (out.get("item") or out.get("category_name") or "").strip()
        if not name:
            return out
        cat = getattr(self, "_item_by_name", {}).get(name.lower())
        if cat is not None and getattr(cat, "_id", None) is not None:
            out["category_id"] = cat._id
            out["item"] = cat.name
            out["category_name"] = cat.name
        return out

    def _on_edit_state_changed(self, editing: bool, dirty_count: int) -> None:
        if editing:
            self._edit_btn.setText("  Cancel Edit")
            try:
                self._edit_btn.setIcon(qta.icon("mdi.close", color=_T2))
            except Exception:
                pass
        else:
            self._edit_btn.setText("  Edit")
            try:
                self._edit_btn.setIcon(qta.icon("mdi.pencil", color=_T2))
            except Exception:
                pass
        self._save_btn.setEnabled(editing and dirty_count > 0)
        if dirty_count:
            self._save_btn.setText(f"  Save ({dirty_count})")
        else:
            self._save_btn.setText("  Save")

    def _on_toggle_edit(self) -> None:
        if self._table.is_edit_mode():
            if self._table.has_dirty():
                resp = QMessageBox.question(
                    self, "Discard changes?",
                    "Exit edit mode and discard unsaved cell changes?",
                    QMessageBox.Discard | QMessageBox.Cancel,
                    QMessageBox.Cancel,
                )
                if resp == QMessageBox.Cancel:
                    return
            self._table.exit_edit_mode(discard=True)
            self._filters_loaded = False
            self._reset_and_load(keep_col_filters=True)
            return
        asyncio.ensure_future(self._enter_edit_mode_async())

    async def _enter_edit_mode_async(self) -> None:
        await self._ensure_item_names()
        fleet: set = set()
        fleet_kinds: dict = {}
        labels = None
        try:
            from tahmeed.services.truck_service import get_fleet_numbers, get_fleet_kinds
            from tahmeed.services.settings_service import get_setting
            from tahmeed.services.truck_format import (
                DEFAULT_PLACE_LABELS, merge_allowed_labels,
            )
            fleet = set(await get_fleet_numbers() or set())
            try:
                fleet_kinds = await get_fleet_kinds() or {}
            except Exception:
                fleet_kinds = {}
            try:
                raw = await get_setting("allowed_truck_labels")
                if isinstance(raw, list) and raw:
                    labels = merge_allowed_labels(raw, DEFAULT_PLACE_LABELS)
                else:
                    labels = set(DEFAULT_PLACE_LABELS)
            except Exception:
                labels = set(DEFAULT_PLACE_LABELS)
        except Exception:
            fleet = set()
            labels = None
        role = getattr(self._user, "role", "") if self._user else ""
        self._table.set_lookups(
            item_names=self._item_names,
            fleet_numbers=fleet,
            fleet_kinds=fleet_kinds,
            allowed_truck_labels=labels,
            default_year=self._year,
            can_add_fleet=role in ("admin", "accountant"),
        )
        self._table.enter_edit_mode()

    def _on_save_edits(self) -> None:
        if self._edit_in_flight or not self._table.has_dirty():
            return
        asyncio.ensure_future(self._save_edits_async())

    async def _save_edits_async(self) -> None:
        from bson import ObjectId
        from tahmeed.services.accountant_service import update_master_transaction

        self._table._commit_open_editor()
        dirty = self._table.dirty_rows()
        if not dirty:
            return
        await self._ensure_item_names()
        self._edit_in_flight = True
        self._loading_overlay.show_loading(f"Saving {len(dirty)} row(s)…")
        saved = 0
        errors: List[str] = []
        try:
            for row in dirty:
                tx = self._table.tx_at(row)
                if tx is None or tx._id is None:
                    continue
                updates = self._table.updates_from_row(row)
                if not updates:
                    continue
                try:
                    ok = await update_master_transaction(
                        ObjectId(str(tx._id)),
                        self._with_category_id(updates),
                        self._accountant_id(),
                    )
                    if ok:
                        saved += 1
                    else:
                        errors.append(f"Row {row + 1}: not updated")
                except Exception as exc:
                    errors.append(f"Row {row + 1}: {exc}")
            if errors:
                QMessageBox.warning(
                    self, "Partial Save",
                    f"Saved {saved} row(s).\n\n" + "\n".join(errors[:8]),
                )
            self._table.exit_edit_mode(discard=False)
            self._filters_loaded = False
            self._reset_and_load(keep_col_filters=True)
        finally:
            self._edit_in_flight = False
            self._loading_overlay.hide_loading()

    def _on_bulk_set_item(self) -> None:
        if self._edit_in_flight:
            return
        ids = self._table.selected_tx_ids_for_item_bulk()
        if not ids:
            QMessageBox.information(
                self, "Bulk Set Item",
                "Select one or more rows (or cells) first.",
            )
            return
        asyncio.ensure_future(self._bulk_set_item_async(ids))

    async def _bulk_set_item_async(self, ids: list) -> None:
        from tahmeed.services.accountant_service import bulk_update_master_transactions

        items = await self._ensure_item_names()
        dlg = _MasterBulkItemDialog(len(ids), items, parent=self)
        if dlg.exec() != QDialog.Accepted:
            return
        name = dlg.item_name()
        if not name:
            return
        updates = self._with_category_id({"item": name, "category_name": name})
        self._edit_in_flight = True
        self._loading_overlay.show_loading(f"Updating item on {len(ids)} rows…")
        try:
            n = await bulk_update_master_transactions(
                ids, updates, self._accountant_id(),
            )
            if n == 0:
                QMessageBox.warning(
                    self, "Not Saved",
                    "No rows were updated. They may no longer be in Master.",
                )
                return
            if self._table.is_edit_mode():
                self._table.exit_edit_mode(discard=False)
            self._filters_loaded = False
            self._reset_and_load(keep_col_filters=True)
        except Exception as exc:
            QMessageBox.critical(self, "Save Failed", str(exc))
        finally:
            self._edit_in_flight = False
            self._loading_overlay.hide_loading()

    # ── Data loading ───────────────────────────────────────────────────────

    async def _load_initial(self, generation: int) -> None:
        self._loading = True
        self._loading_overlay.show_loading("Loading master expenses…")
        self._update_status()
        try:
            from tahmeed.services.accountant_service import (
                get_master_transactions, count_master_transactions,
                get_cashier_names,
            )

            await self._reload_year_options()
            if generation != self._reload_generation:
                return

            kw = self._filter_kw()
            txs, total = await asyncio.gather(
                get_master_transactions(
                    **kw, sort_field=self._sort_field,
                    sort_asc=self._sort_asc, limit=_SCROLL_CHUNK, skip=0,
                ),
                count_master_transactions(**kw),
            )
            if generation != self._reload_generation:
                return

            self._total = total
            cashier_ids = [tx.cashier_id for tx in txs if tx.cashier_id]
            cashier_names = (
                await get_cashier_names(cashier_ids) if cashier_ids else {}
            )
            if generation != self._reload_generation:
                return

            if txs:
                self._table.populate(txs, 0, cashier_names)
            self._loaded = len(txs)
            self._count_lbl.setText(f"{total:,} records")
            self._update_status()

            asyncio.ensure_future(self._load_background(generation, kw))
        except Exception as exc:
            if generation == self._reload_generation:
                self._table.clear_rows()
                self._status_lbl.setText(f"Failed to load: {exc}")
        finally:
            self._loading = False
            self._loading_overlay.hide_loading()
            if generation == self._reload_generation:
                self._update_status()

    async def _load_background(self, generation: int, kw: dict) -> None:
        try:
            from tahmeed.services.accountant_service import (
                get_master_totals, get_master_month_totals,
                get_master_trucks, get_master_categories,
                get_master_column_values,
            )
            from tahmeed.ui.accountant.master_ledger_table import _COL_FIELD

            if not self._filters_loaded:
                totals, month_data, trucks, cats = await asyncio.gather(
                    get_master_totals(**kw),
                    get_master_month_totals(self._year),
                    get_master_trucks(self._year),
                    get_master_categories(self._year),
                )
            else:
                totals = await get_master_totals(**kw)
                month_data = trucks = cats = None

            if generation != self._reload_generation:
                return

            self._footer.update_totals(totals["tzs"], totals["usd"], self._total)
            if not self._filters_loaded and month_data is not None:
                self._filter_bar.populate_trucks(trucks)
                self._filter_bar.populate_categories(cats)
                self._month_bar.update_totals(month_data)
                self._filters_loaded = True

            # Distinct ▾ values for the whole selected range (cascaded).
            base_kw = {
                k: v for k, v in kw.items()
                if k != "column_filters"
            }
            col_filters = kw.get("column_filters") or {}
            fields = sorted(set(_COL_FIELD.values()))
            results = await asyncio.gather(*[
                get_master_column_values(
                    field,
                    **base_kw,
                    column_filters=col_filters or None,
                )
                for field in fields
            ])
            if generation != self._reload_generation:
                return
            field_to_col = {f: c for c, f in _COL_FIELD.items()}
            cache = {
                field_to_col[field]: set(vals)
                for field, vals in zip(fields, results)
                if field in field_to_col
            }
            self._table.set_column_value_cache(cache)
        except Exception:
            pass

    async def _load_more(self) -> None:
        if self._scroll_loading or self._loading:
            return
        if self._table.scroll_frozen():
            return
        if self._loaded >= self._total:
            return
        self._scroll_loading = True
        self._update_status()
        try:
            from tahmeed.services.accountant_service import (
                get_master_transactions, get_cashier_names,
            )

            kw = self._filter_kw()
            gen = self._reload_generation
            txs = await get_master_transactions(
                **kw, sort_field=self._sort_field,
                sort_asc=self._sort_asc, limit=_SCROLL_CHUNK, skip=self._loaded,
            )
            if gen != self._reload_generation:
                return

            cashier_ids = [tx.cashier_id for tx in txs if tx.cashier_id]
            cashier_names = (
                await get_cashier_names(cashier_ids) if cashier_ids else {}
            )
            if gen != self._reload_generation:
                return

            if txs:
                self._table.append_rows(txs, self._loaded, cashier_names)
                self._loaded += len(txs)
        except Exception:
            pass
        finally:
            self._scroll_loading = False
            self._update_status()

    # ── Excel Export ───────────────────────────────────────────────────────

    def _on_export_filtered(self) -> None:
        if self._export_in_flight:
            return
        asyncio.ensure_future(self._do_export(all_records=False))

    def _on_export_all(self) -> None:
        if self._export_in_flight:
            return
        asyncio.ensure_future(self._do_export(all_records=True))

    async def _fetch_export_page(self, kw: dict, *, limit: int, skip: int) -> List[Transaction]:
        from tahmeed.services.accountant_service import get_master_transactions

        return await get_master_transactions(
            **kw,
            sort_field=self._sort_field,
            sort_asc=self._sort_asc,
            limit=limit,
            skip=skip,
        )

    async def _do_export(self, *, all_records: bool) -> None:
        if self._export_in_flight:
            return
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            await notify_export_error(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n\nRun: pip install openpyxl",
            )
            return

        from tahmeed.services.accountant_service import get_cashier_names

        self._export_in_flight = True
        overlay = attach_export_overlay(self)
        kw = self._export_scope_kw(all_records=all_records)
        mode_label = "All" if all_records else "Filtered"
        suffix = mode_label
        try:
            show_export_busy(
                overlay, f"Loading {mode_label.lower()} master expenses…", maximum=0,
            )
            try:
                txs = await fetch_records_with_progress(
                    overlay,
                    lambda *, limit, skip: self._fetch_export_page(
                        kw, limit=limit, skip=skip,
                    ),
                    phase=f"Loading {mode_label.lower()} records",
                )
            except Exception as exc:
                await notify_export_error(self, "Export Error", f"Failed to fetch data: {exc}")
                return
            finally:
                hide_export_busy(self)

            if not txs:
                await notify_export_info(
                    self, "Export",
                    "No records to export for the selected scope.",
                )
                return

            export_cashier_ids = [tx.cashier_id for tx in txs if tx.cashier_id]
            export_cashier_names = (
                await get_cashier_names(export_cashier_ids) if export_cashier_ids else {}
            )

            month_tag = dict(_MONTHS).get(self._month, "All")
            default = f"Master_Expenses_FY{self._year}_{month_tag}_{suffix}.xlsx"
            path = await pick_export_path(
                self, f"Export Master Expenses ({suffix})", default,
            )
            if not path:
                return
            path = normalize_xlsx_path(path)

            year = self._year
            month = self._month
            month_label = dict(_MONTHS).get(month, "All Months")
            scope_note = (
                f"All records — FY {year}  |  {month_label}"
                if all_records
                else f"Filtered view — FY {year}  |  {month_label}"
            )
            total = len(txs)
            fast = total >= FAST_STYLE_ROW_LIMIT

            def _write(progress_cb) -> None:
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f"Master Expenses FY{year}"

                ws.merge_cells("A1:M1")
                ws["A1"] = "TAHMEED COACH TZ LTD"
                ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1B2B4B")
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 22

                ws.merge_cells("A2:M2")
                ws["A2"] = f"Master Expenses Report — {scope_note}"
                ws["A2"].font = Font(name="Segoe UI", bold=True, size=11, color="374151")
                ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 18

                ws.merge_cells("A3:M3")
                ws["A3"] = f"Exported: {datetime.now().strftime('%d %b %Y  %H:%M')}"
                ws["A3"].font = Font(name="Segoe UI", italic=True, size=9, color="9CA3AF")
                ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[3].height = 15

                ws.append([])

                all_col_names = [c[0] for c in _COLS]
                ws.append(all_col_names)
                hdr_row = ws.max_row
                ws.row_dimensions[hdr_row].height = 18
                grey_fill = PatternFill("solid", fgColor="F1F5F9")
                hdr_border = Border(bottom=Side(style="medium", color="9CA3AF"))
                for cell in ws[hdr_row]:
                    cell.font = Font(name="Segoe UI", bold=True, size=10, color="6B7280")
                    cell.fill = grey_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = hdr_border

                tzs_total = 0.0
                usd_total = 0.0
                alt_fill = PatternFill("solid", fgColor="F9FAFB")
                white_fill = PatternFill("solid", fgColor="FFFFFF")
                red_font = Font(name="Cascadia Code", size=10, color="DC2626")
                mono_font = Font(name="Cascadia Code", size=10)
                amber_font = Font(name="Segoe UI", bold=True, size=10, color="D97706")
                rcpt_green = Font(name="Segoe UI", bold=True, size=10, color="16A34A")
                rcpt_red = Font(name="Segoe UI", bold=True, size=10, color="DC2626")

                for i, tx in enumerate(txs):
                    date_str = _fmt_tx_date(tx.date)
                    item_str = tx.item or tx.category_name or ""
                    ref_str = _ref_float_display(tx)
                    cashier_str = (
                        _short_name(export_cashier_names.get(tx.cashier_id, ""))
                        if tx.cashier_id else ""
                    )
                    rcpt_text, _ = _receipt_text(tx.receipt_status)
                    tzs_txt, usd_txt = _amount_cells(tx)
                    tzs_export = "" if tzs_txt == "—" else tzs_txt
                    usd_export = "" if usd_txt == "—" else usd_txt

                    cur = _currency_key(tx)
                    if _is_tzs(cur):
                        tzs_total += tx.amount
                    elif cur == "USD":
                        usd_total += tx.amount

                    ws.append([
                        i + 1, date_str, item_str, tx.description or "",
                        tx.truck_number or "", tx.memo or "", ref_str,
                        tzs_export, usd_export,
                        rcpt_text, tx.ownership or "", tx.approver or "", cashier_str,
                    ])
                    if not fast:
                        r = ws.max_row
                        ws.row_dimensions[r].height = 16
                        fill = alt_fill if i % 2 else white_fill
                        for cell in ws[r]:
                            cell.fill = fill
                            cell.alignment = Alignment(vertical="center")

                        for col_idx, txt in (
                            (_COL_TZS + 1, tzs_export), (_COL_USD + 1, usd_export),
                        ):
                            if not txt:
                                continue
                            c = ws.cell(r, col_idx)
                            c.font = red_font if tx.amount < 0 else mono_font
                            c.alignment = Alignment(horizontal="right", vertical="center")

                        if ref_str:
                            ws.cell(r, _COL_REF + 1).font = amber_font

                        rcpt_fonts = {
                            "Received": rcpt_green, "Pending": amber_font, "No Receipt": rcpt_red,
                        }
                        rf = rcpt_fonts.get(rcpt_text)
                        if rf:
                            ws.cell(r, _COL_RCPT + 1).font = rf
                        ws.cell(r, _COL_RCPT + 1).alignment = Alignment(
                            horizontal="center", vertical="center",
                        )

                    if progress_cb and (
                        (i + 1) % PROGRESS_EVERY == 0 or i + 1 == total
                    ):
                        progress_cb(i + 1, "Writing rows")

                ws.append([])
                ws.append([
                    "", "", "", "TOTAL", "", "", "",
                    _fmt_num(tzs_total, "", 0) if tzs_total else "",
                    _fmt_num(usd_total, "", 2) if usd_total else "",
                    "", "", "", "",
                ])
                total_r = ws.max_row
                ws.row_dimensions[total_r].height = 18
                total_fill = PatternFill("solid", fgColor="EFF6FF")
                for cell in ws[total_r]:
                    cell.fill = total_fill
                ws.cell(total_r, 4).font = Font(name="Segoe UI", bold=True, size=11)
                for col_idx, total_val in (
                    (_COL_TZS + 1, tzs_total), (_COL_USD + 1, usd_total),
                ):
                    if not total_val:
                        continue
                    c = ws.cell(total_r, col_idx)
                    c.font = Font(
                        name="Cascadia Code", bold=True, size=11,
                        color="DC2626" if total_val < 0 else "111827",
                    )
                    c.alignment = Alignment(horizontal="right", vertical="center")

                col_widths = [7, 10, 16, 35, 13, 20, 16, 14, 12, 14, 14, 14, 14]
                for idx, w in enumerate(col_widths, 1):
                    ws.column_dimensions[ws.cell(1, idx).column_letter].width = w

                ws.freeze_panes = ws.cell(hdr_row + 1, 5)
                if progress_cb:
                    progress_cb(total, "Saving file")
                wb.save(path)

            try:
                await run_export_write(overlay, total, _write)
            except Exception as exc:
                await notify_export_error(self, "Save Error", f"Could not save file:\n{exc}")
                return
            finally:
                hide_export_busy(self)

            if not export_file_ready(path):
                await notify_export_error(
                    self, "Save Error", f"The file could not be saved:\n{path}",
                )
                return

            await notify_export_info(
                self, "Export Complete",
                f"Exported {len(txs):,} {suffix.lower()} records to:\n{path}",
            )
        finally:
            hide_export_busy(self)
            self._export_in_flight = False

    # ── Excel Import ─────────────────────────────────────────────────────

    def _on_import(self) -> None:
        if self._import_in_flight:
            return
        asyncio.ensure_future(self._do_import())

    async def _do_import(self) -> None:
        if self._import_in_flight:
            return
        self._import_in_flight = True
        try:
            await self._do_import_body()
        finally:
            self._import_in_flight = False

    async def _do_import_body(self) -> None:
        from pathlib import Path

        from tahmeed.services.category_service import get_all_categories
        from tahmeed.services.master_import_service import (
            apply_mapping_to_preview,
            commit_master_import,
            preview_master_import,
        )
        from tahmeed.ui.dialogs.description_mapping_flow import prompt_unmapped_descriptions

        default_dir = str(Path(__file__).resolve().parents[3])
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Master Expenses",
            default_dir,
            "Excel Files (*.xlsx)",
        )
        if not path:
            return

        self._loading_overlay.show_loading("Reading master Excel file…")
        try:
            categories = await get_all_categories()
        except Exception as exc:
            self._loading_overlay.hide_loading()
            QMessageBox.critical(self, "Import Error", f"Could not load items:\n{exc}")
            return

        if not categories:
            categories = []

        try:
            preview = await preview_master_import(path)
        except Exception as exc:
            self._loading_overlay.hide_loading()
            QMessageBox.critical(self, "Import Error", f"Could not read workbook:\n{exc}")
            return

        if not preview.rows:
            self._loading_overlay.hide_loading()
            QMessageBox.information(self, "Import", "No expense rows found in the selected file.")
            return

        if preview.unmapped:
            self._loading_overlay.hide_loading()
            ok = await prompt_unmapped_descriptions(
                preview,
                categories,
                self,
                allow_skip=False,
                apply_mapping=apply_mapping_to_preview,
            )
            if not ok:
                QMessageBox.information(
                    self,
                    "Import Cancelled",
                    "No records were imported.",
                )
                return

        self._loading_overlay.show_loading("Importing master expenses…")
        verified_by = self._user._id if self._user else None
        try:
            result = await commit_master_import(
                preview,
                verified_by=verified_by,
                skip_duplicates=True,
            )
        except Exception as exc:
            self._loading_overlay.hide_loading()
            QMessageBox.critical(self, "Import Error", f"Import failed:\n{exc}")
            return

        self._loading_overlay.hide_loading()
        skipped_note = ""
        if preview.skipped:
            skipped_note = f"\nSkipped {preview.skipped:,} blank/invalid rows."
        dup_note = ""
        if result["duplicates_skipped"]:
            dup_note = f"\nSkipped {result['duplicates_skipped']:,} duplicate serial(s)."

        QMessageBox.information(
            self,
            "Import Complete",
            f"Imported {result['inserted']:,} master expense record(s)."
            f"{dup_note}{skipped_note}",
        )
        self.refresh()
