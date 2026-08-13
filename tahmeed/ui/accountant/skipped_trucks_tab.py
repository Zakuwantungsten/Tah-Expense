"""Skipped-trucks follow-up tab for Separate Expenses / Fuel imports."""

from __future__ import annotations

import asyncio
import csv
from datetime import datetime
from typing import List

import qtawesome as qta
from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from tahmeed.services.import_truck_check import (
    combo_suffix_of,
    leading_truck_of,
    resolve_truck_cell,
    truck_field_for,
)
from tahmeed.services.truck_format import DEFAULT_PLACE_LABELS, normalize_truck_number
from tahmeed.ui.dialog_theme import (
    show_critical,
    show_info,
    show_question,
    show_warning,
)
from tahmeed.ui.dialogs.truck_correction_dialog import TruckCorrectionDialog, TruckIssue

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _HAS_OPENPYXL = True
except ImportError:
    openpyxl = None  # type: ignore
    _HAS_OPENPYXL = False

_BG = "#F4F6F8"
_WHITE = "#FFFFFF"
_BORDER = "#E5E7EB"
_T1 = "#111827"
_T2 = "#6B7280"
_BLUE = "#0077C5"

_HEADERS = [
    "", "File row", "Skipped", "Truck", "Original", "Reason",
    "Receipt / Ledger", "Date", "Plaza / Details", "Amount",
    "Source file", "Sheet", "Upload id",
]

_COL_CHK = 0

_EXPORT_HEADERS = [
    "File Row",
    "Skipped At",
    "Truck",
    "Original Truck",
    "Reason",
    "Source File",
    "Sheet",
    "Upload ID",
    "Ledger / Receipt",
    "Payment / Date",
    "Type",
    "Amount",
    "Details",
]


def _lbl(text: str, *, size: int = 12, weight: int = 400, color: str = _T1) -> QLabel:
    lab = QLabel(text)
    lab.setStyleSheet(
        f"color:{color};font-size:{size}px;font-weight:{weight};"
        "font-family:'Segoe UI';border:none;background:transparent;"
    )
    return lab


def _btn(text: str, icon: str = "", *, primary: bool = True, height: int = 32) -> QPushButton:
    b = QPushButton(text)
    b.setCursor(Qt.PointingHandCursor)
    b.setFixedHeight(height)
    if icon:
        try:
            b.setIcon(qta.icon(icon, color="#FFF" if primary else _T1))
        except Exception:
            pass
    if primary:
        b.setStyleSheet(
            f"QPushButton{{background:{_BLUE};color:#FFF;border:none;border-radius:5px;"
            f"font-size:12px;font-weight:600;padding:0 12px;}}"
            "QPushButton:hover{background:#005EA3;}"
            "QPushButton:disabled{background:#93C5FD;}"
        )
    else:
        b.setStyleSheet(
            f"QPushButton{{background:{_WHITE};color:{_T1};border:1px solid {_BORDER};"
            f"border-radius:5px;font-size:12px;font-weight:600;padding:0 12px;}}"
            f"QPushButton:hover{{background:{_BG};}}"
        )
    return b


def _cell(text: str) -> QTableWidgetItem:
    item = QTableWidgetItem(str(text) if text is not None else "")
    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
    return item


def _reason_label(reason: str) -> str:
    if reason == "not_in_registry":
        return "Not in registry"
    if reason == "invalid_format":
        return "Invalid format"
    return reason or ""


def _record_ledger(rec: dict) -> str:
    return str(
        rec.get("ledger_id")
        or rec.get("receipt_no")
        or rec.get("lpo_no")
        or rec.get("serial")
        or rec.get("ticket_no")
        or ""
    )


def _record_date(rec: dict) -> str:
    return str(
        rec.get("payment_date")
        or rec.get("toll_date")
        or rec.get("date")
        or rec.get("sales_date")
        or ""
    )


def _record_details(rec: dict) -> str:
    return str(
        rec.get("toll_plaza")
        or rec.get("transaction_details")
        or rec.get("description")
        or rec.get("details")
        or rec.get("heading_to")
        or rec.get("client_name")
        or ""
    )


def _record_amount(rec: dict) -> str:
    return str(rec.get("amount") or rec.get("tender_amount") or "")


def _source_row_label(doc: dict) -> str:
    row = doc.get("source_row")
    if row is None or row == "":
        return "—"
    try:
        return str(int(row))
    except (TypeError, ValueError):
        return str(row)


def _export_row_values(doc: dict) -> List[str]:
    """Flatten a skipped row into export columns for follow-up tracking."""
    skipped_at = doc.get("skipped_at")
    if isinstance(skipped_at, datetime):
        when = skipped_at.strftime("%d %b %Y  %H:%M")
    else:
        when = str(skipped_at or "")
    rec = doc.get("record") or {}
    return [
        _source_row_label(doc),
        when,
        str(doc.get("truck_value") or ""),
        str(doc.get("original_truck") or ""),
        _reason_label(str(doc.get("reason") or "")),
        str(doc.get("source_filename") or ""),
        str(doc.get("sheet_label") or ""),
        str(doc.get("target_upload_id") or ""),
        _record_ledger(rec),
        _record_date(rec),
        str(rec.get("transaction_type") or rec.get("type") or ""),
        _record_amount(rec),
        _record_details(rec),
    ]


def _write_skipped_csv(path: str, docs: List[dict]) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(_EXPORT_HEADERS)
        for doc in docs:
            writer.writerow(_export_row_values(doc))


def _write_skipped_xlsx(path: str, docs: List[dict]) -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Skipped Trucks"

    thin = Side(border_style="thin", color="E5E7EB")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="EFF6FF")
    hdr_font = Font(name="Calibri", bold=True, size=11, color="1E3A5F")
    alt_fill = PatternFill("solid", fgColor="F9FAFB")
    nrm_font = Font(name="Calibri", size=11)

    for c, label in enumerate(_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, doc in enumerate(docs, 2):
        for ci, val in enumerate(_export_row_values(doc), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = nrm_font
            cell.border = bdr
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = alt_fill

    widths = [10, 18, 14, 14, 16, 28, 12, 38, 16, 20, 14, 12, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    wb.save(path)


class SkippedTrucksTab(QWidget):
    """List parked import rows; edit truck and re-upload into the original batch."""

    changed = Signal()

    def __init__(self, feed_key: str, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._feed_key = feed_key
        self._truck_field = truck_field_for(feed_key) or "truck_no"
        self._rows: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        hint = _lbl(
            "Rows skipped during import because the truck was unknown or mistyped. "
            "File row is the 1-based data row in that upload (after the header). "
            "Tick one or more rows (or Select all), edit truck numbers, then "
            "Re-upload — each truck is checked against the fleet registry again "
            "before it rejoins the original upload batch.",
            size=11,
            color=_T2,
        )
        hint.setWordWrap(True)
        vl.addWidget(hint)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search = QLineEdit()
        self._search.setPlaceholderText(
            "Search truck, receipt, plaza, file, row #…"
        )
        self._search.setFixedWidth(300)
        self._search.setStyleSheet(
            f"QLineEdit{{background:{_WHITE};border:1px solid {_BORDER};border-radius:5px;"
            f"padding:0 10px;min-height:32px;font-size:12px;}}"
        )
        self._search.returnPressed.connect(self.refresh)
        tbl.addWidget(self._search)

        refresh_btn = _btn("Refresh", "mdi.refresh", primary=False)
        refresh_btn.clicked.connect(self.refresh)
        tbl.addWidget(refresh_btn)

        sel_all_btn = _btn("Select all", primary=False)
        sel_all_btn.clicked.connect(lambda: self._set_all_checked(True))
        tbl.addWidget(sel_all_btn)

        clear_sel_btn = _btn("Clear selection", primary=False)
        clear_sel_btn.clicked.connect(lambda: self._set_all_checked(False))
        tbl.addWidget(clear_sel_btn)

        export_xlsx_btn = _btn("Export Excel", "mdi.file-excel-outline", primary=False)
        export_xlsx_btn.clicked.connect(lambda: self._export("xlsx"))
        tbl.addWidget(export_xlsx_btn)

        export_csv_btn = _btn("Export CSV", "mdi.file-delimited-outline", primary=False)
        export_csv_btn.clicked.connect(lambda: self._export("csv"))
        tbl.addWidget(export_csv_btn)

        tbl.addStretch()

        edit_btn = _btn("Edit truck", primary=False)
        edit_btn.clicked.connect(self._edit_selected)
        tbl.addWidget(edit_btn)

        reup_btn = _btn("Re-upload selected", "mdi.upload-outline")
        reup_btn.clicked.connect(self._reupload_selected)
        tbl.addWidget(reup_btn)

        del_btn = _btn("Delete", "mdi.trash-can-outline", primary=False)
        del_btn.clicked.connect(self._delete_selected)
        tbl.addWidget(del_btn)
        vl.addWidget(tb)

        self._table = QTableWidget(0, len(_HEADERS))
        self._table.setHorizontalHeaderLabels(_HEADERS)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.setStyleSheet(
            f"QTableWidget{{background:{_WHITE};border:1px solid {_BORDER};"
            f"border-radius:6px;gridline-color:{_BORDER};}}"
            f"QHeaderView::section{{background:{_BG};color:{_T2};font-weight:600;"
            f"border:none;border-bottom:1px solid {_BORDER};padding:6px;}}"
        )
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(_COL_CHK, QHeaderView.Fixed)
        self._table.setColumnWidth(_COL_CHK, 36)
        hdr.setStretchLastSection(True)
        self._table.itemChanged.connect(self._on_item_changed)
        self._table.cellClicked.connect(self._on_cell_clicked)
        vl.addWidget(self._table, 1)

        self._status = _lbl("", size=11, color=_T2)
        vl.addWidget(self._status)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc

        search = self._search.text().strip()
        try:
            total = await svc.count_skipped_import_rows(self._feed_key, search)
            self._rows = await svc.list_skipped_import_rows(
                self._feed_key, search=search, limit=500, skip=0
            )
        except Exception as exc:
            show_critical(self, "Skipped", str(exc))
            return
        self._fill()
        selected = self._checked_count()
        if selected:
            self._status.setText(
                f"{total:,} skipped row(s)  ·  {selected:,} selected"
            )
        else:
            self._status.setText(f"{total:,} skipped row(s)")

    def _fill(self) -> None:
        t = self._table
        t.blockSignals(True)
        t.setRowCount(0)
        for doc in self._rows:
            r = t.rowCount()
            t.insertRow(r)
            skipped_at = doc.get("skipped_at")
            if isinstance(skipped_at, datetime):
                when = skipped_at.strftime("%d %b %Y  %H:%M")
            else:
                when = str(skipped_at or "—")
            rec = doc.get("record") or {}
            upload_id = str(doc.get("target_upload_id") or "")
            upload_short = (upload_id[:8] + "…") if len(upload_id) > 8 else upload_id

            chk = QTableWidgetItem("")
            chk.setFlags(
                Qt.ItemIsEnabled | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable
            )
            chk.setCheckState(Qt.Unchecked)
            chk.setData(Qt.UserRole, str(doc.get("_id") or ""))
            t.setItem(r, _COL_CHK, chk)

            vals = [
                _source_row_label(doc),
                when,
                doc.get("truck_value") or "",
                doc.get("original_truck") or "",
                _reason_label(str(doc.get("reason") or "")),
                _record_ledger(rec),
                _record_date(rec),
                _record_details(rec),
                _record_amount(rec),
                doc.get("source_filename") or "",
                doc.get("sheet_label") or "—",
                upload_short,
            ]
            for c, val in enumerate(vals):
                t.setItem(r, c + 1, _cell(val))
        t.blockSignals(False)
        self._update_selection_status()

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if item is None or item.column() != _COL_CHK:
            return
        self._update_selection_status()

    def _on_cell_clicked(self, row: int, col: int) -> None:
        """Clicking a data cell toggles that row's checkbox for easier multi-select."""
        if col == _COL_CHK:
            return
        chk = self._table.item(row, _COL_CHK)
        if chk is None:
            return
        new_state = (
            Qt.Unchecked if chk.checkState() == Qt.Checked else Qt.Checked
        )
        self._table.blockSignals(True)
        chk.setCheckState(new_state)
        self._table.blockSignals(False)
        self._update_selection_status()

    def _set_all_checked(self, checked: bool) -> None:
        state = Qt.Checked if checked else Qt.Unchecked
        t = self._table
        t.blockSignals(True)
        for r in range(t.rowCount()):
            chk = t.item(r, _COL_CHK)
            if chk is not None:
                chk.setCheckState(state)
        t.blockSignals(False)
        self._update_selection_status()

    def _checked_count(self) -> int:
        n = 0
        for r in range(self._table.rowCount()):
            chk = self._table.item(r, _COL_CHK)
            if chk is not None and chk.checkState() == Qt.Checked:
                n += 1
        return n

    def _update_selection_status(self) -> None:
        total = len(self._rows)
        selected = self._checked_count()
        base = f"{total:,} skipped row(s)"
        if selected:
            self._status.setText(f"{base}  ·  {selected:,} selected")
        else:
            self._status.setText(base)

    def _export(self, fmt: str) -> None:
        if not self._rows:
            show_info(
                self, "Export", "No skipped rows to export. Refresh or clear search.",
            )
            return
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        default = f"skipped_trucks_{self._feed_key}_{stamp}.{fmt}"
        if fmt == "xlsx":
            filt = "Excel Files (*.xlsx)"
        else:
            filt = "CSV Files (*.csv)"
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Skipped Trucks", default, filt,
        )
        if not path:
            return
        if fmt == "xlsx" and not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        elif fmt == "csv" and not path.lower().endswith(".csv"):
            path += ".csv"
        try:
            if fmt == "xlsx":
                _write_skipped_xlsx(path, self._rows)
            else:
                _write_skipped_csv(path, self._rows)
        except Exception as exc:
            show_critical(self, "Export Error", str(exc))
            return
        show_info(
            self,
            "Export Complete",
            f"Exported {len(self._rows):,} skipped row(s) to:\n{path}",
        )

    def _selected_ids(self) -> List[str]:
        ids: List[str] = []
        for r in range(self._table.rowCount()):
            chk = self._table.item(r, _COL_CHK)
            if chk is None or chk.checkState() != Qt.Checked:
                continue
            rid = chk.data(Qt.UserRole)
            if rid:
                ids.append(str(rid))
        return ids

    def _selected_docs(self) -> List[dict]:
        ids = set(self._selected_ids())
        return [d for d in self._rows if str(d.get("_id") or "") in ids]

    def _edit_selected(self) -> None:
        docs = self._selected_docs()
        if not docs:
            show_info(
                self,
                "Edit truck",
                "Select one or more skipped rows (tick the checkbox, or click a row).",
            )
            return
        asyncio.ensure_future(self._do_edit(docs))

    async def _do_edit(self, docs: List[dict]) -> None:
        """Open the fleet correction dialog for the selected skipped rows."""
        from tahmeed.services import accountant_service as svc
        from tahmeed.services.truck_service import get_fleet_numbers, add_fleet_by_collection
        from tahmeed.services.truck_format import merge_allowed_labels
        from tahmeed.services import settings_service

        try:
            fleet = await get_fleet_numbers()
        except Exception as exc:
            show_critical(
                self,
                "Edit truck",
                "Could not load the fleet registry.\n\n"
                f"{exc}\n\nNothing was changed.",
            )
            return
        try:
            from tahmeed.services.truck_service import get_fleet_kinds
            fleet_kinds = await get_fleet_kinds()
        except Exception:
            fleet_kinds = {}
        try:
            stored = await settings_service.get_setting("allowed_truck_labels")
        except Exception:
            stored = []
        labels = merge_allowed_labels(DEFAULT_PLACE_LABELS, stored or [])

        issues: List[TruckIssue] = []
        for i, doc in enumerate(docs):
            raw = str(
                doc.get("truck_value")
                or doc.get("original_truck")
                or ""
            ).strip()
            truck_only = leading_truck_of(raw)
            norm = normalize_truck_number(truck_only, allowed_labels=labels)
            kind = (
                "invalid_format"
                if norm.status == "invalid"
                else "not_in_registry"
            )
            src = _source_row_label(doc)
            row_label = (
                f"File row {src}"
                if src != "—"
                else f"Skipped row {i + 1}"
            )
            issues.append(TruckIssue(
                row=i,
                original=raw,
                kind=kind,
                row_label=row_label,
                combo_suffix=combo_suffix_of(raw),
            ))

        dlg = TruckCorrectionDialog(
            issues,
            fleet,
            can_add=True,
            allowed_labels=labels,
            import_mode=True,
            fleet_kinds=fleet_kinds,
            heading="Edit truck numbers",
            intro=(
                "Correct the selected skipped truck number(s). "
                "Type to look up the fleet registry, then Apply, Allow anyway, "
                "Add to registry, or Skip a row to leave it unchanged. "
                "Applying one fix can also update every other selected row "
                "with the same original value."
            ),
            parent=self,
        )
        dlg.exec()

        for kind, number in dlg.pending_registry_adds:
            try:
                await add_fleet_by_collection(kind, number)
            except Exception as exc:
                show_warning(
                    self, "Registry", f"Could not add {number}:\n{exc}"
                )

        if dlg.new_labels:
            try:
                merged = merge_allowed_labels(labels, dlg.new_labels)
                await settings_service.set_setting(
                    "allowed_truck_labels", sorted(merged)
                )
            except Exception:
                pass

        by_row = {iss.row: iss for iss in dlg.issues}
        updated = 0
        for i, doc in enumerate(docs):
            iss = by_row.get(i)
            if iss is None or iss.omit_row or (iss.skip and not iss.corrected):
                continue
            value = (iss.corrected or iss.original or "").strip()
            if not value:
                continue
            try:
                await svc.update_skipped_import_truck(str(doc["_id"]), value)
                updated += 1
            except Exception as exc:
                show_critical(self, "Edit truck", str(exc))
                await self._load()
                return

        await self._load()
        if updated:
            show_info(
                self,
                "Edit truck",
                f"Updated truck number on {updated:,} skipped row(s).",
            )
        else:
            show_info(
                self,
                "Edit truck",
                "No truck numbers were changed.",
            )

    def _delete_selected(self) -> None:
        ids = self._selected_ids()
        if not ids:
            show_info(
                self,
                "Delete",
                "Select one or more skipped rows to delete.",
            )
            return
        if show_question(
            self,
            "Delete skipped",
            f"Permanently delete {len(ids):,} skipped row(s)?",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._do_delete(ids))

    async def _do_delete(self, ids: List[str]) -> None:
        from tahmeed.services import accountant_service as svc

        try:
            await svc.delete_skipped_import_rows(ids)
        except Exception as exc:
            show_critical(self, "Delete", str(exc))
            return
        self.changed.emit()
        await self._load()

    def _reupload_selected(self) -> None:
        docs = self._selected_docs()
        if not docs:
            show_info(
                self,
                "Re-upload",
                "Select one or more skipped rows (tick checkboxes, or Select all).",
            )
            return
        asyncio.ensure_future(self._do_reupload(docs))

    async def _do_reupload(self, docs: List[dict]) -> None:
        """Re-check every selected truck against the fleet, then re-upload passers."""
        from tahmeed.services import accountant_service as svc
        from tahmeed.services.import_truck_check import scan_import_trucks
        from tahmeed.services.truck_service import get_fleet_numbers, add_fleet_by_collection
        from tahmeed.services.truck_format import merge_allowed_labels
        from tahmeed.services import settings_service

        try:
            fleet = await get_fleet_numbers()
        except Exception as exc:
            show_critical(
                self,
                "Re-upload",
                "Could not load the fleet registry to verify trucks.\n\n"
                f"{exc}\n\nNothing was re-uploaded.",
            )
            return
        try:
            from tahmeed.services.truck_service import get_fleet_kinds
            fleet_kinds = await get_fleet_kinds()
        except Exception:
            fleet_kinds = {}
        try:
            stored = await settings_service.get_setting("allowed_truck_labels")
        except Exception:
            stored = []
        labels = merge_allowed_labels(DEFAULT_PLACE_LABELS, stored or [])

        field = docs[0].get("truck_field") or self._truck_field
        work_rows: List[dict] = []
        for doc in docs:
            row_field = doc.get("truck_field") or field
            rec = dict(doc.get("record") or {})
            rec[row_field] = str(doc.get("truck_value") or rec.get(row_field, "") or "")
            # Keep a stable field name for the shared scanner
            if row_field != field:
                rec[field] = rec[row_field]
            work_rows.append(rec)

        scan = scan_import_trucks(
            work_rows, field, fleet, allowed_labels=labels,
        )

        by_issue: dict = {}
        if scan.issues:
            dlg = TruckCorrectionDialog(
                scan.issues,
                fleet,
                can_add=True,
                allowed_labels=labels,
                import_mode=True,
                fleet_kinds=fleet_kinds,
                parent=self,
            )
            dlg.exec()
            for kind, number in dlg.pending_registry_adds:
                try:
                    await add_fleet_by_collection(kind, number)
                    fleet.add(number)
                except Exception as exc:
                    show_warning(
                        self, "Registry", f"Could not add {number}:\n{exc}"
                    )
            by_issue = {iss.row: iss for iss in dlg.issues}
            # Apply dialog outcomes onto working rows
            for i, row in enumerate(work_rows):
                iss = by_issue.get(i)
                if iss is None:
                    continue
                if iss.omit_row or (iss.skip and not iss.corrected):
                    continue
                value = (iss.corrected or iss.original or "").strip()
                if value:
                    row[field] = value

            # Refresh fleet after possible registry adds
            try:
                fleet = await get_fleet_numbers()
            except Exception:
                pass

        ready_ids: List[str] = []
        blocked = 0
        for i, doc in enumerate(docs):
            iss = by_issue.get(i)
            if iss is not None and (
                iss.omit_row or (iss.skip and not iss.corrected)
            ):
                blocked += 1
                continue

            value = str(work_rows[i].get(field) or "").strip()
            if not value:
                blocked += 1
                continue

            allow_anyway = bool(iss and iss.allow_anyway)
            status, canonical = resolve_truck_cell(value, fleet, labels)
            if status in ("ok", "place_label"):
                await svc.update_skipped_import_truck(str(doc["_id"]), canonical)
                ready_ids.append(str(doc["_id"]))
                continue

            if allow_anyway:
                # Explicit override from the correction dialog only
                await svc.update_skipped_import_truck(str(doc["_id"]), value)
                ready_ids.append(str(doc["_id"]))
                continue

            blocked += 1

        if not ready_ids:
            show_info(
                self,
                "Re-upload",
                "No rows passed the fleet registry check.\n"
                "Fix the truck number(s), add them to the registry, "
                "or use Allow anyway in the correction dialog.",
            )
            await self._load()
            return

        try:
            saved = await svc.reupload_skipped_import_rows(ready_ids)
        except Exception as exc:
            show_critical(self, "Re-upload", str(exc))
            await self._load()
            return

        msg = (
            f"Re-uploaded {saved:,} row(s) into their original upload batch(es) "
            "after fleet registry verification."
        )
        if blocked:
            msg += f"\n{blocked:,} row(s) stayed in Skipped (still unresolved)."
        show_info(self, "Re-upload complete", msg)
        self.changed.emit()
        await self._load()
