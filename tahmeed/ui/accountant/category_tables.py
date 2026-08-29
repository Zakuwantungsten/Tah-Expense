"""AccountantDashboard — Category Sub-Tables.

One reusable, read-only view per cashier-fed category (LATRA, C28, Mileage,
Council Fees, …). Each shows the *verified* transactions whose
``category_name`` matches, using the exact same column set as the cashier's
DailyRegister (excel_grid.py):

    S/NO · DATE · ITEM · DESCRIPTION · TRUCK NO. · MEMO · NOTES
         · TZS · RECEIPT · OWNERSHIP · APR BY

Design matches the Master Expenses / Verify grid: slate zebra stripes,
muted grey headers, colored receipt pill, infinite scroll.

The widget reuses the existing Master Expenses service queries
(``get_master_transactions`` etc.) with a fixed ``category`` filter — no data
duplication, the table is just a scoped query over the transactions
collection.
"""

from __future__ import annotations

import asyncio
from datetime import datetime
from typing import Dict, List, Optional

import qtawesome as qta

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame, QLabel,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QLineEdit, QComboBox, QPushButton, QMessageBox, QFileDialog,
    QAbstractItemView, QDateEdit,
)
from PySide6.QtCore import Qt, Signal, QTimer, QSize, QDate
from PySide6.QtGui import QFont, QColor, QBrush

from tahmeed.app_state import app_state
from tahmeed.models.transaction import Transaction
from tahmeed.services.category_service import item_key
from tahmeed.ui.accountant.date_filters import (
    add_from_to_editors, read_from_to, sync_from_to,
)
from tahmeed.ui.widgets.column_persistence import bind_column_width_persistence
from tahmeed.ui.widgets.export_runner import (
    attach_export_overlay,
    export_file_ready,
    fetch_records_with_progress,
    hide_export_busy,
    normalize_xlsx_path,
    notify_export_error,
    notify_export_info,
    pick_export_path,
    run_export_write,
    show_export_busy,
    FAST_STYLE_ROW_LIMIT,
    PROGRESS_EVERY,
)
from tahmeed.ui.widgets.loading_overlay import LoadingOverlay
from tahmeed.ui.accountant.feed_sort_helpers import (
    CATEGORY_TABLE_SORT, wire_feed_table_sort, sort_kw, reset_feed_sort,
)

# ── Design tokens (match accountant dashboard palette) ──────────────────────────
_WHITE      = "#FFFFFF"
_BG         = "#F4F6F8"
_BORDER     = "#E5E7EB"
_BLUE       = "#0077C5"
_BLUE_L     = "#E8F4FD"   # row selection highlight (matches Master / Verify)
_STRIPE     = "#F1F5F9"   # slate zebra stripe (matches Master / Verify)
_GREEN      = "#16A34A"
_GREEN_L    = "#DCFCE7"
_AMBER      = "#D97706"
_AMBER_L    = "#FEF3C7"
_RED        = "#DC2626"
_RED_L      = "#FEE2E2"
_T1         = "#111827"
_T2         = "#6B7280"
_TM         = "#9CA3AF"
_HDR_BG     = "#F1F5F9"

_SCROLL_CHUNK = 50
_ROW_H        = 28
_HDR_H      = 28

# (label, pixel width, right-aligned?, mono?)
_COLS = [
    ("S/NO",        52,  "center", False),
    ("DATE",        100, "left",   False),
    ("ITEM",        110, "left",   False),
    ("DESCRIPTION", 260, "left",   False),
    ("TRUCK NO.",   95,  "left",   False),
    ("MEMO",        140, "left",   False),
    ("NOTES",       56,  "center", False),
    ("TZS",         120, "right",  False),
    ("RECEIPT",     110, "center", False),
    ("OWNERSHIP",   95,  "left",   False),
    ("APR BY",      90,  "left",   False),
]

_MONTHS = [
    (0, "All Months"),
    (1, "Jan"), (2, "Feb"), (3, "Mar"), (4, "Apr"), (5, "May"), (6, "Jun"),
    (7, "Jul"), (8, "Aug"), (9, "Sep"), (10, "Oct"), (11, "Nov"), (12, "Dec"),
]

_RECEIPT_MAP = {
    "received": ("Received",   _GREEN, _GREEN_L),
    "pending":  ("Pending",    _AMBER, _AMBER_L),
    "missing":  ("No Receipt", _RED,   _RED_L),
}

# ── Category registry: sidebar key → (title / category_name, mdi icon) ──────────
#  The title doubles as the category_name filter (matches how categories are
#  stored on transactions and shown in the sidebar / verify inbox).
CATEGORY_DEFS: Dict[str, tuple] = {
    "mileage":         ("Mileage",                "mdi.road-variant"),
    "latra":           ("LATRA",                  "mdi.card-account-details-outline"),
    "c28":             ("C28",                    "mdi.file-document-outline"),
    "c40":             ("C40",                    "mdi.file-document-outline"),
    "carbon_permit":   ("Carbon & Permit",        "mdi.leaf"),
    "diesel_cash":     ("Diesel Cash",            "mdi.gas-station-outline"),
    "council_fees":    ("Council Fees",           "mdi.city-variant"),
    "return_weigh":    ("Return & Weighbridge",   "mdi.scale"),
    "parking_petroda": ("Parking Petroda",        "mdi.parking"),
    "backload":        ("Backload Facilitation",  "mdi.truck-delivery"),
    "rope_sealing":    ("Rope & Sealing",         "mdi.link-variant"),
    "radiation":       ("Radiation Taxes",        "mdi.radioactive"),
    "health_fee":      ("Health Fee",             "mdi.hospital-box"),
    "halmashauri":     ("Halmashauri Parking",    "mdi.parking"),
}

_TABLE_SS = (
    f"QTableWidget {{"
    f"  background: {_WHITE};"
    f"  gridline-color: {_BORDER};"
    f"  font-size: 11px; font-family:'Segoe UI';"
    f"  color: {_T1}; border: none;"
    f"}}"
    f"QTableWidget::item {{ padding: 2px 8px; border: none; }}"
    f"QTableWidget::item:selected {{ background: {_BLUE_L}; color: {_T1}; }}"
    f"QHeaderView::section {{"
    f"  background: {_HDR_BG}; color: {_T2};"
    f"  font-size: 10px; font-weight: 600; font-family:'Segoe UI';"
    f"  border: none; border-bottom: 1px solid {_BORDER};"
    f"  border-right: 1px solid {_BORDER}; padding: 0 8px; min-height: {_HDR_H}px;"
    f"}}"
    f"QHeaderView::section:hover {{ background: #E2E8F0; }}"
    f"QScrollBar:horizontal {{ background: {_BG}; height: 8px; margin: 0; }}"
    f"QScrollBar::handle:horizontal {{ background: #D1D5DB; border-radius: 4px; min-width: 24px; }}"
    f"QScrollBar:vertical {{ background: {_BG}; width: 8px; margin: 0; }}"
    f"QScrollBar::handle:vertical {{ background: #D1D5DB; border-radius: 4px; min-height: 24px; }}"
    f"QScrollBar::add-line, QScrollBar::sub-line {{ width: 0; height: 0; }}"
)


# ── Helpers ─────────────────────────────────────────────────────────────────────

def _lbl(text: str = "", size: int = 13, weight: int = 400, color: str = _T1) -> QLabel:
    w = QLabel(text)
    w.setStyleSheet(
        f"color: {color}; font-size: {size}px; font-weight: {weight};"
        " font-family:'Segoe UI'; background: transparent;"
    )
    return w


def _input_ss() -> str:
    return (
        f"QLineEdit, QComboBox, QDateEdit {{"
        f"  border: 1px solid {_BORDER}; border-radius: 5px;"
        f"  background: {_WHITE}; color: {_T1}; font-size: 12px;"
        "  font-family:'Segoe UI'; padding: 0 8px;"
        "  min-height: 32px; max-height: 32px; }"
        f"QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{ border-color: {_BLUE}; }}"
        "QComboBox::drop-down { border: none; width: 20px; }"
    )


def _btn(text: str, icon_name: str = "", primary: bool = False) -> QPushButton:
    b = QPushButton(f"  {text}" if icon_name else text)
    b.setCursor(Qt.PointingHandCursor)
    b.setFixedHeight(32)
    if icon_name:
        try:
            b.setIcon(qta.icon(icon_name, color="#FFFFFF" if primary else _T2))
            b.setIconSize(QSize(15, 15))
        except Exception:
            pass
    if primary:
        b.setStyleSheet(
            f"QPushButton {{ background: {_BLUE}; color: #FFF; border: none;"
            " border-radius: 5px; font-size: 12px; font-weight: 600;"
            " font-family:'Segoe UI'; padding: 0 12px; }"
            "QPushButton:hover { background: #005EA3; }"
            "QPushButton:disabled { background: #93C5FD; }"
        )
    else:
        b.setStyleSheet(
            f"QPushButton {{ background: {_WHITE}; color: {_T1};"
            f" border: 1px solid {_BORDER}; border-radius: 5px; font-size: 12px;"
            " font-family:'Segoe UI'; padding: 0 12px; }"
            f"QPushButton:hover {{ background: {_BG}; }}"
        )
    return b


def _set_cell(table: QTableWidget, row: int, col: int, text: str,
              align: str = "left", bg: str = _WHITE, color: str = _T1,
              mono: bool = False) -> None:
    item = QTableWidgetItem(text)
    flag = {"left": Qt.AlignLeft, "right": Qt.AlignRight, "center": Qt.AlignHCenter}[align]
    item.setTextAlignment(int(flag) | int(Qt.AlignVCenter))
    item.setBackground(QBrush(QColor(bg)))
    item.setForeground(QBrush(QColor(color)))
    item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
    if mono:
        f = QFont("Cascadia Code", 11)
        f.setStyleHint(QFont.Monospace)
        item.setFont(f)
    table.setItem(row, col, item)


def _receipt_text(status: str) -> tuple:
    """Return (display text, text color) for a receipt status — no pill/background."""
    text, fg, _bg = _RECEIPT_MAP.get(status, ("—", _TM, ""))
    return text, fg


def _fmt_short(amount: float) -> str:
    a = abs(amount)
    if a >= 1_000_000:
        return f"{amount / 1_000_000:.1f}M"
    if a >= 1_000:
        return f"{amount / 1_000:.0f}K"
    return f"{amount:,.0f}"


# ── CategoryTableWidget ──────────────────────────────────────────────────────────

class CategoryTableWidget(QWidget):
    """Read-only verified-transactions table scoped to a single category."""

    def __init__(self, category_name: str, title: str, icon_name: str,
                 description_filter: str = "",
                 parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._category = category_name
        self._title = title
        self._icon = icon_name
        self._description_filter = description_filter

        self._year = app_state.fiscal_year
        self._month = 0
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._scroll_loading = False
        self._reload_generation = 0
        self._export_in_flight = False

        self._debounce = QTimer(self)
        self._debounce.setSingleShot(True)
        self._debounce.setInterval(350)
        self._debounce.timeout.connect(self._reset_and_reload)

        self._build()

    # ── Build ───────────────────────────────────────────────────────────────
    def _build(self) -> None:
        self.setStyleSheet(f"background: {_BG};")
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Title bar ──────────────────────────────────────────────────────
        title_bar = QFrame()
        title_bar.setFixedHeight(52)
        title_bar.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-bottom: 1px solid {_BORDER}; }}"
        )
        tb = QHBoxLayout(title_bar)
        tb.setContentsMargins(20, 0, 20, 0)
        tb.setSpacing(12)

        try:
            icon_lbl = QLabel()
            icon_lbl.setFixedSize(22, 22)
            icon_lbl.setPixmap(qta.icon(self._icon, color=_BLUE).pixmap(22, 22))
            icon_lbl.setStyleSheet("background: transparent;")
            tb.addWidget(icon_lbl)
        except Exception:
            pass

        tb.addWidget(_lbl(self._title, size=16, weight=700))
        self._count_lbl = _lbl("", size=12, color=_T2)
        tb.addWidget(self._count_lbl)
        tb.addStretch()

        tb.addWidget(_lbl("FY", size=12, color=_T2))
        self._fy_cb = QComboBox()
        current_yr = datetime.now().year
        for yr in range(current_yr - 3, current_yr + 2):
            self._fy_cb.addItem(str(yr), yr)
        idx = self._fy_cb.findData(self._year)
        self._fy_cb.setCurrentIndex(idx if idx >= 0 else self._fy_cb.count() - 2)
        self._fy_cb.setFixedWidth(80)
        self._fy_cb.setStyleSheet(_input_ss())
        self._fy_cb.currentIndexChanged.connect(self._on_year_changed)
        tb.addWidget(self._fy_cb)

        refresh_btn = QPushButton()
        refresh_btn.setFixedSize(32, 32)
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(
            "QPushButton { background: transparent; border: none; border-radius: 4px; }"
            f"QPushButton:hover {{ background: {_BG}; }}"
        )
        try:
            refresh_btn.setIcon(qta.icon("mdi.refresh", color=_T2))
            refresh_btn.setIconSize(QSize(18, 18))
        except Exception:
            refresh_btn.setText("↻")
        refresh_btn.clicked.connect(self.refresh)
        tb.addWidget(refresh_btn)

        root.addWidget(title_bar)

        # ── Filter bar ─────────────────────────────────────────────────────
        filter_bar = QFrame()
        filter_bar.setFixedHeight(52)
        filter_bar.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-bottom: 1px solid {_BORDER}; }}"
        )
        fl = QHBoxLayout(filter_bar)
        fl.setContentsMargins(16, 0, 16, 0)
        fl.setSpacing(10)

        try:
            si = QLabel()
            si.setFixedSize(16, 16)
            si.setPixmap(qta.icon("mdi.magnify", color=_TM).pixmap(16, 16))
            si.setStyleSheet("background: transparent;")
            fl.addWidget(si)
        except Exception:
            pass

        self._search = QLineEdit()
        self._search.setPlaceholderText("Search description or truck…")
        self._search.setFixedWidth(240)
        self._search.setStyleSheet(_input_ss())
        self._search.textChanged.connect(self._on_filter_changed)
        fl.addWidget(self._search)

        self._month_cb = QComboBox()
        for idx_m, label in _MONTHS:
            self._month_cb.addItem(label, idx_m)
        self._month_cb.setFixedWidth(120)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.currentIndexChanged.connect(self._on_month_changed)
        fl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            fl, self._on_filter_changed, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=False,
        )
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=False)

        self._rcpt_cb = QComboBox()
        for label, val in [("All Receipts", "all"), ("Received", "received"),
                           ("Pending", "pending"), ("No Receipt", "missing")]:
            self._rcpt_cb.addItem(label, val)
        self._rcpt_cb.setFixedWidth(128)
        self._rcpt_cb.setStyleSheet(_input_ss())
        self._rcpt_cb.currentIndexChanged.connect(self._on_filter_changed)
        fl.addWidget(self._rcpt_cb)

        clear_btn = _btn("Clear", "mdi.filter-remove-outline", primary=False)
        clear_btn.setToolTip("Clear search, date, receipt filters, and column sort.")
        clear_btn.clicked.connect(self._clear_filters)
        fl.addWidget(clear_btn)

        fl.addStretch()

        export_btn = _btn("Export Excel", "mdi.microsoft-excel")
        export_btn.clicked.connect(self._on_export)
        fl.addWidget(export_btn)

        root.addWidget(filter_bar)

        # ── Table ──────────────────────────────────────────────────────────
        self._table = QTableWidget(0, len(_COLS))
        self._table.setHorizontalHeaderLabels([c[0] for c in _COLS])
        self._table.setStyleSheet(_TABLE_SS)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(_ROW_H)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.setShowGrid(True)
        hdr = self._table.horizontalHeader()
        hdr.setSectionsMovable(False)
        hdr.setStretchLastSection(True)
        for i, (_, width, _a, _m) in enumerate(_COLS):
            self._table.setColumnWidth(i, width)
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        bind_column_width_persistence(
            self._table,
            f"category_{item_key(self._category)}",
            [c[1] for c in _COLS],
        )
        self._sort_state = wire_feed_table_sort(
            self._table,
            CATEGORY_TABLE_SORT,
            default_field="date",
            default_asc=False,
            on_sort_changed=self._on_sort_changed,
        )
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        root.addWidget(self._table, 1)

        # ── Footer totals ──────────────────────────────────────────────────
        footer = QFrame()
        footer.setFixedHeight(40)
        footer.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 2px solid {_BORDER}; }}"
        )
        fol = QHBoxLayout(footer)
        fol.setContentsMargins(16, 0, 16, 0)
        fol.setSpacing(20)
        fol.addWidget(_lbl("TOTAL (filtered view)", size=11, weight=700, color=_T2))
        self._tzs_lbl = _lbl("TZS  —", size=13, weight=700)
        fol.addWidget(self._tzs_lbl)

        sep = QFrame()
        sep.setFrameShape(QFrame.VLine)
        sep.setFixedWidth(1)
        sep.setStyleSheet(f"background: {_BORDER};")
        fol.addWidget(sep)

        self._usd_lbl = _lbl("USD  —", size=13, weight=700)
        fol.addWidget(self._usd_lbl)
        fol.addStretch()
        self._footer_count = _lbl("", size=11, color=_T2)
        fol.addWidget(self._footer_count)
        root.addWidget(footer)

        # ── Scroll status ──────────────────────────────────────────────────
        status_bar = QFrame()
        status_bar.setFixedHeight(36)
        status_bar.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 1px solid {_BORDER}; }}"
        )
        sl = QHBoxLayout(status_bar)
        sl.setContentsMargins(16, 0, 16, 0)
        self._scroll_status = _lbl("—", size=12, color=_T2)
        sl.addWidget(self._scroll_status)
        sl.addStretch()
        root.addWidget(status_bar)

        self._loading_overlay = LoadingOverlay(self, "Loading…")

    # ── Public API ───────────────────────────────────────────────────────────
    def refresh(self) -> None:
        self._reset_and_reload()

    def _reset_and_reload(self) -> None:
        self._reload_generation += 1
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        self._update_status()
        asyncio.ensure_future(self._reload())

    # ── Filter helpers ─────────────────────────────────────────────────────
    def _search_text(self) -> str:
        return self._search.text().strip()

    def _receipt_filter(self) -> str:
        return self._rcpt_cb.currentData() or "all"

    def _filter_kw(self) -> dict:
        date_from, date_to = self._date_filters()
        return dict(
            year=self._year, month=self._month,
            search=self._search_text(), truck="",
            category=self._category, receipt=self._receipt_filter(),
            description=self._description_filter,
            date_from=date_from, date_to=date_to,
        )

    def _on_sort_changed(self, field: str, asc: bool) -> None:
        self._reset_and_reload()

    # ── Event handlers ─────────────────────────────────────────────────────
    def _on_year_changed(self) -> None:
        yr = self._fy_cb.currentData()
        if yr and yr != self._year:
            self._year = yr
            app_state.fiscal_year = yr
            sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=False)
            self._reset_and_reload()

    def _on_month_changed(self) -> None:
        self._month = self._month_cb.currentData() or 0
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=False)
        self._reset_and_reload()

    def _date_filters(self):
        return read_from_to(self._from_date, self._to_date, optional=False)

    def _on_filter_changed(self) -> None:
        self._debounce.start()

    def _clear_filters(self) -> None:
        self._search.blockSignals(True)
        self._month_cb.blockSignals(True)
        self._rcpt_cb.blockSignals(True)
        try:
            self._search.clear()
            self._month_cb.setCurrentIndex(0)
            self._rcpt_cb.setCurrentIndex(0)
        finally:
            self._search.blockSignals(False)
            self._month_cb.blockSignals(False)
            self._rcpt_cb.blockSignals(False)
        self._month = 0
        sync_from_to(self._from_date, self._to_date, self._year, 0, optional=False)
        reset_feed_sort(self._sort_state)
        self._reset_and_reload()

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _update_status(self) -> None:
        if self._loading and self._loaded == 0:
            self._scroll_status.setText("Loading…")
        elif self._total == 0:
            self._scroll_status.setText("No records match the current filters.")
        elif self._loaded >= self._total:
            self._scroll_status.setText(f"Showing all {self._total:,} records")
        else:
            self._scroll_status.setText(
                f"Showing {self._loaded:,} of {self._total:,}  •  Scroll down for more"
            )

    # ── Reload ─────────────────────────────────────────────────────────────
    async def _reload(self) -> None:
        gen = self._reload_generation
        self._loading = True
        self._loading_overlay.show_loading(f"Loading {self._title}…")
        self._update_status()
        try:
            from tahmeed.services.accountant_service import (
                get_master_transactions, count_master_transactions, get_master_totals,
            )
            kw = self._filter_kw()
            txs, total, totals = await asyncio.gather(
                get_master_transactions(
                    **kw, **sort_kw(self._sort_state),
                    limit=_SCROLL_CHUNK, skip=0,
                ),
                count_master_transactions(**kw),
                get_master_totals(**kw),
            )
            if gen != self._reload_generation:
                return
            self._total = total
            self._populate(txs, 0, append=False)
            self._loaded = len(txs)
            self._tzs_lbl.setText(f"TZS  {totals['tzs']:,.0f}")
            usd = totals["usd"]
            self._usd_lbl.setText(f"USD  ${usd:,.2f}" if usd else "USD  —")
            self._footer_count.setText(f"{total:,} records")
            self._count_lbl.setText(f"{total:,} records")
        except Exception as exc:
            if gen == self._reload_generation:
                self._table.setRowCount(0)
                self._scroll_status.setText(f"Failed to load: {exc}")
        finally:
            if gen == self._reload_generation:
                self._loading = False
                self._loading_overlay.hide_loading()
                self._update_status()

    async def _load_more(self) -> None:
        if self._scroll_loading or self._loading:
            return
        if self._loaded >= self._total:
            return
        self._scroll_loading = True
        self._update_status()
        try:
            from tahmeed.services.accountant_service import get_master_transactions

            gen = self._reload_generation
            kw = self._filter_kw()
            txs = await get_master_transactions(
                **kw, **sort_kw(self._sort_state),
                limit=_SCROLL_CHUNK, skip=self._loaded,
            )
            if gen != self._reload_generation:
                return
            if txs:
                self._populate(txs, self._loaded, append=True)
                self._loaded += len(txs)
        except Exception:
            pass
        finally:
            self._scroll_loading = False
            self._update_status()

    def _populate(self, txs: List[Transaction], skip: int, *, append: bool = False) -> None:
        if not append:
            self._table.setRowCount(0)
        start_row = self._table.rowCount()
        for i, tx in enumerate(txs):
            r = start_row + i
            self._table.insertRow(r)
            row_bg = _STRIPE if (skip + i) % 2 else _WHITE
            _set_cell(self._table, r, 0, str(skip + i + 1), "center", row_bg)
            _set_cell(self._table, r, 1,
                      tx.date.strftime("%d %b %Y") if tx.date else "—", "left", row_bg)
            _set_cell(self._table, r, 2, tx.item or "—", "left", row_bg)
            _set_cell(self._table, r, 3, tx.description or "—", "left", row_bg)
            _set_cell(self._table, r, 4, tx.truck_number or "—", "left", row_bg)
            _set_cell(self._table, r, 5, tx.memo or "—", "left", row_bg)
            _set_cell(self._table, r, 6, "✓" if tx.notes_flag else "—", "center",
                      row_bg, color=_BLUE if tx.notes_flag else _TM)

            if tx.currency == "TZS":
                tzs_txt = f"{tx.amount:,.0f}"
                tzs_col = _T1
            else:
                tzs_txt, tzs_col = "—", _TM
            _set_cell(self._table, r, 7, tzs_txt, "right", row_bg, color=tzs_col)

            rcpt_text, rcpt_fg = _receipt_text(tx.receipt_status or "pending")
            _set_cell(self._table, r, 8, rcpt_text, "center", row_bg, color=rcpt_fg)

            _set_cell(self._table, r, 9, tx.ownership or "—", "left", row_bg)
            _set_cell(self._table, r, 10, tx.approver or "—", "left", row_bg)
            self._table.setRowHeight(r, _ROW_H)

    # ── Excel export ───────────────────────────────────────────────────────
    def _on_export(self) -> None:
        if self._export_in_flight:
            return
        asyncio.ensure_future(self._do_export())

    async def _do_export(self) -> None:
        if self._export_in_flight:
            return
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            await notify_export_error(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n\nRun: pip install openpyxl",
            )
            return

        from tahmeed.services.accountant_service import get_master_transactions

        self._export_in_flight = True
        overlay = attach_export_overlay(self)
        date_from, date_to = self._date_filters()
        kw = dict(
            year=self._year, month=self._month,
            search=self._search_text(), truck="",
            category=self._category, receipt=self._receipt_filter(),
            description=self._description_filter,
            date_from=date_from, date_to=date_to,
        )
        try:
            show_export_busy(overlay, f"Loading {self._title}…", maximum=0)
            try:
                txs = await fetch_records_with_progress(
                    overlay,
                    lambda *, limit, skip: get_master_transactions(
                        **kw, **sort_kw(self._sort_state), limit=limit, skip=skip,
                    ),
                    phase=f"Loading {self._title}",
                )
            except Exception as exc:
                await notify_export_error(self, "Export Error", f"Failed to fetch data: {exc}")
                return
            finally:
                hide_export_busy(self)

            if not txs:
                await notify_export_info(self, "Export", "No records to export.")
                return

            from tahmeed.services.export_restriction_service import (
                filter_transactions_for_export,
            )

            txs = await filter_transactions_for_export(
                txs, surface="category_tables", fmt="excel",
            )
            if not txs:
                await notify_export_info(
                    self, "Export",
                    "All records belong to items restricted from Excel export.",
                )
                return

            month_tag = dict(_MONTHS).get(self._month, "All").replace(" ", "")
            safe_title = self._title.replace(" ", "_").replace("&", "and").replace("/", "-")
            default = f"{safe_title}_FY{self._year}_{month_tag}.xlsx"
            path = await pick_export_path(self, f"Export {self._title}", default)
            if not path:
                return
            path = normalize_xlsx_path(path)

            title = self._title
            year = self._year
            month = self._month
            total = len(txs)
            fast = total >= FAST_STYLE_ROW_LIMIT

            def _write(progress_cb) -> None:
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = title[:28]

                ws.merge_cells("A1:K1")
                ws["A1"] = "TAHMEED COACH TZ LTD"
                ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1B2B4B")
                ws["A1"].alignment = Alignment(horizontal="center")
                month_label = dict(_MONTHS).get(month, "All Months")
                ws.merge_cells("A2:K2")
                ws["A2"] = f"{title} — FY {year}  |  {month_label}"
                ws["A2"].font = Font(name="Segoe UI", bold=True, size=11, color="374151")
                ws["A2"].alignment = Alignment(horizontal="center")
                ws.append([])

                headers = [c[0] for c in _COLS]
                ws.append(headers)
                hdr_row = ws.max_row
                grey = PatternFill("solid", fgColor="F1F5F9")
                for cell in ws[hdr_row]:
                    cell.font = Font(name="Segoe UI", bold=True, size=10, color="6B7280")
                    cell.fill = grey
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                stripe = PatternFill("solid", fgColor="F1F5F9")
                white = PatternFill("solid", fgColor="FFFFFF")
                body_font = Font(name="Segoe UI", size=10)
                tzs_total = usd_total = 0.0
                for i, tx in enumerate(txs):
                    if tx.currency == "TZS":
                        tzs_val, usd_val = tx.amount, None
                        tzs_total += tx.amount
                    else:
                        tzs_val, usd_val = None, tx.amount
                        usd_total += tx.amount
                    receipt_str = {
                        "received": "Received", "pending": "Pending",
                        "missing": "No Receipt",
                    }.get(tx.receipt_status or "", "")
                    ws.append([
                        i + 1,
                        tx.date.strftime("%d-%b-%Y") if tx.date else "",
                        tx.item or "", tx.description or "", tx.truck_number or "",
                        tx.memo or "", "Yes" if tx.notes_flag else "",
                        tzs_val, receipt_str, tx.ownership or "", tx.approver or "",
                    ])
                    if not fast:
                        r = ws.max_row
                        fill = stripe if i % 2 else white
                        for cell in ws[r]:
                            cell.fill = fill
                            cell.alignment = Alignment(vertical="center")
                        c = ws.cell(r, 8)
                        if tzs_val is not None:
                            c.font = body_font
                            c.number_format = "#,##0"
                            c.alignment = Alignment(horizontal="right", vertical="center")

                    if progress_cb and (
                        (i + 1) % PROGRESS_EVERY == 0 or i + 1 == total
                    ):
                        progress_cb(i + 1, "Writing rows")

                ws.append([])
                ws.append(["", "", "", "TOTAL", "", "", "", tzs_total or "", "", "", ""])
                total_r = ws.max_row
                ws.cell(total_r, 4).font = Font(name="Segoe UI", bold=True, size=11)
                if tzs_total:
                    c = ws.cell(total_r, 8)
                    c.font = Font(name="Segoe UI", bold=True, size=11, color="111827")
                    c.number_format = "#,##0"
                    c.alignment = Alignment(horizontal="right", vertical="center")

                widths = [6, 12, 14, 34, 12, 20, 7, 15, 13, 13, 12]
                for idx, w in enumerate(widths, 1):
                    ws.column_dimensions[ws.cell(1, idx).column_letter].width = w
                ws.freeze_panes = ws.cell(hdr_row + 1, 1)
                if progress_cb:
                    progress_cb(total, "Saving file")
                wb.save(path)

            try:
                await run_export_write(overlay, total, _write)
            except Exception as exc:
                await notify_export_error(self, "Save Error", f"Could not save file:\n{exc}")
                return
            finally:
                hide_export_busy(self)

            if not export_file_ready(path):
                await notify_export_error(
                    self, "Save Error", f"The file could not be saved:\n{path}",
                )
                return

            await notify_export_info(
                self, "Export Complete",
                f"Exported {len(txs):,} records to:\n{path}",
            )
        finally:
            hide_export_busy(self)
            self._export_in_flight = False
