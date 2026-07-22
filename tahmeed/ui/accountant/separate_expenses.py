"""AccountantDashboard — Separate Expenses widgets (ASK 8).

Covers all eight views under the SEPARATE EXPENSES sidebar section:
  TollPlazaWidget      — import from Dot Com Zambia xlsx/csv, dedup by Receipt No
  ParkingCongoWidget   — import from Congo transporter ledger, dedup by Serial
  CongoExpensesWidget  — Excel import (last sheet), upload browse + detail view
  AhmedKimviWidget     — Excel import (last sheet), upload browse + detail view
  ZambiaParkingWidget  — weekly statement import, upload browse + detail view
  AfritrackWidget      — placeholder stub
  ThirdPartyWidget     — placeholder stub
  ComesaWidget         — placeholder stub
"""

from __future__ import annotations

import asyncio
import csv
import io
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from tahmeed.ui.accountant.date_filters import (
    add_from_to_editors, read_from_to, sync_from_to,
)

import qtawesome as qta

from PySide6.QtCore import Qt, QSize, QTimer, Signal
from PySide6.QtGui import QColor, QDragEnterEvent, QDropEvent, QFont
from PySide6.QtWidgets import (
    QAbstractItemView, QApplication, QComboBox, QDialog, QDialogButtonBox,
    QFileDialog, QFormLayout, QFrame, QGridLayout, QHBoxLayout, QHeaderView,
    QLabel, QLineEdit, QMenu, QMessageBox, QPushButton, QSizePolicy,
    QStackedWidget, QTableWidget, QTableWidgetItem, QVBoxLayout,
    QWidget, QDateEdit,
)

# ── openpyxl (optional — fallback to csv-only if missing) ──────────────────────
try:
    import openpyxl
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── Design tokens (match dashboard palette) ────────────────────────────────────
_WHITE   = "#FFFFFF"
_BG      = "#F4F6F8"
_BORDER  = "#E5E7EB"
_BLUE    = "#0077C5"
_BLUE_L  = "#E8F4FD"
_GREEN   = "#16A34A"
_GREEN_L = "#DCFCE7"
_AMBER   = "#D97706"
_AMBER_L = "#FEF3C7"
_RED     = "#DC2626"
_RED_L   = "#FEE2E2"
_T1      = "#111827"
_T2      = "#6B7280"
_TM      = "#9CA3AF"
_HDR_BG  = "#F1F5F9"
_ALT_ROW = "#F9FAFB"
_NAVY    = "#1B2B4B"

_PAGE_SIZES = [25, 50, 100]
_ROW_H      = 28
_HDR_H      = 26
_ROW_EVEN   = "#FFFFFF"
_ROW_ODD    = "#F1F5F9"   # slate-100 — manual stripe (readable on all row types)
_CREDIT_ROW_BG = "#E6F4EC"  # soft mint for money-in / credit rows
_CREDIT_FG       = "#047857"

# ── Afritrack / QuickBooks-inspired palette (used only for AfritrackWidget) ───
_QB_HDR_BG   = "#EFF6FF"   # very light blue header
_QB_HDR_FG   = "#1E3A5F"   # dark navy header text
_QB_FORM_BG  = "#EFF6FF"   # formula-cell background
_QB_FORM_FG  = "#1D4ED8"   # formula-cell text (blue)
_QB_VNEG_BG  = "#FEF2F2"   # negative variance background
_QB_VNEG_FG  = "#DC2626"   # negative variance text
_QB_VPOS_FG  = "#15803D"   # positive variance text
_QB_VZRO_FG  = "#9CA3AF"   # zero variance text
_QB_SEL_BG   = "#DBEAFE"   # selection highlight
_QB_SEL_FG   = "#1E3A5F"   # selection text
_QB_FOOT_BG  = "#F8FAFC"   # footer background
_QB_BOLD_FG  = "#0F172A"   # bold row text
_QB_RED_DARK = "#B91C1C"   # red totals text
_QB_RED_BG   = "#FEF2F2"   # red totals background


# ═══════════════════════════════════════════════════════════════════════════════
#  Primitive helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _lbl(text: str = "", size: int = 13, weight: int = 400,
         color: str = _T1) -> QLabel:
    w = QLabel(text)
    w.setStyleSheet(
        f"color:{color};font-size:{size}px;font-weight:{weight};"
        "font-family:'Segoe UI';background:transparent;"
    )
    return w


def _btn(text: str, icon: str = "", primary: bool = True,
         danger: bool = False, height: int = 34) -> QPushButton:
    b = QPushButton(f"  {text}" if icon else text)
    b.setCursor(Qt.PointingHandCursor)
    b.setFixedHeight(height)
    if icon:
        try:
            b.setIcon(qta.icon(icon, color="#FFFFFF" if (primary or danger) else _T1))
            b.setIconSize(QSize(16, 16))
        except Exception:
            pass
    if danger:
        ss = (f"QPushButton{{background:{_RED};color:#FFF;border:none;border-radius:5px;"
              f"font-size:12px;font-weight:600;font-family:'Segoe UI';padding:0 14px;}}"
              f"QPushButton:hover{{background:#B91C1C;}}"
              f"QPushButton:disabled{{background:#FCA5A5;}}")
    elif primary:
        ss = (f"QPushButton{{background:{_BLUE};color:#FFF;border:none;border-radius:5px;"
              f"font-size:12px;font-weight:600;font-family:'Segoe UI';padding:0 14px;}}"
              f"QPushButton:hover{{background:#005EA3;}}"
              f"QPushButton:disabled{{background:#93C5FD;}}")
    else:
        ss = (f"QPushButton{{background:{_WHITE};color:{_T1};border:1px solid {_BORDER};"
              f"border-radius:5px;font-size:12px;font-family:'Segoe UI';padding:0 14px;}}"
              f"QPushButton:hover{{background:{_BG};}}"
              f"QPushButton:disabled{{color:{_TM};}}")
    b.setStyleSheet(ss)
    return b


def _input_ss() -> str:
    return (
        f"QLineEdit,QComboBox,QDateEdit{{border:1px solid {_BORDER};border-radius:5px;"
        f"background:{_WHITE};color:{_T1};font-size:12px;"
        f"font-family:'Segoe UI';padding:0 8px;"
        f"min-height:32px;max-height:32px;}}"
        f"QLineEdit:focus,QComboBox:focus,QDateEdit:focus{{border-color:{_BLUE};}}"
        "QComboBox::drop-down{border:none;width:20px;}"
    )


def _hsep() -> QFrame:
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setFixedHeight(1)
    f.setStyleSheet(f"background:{_BORDER};")
    return f


def _card(widget: QWidget) -> QFrame:
    f = QFrame()
    f.setStyleSheet(
        f"QFrame{{background:{_WHITE};border:1px solid {_BORDER};"
        "border-radius:6px;}}"
    )
    vl = QVBoxLayout(f)
    vl.setContentsMargins(0, 0, 0, 0)
    vl.setSpacing(0)
    vl.addWidget(widget)
    return f


def _table_style() -> str:
    return (
        f"QTableWidget{{background:{_WHITE};gridline-color:{_BORDER};"
        "border:none;font-size:11px;font-family:'Segoe UI';}}"
        f"QTableWidget::item{{padding:2px 8px;color:{_T1};}}"
        f"QTableWidget::item:selected{{background:{_BLUE_L};color:{_T1};}}"
        f"QHeaderView::section{{background:{_HDR_BG};color:{_T2};"
        "font-size:10px;font-weight:600;font-family:'Segoe UI';"
        f"border:none;border-bottom:1px solid {_BORDER};"
        f"padding:0 8px;height:{_HDR_H}px;}}"
        "QScrollBar:vertical{width:8px;background:transparent;}"
        f"QScrollBar::handle:vertical{{background:#D1D5DB;border-radius:4px;}}"
    )


def _stripe_bg(row_idx: int) -> str:
    return _ROW_ODD if row_idx % 2 else _ROW_EVEN


def _apply_row_bg(t: QTableWidget, row: int, bg: str) -> None:
    for col in range(t.columnCount()):
        item = t.item(row, col)
        if item:
            item.setBackground(QColor(bg))


def _finish_table_row(
    t: QTableWidget, row: int, bg: str | None = None,
) -> None:
    """Apply stripe (or custom) background and compact row height."""
    _apply_row_bg(t, row, bg if bg else _stripe_bg(row))
    t.setRowHeight(row, _ROW_H)


def _make_table(headers: List[str]) -> QTableWidget:
    t = QTableWidget(0, len(headers))
    t.setHorizontalHeaderLabels(headers)
    t.setEditTriggers(QAbstractItemView.NoEditTriggers)
    t.setSelectionBehavior(QAbstractItemView.SelectRows)
    t.setAlternatingRowColors(False)
    t.verticalHeader().setVisible(False)
    t.verticalHeader().setDefaultSectionSize(_ROW_H)
    t.horizontalHeader().setStretchLastSection(True)
    t.setStyleSheet(_table_style())
    return t


def _cell(text: str, align: Qt.AlignmentFlag = Qt.AlignLeft | Qt.AlignVCenter,
          mono: bool = False, color: str = "") -> QTableWidgetItem:
    item = QTableWidgetItem(str(text) if text is not None else "—")
    item.setTextAlignment(align)
    if mono:
        item.setFont(QFont("Cascadia Code", 11))
    if color:
        item.setForeground(QColor(color))
    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
    return item


def _fmt_date(dt) -> str:
    if isinstance(dt, datetime):
        return dt.strftime("%d %b %y")
    return str(dt) if dt else "—"


def _fmt_num(v, prefix: str = "", decimals: int = 2) -> str:
    if v is None:
        return "—"
    try:
        return f"{prefix}{float(v):,.{decimals}f}"
    except Exception:
        return str(v)


def _kimvi_fmt_amount(amt: float) -> str:
    """Format signed USD — negative = money in, positive = money out."""
    if amt is None:
        return "—"
    try:
        v = float(amt)
    except (TypeError, ValueError):
        return str(amt)
    if v < 0:
        return f"({abs(v):,.0f})"
    return f"{v:,.0f}"


# ═══════════════════════════════════════════════════════════════════════════════
#  Shared page title + toolbar wrapper
# ═══════════════════════════════════════════════════════════════════════════════

class _PageHeader(QWidget):
    """Title row with optional right-aligned buttons."""

    def __init__(self, title: str, icon_name: str = "mdi.cash-multiple",
                 parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setStyleSheet("background:transparent;")
        hl = QHBoxLayout(self)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(10)

        try:
            icon_lbl = QLabel()
            icon_lbl.setPixmap(qta.icon(icon_name, color=_BLUE).pixmap(22, 22))
            icon_lbl.setFixedSize(22, 22)
            icon_lbl.setStyleSheet("background:transparent;")
            hl.addWidget(icon_lbl)
        except Exception:
            pass

        title_lbl = _lbl(title, size=18, weight=700)
        hl.addWidget(title_lbl)
        hl.addStretch()
        self._right = hl

    def add_right(self, widget: QWidget) -> None:
        self._right.addWidget(widget)


# ═══════════════════════════════════════════════════════════════════════════════
#  Pagination footer
# ═══════════════════════════════════════════════════════════════════════════════

class _PaginationBar(QWidget):
    page_changed = Signal(int)
    size_changed = Signal(int)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setStyleSheet("background:transparent;")
        self._page = 1
        self._total = 0
        self._page_size = 25

        hl = QHBoxLayout(self)
        hl.setContentsMargins(0, 4, 0, 4)
        hl.setSpacing(8)

        self._size_cb = QComboBox()
        self._size_cb.setFixedWidth(80)
        for s in _PAGE_SIZES:
            self._size_cb.addItem(f"Show {s}", s)
        self._size_cb.setStyleSheet(_input_ss())
        self._size_cb.currentIndexChanged.connect(self._on_size_change)
        hl.addWidget(self._size_cb)

        self._info_lbl = _lbl("", size=11, color=_T2)
        hl.addWidget(self._info_lbl)
        hl.addStretch()

        self._prev_btn = _btn("← Prev", primary=False, height=30)
        self._prev_btn.clicked.connect(self._prev)
        hl.addWidget(self._prev_btn)

        self._page_lbl = _lbl("Page 1 of 1", size=11, color=_T2)
        hl.addWidget(self._page_lbl)

        self._next_btn = _btn("Next →", primary=False, height=30)
        self._next_btn.clicked.connect(self._next)
        hl.addWidget(self._next_btn)

    def set_total(self, total: int, page_size: int, current_page: int) -> None:
        self._total = total
        self._page_size = page_size
        self._page = current_page
        pages = max(1, (total + page_size - 1) // page_size)
        self._page_lbl.setText(f"Page {current_page} of {pages}")
        self._info_lbl.setText(f"{total:,} records")
        self._prev_btn.setEnabled(current_page > 1)
        self._next_btn.setEnabled(current_page < pages)

    def _prev(self) -> None:
        if self._page > 1:
            self._page -= 1
            self.page_changed.emit(self._page)

    def _next(self) -> None:
        pages = max(1, (self._total + self._page_size - 1) // self._page_size)
        if self._page < pages:
            self._page += 1
            self.page_changed.emit(self._page)

    def _on_size_change(self, idx: int) -> None:
        self._page = 1
        self._page_size = self._size_cb.currentData()
        self.size_changed.emit(self._page_size)


# ═══════════════════════════════════════════════════════════════════════════════
#  Footer totals bar
# ═══════════════════════════════════════════════════════════════════════════════

class _TotalsBar(QFrame):
    def __init__(self, labels: List[Tuple[str, str]], parent: QWidget | None = None) -> None:
        """labels = [(key, display_prefix), ...]  e.g. [("usd", "USD"), ("kwacha", "ZMW")]"""
        super().__init__(parent)
        self.setFixedHeight(38)
        self.setStyleSheet(
            f"QFrame{{background:{_HDR_BG};border-top:1px solid {_BORDER};"
            "border-bottom-left-radius:6px;border-bottom-right-radius:6px;}}"
        )
        hl = QHBoxLayout(self)
        hl.setContentsMargins(16, 0, 16, 0)
        hl.setSpacing(32)

        self._lbl_map: Dict[str, QLabel] = {}
        for key, prefix in labels:
            sub = QWidget()
            sub.setStyleSheet("background:transparent;")
            sub_hl = QHBoxLayout(sub)
            sub_hl.setContentsMargins(0, 0, 0, 0)
            sub_hl.setSpacing(4)
            sub_hl.addWidget(_lbl("TOTAL:", size=11, weight=600, color=_T2))
            val = _lbl("—", size=12, weight=700, color=_T1)
            sub_hl.addWidget(val)
            self._lbl_map[key] = val
            hl.addWidget(sub)
        hl.addStretch()

    def set_total(self, key: str, value: float, prefix: str = "") -> None:
        if key in self._lbl_map:
            self._lbl_map[key].setText(_fmt_num(value, prefix=prefix))


# ═══════════════════════════════════════════════════════════════════════════════
#  Import Dialog — shared by TollPlaza / ParkingCongo / ZambiaParking
# ═══════════════════════════════════════════════════════════════════════════════

class _DropZone(QFrame):
    """Drag-and-drop file drop target."""
    file_dropped = Signal(str)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setMinimumHeight(100)
        self.setStyleSheet(
            f"QFrame{{border:2px dashed {_BORDER};border-radius:8px;"
            f"background:{_BG};}}"
            f"QFrame:hover{{border-color:{_BLUE};background:{_BLUE_L};}}"
        )
        vl = QVBoxLayout(self)
        vl.setAlignment(Qt.AlignCenter)
        try:
            icon_lbl = QLabel()
            icon_lbl.setPixmap(qta.icon("mdi.upload-outline", color=_TM).pixmap(32, 32))
            icon_lbl.setAlignment(Qt.AlignCenter)
            icon_lbl.setStyleSheet("background:transparent;border:none;")
            vl.addWidget(icon_lbl)
        except Exception:
            pass
        vl.addWidget(_lbl("Drag & drop .xlsx or .csv here", size=13, color=_T2))
        self._path_lbl = _lbl("", size=11, color=_TM)
        self._path_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._path_lbl)

    def dragEnterEvent(self, event: QDragEnterEvent) -> None:
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            self.setStyleSheet(
                f"QFrame{{border:2px dashed {_BLUE};border-radius:8px;"
                f"background:{_BLUE_L};}}"
            )

    def dragLeaveEvent(self, event) -> None:
        self.setStyleSheet(
            f"QFrame{{border:2px dashed {_BORDER};border-radius:8px;"
            f"background:{_BG};}}"
            f"QFrame:hover{{border-color:{_BLUE};background:{_BLUE_L};}}"
        )

    def dropEvent(self, event: QDropEvent) -> None:
        urls = event.mimeData().urls()
        if urls:
            path = urls[0].toLocalFile()
            self._path_lbl.setText(path)
            self.file_dropped.emit(path)
        self.dragLeaveEvent(event)

    def set_path(self, path: str) -> None:
        self._path_lbl.setText(Path(path).name)


def _read_file_rows(
    path: str, header_row: int = 0
) -> Tuple[List[str], List[List[Any]]]:
    """Return (headers, data_rows) from an xlsx or csv file.

    header_row: 0-based index of the row containing column names.
    Rows before header_row are skipped entirely (e.g. title rows).
    """
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xls") and _HAS_OPENPYXL:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= header_row:
            return [], []
        headers = [str(c) if c is not None else "" for c in rows[header_row]]
        data = [
            [str(c) if c is not None else "" for c in r]
            for r in rows[header_row + 1:]
            if any(c is not None for c in r)
        ]
        return headers, data
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if len(rows) <= header_row:
            return [], []
        return rows[header_row], rows[header_row + 1:]


def _read_workbook_rows(path: str) -> List[List[Any]]:
    """Return every row from the active worksheet or CSV file."""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xls") and _HAS_OPENPYXL:
        wb = openpyxl.load_workbook(path, data_only=True)
        return list(wb.active.iter_rows(values_only=True))
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))


def _detect_header_row(
    rows: List[List[Any]],
    col_map: Dict[str, List[str]],
    max_scan: int = 30,
) -> int:
    """Pick the row whose cells best match the expected import column headers."""
    best_idx = 0
    best_score = -1
    for i, row in enumerate(rows[:max_scan]):
        hdr_lower = {
            str(c).strip().lower()
            for c in row
            if c is not None and str(c).strip()
        }
        if not hdr_lower:
            continue
        score = sum(
            1 for cands in col_map.values()
            if any(c.lower() in hdr_lower for c in cands)
        )
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _parse_rows_from_workbook(
    path: str,
    header_row: int,
    auto_header_row: bool,
    col_map: Dict[str, List[str]],
) -> Tuple[List[str], List[List[Any]]]:
    """Read a workbook and return (headers, data_rows)."""
    rows = _read_workbook_rows(path)
    if not rows:
        return [], []
    hdr_idx = _detect_header_row(rows, col_map) if auto_header_row else header_row
    if len(rows) <= hdr_idx:
        return [], []
    headers = [str(c) if c is not None else "" for c in rows[hdr_idx]]
    data = [
        [str(c) if c is not None else "" for c in r]
        for r in rows[hdr_idx + 1:]
        if any(c is not None and str(c).strip() for c in r)
    ]
    return headers, data


class ImportDialog(QDialog):
    """
    Generic file-import dialog.

    Parameters
    ----------
    feed_type       : str   "toll_plaza" | "parking_congo" | "zambia_parking"
    dedup_key       : str   column key used for duplicate detection
    preview_headers : list  column display names for the preview table
    col_map         : dict  maps expected_key → list of candidate header names
                            (case-insensitive, first match wins)
    save_fn         : coroutine  async fn(records: list[dict]) → int
    exist_fn        : coroutine  async fn(keys: list[str]) → set[str]
    header_row      : int   0-based row index of the column headers in the file
                            (default 0; use 1 for files with a title row above headers)
    auto_header_row : bool  scan the sheet to find the header row automatically
    template_title  : str   optional title row for downloadable template xlsx
    template_headers: list  column headers for downloadable template (enables button)
    template_filename: str  suggested filename for template download
    """

    imported = Signal(int)

    def __init__(
        self,
        feed_type: str,
        dedup_key: str,
        preview_headers: List[str],
        col_map: Dict[str, List[str]],
        save_fn,
        exist_fn,
        header_row: int = 0,
        auto_header_row: bool = False,
        template_title: str = "",
        template_headers: Optional[List[str]] = None,
        template_filename: str = "import_template.xlsx",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._feed_type       = feed_type
        self._dedup_key       = dedup_key
        self._preview_headers = preview_headers
        self._col_map         = col_map
        self._save_fn         = save_fn
        self._exist_fn        = exist_fn
        self._header_row      = header_row
        self._auto_header_row = auto_header_row
        self._template_title  = template_title
        self._template_headers = template_headers or []
        self._template_filename = template_filename

        # A fresh UUID tags every record from this import session so we can
        # later group/browse records by upload batch.
        self._upload_id:      str = str(uuid.uuid4())
        self._source_filename: str = ""

        self._raw_headers: List[str] = []
        self._all_rows:    List[dict] = []
        self._new_rows:    List[dict] = []

        self.setWindowTitle(f"Import — {feed_type.replace('_', ' ').title()}")
        self.setMinimumWidth(680)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        # Drop zone
        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        # Browse + optional template download
        browse_row = QWidget()
        browse_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(browse_row)
        brl.setContentsMargins(0, 0, 0, 0)
        brl.setSpacing(8)
        if self._template_headers:
            tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
            tmpl_btn.clicked.connect(self._download_template)
            brl.addWidget(tmpl_btn)
        browse_btn = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse_btn.clicked.connect(self._browse)
        brl.addStretch()
        brl.addWidget(browse_btn)
        vl.addWidget(browse_row)

        # Stats row
        self._stats_lbl = _lbl("No file loaded.", size=12, color=_T2)
        vl.addWidget(self._stats_lbl)

        vl.addWidget(_hsep())

        # Preview table
        preview_title = _lbl("Preview (first 10 rows)", size=12, weight=600)
        vl.addWidget(preview_title)

        self._preview_tbl = _make_table(self._preview_headers)
        self._preview_tbl.setMinimumHeight(200)
        vl.addWidget(self._preview_tbl)

        vl.addWidget(_hsep())

        # Buttons
        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.setSpacing(8)
        bbl.addStretch()

        cancel_btn = _btn("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        bbl.addWidget(cancel_btn)

        self._import_btn = _btn("Import Records", "mdi.check-circle-outline")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)

        vl.addWidget(btn_row)

    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open File", "", "Excel / CSV (*.xlsx *.xls *.csv)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _download_template(self) -> None:
        if not self._template_headers:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            self._template_filename,
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, self._template_title, self._template_headers)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _on_file(self, path: str) -> None:
        self._stats_lbl.setText("Reading file…")
        self._source_filename = Path(path).name
        try:
            if self._auto_header_row:
                headers, rows = _parse_rows_from_workbook(
                    path, self._header_row, True, self._col_map,
                )
            else:
                headers, rows = _read_file_rows(path, self._header_row)
        except Exception as exc:
            self._stats_lbl.setText(f"Error reading file: {exc}")
            return

        self._raw_headers = headers
        # Map columns
        hdr_lower = {h.strip().lower(): i for i, h in enumerate(headers)}

        def _find(candidates: List[str]) -> Optional[int]:
            for c in candidates:
                idx = hdr_lower.get(c.lower())
                if idx is not None:
                    return idx
            return None

        field_idxs: Dict[str, Optional[int]] = {
            key: _find(cands) for key, cands in self._col_map.items()
        }

        records: List[dict] = []
        for row in rows:
            rec: dict = {
                "_raw": row,
                "feed_type":       self._feed_type,
                "upload_id":       self._upload_id,
                "source_filename": self._source_filename,
            }
            for key, idx in field_idxs.items():
                rec[key] = row[idx].strip() if (idx is not None and idx < len(row)) else ""
            records.append(rec)

        self._all_rows = records
        dedup_vals = [r.get(self._dedup_key, "") for r in records if r.get(self._dedup_key)]
        asyncio.ensure_future(self._check_dupes(records, dedup_vals))

    async def _check_dupes(self, records: List[dict], keys: List[str]) -> None:
        try:
            existing: set = await self._exist_fn(keys)
        except Exception:
            existing = set()

        self._new_rows = [r for r in records
                          if r.get(self._dedup_key) not in existing]
        dupe_count = len(records) - len(self._new_rows)

        self._stats_lbl.setText(
            f"New records: {len(self._new_rows):,}     "
            f"Duplicates (skipped): {dupe_count:,}"
        )
        self._import_btn.setEnabled(bool(self._new_rows))
        self._import_btn.setText(f"Import {len(self._new_rows):,} Records")
        self._fill_preview(self._new_rows[:10])

    def _fill_preview(self, rows: List[dict]) -> None:
        t = self._preview_tbl
        t.setRowCount(0)
        keys = list(self._col_map.keys())
        for row in rows:
            r = t.rowCount()
            t.insertRow(r)
            for c, key in enumerate(keys):
                if c < t.columnCount():
                    t.setItem(r, c, _cell(row.get(key, "")))
            _finish_table_row(t, r)

    def _do_import(self) -> None:
        self._import_btn.setEnabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        try:
            saved = await self._save_fn(self._new_rows)
            self.imported.emit(saved)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(self._new_rows):,} Records")


# ═══════════════════════════════════════════════════════════════════════════════
#  1. TollPlazaWidget
# ═══════════════════════════════════════════════════════════════════════════════

def _write_xlsx_template(path: str, title: str, headers: List[str]) -> None:
    """Write an empty import template workbook with a title row and header row."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to generate Excel templates.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import"
    if title:
        ws.append([title])
    ws.append(headers)
    wb.save(path)


_TOLL_COL_MAP = {
    "toll_date":    ["toll date", "date", "transaction date"],
    "toll_plaza":   ["toll plaza", "plaza", "station"],
    "client_name":  ["client name", "client", "account name"],
    "card_no":      ["card no", "card number", "card"],
    "vehicle_reg":  ["vehicle reg", "vehicle", "vehicle registration", "plate"],
    "vehicle_class":["vehicle class", "class"],
    "tender_amount":["tender", "amount", "tender amount", "zmw"],
    "receipt_no":   ["receipt no", "receipt", "receipt number", "receipt no."],
    "device":       ["device code", "device", "device id"],
    "lane":         ["lane"],
    "cashier_name": ["cashier", "cashier name", "operator"],
}

# Column headers shown in the full per-record detail table
_TOLL_DETAIL_HEADERS = [
    "TOLL DATE", "TOLL PLAZA", "CLIENT NAME", "CARD NO", "VEHICLE REG",
    "CLASS", "TENDER (ZMW)", "RECEIPT NO", "DEVICE", "LANE", "CASHIER",
]

# Summary columns shown in the upload-list browse table
_TOLL_BROWSE_HEADERS = [
    "UPLOAD DATE", "FILE NAME", "RECORDS", "TOTAL (ZMW)", "DATE RANGE",
]

_TOLL_TEMPLATE_TITLE = "Dot Com Zambia — Toll Plaza Export"

_TOLL_MONTHS: List[Tuple[str, int]] = [
    ("All Months", 0),
    ("January", 1), ("February", 2), ("March", 3), ("April", 4),
    ("May", 5), ("June", 6), ("July", 7), ("August", 8),
    ("September", 9), ("October", 10), ("November", 11), ("December", 12),
]

_TOLL_SCROLL_CHUNK = 50
_SCROLL_CHUNK = 50


def _toll_fill_detail_row(t: QTableWidget, r: int, rec: dict) -> None:
    """Populate one Toll Plaza record row (shared by detail + all-entries views)."""
    t.setItem(r,  0, _cell(str(rec.get("toll_date",    "") or "")))
    t.setItem(r,  1, _cell(str(rec.get("toll_plaza",   "") or "")))
    t.setItem(r,  2, _cell(str(rec.get("client_name",  "") or "")))
    t.setItem(r,  3, _cell(str(rec.get("card_no",      "") or "")))
    t.setItem(r,  4, _cell(str(rec.get("vehicle_reg",  "") or "")))
    t.setItem(r,  5, _cell(str(rec.get("vehicle_class","") or "")))
    t.setItem(r,  6, _cell(
        _fmt_num(rec.get("tender_amount"), "", 0),
        align=Qt.AlignRight | Qt.AlignVCenter,
    ))
    t.setItem(r,  7, _cell(str(rec.get("receipt_no",   "") or "")))
    t.setItem(r,  8, _cell(str(rec.get("device",       "") or "")))
    t.setItem(r,  9, _cell(str(rec.get("lane",         "") or ""),
                           align=Qt.AlignCenter | Qt.AlignVCenter))
    t.setItem(r, 10, _cell(str(rec.get("cashier_name", "") or "")))
    _finish_table_row(t, r)


class _SegmentTabBar(QWidget):
    """Two-or-more-option segmented tab control."""

    tab_changed = Signal(int)

    def __init__(self, labels: List[str], parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._index = 0
        self._buttons: List[QPushButton] = []
        self.setStyleSheet("background:transparent;")
        hl = QHBoxLayout(self)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(4)
        for i, label in enumerate(labels):
            btn = QPushButton(label)
            btn.setCursor(Qt.PointingHandCursor)
            btn.setFixedHeight(32)
            btn.clicked.connect(lambda checked=False, idx=i: self.set_index(idx))
            self._buttons.append(btn)
            hl.addWidget(btn)
        hl.addStretch()
        self._apply_styles()

    def set_index(self, idx: int, emit: bool = True) -> None:
        if idx < 0 or idx >= len(self._buttons):
            return
        self._index = idx
        self._apply_styles()
        if emit:
            self.tab_changed.emit(idx)

    def current_index(self) -> int:
        return self._index

    def _apply_styles(self) -> None:
        active = (
            f"QPushButton{{background:{_BLUE};color:#FFF;border:none;border-radius:5px;"
            f"font-size:12px;font-weight:600;font-family:'Segoe UI';padding:0 16px;}}"
            f"QPushButton:hover{{background:#005EA3;}}"
        )
        inactive = (
            f"QPushButton{{background:{_WHITE};color:{_T1};border:1px solid {_BORDER};"
            f"border-radius:5px;font-size:12px;font-family:'Segoe UI';padding:0 16px;}}"
            f"QPushButton:hover{{background:{_BG};}}"
        )
        for i, btn in enumerate(self._buttons):
            btn.setStyleSheet(active if i == self._index else inactive)


def _populate_year_combo(cb: QComboBox, years: List[int], selected_year: int) -> int:
    """Rebuild a year filter combo from uploaded data; return the active year."""
    cb.blockSignals(True)
    cb.clear()
    cb.addItem("All Years", 0)
    for year in years:
        cb.addItem(str(year), year)
    active = selected_year
    if selected_year > 0:
        idx = cb.findData(selected_year)
        if idx >= 0:
            cb.setCurrentIndex(idx)
        else:
            active = 0
            cb.setCurrentIndex(0)
    cb.blockSignals(False)
    return active


class _TollAllEntries(QWidget):
    """Flat, filterable list of every Toll Plaza record — infinite scroll."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._search  = ""
        self._year    = 0
        self._month   = 0
        self._loaded  = 0
        self._total   = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search vehicle, plaza, receipt, cashier…")
        self._search_edit.setFixedWidth(280)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        tbl.addStretch()
        vl.addWidget(tb)

        self._totals = _TotalsBar([("zmw", "ZMW "), ("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_table(_TOLL_DETAIL_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            years = await svc.get_toll_plaza_available_years()
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_toll_plaza_all_totals(self._search, self._year, month, **self._date_kw()),
                svc.get_toll_plaza_all_records(
                    self._search, self._year, month,
                    limit=_TOLL_SCROLL_CHUNK, skip=0, **self._date_kw(),
                ),
                svc.count_toll_plaza_all_records(self._search, self._year, month, **self._date_kw()),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return

        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("zmw",   float(totals.get("total_zmw", 0)), "ZMW ")
        self._totals.set_total("count", int(totals.get("count", 0)), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_toll_plaza_all_records(
                self._search, self._year, month,
                limit=_TOLL_SCROLL_CHUNK, skip=self._loaded, **self._date_kw(),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return

        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _toll_fill_detail_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()


# ─────────────────────────────────────────────────────────────────────────────
#  Upload Browse sub-widget — one row per import batch
# ─────────────────────────────────────────────────────────────────────────────

class _TollUploadBrowse(QWidget):
    """Table of every Toll Plaza import batch. Clicking a row drills into it."""

    upload_clicked = Signal(object)   # emits the upload summary dict
    delete_clicked = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        # Totals strip
        self._totals = _TotalsBar([("zmw", "ZMW "), ("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_table(_TOLL_BROWSE_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 160)
        self._table.setColumnWidth(1, 240)
        self._table.setColumnWidth(2, 80)
        self._table.setColumnWidth(3, 120)
        self._table.setStyleSheet(_table_style())
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        hint = _lbl(
            "Click any row to view its records · right-click to delete an upload.",
            size=11, color=_TM,
        )
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        uploads = await svc.get_toll_plaza_uploads()
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        total_zmw = 0.0
        total_recs = 0
        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)
            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count = int(up.get("record_count", 0))
            zmw   = float(up.get("total_zmw", 0))
            min_d = str(up.get("min_toll_date", "") or "").strip()
            max_d = str(up.get("max_toll_date", "") or "").strip()
            date_range = f"{min_d[:10]} — {max_d[:10]}" if min_d else "—"

            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 2, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 3, _cell(f"{zmw:,.0f}", align=Qt.AlignRight | Qt.AlignVCenter))
            t.setItem(r, 4, _cell(date_range))
            _finish_table_row(t, r)
            total_zmw  += zmw
            total_recs += count

        self._totals.set_total("zmw",   total_zmw,  "ZMW ")
        self._totals.set_total("count", total_recs, "")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


# ─────────────────────────────────────────────────────────────────────────────
#  Upload Detail sub-widget — all records for one import batch
# ─────────────────────────────────────────────────────────────────────────────

class _TollUploadDetail(QWidget):
    """Full record table for a single Toll Plaza upload batch."""

    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._search    = ""
        self._loaded    = 0
        self._total     = 0
        self._loading   = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        # ── Breadcrumb / back row ──────────────────────────────────────────
        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)

        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)

        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()

        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        vl.addWidget(nav)

        # ── Info strip ────────────────────────────────────────────────────
        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        # ── Toolbar ───────────────────────────────────────────────────────
        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search vehicle, plaza, receipt, cashier…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        tbl.addStretch()
        vl.addWidget(tb)

        # ── Full record table ─────────────────────────────────────────────
        self._table = _make_table(_TOLL_DETAIL_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([("zmw", "ZMW "), ("count", "Records: ")])
        vl.addWidget(self._totals)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        filename   = upload_doc.get("source_filename") or "Unknown file"
        count      = int(upload_doc.get("record_count", 0))
        zmw        = float(upload_doc.get("total_zmw", 0))
        import_dt  = upload_doc.get("import_date")
        date_str   = (
            import_dt.strftime("%d %b %Y")
            if isinstance(import_dt, datetime) else ""
        )
        self._crumb_lbl.setText(f"Uploads  ›  {filename}")
        self._info_lbl.setText(
            f"{filename}   •   {count:,} records   •   ZMW {zmw:,.0f}   •   {date_str}"
        )
        self._totals.set_total("zmw",   zmw,   "ZMW ")
        self._totals.set_total("count", count, "")

        # Reset search + pagination for this upload
        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._reset_and_load()

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            recs, total = await asyncio.gather(
                svc.get_toll_plaza_upload_records(
                    self._upload_id, self._search, _SCROLL_CHUNK, 0,
                ),
                svc.count_toll_plaza_upload_records(self._upload_id, self._search),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("count", total, "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            recs = await svc.get_toll_plaza_upload_records(
                self._upload_id, self._search, _SCROLL_CHUNK, self._loaded,
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _toll_fill_detail_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()


# ─────────────────────────────────────────────────────────────────────────────
#  1. TollPlazaWidget — master/detail shell
# ─────────────────────────────────────────────────────────────────────────────

class TollPlazaWidget(QWidget):
    """
    Toll Plaza main page.

    All Entries tab: every record across uploads with month/year filters and
    infinite scroll.  Uploads tab: batch browse list; clicking a row drills
    into the per-record detail view for that upload.
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader("Toll Plaza", "mdi.boom-gate")
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        self._import_btn = _btn("Import from Dot Com Zambia", "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Uploads"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _TollAllEntries()
        self._main_stack.addWidget(self._all_entries)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _TollUploadBrowse()
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _TollUploadDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        self._tabs.tab_changed.connect(self._main_stack.setCurrentIndex)
        vl.addWidget(self._main_stack, 1)

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._show_browse()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(1, emit=False)
        self._main_stack.setCurrentIndex(1)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            "Toll_Plaza_Import_Template.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, _TOLL_TEMPLATE_TITLE, _TOLL_DETAIL_HEADERS)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        dlg = ImportDialog(
            feed_type="toll_plaza",
            dedup_key="receipt_no",
            preview_headers=_TOLL_DETAIL_HEADERS,
            col_map=_TOLL_COL_MAP,
            save_fn=svc.save_imported_feed,
            exist_fn=svc.get_existing_feed_keys,
            header_row=1,
            template_title=_TOLL_TEMPLATE_TITLE,
            template_headers=_TOLL_DETAIL_HEADERS,
            template_filename="Toll_Plaza_Import_Template.xlsx",
            parent=self,
        )
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        QMessageBox.information(self, "Import Complete", f"Imported {n:,} new records.")
        self._all_entries.refresh()
        self._show_browse()
        self._tabs.set_index(1)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        upload_id = str(upload_doc.get("_id") or "")
        if not upload_id:
            return
        filename = upload_doc.get("source_filename") or "Unknown file"
        count = int(upload_doc.get("record_count", 0))
        if QMessageBox.question(
            self,
            "Delete Upload",
            f"Delete \"{filename}\" and all {count:,} records in this upload?\n\n"
            "This cannot be undone.",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(upload_id))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            deleted = await svc.delete_toll_plaza_upload(upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        if deleted <= 0:
            QMessageBox.warning(self, "Delete Upload", "No records were deleted.")
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._show_browse()


# ═══════════════════════════════════════════════════════════════════════════════
#  2. ParkingCongoWidget
# ═══════════════════════════════════════════════════════════════════════════════

_PCONGO_COL_MAP = {
    "sn":                  ["#", "sn", "s/n", "no"],
    "ledger_id":           ["ledger id", "ledger_id", "id"],
    "payment_date":        ["payment date", "date", "payment_date"],
    "transaction_type":    ["type", "transaction type", "trans type"],
    "amount":              ["amount", "amt"],
    "running_bal":         ["running bal", "running balance", "bal", "balance"],
    "cashier":             ["cashier"],
    "vehicle_no":          ["vehicle #", "vehicle no", "vehicle", "plate"],
    "direction":           ["direction", "dir"],
    "gate_in":             ["gate in", "in"],
    "transaction_details": ["transaction details", "details", "trans details", "ticket"],
}

_PCONGO_DETAIL_HEADERS = [
    "#", "LEDGER ID", "PAYMENT DATE", "TYPE", "AMOUNT", "RUNNING BAL",
    "CASHIER", "VEHICLE #", "DIRECTION", "GATE IN", "TRANSACTION DETAILS",
]

_PCONGO_TEMPLATE_HEADERS = [
    "#", "LEDGER ID", "PAYMENT DATE", "TYPE", "AMOUNT", "RUNNING BAL",
    "CASHIER", "VEHICLE #", "DIRECTION", "GATE IN", "TICKET",
]

_PCONGO_BROWSE_HEADERS = [
    "UPLOAD DATE", "FILE NAME", "RECORDS", "DATE RANGE",
]

_PCONGO_SCROLL_CHUNK = 50


def _pcongo_amount_color(amt_str: str) -> str:
    """Return green for credits (positive), red for debits (negative)."""
    try:
        return _GREEN if float(amt_str) > 0 else _RED
    except (ValueError, TypeError):
        return _T1


def _pcongo_fill_detail_row(t: QTableWidget, r: int, rec: dict) -> None:
    """Populate one Parking Congo record row (shared by detail + all-entries views)."""
    amt_str = str(rec.get("amount", "") or "")
    amt_color = _pcongo_amount_color(amt_str)
    try:
        amt_display = _fmt_num(float(amt_str), "", 2)
    except (ValueError, TypeError):
        amt_display = amt_str or "—"

    t.setItem(r,  0, _cell(str(rec.get("sn", "") or "")))
    t.setItem(r,  1, _cell(str(rec.get("ledger_id", "") or "")))
    t.setItem(r,  2, _cell(str(rec.get("payment_date", "") or "")))
    t.setItem(r,  3, _cell(str(rec.get("transaction_type", "") or ""), color=amt_color))
    t.setItem(r,  4, _cell(
        amt_display, align=Qt.AlignRight | Qt.AlignVCenter, color=amt_color,
    ))
    t.setItem(r,  5, _cell(
        str(rec.get("running_bal", "") or ""),
        align=Qt.AlignRight | Qt.AlignVCenter,
    ))
    t.setItem(r,  6, _cell(str(rec.get("cashier", "") or "")))
    t.setItem(r,  7, _cell(str(rec.get("vehicle_no", "") or "")))
    t.setItem(r,  8, _cell(str(rec.get("direction", "") or "")))
    t.setItem(r,  9, _cell(str(rec.get("gate_in", "") or "")))
    t.setItem(r, 10, _cell(str(rec.get("transaction_details", "") or "")))
    _finish_table_row(t, r)


class _ParkingCongoAllEntries(QWidget):
    """Flat, filterable list of every Parking Congo record — infinite scroll."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._search  = ""
        self._year    = 0
        self._month   = 0
        self._loaded  = 0
        self._total   = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search vehicle, ledger ID, type, cashier…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        tbl.addStretch()
        vl.addWidget(tb)

        self._totals = _TotalsBar([("amount", "Total: "), ("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_table(_PCONGO_DETAIL_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            years = await svc.get_parking_congo_available_years()
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_parking_congo_all_totals(self._search, self._year, month, **self._date_kw()),
                svc.get_parking_congo_all_records(
                    self._search, self._year, month,
                    limit=_PCONGO_SCROLL_CHUNK, skip=0, **self._date_kw(),
                ),
                svc.count_parking_congo_all_records(self._search, self._year, month, **self._date_kw()),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return

        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("amount", float(totals.get("amount", 0)), "")
        self._totals.set_total("count",  int(totals.get("count", 0)), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_parking_congo_all_records(
                self._search, self._year, month,
                limit=_PCONGO_SCROLL_CHUNK, skip=self._loaded, **self._date_kw(),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return

        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _pcongo_fill_detail_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()


# ─────────────────────────────────────────────────────────────────────────────
#  Upload Browse sub-widget — one row per import batch
# ─────────────────────────────────────────────────────────────────────────────

class _ParkingCongoUploadBrowse(QWidget):
    """Table of every Parking Congo import batch. Clicking a row drills into it."""

    upload_clicked = Signal(object)
    delete_clicked = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        self._table = _make_table(_PCONGO_BROWSE_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 180)
        self._table.setColumnWidth(1, 260)
        self._table.setColumnWidth(2, 80)
        self._table.setStyleSheet(_table_style())
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        hint = _lbl(
            "Click any row to view its records · right-click to delete an upload.",
            size=11, color=_TM,
        )
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        uploads = await svc.get_parking_congo_uploads()
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)
            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count  = int(up.get("record_count", 0))
            min_d  = str(up.get("min_date", "") or "").strip()
            max_d  = str(up.get("max_date", "") or "").strip()
            date_range = f"{min_d[:10]} — {max_d[:10]}" if min_d else "—"

            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 2, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 3, _cell(date_range))
            _finish_table_row(t, r)

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


# ─────────────────────────────────────────────────────────────────────────────
#  Upload Detail sub-widget — all records for one import batch
# ─────────────────────────────────────────────────────────────────────────────

class _ParkingCongoUploadDetail(QWidget):
    """Full record table for a single Parking Congo upload batch."""

    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._search    = ""
        self._page      = 1
        self._page_size = 50
        self._total     = 0
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)

        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)

        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()

        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search vehicle, ledger ID, type, cashier…")
        self._search_edit.setFixedWidth(320)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        tbl.addStretch()
        vl.addWidget(tb)

        self._table = _make_table(_PCONGO_DETAIL_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([("count", "Records: ")])
        vl.addWidget(self._totals)

        self._pager = _PaginationBar()
        self._pager.page_changed.connect(self._go_page)
        self._pager.size_changed.connect(self._on_page_size)
        vl.addWidget(self._pager)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        filename  = upload_doc.get("source_filename") or "Unknown file"
        count     = int(upload_doc.get("record_count", 0))
        import_dt = upload_doc.get("import_date")
        date_str  = (
            import_dt.strftime("%d %b %Y")
            if isinstance(import_dt, datetime) else ""
        )
        self._crumb_lbl.setText(f"Uploads  ›  {filename}")
        self._info_lbl.setText(
            f"{filename}   •   {count:,} records   •   {date_str}"
        )
        self._totals.set_total("count", count, "")

        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._page = 1
        asyncio.ensure_future(self._load())

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id:
            return
        skip = (self._page - 1) * self._page_size
        recs, total = await asyncio.gather(
            svc.get_parking_congo_upload_records(
                self._upload_id, self._search, self._page_size, skip
            ),
            svc.count_parking_congo_upload_records(self._upload_id, self._search),
        )
        self._total = total
        self._fill_table(recs)
        self._pager.set_total(total, self._page_size, self._page)
        self._totals.set_total("count", total, "")

    def _fill_table(self, recs: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        for rec in recs:
            r = t.rowCount()
            t.insertRow(r)
            _pcongo_fill_detail_row(t, r, rec)

    def _on_search(self, text: str) -> None:
        self._search = text
        self._page   = 1
        asyncio.ensure_future(self._load())

    def _go_page(self, page: int) -> None:
        self._page = page
        asyncio.ensure_future(self._load())

    def _on_page_size(self, size: int) -> None:
        self._page_size = size
        self._page      = 1
        asyncio.ensure_future(self._load())


# ─────────────────────────────────────────────────────────────────────────────
#  2. ParkingCongoWidget — master/detail shell
# ─────────────────────────────────────────────────────────────────────────────

class ParkingCongoWidget(QWidget):
    """
    Parking Congo main page.

    All Entries tab: every record across uploads with month/year filters and
    infinite scroll.  Uploads tab: batch browse list; clicking a row drills
    into the per-record detail view for that upload.
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader("Parking Congo", "mdi.parking")
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        self._import_btn = _btn("Import from Congo Ledger", "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Uploads"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _ParkingCongoAllEntries()
        self._main_stack.addWidget(self._all_entries)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _ParkingCongoUploadBrowse()
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _ParkingCongoUploadDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        self._tabs.tab_changed.connect(self._main_stack.setCurrentIndex)
        vl.addWidget(self._main_stack, 1)

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._show_browse()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(1, emit=False)
        self._main_stack.setCurrentIndex(1)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            "Parking_Congo_Import_Template.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, "", _PCONGO_TEMPLATE_HEADERS)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        dlg = ImportDialog(
            feed_type="parking_congo",
            dedup_key="ledger_id",
            preview_headers=_PCONGO_DETAIL_HEADERS,
            col_map=_PCONGO_COL_MAP,
            save_fn=svc.save_imported_feed,
            exist_fn=svc.get_existing_feed_keys,
            auto_header_row=True,
            template_title="",
            template_headers=_PCONGO_TEMPLATE_HEADERS,
            template_filename="Parking_Congo_Import_Template.xlsx",
            parent=self,
        )
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        QMessageBox.information(self, "Import Complete", f"Imported {n:,} new records.")
        self._all_entries.refresh()
        self._show_browse()
        self._tabs.set_index(1)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        upload_id = str(upload_doc.get("_id") or "")
        if not upload_id:
            return
        filename = upload_doc.get("source_filename") or "Unknown file"
        count = int(upload_doc.get("record_count", 0))
        if QMessageBox.question(
            self,
            "Delete Upload",
            f"Delete \"{filename}\" and all {count:,} records in this upload?\n\n"
            "This cannot be undone.",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(upload_id))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            deleted = await svc.delete_parking_congo_upload(upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        if deleted <= 0:
            QMessageBox.warning(self, "Delete Upload", "No records were deleted.")
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._show_browse()


# ═══════════════════════════════════════════════════════════════════════════════
#  3. CongoExpensesWidget  — Excel import (last sheet), upload browse + detail
# ═══════════════════════════════════════════════════════════════════════════════

_CONGO_HEADERS = [
    "S/NO", "DATE", "LPO NO", "TRUCK NO", "DESCRIPTION", "AMOUNT (USD)",
]

_CONGO_BROWSE_HEADERS = [
    "UPLOAD DATE", "FILE NAME", "SHEET", "RECORDS", "BALANCE (USD)", "DATE RANGE",
]


def _parse_congo_sheet(ws) -> List[dict]:
    """Parse one Congo Expenses worksheet into row dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    records: List[dict] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(c is not None for c in row):
            continue

        s_no        = row[0] if len(row) > 0 else None
        dt          = row[1] if len(row) > 1 else None
        lpo         = row[2] if len(row) > 2 else None
        truck       = row[3] if len(row) > 3 else None
        description = row[4] if len(row) > 4 else None
        amount      = row[5] if len(row) > 5 else None

        desc_str = str(description).strip() if description is not None else ""
        if desc_str.upper() == "TOTAL":
            continue

        try:
            amt = float(amount) if amount is not None else 0.0
        except (TypeError, ValueError):
            amt = 0.0

        if isinstance(dt, datetime):
            date_str     = dt.strftime("%d %b %Y")
            expense_date = dt
        else:
            date_str     = str(dt).strip() if dt is not None else ""
            expense_date = None

        lpo_str   = str(lpo).strip() if lpo is not None else ""
        truck_str = str(truck).strip() if truck is not None else ""

        serial_no: Optional[int] = None
        if s_no is not None and str(s_no).strip():
            try:
                serial_no = int(float(s_no))
            except (TypeError, ValueError):
                serial_no = None

        records.append({
            "serial_no":    serial_no,
            "date_str":     date_str,
            "expense_date": expense_date,
            "lpo_no":       lpo_str,
            "truck_no":     truck_str,
            "description":  desc_str,
            "amount_usd":   amt,
            "is_advance":   serial_no is None,
            "row_index":    row_idx,
        })

    return records


def _parse_congo_last_sheet(path: str) -> Tuple[str, List[dict]]:
    """Read only the last worksheet from a Congo Expenses workbook."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Congo Expenses Excel files.")

    wb = openpyxl.load_workbook(path, data_only=True)
    if not wb.sheetnames:
        return "", []

    sheet_name = wb.sheetnames[-1]
    return sheet_name, _parse_congo_sheet(wb[sheet_name])


def _parse_congo_all_sheets(path: str) -> List[dict]:
    """Parse every worksheet in a Congo Expenses workbook."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Congo Expenses Excel files.")

    wb = openpyxl.load_workbook(path, data_only=True)
    batches: List[dict] = []
    for sheet_name in wb.sheetnames:
        records = _parse_congo_sheet(wb[sheet_name])
        if records:
            batches.append({"sheet_label": sheet_name, "records": records})
    return batches


def _congo_fill_row(t: QTableWidget, r: int, rec: dict) -> None:
    """Populate one Congo Expenses row — same signed-amount styling as Kimvi."""
    serial = rec.get("serial_no")
    amt    = rec.get("amount_usd", 0)
    try:
        amt_f = float(amt) if amt is not None else 0.0
    except (TypeError, ValueError):
        amt_f = 0.0

    is_in   = amt_f < 0
    row_bg  = _CREDIT_ROW_BG if is_in else _stripe_bg(r)
    amt_clr = _CREDIT_FG if is_in else _T1

    t.setItem(r, 0, _cell("—" if serial is None else str(serial)))
    t.setItem(r, 1, _cell(rec.get("date_str", "")))
    t.setItem(r, 2, _cell(rec.get("lpo_no", "")))
    t.setItem(r, 3, _cell(rec.get("truck_no", "")))
    t.setItem(r, 4, _cell(rec.get("description", "")))
    t.setItem(r, 5, _cell(
        _kimvi_fmt_amount(amt_f),
        color=amt_clr,
        align=Qt.AlignRight | Qt.AlignVCenter,
    ))
    _apply_row_bg(t, r, row_bg)
    t.setRowHeight(r, _ROW_H)


def _make_congo_summary_cards() -> Tuple[QFrame, QLabel, QLabel, QLabel]:
    """Money In / Money Out / Balance summary strip for Congo views."""
    frame = QFrame()
    frame.setStyleSheet(
        f"QFrame{{background:{_BLUE_L};border:1px solid #BFDBFE;"
        "border-radius:6px;padding:8px;}}"
    )
    sl = QHBoxLayout(frame)
    sl.setContentsMargins(12, 8, 12, 8)
    sl.setSpacing(24)
    in_lbl  = _lbl("Money In: —", size=12, weight=600, color=_GREEN)
    out_lbl = _lbl("Money Out: —", size=12, color=_T2)
    bal_lbl = _lbl("Balance: —", size=12, weight=600, color=_T1)
    sl.addWidget(in_lbl)
    sl.addWidget(out_lbl)
    sl.addWidget(bal_lbl)
    sl.addStretch()
    return frame, in_lbl, out_lbl, bal_lbl


def _set_congo_summary(in_lbl: QLabel, out_lbl: QLabel, bal_lbl: QLabel,
                       money_in: float, money_out: float) -> None:
    balance = money_in + money_out
    in_lbl.setText(f"Money In: USD {_kimvi_fmt_amount(money_in)}")
    out_lbl.setText(f"Money Out: USD {_kimvi_fmt_amount(money_out)}")
    bal_color = _CREDIT_FG if balance < 0 else (_RED if balance > 0 else _T1)
    bal_lbl.setText(f"Balance: USD {_kimvi_fmt_amount(balance)}")
    bal_lbl.setStyleSheet(
        f"color:{bal_color};font-size:12px;font-weight:600;"
        "font-family:'Segoe UI';background:transparent;"
    )


class _CongoEntriesBase(QWidget):
    """Shared flat-list view for Congo All Entries / Money In tabs."""

    def __init__(self, money_in_only: bool = False, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._money_in_only = money_in_only
        self._search  = ""
        self._year    = 0
        self._month   = 0
        self._loaded  = 0
        self._total   = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search LPO, truck, description, date…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        tbl.addStretch()
        vl.addWidget(tb)

        self._summary_frame, self._in_lbl, self._out_lbl, self._bal_lbl = (
            _make_congo_summary_cards()
        )
        vl.addWidget(self._summary_frame)

        self._totals = _TotalsBar([("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_kimvi_table(_CONGO_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            years = await svc.get_congo_available_years()
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_congo_all_totals(
                    self._search, self._year, month,
                    money_in_only=self._money_in_only, **self._date_kw(),
                ),
                svc.get_congo_all_records(
                    self._search, self._year, month,
                    money_in_only=self._money_in_only,
                    limit=_SCROLL_CHUNK, skip=0, **self._date_kw(),
                ),
                svc.count_congo_all_records(
                    self._search, self._year, month,
                    money_in_only=self._money_in_only, **self._date_kw(),
                ),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return

        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        _set_congo_summary(
            self._in_lbl, self._out_lbl, self._bal_lbl,
            float(totals.get("money_in", 0)),
            float(totals.get("money_out", 0)),
        )
        self._totals.set_total("count", int(totals.get("count", 0)), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_congo_all_records(
                self._search, self._year, month,
                money_in_only=self._money_in_only,
                limit=_SCROLL_CHUNK, skip=self._loaded, **self._date_kw(),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _congo_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()


class _CongoAllEntries(_CongoEntriesBase):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(money_in_only=False, parent=parent)


class _CongoMoneyInEntries(_CongoEntriesBase):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(money_in_only=True, parent=parent)


class CongoImportDialog(QDialog):
    """Import the last worksheet from a Congo Expenses workbook."""

    imported = Signal(int)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id       = str(uuid.uuid4())
        self._source_filename = ""
        self._sheet_label     = ""
        self._records: List[dict] = []
        self._already_exists  = False

        self.setWindowTitle("Import — Congo Expenses")
        self.setMinimumWidth(720)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        hint = _lbl(
            "Only the last sheet in the workbook will be imported "
            "(one LPO period per upload).",
            size=12, color=_T2,
        )
        vl.addWidget(hint)

        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        browse_row = QWidget()
        browse_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(browse_row)
        brl.setContentsMargins(0, 0, 0, 0)
        browse_btn = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse_btn.clicked.connect(self._browse)
        brl.addStretch()
        brl.addWidget(browse_btn)
        vl.addWidget(browse_row)

        self._stats_lbl = _lbl("No file loaded.", size=12, color=_T2)
        vl.addWidget(self._stats_lbl)
        vl.addWidget(_hsep())

        preview_title = _lbl("Preview (first 10 rows)", size=12, weight=600)
        vl.addWidget(preview_title)

        self._preview_tbl = _make_kimvi_table(_CONGO_HEADERS)
        self._preview_tbl.setMinimumHeight(200)
        vl.addWidget(self._preview_tbl)
        vl.addWidget(_hsep())

        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.addStretch()

        cancel_btn = _btn("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        bbl.addWidget(cancel_btn)

        self._import_btn = _btn("Import Sheet", "mdi.check-circle-outline")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)
        vl.addWidget(btn_row)

    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Congo Expenses Workbook", "", "Excel (*.xlsx *.xls)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _on_file(self, path: str) -> None:
        self._stats_lbl.setText("Reading file…")
        self._source_filename = Path(path).name
        try:
            sheet_label, records = _parse_congo_last_sheet(path)
        except Exception as exc:
            self._stats_lbl.setText(f"Error reading file: {exc}")
            self._import_btn.setEnabled(False)
            return

        self._sheet_label = sheet_label
        self._records = records
        asyncio.ensure_future(self._check_sheet(sheet_label, records))

    async def _check_sheet(self, sheet_label: str, records: List[dict]) -> None:
        from tahmeed.services import accountant_service as svc

        try:
            exists = await svc.congo_sheet_exists(sheet_label)
        except Exception:
            exists = False

        self._already_exists = exists
        money_in  = sum(r["amount_usd"] for r in records if r["amount_usd"] < 0)
        money_out = sum(r["amount_usd"] for r in records if r["amount_usd"] > 0)
        balance   = money_in + money_out

        if exists:
            self._stats_lbl.setText(
                f"Sheet \"{sheet_label}\" was already uploaded — import blocked."
            )
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Already Uploaded")
        elif not records:
            self._stats_lbl.setText(
                f"Last sheet \"{sheet_label}\" has no data rows to import."
            )
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Import Sheet")
        else:
            self._stats_lbl.setText(
                f"Last sheet: {sheet_label}     "
                f"Rows: {len(records):,}     "
                f"In: USD {money_in:,.0f}     "
                f"Out: USD {money_out:,.0f}     "
                f"Balance: USD {balance:,.0f}"
            )
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(records):,} Rows")

        t = self._preview_tbl
        t.setRowCount(0)
        for rec in records[:10]:
            row = t.rowCount()
            t.insertRow(row)
            _congo_fill_row(t, row, rec)

    def _do_import(self) -> None:
        if self._already_exists or not self._records:
            return
        self._import_btn.setEnabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        from tahmeed.services import accountant_service as svc

        docs = []
        for rec in self._records:
            doc = dict(rec)
            doc["sheet_label"]     = self._sheet_label
            doc["upload_id"]       = self._upload_id
            doc["source_filename"] = self._source_filename
            docs.append(doc)

        try:
            saved = await svc.save_congo_import(docs)
            self.imported.emit(saved)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(self._records):,} Rows")


class _CongoExpUploadBrowse(QWidget):
    upload_clicked = Signal(object)
    delete_clicked = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        self._totals = _TotalsBar([
            ("balance", "Balance USD "),
            ("count",   "Total records: "),
        ])
        vl.addWidget(self._totals)

        self._table = _make_kimvi_table(_CONGO_BROWSE_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 160)
        self._table.setColumnWidth(1, 220)
        self._table.setColumnWidth(2, 80)
        self._table.setColumnWidth(3, 80)
        self._table.setColumnWidth(4, 120)
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        hint = _lbl(
            "Click any row to view its records · right-click to delete an upload.",
            size=11, color=_TM,
        )
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        uploads = await svc.get_congo_uploads()
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        total_balance = 0.0
        total_recs    = 0

        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)

            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count   = int(up.get("record_count", 0))
            balance = float(up.get("balance_usd", 0))

            min_d = up.get("min_expense_date")
            max_d = up.get("max_expense_date")
            if isinstance(min_d, datetime) and isinstance(max_d, datetime):
                date_range = f"{min_d.strftime('%d %b %Y')} — {max_d.strftime('%d %b %Y')}"
            else:
                date_range = "—"

            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 2, _cell(up.get("sheet_label") or "—", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 3, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            bal_color = _KIMVI_IN_FG if balance < 0 else (_RED if balance > 0 else "")
            t.setItem(r, 4, _cell(
                _kimvi_fmt_amount(balance),
                color=bal_color,
                align=Qt.AlignRight | Qt.AlignVCenter,
            ))
            t.setItem(r, 5, _cell(date_range))
            _finish_table_row(t, r)

            total_balance += balance
            total_recs    += count

        self._totals.set_total("balance", total_balance, "Balance USD ")
        self._totals.set_total("count",   total_recs,    "")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


class _CongoExpUploadDetail(QWidget):
    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._search    = ""
        self._loaded    = 0
        self._total     = 0
        self._loading   = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)

        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)

        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()

        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        self._summary_frame, self._in_lbl, self._out_lbl, self._bal_lbl = (
            _make_congo_summary_cards()
        )
        vl.addWidget(self._summary_frame)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search LPO, truck, description, date…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        tbl.addStretch()
        vl.addWidget(tb)

        self._table = _make_kimvi_table(_CONGO_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([("count", "Records: ")])
        vl.addWidget(self._totals)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        filename    = upload_doc.get("source_filename") or "Unknown file"
        sheet_label = upload_doc.get("sheet_label") or "—"
        count       = int(upload_doc.get("record_count", 0))
        money_in    = float(upload_doc.get("money_in", 0))
        money_out   = float(upload_doc.get("money_out", 0))
        balance     = float(upload_doc.get("balance_usd", 0))
        import_dt   = upload_doc.get("import_date")
        date_str    = (
            import_dt.strftime("%d %b %Y")
            if isinstance(import_dt, datetime) else ""
        )

        self._crumb_lbl.setText(f"Uploads  ›  {filename}  ›  Sheet {sheet_label}")
        self._info_lbl.setText(
            f"Sheet {sheet_label}   •   {count:,} rows   •   {date_str}"
        )
        _set_congo_summary(self._in_lbl, self._out_lbl, self._bal_lbl, money_in, money_out)
        self._totals.set_total("count", count, "")

        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._reset_and_load()

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            recs, total = await asyncio.gather(
                svc.get_congo_upload_records(
                    self._upload_id, self._search, _SCROLL_CHUNK, 0,
                ),
                svc.count_congo_upload_records(self._upload_id, self._search),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("count", total, "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            recs = await svc.get_congo_upload_records(
                self._upload_id, self._search, _SCROLL_CHUNK, self._loaded,
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._totals.set_total("count", self._total, "")
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _congo_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()


class CongoExpensesWidget(QWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader("Congo Expenses", "mdi.map-marker")
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        bulk_btn = _btn("Bulk Import All Sheets", "mdi.file-multiple-outline", primary=False)
        bulk_btn.clicked.connect(self._open_bulk_import)
        self._import_btn = _btn("Import Latest Sheet", "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(bulk_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Money In", "Uploads"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _CongoAllEntries()
        self._main_stack.addWidget(self._all_entries)

        self._money_in = _CongoMoneyInEntries()
        self._main_stack.addWidget(self._money_in)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _CongoExpUploadBrowse()
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _CongoExpUploadDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        self._tabs.tab_changed.connect(self._main_stack.setCurrentIndex)
        vl.addWidget(self._main_stack, 1)

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(2, emit=False)
        self._main_stack.setCurrentIndex(2)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            "Congo_Expenses_Import_Template.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, "", _CONGO_HEADERS)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        dlg = CongoImportDialog(parent=self)
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        QMessageBox.information(self, "Import Complete", f"Imported {n:,} rows.")
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()
        self._tabs.set_index(2)

    def _open_bulk_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        dlg = BulkSheetImportDialog(
            title="Bulk Import — Congo Expenses",
            hint=(
                "Every sheet in the workbook will be checked. "
                "New sheets are imported; sheets already uploaded are skipped."
            ),
            parse_all_fn=_parse_congo_all_sheets,
            existing_fn=svc.congo_existing_sheet_labels,
            save_fn=svc.save_congo_import,
            parent=self,
        )
        dlg.imported.connect(self._on_bulk_imported)
        dlg.exec()

    def _on_bulk_imported(self, sheets: int, rows: int) -> None:
        QMessageBox.information(
            self, "Bulk Import Complete",
            f"Imported {sheets:,} sheet{'s' if sheets != 1 else ''} ({rows:,} rows).",
        )
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()
        self._tabs.set_index(2)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        upload_id = str(upload_doc.get("_id") or "")
        if not upload_id:
            return
        filename = upload_doc.get("source_filename") or "Unknown file"
        sheet_label = upload_doc.get("sheet_label") or "—"
        count = int(upload_doc.get("record_count", 0))
        if QMessageBox.question(
            self,
            "Delete Upload",
            f"Delete sheet \"{sheet_label}\" from \"{filename}\" "
            f"and all {count:,} records?\n\nThis cannot be undone.",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(upload_id))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            deleted = await svc.delete_congo_upload(upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        if deleted <= 0:
            QMessageBox.warning(self, "Delete Upload", "No records were deleted.")
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()


# ═══════════════════════════════════════════════════════════════════════════════
#  4. AhmedKimviWidget  — Excel import (last sheet), upload browse + detail
# ═══════════════════════════════════════════════════════════════════════════════

_KIMVI_HEADERS = ["S/NO", "DATE", "TRUCK NO", "PARTICULARS", "AMOUNT (USD)"]

_KIMVI_BROWSE_HEADERS = [
    "UPLOAD DATE", "FILE NAME", "SHEET", "RECORDS", "BALANCE (USD)", "DATE RANGE",
]

_KIMVI_IN_FG = _CREDIT_FG
_make_kimvi_table = _make_table


def _parse_kimvi_sheet(ws) -> List[dict]:
    """Parse one Ahmed Kimvi worksheet into row dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    records: List[dict] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not any(c is not None for c in row):
            continue

        s_no        = row[0] if len(row) > 0 else None
        dt          = row[1] if len(row) > 1 else None
        truck       = row[2] if len(row) > 2 else None
        particulars = row[3] if len(row) > 3 else None
        amount      = row[4] if len(row) > 4 else None

        particulars_str = str(particulars).strip() if particulars is not None else ""
        if particulars_str.upper() == "TOTAL":
            continue

        try:
            amt = float(amount) if amount is not None else 0.0
        except (TypeError, ValueError):
            amt = 0.0

        if isinstance(dt, datetime):
            date_str     = dt.strftime("%d %b %Y")
            expense_date = dt
        else:
            date_str     = str(dt).strip() if dt is not None else ""
            expense_date = None

        truck_str = str(truck).strip() if truck is not None else ""

        serial_no: Optional[int] = None
        if s_no is not None and str(s_no).strip():
            try:
                serial_no = int(float(s_no))
            except (TypeError, ValueError):
                serial_no = None

        records.append({
            "serial_no":    serial_no,
            "date_str":     date_str,
            "expense_date": expense_date,
            "truck_no":     truck_str,
            "description":  particulars_str,
            "amount_usd":   amt,
            "is_advance":   serial_no is None,
            "row_index":    row_idx,
        })

    return records


def _parse_kimvi_last_sheet(path: str) -> Tuple[str, List[dict]]:
    """Read only the last worksheet from an Ahmed Kimvi expenses workbook."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Ahmed Kimvi Excel files.")

    wb = openpyxl.load_workbook(path, data_only=True)
    if not wb.sheetnames:
        return "", []

    sheet_name = wb.sheetnames[-1]
    return sheet_name, _parse_kimvi_sheet(wb[sheet_name])


def _parse_kimvi_all_sheets(path: str) -> List[dict]:
    """Parse every worksheet in an Ahmed Kimvi expenses workbook."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Ahmed Kimvi Excel files.")

    wb = openpyxl.load_workbook(path, data_only=True)
    batches: List[dict] = []
    for sheet_name in wb.sheetnames:
        records = _parse_kimvi_sheet(wb[sheet_name])
        if records:
            batches.append({"sheet_label": sheet_name, "records": records})
    return batches


_BULK_SHEET_HEADERS = ["SHEET", "RECORDS", "BALANCE (USD)", "STATUS"]


class BulkSheetImportDialog(QDialog):
    """Import every new worksheet from a multi-sheet expenses workbook."""

    imported = Signal(int, int)   # (sheets_imported, total_rows)

    def __init__(
        self,
        title: str,
        hint: str,
        parse_all_fn,
        existing_fn,
        save_fn,
        balance_fn=None,
        balance_header: str = "BALANCE (USD)",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._title          = title
        self._parse_all_fn   = parse_all_fn
        self._existing_fn    = existing_fn
        self._save_fn        = save_fn
        self._balance_fn     = balance_fn or (
            lambda records: sum(r.get("amount_usd", 0) for r in records)
        )
        self._balance_header = balance_header
        self._source_filename = ""
        self._pending: List[dict] = []
        self._skipped_count = 0

        self.setWindowTitle(title)
        self.setMinimumWidth(720)
        self.setMinimumHeight(480)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()
        self._hint_lbl.setText(hint)

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        self._hint_lbl = _lbl("", size=12, color=_T2)
        vl.addWidget(self._hint_lbl)

        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        browse_row = QWidget()
        browse_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(browse_row)
        brl.setContentsMargins(0, 0, 0, 0)
        browse_btn = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse_btn.clicked.connect(self._browse)
        brl.addStretch()
        brl.addWidget(browse_btn)
        vl.addWidget(browse_row)

        self._stats_lbl = _lbl("No file loaded.", size=12, color=_T2)
        vl.addWidget(self._stats_lbl)
        vl.addWidget(_hsep())

        preview_title = _lbl("Sheets in workbook", size=12, weight=600)
        vl.addWidget(preview_title)

        bulk_headers = ["SHEET", "RECORDS", self._balance_header, "STATUS"]
        self._preview_tbl = _make_kimvi_table(bulk_headers)
        self._preview_tbl.setMinimumHeight(260)
        vl.addWidget(self._preview_tbl, 1)
        vl.addWidget(_hsep())

        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.addStretch()

        cancel_btn = _btn("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        bbl.addWidget(cancel_btn)

        self._import_btn = _btn("Import All New Sheets", "mdi.check-all")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)
        vl.addWidget(btn_row)

    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, self._title, "", "Excel (*.xlsx *.xls)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _on_file(self, path: str) -> None:
        self._stats_lbl.setText("Reading file…")
        self._source_filename = Path(path).name
        try:
            batches = self._parse_all_fn(path)
        except Exception as exc:
            self._stats_lbl.setText(f"Error reading file: {exc}")
            self._import_btn.setEnabled(False)
            self._pending = []
            self._preview_tbl.setRowCount(0)
            return

        asyncio.ensure_future(self._check_batches(batches))

    async def _check_batches(self, batches: List[dict]) -> None:
        labels = [b["sheet_label"] for b in batches]
        try:
            existing: set = await self._existing_fn(labels)
        except Exception:
            existing = set()

        self._pending = []
        self._skipped_count = 0
        preview_rows: List[dict] = []

        for batch in batches:
            label   = batch["sheet_label"]
            records = batch["records"]
            balance = self._balance_fn(records)
            if label in existing:
                status = "Already uploaded"
                self._skipped_count += 1
            else:
                status = "New"
                self._pending.append({
                    "sheet_label": label,
                    "records":     records,
                    "upload_id":   str(uuid.uuid4()),
                })
            preview_rows.append({
                "sheet_label": label,
                "count":       len(records),
                "balance":     balance,
                "status":      status,
            })

        new_sheets = len(self._pending)
        new_rows   = sum(len(b["records"]) for b in self._pending)

        if not batches:
            self._stats_lbl.setText("Workbook has no sheets with data to import.")
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Import All New Sheets")
        elif new_sheets == 0:
            self._stats_lbl.setText(
                f"All {len(batches):,} sheets were already uploaded — nothing new to import."
            )
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Nothing to Import")
        else:
            self._stats_lbl.setText(
                f"Sheets in file: {len(batches):,}     "
                f"New: {new_sheets:,} ({new_rows:,} rows)     "
                f"Skipped: {self._skipped_count:,}"
            )
            self._import_btn.setEnabled(True)
            self._import_btn.setText(
                f"Import {new_sheets:,} Sheet{'s' if new_sheets != 1 else ''} "
                f"({new_rows:,} rows)"
            )

        self._fill_preview(preview_rows)

    def _fill_preview(self, rows: List[dict]) -> None:
        t = self._preview_tbl
        t.setRowCount(0)
        for i, row in enumerate(rows):
            r = t.rowCount()
            t.insertRow(r)
            status = row["status"]
            is_skip = status == "Already uploaded"
            status_color = _TM if is_skip else _GREEN

            t.setItem(r, 0, _cell(row["sheet_label"]))
            t.setItem(r, 1, _cell(
                f"{row['count']:,}", align=Qt.AlignCenter | Qt.AlignVCenter,
            ))
            balance = row["balance"]
            bal_color = _KIMVI_IN_FG if balance < 0 else (_RED if balance > 0 else "")
            t.setItem(r, 2, _cell(
                _kimvi_fmt_amount(balance),
                mono=True,
                color=bal_color,
                align=Qt.AlignRight | Qt.AlignVCenter,
            ))
            t.setItem(r, 3, _cell(status, color=status_color))
            _finish_table_row(t, r, _stripe_bg(i))

    def _do_import(self) -> None:
        if not self._pending:
            return
        self._import_btn.setEnabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        sheets_imported = 0
        total_rows      = 0

        try:
            for batch in self._pending:
                docs = []
                for rec in batch["records"]:
                    doc = dict(rec)
                    doc["sheet_label"]     = batch["sheet_label"]
                    doc["upload_id"]       = batch["upload_id"]
                    doc["source_filename"] = self._source_filename
                    docs.append(doc)
                saved = await self._save_fn(docs)
                if saved:
                    sheets_imported += 1
                    total_rows      += saved

            self.imported.emit(sheets_imported, total_rows)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Bulk Import Error", str(exc))
            new_rows = sum(len(b["records"]) for b in self._pending)
            self._import_btn.setEnabled(True)
            self._import_btn.setText(
                f"Import {len(self._pending):,} Sheets ({new_rows:,} rows)"
            )


class KimviImportDialog(QDialog):
    """Import the last worksheet from an Ahmed Kimvi expenses workbook."""

    imported = Signal(int)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id       = str(uuid.uuid4())
        self._source_filename = ""
        self._sheet_label     = ""
        self._records: List[dict] = []
        self._already_exists  = False

        self.setWindowTitle("Import — Ahmed Kimvi (Klesa)")
        self.setMinimumWidth(680)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        hint = _lbl(
            "Only the last sheet in the workbook will be imported "
            "(one visit period per upload).",
            size=12, color=_T2,
        )
        vl.addWidget(hint)

        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        browse_row = QWidget()
        browse_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(browse_row)
        brl.setContentsMargins(0, 0, 0, 0)
        browse_btn = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse_btn.clicked.connect(self._browse)
        brl.addStretch()
        brl.addWidget(browse_btn)
        vl.addWidget(browse_row)

        self._stats_lbl = _lbl("No file loaded.", size=12, color=_T2)
        vl.addWidget(self._stats_lbl)

        vl.addWidget(_hsep())

        preview_title = _lbl("Preview (first 10 rows)", size=12, weight=600)
        vl.addWidget(preview_title)

        self._preview_tbl = _make_kimvi_table(_KIMVI_HEADERS)
        self._preview_tbl.setMinimumHeight(200)
        vl.addWidget(self._preview_tbl)

        vl.addWidget(_hsep())

        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.addStretch()

        cancel_btn = _btn("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        bbl.addWidget(cancel_btn)

        self._import_btn = _btn("Import Sheet", "mdi.check-circle-outline")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)

        vl.addWidget(btn_row)

    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Ahmed Kimvi Workbook", "", "Excel (*.xlsx *.xls)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _on_file(self, path: str) -> None:
        self._stats_lbl.setText("Reading file…")
        self._source_filename = Path(path).name
        try:
            sheet_label, records = _parse_kimvi_last_sheet(path)
        except Exception as exc:
            self._stats_lbl.setText(f"Error reading file: {exc}")
            self._import_btn.setEnabled(False)
            return

        self._sheet_label = sheet_label
        self._records = records
        asyncio.ensure_future(self._check_sheet(sheet_label, records))

    async def _check_sheet(self, sheet_label: str, records: List[dict]) -> None:
        from tahmeed.services import accountant_service as svc

        try:
            exists = await svc.kimvi_sheet_exists(sheet_label)
        except Exception:
            exists = False

        self._already_exists = exists
        money_in  = sum(r["amount_usd"] for r in records if r["amount_usd"] < 0)
        money_out = sum(r["amount_usd"] for r in records if r["amount_usd"] > 0)
        balance   = money_in + money_out

        if exists:
            self._stats_lbl.setText(
                f"Sheet \"{sheet_label}\" was already uploaded — import blocked."
            )
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Already Uploaded")
        elif not records:
            self._stats_lbl.setText(
                f"Last sheet \"{sheet_label}\" has no data rows to import."
            )
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Import Sheet")
        else:
            self._stats_lbl.setText(
                f"Last sheet: {sheet_label}     "
                f"Rows: {len(records):,}     "
                f"In: USD {money_in:,.0f}     "
                f"Out: USD {money_out:,.0f}     "
                f"Balance: USD {balance:,.0f}"
            )
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(records):,} Rows")

        self._fill_preview(records[:10])

    def _fill_preview(self, rows: List[dict]) -> None:
        t = self._preview_tbl
        t.setRowCount(0)
        for rec in rows:
            r = t.rowCount()
            t.insertRow(r)
            _kimvi_fill_row(t, r, rec)

    def _do_import(self) -> None:
        if self._already_exists or not self._records:
            return
        self._import_btn.setEnabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        from tahmeed.services import accountant_service as svc

        docs = []
        for rec in self._records:
            doc = dict(rec)
            doc["sheet_label"]     = self._sheet_label
            doc["upload_id"]       = self._upload_id
            doc["source_filename"] = self._source_filename
            docs.append(doc)

        try:
            saved = await svc.save_kimvi_import(docs)
            self.imported.emit(saved)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(self._records):,} Rows")


def _kimvi_fill_row(t: QTableWidget, r: int, rec: dict) -> None:
    """Populate one Ahmed Kimvi row — green tint for money-in, slate stripe otherwise."""
    serial = rec.get("serial_no")
    amt    = rec.get("amount_usd", 0)
    try:
        amt_f = float(amt) if amt is not None else 0.0
    except (TypeError, ValueError):
        amt_f = 0.0

    is_in   = amt_f < 0
    row_bg  = _CREDIT_ROW_BG if is_in else _stripe_bg(r)
    amt_clr = _CREDIT_FG if is_in else _T1

    t.setItem(r, 0, _cell("—" if serial is None else str(serial)))
    t.setItem(r, 1, _cell(rec.get("date_str", "")))
    t.setItem(r, 2, _cell(rec.get("truck_no", "")))
    t.setItem(r, 3, _cell(rec.get("description", "")))
    t.setItem(r, 4, _cell(
        _kimvi_fmt_amount(amt_f),
        color=amt_clr,
        align=Qt.AlignRight | Qt.AlignVCenter,
    ))
    _apply_row_bg(t, r, row_bg)
    t.setRowHeight(r, _ROW_H)


class _KimviEntriesBase(QWidget):
    """Shared flat-list view for Ahmed Kimvi All Entries / Money In tabs."""

    def __init__(self, money_in_only: bool = False, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._money_in_only = money_in_only
        self._search  = ""
        self._year    = 0
        self._month   = 0
        self._loaded  = 0
        self._total   = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search truck, particulars, date…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        tbl.addStretch()
        vl.addWidget(tb)

        self._summary_frame, self._in_lbl, self._out_lbl, self._bal_lbl = (
            _make_congo_summary_cards()
        )
        vl.addWidget(self._summary_frame)

        self._totals = _TotalsBar([("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_kimvi_table(_KIMVI_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            years = await svc.get_kimvi_available_years()
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_kimvi_all_totals(
                    self._search, self._year, month,
                    money_in_only=self._money_in_only, **self._date_kw(),
                ),
                svc.get_kimvi_all_records(
                    self._search, self._year, month,
                    money_in_only=self._money_in_only,
                    limit=_SCROLL_CHUNK, skip=0, **self._date_kw(),
                ),
                svc.count_kimvi_all_records(
                    self._search, self._year, month,
                    money_in_only=self._money_in_only, **self._date_kw(),
                ),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return

        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        _set_congo_summary(
            self._in_lbl, self._out_lbl, self._bal_lbl,
            float(totals.get("money_in", 0)),
            float(totals.get("money_out", 0)),
        )
        self._totals.set_total("count", int(totals.get("count", 0)), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_kimvi_all_records(
                self._search, self._year, month,
                money_in_only=self._money_in_only,
                limit=_SCROLL_CHUNK, skip=self._loaded, **self._date_kw(),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _kimvi_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()


class _KimviAllEntries(_KimviEntriesBase):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(money_in_only=False, parent=parent)


class _KimviMoneyInEntries(_KimviEntriesBase):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(money_in_only=True, parent=parent)


class _KimviUploadBrowse(QWidget):
    """Table of every Ahmed Kimvi import batch."""

    upload_clicked = Signal(object)
    delete_clicked = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        self._totals = _TotalsBar([
            ("balance", "Balance USD "),
            ("count",   "Total records: "),
        ])
        vl.addWidget(self._totals)

        self._table = _make_kimvi_table(_KIMVI_BROWSE_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 160)
        self._table.setColumnWidth(1, 220)
        self._table.setColumnWidth(2, 80)
        self._table.setColumnWidth(3, 80)
        self._table.setColumnWidth(4, 120)
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        hint = _lbl(
            "Click any row to view its records · right-click to delete an upload.",
            size=11, color=_TM,
        )
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        uploads = await svc.get_kimvi_uploads()
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        total_balance = 0.0
        total_recs    = 0

        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)

            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count   = int(up.get("record_count", 0))
            balance = float(up.get("balance_usd", 0))

            min_d = up.get("min_expense_date")
            max_d = up.get("max_expense_date")
            if isinstance(min_d, datetime) and isinstance(max_d, datetime):
                date_range = f"{min_d.strftime('%d %b %Y')} — {max_d.strftime('%d %b %Y')}"
            else:
                date_range = "—"

            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 2, _cell(up.get("sheet_label") or "—", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 3, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            bal_color = _KIMVI_IN_FG if balance < 0 else (_RED if balance > 0 else "")
            t.setItem(r, 4, _cell(
                _kimvi_fmt_amount(balance),
                color=bal_color,
                align=Qt.AlignRight | Qt.AlignVCenter,
            ))
            t.setItem(r, 5, _cell(date_range))
            _finish_table_row(t, r)

            total_balance += balance
            total_recs    += count

        self._totals.set_total("balance", total_balance, "Balance USD ")
        self._totals.set_total("count",   total_recs,    "")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


class _KimviUploadDetail(QWidget):
    """Full record table for a single Ahmed Kimvi import batch."""

    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._search    = ""
        self._loaded    = 0
        self._total     = 0
        self._loading   = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)

        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)

        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()

        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        self._summary_frame, self._in_lbl, self._out_lbl, self._bal_lbl = (
            _make_congo_summary_cards()
        )
        vl.addWidget(self._summary_frame)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search truck, particulars, date…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        tbl.addStretch()
        vl.addWidget(tb)

        self._table = _make_kimvi_table(_KIMVI_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([("count", "Records: ")])
        vl.addWidget(self._totals)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        filename    = upload_doc.get("source_filename") or "Unknown file"
        sheet_label = upload_doc.get("sheet_label") or "—"
        count       = int(upload_doc.get("record_count", 0))
        money_in    = float(upload_doc.get("money_in", 0))
        money_out   = float(upload_doc.get("money_out", 0))
        import_dt   = upload_doc.get("import_date")
        date_str    = (
            import_dt.strftime("%d %b %Y")
            if isinstance(import_dt, datetime) else ""
        )

        self._crumb_lbl.setText(f"Uploads  ›  {filename}  ›  Sheet {sheet_label}")
        self._info_lbl.setText(
            f"Sheet {sheet_label}   •   {count:,} rows   •   {date_str}"
        )
        _set_congo_summary(self._in_lbl, self._out_lbl, self._bal_lbl, money_in, money_out)
        self._totals.set_total("count", count, "")

        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._reset_and_load()

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            recs, total = await asyncio.gather(
                svc.get_kimvi_upload_records(
                    self._upload_id, self._search, _SCROLL_CHUNK, 0,
                ),
                svc.count_kimvi_upload_records(self._upload_id, self._search),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("count", total, "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            recs = await svc.get_kimvi_upload_records(
                self._upload_id, self._search, _SCROLL_CHUNK, self._loaded,
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._totals.set_total("count", self._total, "")
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _kimvi_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()


class AhmedKimviWidget(QWidget):
    """
    Ahmed Kimvi (Klesa) main page.

    Three tabs: All Entries, Money In, and Uploads (browse + drill-down).
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader("Ahmed Kimvi (Klesa)", "mdi.account-cash")
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        bulk_btn = _btn("Bulk Import All Sheets", "mdi.file-multiple-outline", primary=False)
        bulk_btn.clicked.connect(self._open_bulk_import)
        self._import_btn = _btn("Import Latest Sheet", "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(bulk_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Money In", "Uploads"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _KimviAllEntries()
        self._main_stack.addWidget(self._all_entries)

        self._money_in = _KimviMoneyInEntries()
        self._main_stack.addWidget(self._money_in)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _KimviUploadBrowse()
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _KimviUploadDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        self._tabs.tab_changed.connect(self._main_stack.setCurrentIndex)
        vl.addWidget(self._main_stack, 1)

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(2, emit=False)
        self._main_stack.setCurrentIndex(2)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            "Ahmed_Kimvi_Import_Template.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, "", _KIMVI_HEADERS)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        dlg = KimviImportDialog(parent=self)
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        QMessageBox.information(self, "Import Complete", f"Imported {n:,} rows.")
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()
        self._tabs.set_index(2)

    def _open_bulk_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        dlg = BulkSheetImportDialog(
            title="Bulk Import — Ahmed Kimvi (Klesa)",
            hint=(
                "Every sheet in the workbook will be checked. "
                "New sheets are imported; sheets already uploaded are skipped."
            ),
            parse_all_fn=_parse_kimvi_all_sheets,
            existing_fn=svc.kimvi_existing_sheet_labels,
            save_fn=svc.save_kimvi_import,
            parent=self,
        )
        dlg.imported.connect(self._on_bulk_imported)
        dlg.exec()

    def _on_bulk_imported(self, sheets: int, rows: int) -> None:
        QMessageBox.information(
            self, "Bulk Import Complete",
            f"Imported {sheets:,} sheet{'s' if sheets != 1 else ''} ({rows:,} rows).",
        )
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()
        self._tabs.set_index(2)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        upload_id = str(upload_doc.get("_id") or "")
        if not upload_id:
            return
        filename = upload_doc.get("source_filename") or "Unknown file"
        sheet_label = upload_doc.get("sheet_label") or "—"
        count = int(upload_doc.get("record_count", 0))
        if QMessageBox.question(
            self,
            "Delete Upload",
            f"Delete sheet \"{sheet_label}\" from \"{filename}\" "
            f"and all {count:,} records?\n\nThis cannot be undone.",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(upload_id))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            deleted = await svc.delete_kimvi_upload(upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        if deleted <= 0:
            QMessageBox.warning(self, "Delete Upload", "No records were deleted.")
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()


# ═══════════════════════════════════════════════════════════════════════════════
#  5. ZambiaParkingWidget  — weekly statement import, upload browse + detail
# ═══════════════════════════════════════════════════════════════════════════════

_ZAMBIA_PARK_HEADERS = [
    "DATE", "TYPE", "PLATE NUM.", "TICKET NO.", "DEBIT", "CREDIT", "BALANCE", "HEADING TO",
]

_ZAMBIA_BROWSE_HEADERS = [
    "UPLOAD DATE", "FILE NAME", "SHEET", "RECORDS", "CLOSING BAL (ZMW)", "DATE RANGE",
]


def _zambia_to_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _zambia_is_balance_row(type_str: str) -> bool:
    upper = type_str.upper()
    return (
        "CLOSING" in upper
        or "OPENING" in upper
        or upper in ("OB", "OPENING BALANCE")
    )


def _parse_zambia_parking_sheet(ws, sheet_label: str = "") -> List[dict]:
    """Parse one Zambia Parking weekly worksheet into row dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    records: List[dict] = []
    for row_idx, row in enumerate(rows[2:], start=3):
        if not any(c is not None for c in row):
            continue

        dt          = row[0] if len(row) > 0 else None
        typ         = row[1] if len(row) > 1 else None
        plate       = row[2] if len(row) > 2 else None
        ticket      = row[3] if len(row) > 3 else None
        debit       = row[4] if len(row) > 4 else None
        credit      = row[5] if len(row) > 5 else None
        balance     = row[6] if len(row) > 6 else None
        heading     = row[7] if len(row) > 7 else None

        type_str = str(typ).strip() if typ is not None else ""

        if isinstance(dt, datetime):
            date_str         = dt.strftime("%d %b %Y")
            transaction_date = dt
        else:
            date_str         = str(dt).strip() if dt is not None else ""
            transaction_date = None

        ticket_str = ""
        if ticket is not None and str(ticket).strip():
            if isinstance(ticket, float) and ticket == int(ticket):
                ticket_str = str(int(ticket))
            else:
                ticket_str = str(ticket).strip()

        records.append({
            "date":              date_str,
            "transaction_date":  transaction_date,
            "type":              type_str,
            "plate_num":         str(plate).strip() if plate is not None else "",
            "ticket_no":         ticket_str,
            "debit":             _zambia_to_float(debit),
            "credit":            _zambia_to_float(credit),
            "balance":           _zambia_to_float(balance),
            "heading_to":        str(heading).strip() if heading is not None else "",
            "is_balance_row":    _zambia_is_balance_row(type_str),
            "sheet_label":       sheet_label,
            "row_index":         row_idx,
            "feed_type":         "zambia_parking",
        })

    return records


def _parse_zambia_last_sheet(path: str) -> Tuple[str, List[dict]]:
    """Read only the last worksheet from a Zambia Parking workbook."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Zambia Parking Excel files.")

    wb = openpyxl.load_workbook(path, data_only=True)
    if not wb.sheetnames:
        return "", []

    sheet_name = wb.sheetnames[-1]
    return sheet_name, _parse_zambia_parking_sheet(wb[sheet_name], sheet_name)


def _parse_zambia_all_sheets(path: str) -> List[dict]:
    """Parse every worksheet in a Zambia Parking workbook."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Zambia Parking Excel files.")

    wb = openpyxl.load_workbook(path, data_only=True)
    batches: List[dict] = []
    for sheet_name in wb.sheetnames:
        records = _parse_zambia_parking_sheet(wb[sheet_name], sheet_name)
        if records:
            batches.append({"sheet_label": sheet_name, "records": records})
    return batches


def _zambia_batch_balance(records: List[dict]) -> float:
    for rec in reversed(records):
        bal = rec.get("balance")
        if bal is not None:
            try:
                return float(bal)
            except (TypeError, ValueError):
                continue
    return 0.0


def _zambia_fill_row(t: QTableWidget, r: int, rec: dict) -> None:
    """Populate one Zambia Parking row."""
    is_balance = rec.get("is_balance_row") or _zambia_is_balance_row(rec.get("type", ""))
    row_bg     = _BLUE_L if is_balance else None

    t.setItem(r, 0, _cell(rec.get("date", "")))
    t.setItem(r, 1, _cell(rec.get("type", "")))
    t.setItem(r, 2, _cell(rec.get("plate_num", "")))
    t.setItem(r, 3, _cell(rec.get("ticket_no", "")))
    t.setItem(r, 4, _cell(
        _fmt_num(rec.get("debit"), "ZMW ", 0) if rec.get("debit") is not None else "—",
    ))
    credit = rec.get("credit")
    t.setItem(r, 5, _cell(
        _fmt_num(credit, "ZMW ", 0) if credit is not None else "—",
        color=_GREEN if credit else "",
    ))
    t.setItem(r, 6, _cell(
        _fmt_num(rec.get("balance"), "ZMW ", 0) if rec.get("balance") is not None else "—",
    ))
    t.setItem(r, 7, _cell(rec.get("heading_to", "")))
    _finish_table_row(t, r, row_bg)


def _make_zambia_summary_cards() -> Tuple[QFrame, QLabel, QLabel, QLabel]:
    """Money In (credit) / Money Out (debit) / Balance summary strip for Zambia views."""
    frame = QFrame()
    frame.setStyleSheet(
        f"QFrame{{background:{_BLUE_L};border:1px solid #BFDBFE;"
        "border-radius:6px;padding:8px;}}"
    )
    sl = QHBoxLayout(frame)
    sl.setContentsMargins(12, 8, 12, 8)
    sl.setSpacing(24)
    in_lbl  = _lbl("Money In: —", size=12, weight=600, color=_GREEN)
    out_lbl = _lbl("Money Out: —", size=12, color=_T2)
    bal_lbl = _lbl("Balance: —", size=12, weight=600, color=_T1)
    sl.addWidget(in_lbl)
    sl.addWidget(out_lbl)
    sl.addWidget(bal_lbl)
    sl.addStretch()
    return frame, in_lbl, out_lbl, bal_lbl


def _set_zambia_summary(
    in_lbl: QLabel, out_lbl: QLabel, bal_lbl: QLabel,
    total_credit: float, total_debit: float,
) -> None:
    balance = total_credit - total_debit
    in_lbl.setText(f"Money In: ZMW {_fmt_num(total_credit, '', 0)}")
    out_lbl.setText(f"Money Out: ZMW {_fmt_num(total_debit, '', 0)}")
    bal_color = _GREEN if balance > 0 else (_RED if balance < 0 else _T1)
    bal_lbl.setText(f"Balance: ZMW {_fmt_num(balance, '', 0)}")
    bal_lbl.setStyleSheet(
        f"color:{bal_color};font-size:12px;font-weight:600;"
        "font-family:'Segoe UI';background:transparent;"
    )


class _ZambiaParkingEntriesBase(QWidget):
    """Shared flat-list view for Zambia Parking All Entries / Money In tabs."""

    def __init__(self, credit_only: bool = False, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._credit_only = credit_only
        self._search  = ""
        self._year    = 0
        self._month   = 0
        self._loaded  = 0
        self._total   = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search plate, ticket, destination, type…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        tbl.addStretch()
        vl.addWidget(tb)

        self._summary_frame, self._in_lbl, self._out_lbl, self._bal_lbl = (
            _make_zambia_summary_cards()
        )
        vl.addWidget(self._summary_frame)

        self._totals = _TotalsBar([("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_table(_ZAMBIA_PARK_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            years = await svc.get_zambia_parking_available_years()
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_zambia_parking_all_totals(
                    self._search, self._year, month,
                    credit_only=self._credit_only, **self._date_kw(),
                ),
                svc.get_zambia_parking_all_records(
                    self._search, self._year, month,
                    credit_only=self._credit_only,
                    limit=_SCROLL_CHUNK, skip=0, **self._date_kw(),
                ),
                svc.count_zambia_parking_all_records(
                    self._search, self._year, month,
                    credit_only=self._credit_only, **self._date_kw(),
                ),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return

        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        _set_zambia_summary(
            self._in_lbl, self._out_lbl, self._bal_lbl,
            float(totals.get("total_credit", 0)),
            float(totals.get("total_debit", 0)),
        )
        self._totals.set_total("count", int(totals.get("count", 0)), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_zambia_parking_all_records(
                self._search, self._year, month,
                credit_only=self._credit_only,
                limit=_SCROLL_CHUNK, skip=self._loaded, **self._date_kw(),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _zambia_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()


class _ZambiaParkingAllEntries(_ZambiaParkingEntriesBase):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(credit_only=False, parent=parent)


class _ZambiaParkingMoneyInEntries(_ZambiaParkingEntriesBase):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(credit_only=True, parent=parent)


class ZambiaParkingImportDialog(QDialog):
    """Import the last worksheet from a Zambia Parking weekly statement."""

    imported = Signal(int)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id       = str(uuid.uuid4())
        self._source_filename = ""
        self._sheet_label     = ""
        self._records: List[dict] = []
        self._already_exists  = False

        self.setWindowTitle("Import — Zambia Parking")
        self.setMinimumWidth(720)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        hint = _lbl(
            "Only the last sheet in the workbook will be imported "
            "(one weekly statement per upload).",
            size=12, color=_T2,
        )
        vl.addWidget(hint)

        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        browse_row = QWidget()
        browse_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(browse_row)
        brl.setContentsMargins(0, 0, 0, 0)
        browse_btn = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse_btn.clicked.connect(self._browse)
        brl.addStretch()
        brl.addWidget(browse_btn)
        vl.addWidget(browse_row)

        self._stats_lbl = _lbl("No file loaded.", size=12, color=_T2)
        vl.addWidget(self._stats_lbl)
        vl.addWidget(_hsep())

        preview_title = _lbl("Preview (first 10 rows)", size=12, weight=600)
        vl.addWidget(preview_title)

        self._preview_tbl = _make_table(_ZAMBIA_PARK_HEADERS)
        self._preview_tbl.setMinimumHeight(200)
        vl.addWidget(self._preview_tbl)
        vl.addWidget(_hsep())

        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.addStretch()

        cancel_btn = _btn("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        bbl.addWidget(cancel_btn)

        self._import_btn = _btn("Import Sheet", "mdi.check-circle-outline")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)
        vl.addWidget(btn_row)

    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open Zambia Parking Workbook", "", "Excel (*.xlsx *.xls)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _on_file(self, path: str) -> None:
        self._stats_lbl.setText("Reading file…")
        self._source_filename = Path(path).name
        try:
            sheet_label, records = _parse_zambia_last_sheet(path)
        except Exception as exc:
            self._stats_lbl.setText(f"Error reading file: {exc}")
            self._import_btn.setEnabled(False)
            return

        self._sheet_label = sheet_label
        self._records = records
        asyncio.ensure_future(self._check_sheet(sheet_label, records))

    async def _check_sheet(self, sheet_label: str, records: List[dict]) -> None:
        from tahmeed.services import accountant_service as svc

        try:
            exists = await svc.zambia_sheet_exists(sheet_label)
        except Exception:
            exists = False

        self._already_exists = exists
        total_debit  = sum(r.get("debit") or 0 for r in records)
        total_credit = sum(r.get("credit") or 0 for r in records)
        closing_bal  = _zambia_batch_balance(records)

        if exists:
            self._stats_lbl.setText(
                f"Sheet \"{sheet_label}\" was already uploaded — import blocked."
            )
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Already Uploaded")
        elif not records:
            self._stats_lbl.setText(
                f"Last sheet \"{sheet_label}\" has no data rows to import."
            )
            self._import_btn.setEnabled(False)
            self._import_btn.setText("Import Sheet")
        else:
            self._stats_lbl.setText(
                f"Last sheet: {sheet_label}     "
                f"Rows: {len(records):,}     "
                f"Debit: ZMW {total_debit:,.0f}     "
                f"Credit: ZMW {total_credit:,.0f}     "
                f"Closing: ZMW {closing_bal:,.0f}"
            )
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(records):,} Rows")

        t = self._preview_tbl
        t.setRowCount(0)
        for rec in records[:10]:
            row = t.rowCount()
            t.insertRow(row)
            _zambia_fill_row(t, row, rec)

    def _do_import(self) -> None:
        if self._already_exists or not self._records:
            return
        self._import_btn.setEnabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        from tahmeed.services import accountant_service as svc

        docs = []
        for rec in self._records:
            doc = dict(rec)
            doc["sheet_label"]     = self._sheet_label
            doc["upload_id"]       = self._upload_id
            doc["source_filename"] = self._source_filename
            docs.append(doc)

        try:
            saved = await svc.save_imported_feed(docs)
            self.imported.emit(saved)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(self._records):,} Rows")


class _ZambiaParkingUploadBrowse(QWidget):
    upload_clicked = Signal(object)
    delete_clicked = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        self._totals = _TotalsBar([
            ("balance", "Closing bal: ZMW "),
            ("count",   "Total records: "),
        ])
        vl.addWidget(self._totals)

        self._table = _make_table(_ZAMBIA_BROWSE_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 160)
        self._table.setColumnWidth(1, 220)
        self._table.setColumnWidth(2, 80)
        self._table.setColumnWidth(3, 80)
        self._table.setColumnWidth(4, 130)
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        hint = _lbl(
            "Click any row to view its records · right-click to delete an upload.",
            size=11, color=_TM,
        )
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        uploads = await svc.get_zambia_parking_uploads()
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        total_balance = 0.0
        total_recs    = 0

        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)

            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count         = int(up.get("record_count", 0))
            closing_bal   = float(up.get("closing_balance", 0) or 0)
            min_d         = up.get("min_transaction_date")
            max_d         = up.get("max_transaction_date")
            if isinstance(min_d, datetime) and isinstance(max_d, datetime):
                date_range = f"{min_d.strftime('%d %b %Y')} — {max_d.strftime('%d %b %Y')}"
            else:
                date_range = "—"

            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 2, _cell(up.get("sheet_label") or "—", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 3, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 4, _cell(
                _fmt_num(closing_bal, "ZMW ", 0),
                align=Qt.AlignRight | Qt.AlignVCenter,
            ))
            t.setItem(r, 5, _cell(date_range))
            _finish_table_row(t, r)

            total_balance += closing_bal
            total_recs    += count

        self._totals.set_total("balance", total_balance, "Closing bal: ZMW ")
        self._totals.set_total("count",   total_recs,    "")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


class _ZambiaParkingUploadDetail(QWidget):
    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._search    = ""
        self._loaded    = 0
        self._total     = 0
        self._loading   = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)

        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)

        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()

        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        self._summary_frame, self._in_lbl, self._out_lbl, self._bal_lbl = (
            _make_zambia_summary_cards()
        )
        vl.addWidget(self._summary_frame)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search plate, ticket, destination…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        tbl.addStretch()
        vl.addWidget(tb)

        self._table = _make_table(_ZAMBIA_PARK_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([("debit", "ZMW "), ("credit", "CR: ZMW ")])
        vl.addWidget(self._totals)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        filename      = upload_doc.get("source_filename") or "Unknown file"
        sheet_label   = upload_doc.get("sheet_label") or "—"
        count         = int(upload_doc.get("record_count", 0))
        total_debit   = float(upload_doc.get("total_debit", 0) or 0)
        total_credit  = float(upload_doc.get("total_credit", 0) or 0)
        import_dt     = upload_doc.get("import_date")
        date_str      = (
            import_dt.strftime("%d %b %Y")
            if isinstance(import_dt, datetime) else ""
        )

        self._crumb_lbl.setText(f"Uploads  ›  {filename}  ›  {sheet_label}")
        self._info_lbl.setText(
            f"{sheet_label}   •   {count:,} rows   •   {date_str}"
        )
        _set_zambia_summary(self._in_lbl, self._out_lbl, self._bal_lbl, total_credit, total_debit)

        self._totals.set_total("debit",  total_debit,  "ZMW ")
        self._totals.set_total("credit", total_credit, "CR: ZMW ")

        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._reset_and_load()

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(
            f"Showing {self._loaded:,} of {self._total:,}{suffix}"
        )

    async def _load_initial(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            recs, total = await asyncio.gather(
                svc.get_zambia_parking_upload_records(
                    self._upload_id, self._search, _SCROLL_CHUNK, 0,
                ),
                svc.count_zambia_parking_upload_records(self._upload_id, self._search),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            recs = await svc.get_zambia_parking_upload_records(
                self._upload_id, self._search, _SCROLL_CHUNK, self._loaded,
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _zambia_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()


class ZambiaParkingWidget(QWidget):
    """
    Zambia Parking main page.

    Three tabs: All Entries, Money In (credit), and Uploads (browse + drill-down).
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader("Zambia Parking", "mdi.map")
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        bulk_btn = _btn("Bulk Import All Sheets", "mdi.file-multiple-outline", primary=False)
        bulk_btn.clicked.connect(self._open_bulk_import)
        self._import_btn = _btn("Import Latest Sheet", "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(bulk_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Money In", "Uploads"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _ZambiaParkingAllEntries()
        self._main_stack.addWidget(self._all_entries)

        self._money_in = _ZambiaParkingMoneyInEntries()
        self._main_stack.addWidget(self._money_in)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _ZambiaParkingUploadBrowse()
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _ZambiaParkingUploadDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        self._tabs.tab_changed.connect(self._main_stack.setCurrentIndex)
        vl.addWidget(self._main_stack, 1)

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(2, emit=False)
        self._main_stack.setCurrentIndex(2)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            "Zambia_Parking_Import_Template.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, "", _ZAMBIA_PARK_HEADERS)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        dlg = ZambiaParkingImportDialog(parent=self)
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        QMessageBox.information(self, "Import Complete", f"Imported {n:,} rows.")
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()
        self._tabs.set_index(2)

    def _open_bulk_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        dlg = BulkSheetImportDialog(
            title="Bulk Import — Zambia Parking",
            hint=(
                "Every sheet (week) in the workbook will be checked. "
                "New sheets are imported; sheets already uploaded are skipped."
            ),
            parse_all_fn=_parse_zambia_all_sheets,
            existing_fn=svc.zambia_existing_sheet_labels,
            save_fn=svc.save_imported_feed,
            balance_fn=_zambia_batch_balance,
            balance_header="CLOSING BAL (ZMW)",
            parent=self,
        )
        dlg.imported.connect(self._on_bulk_imported)
        dlg.exec()

    def _on_bulk_imported(self, sheets: int, rows: int) -> None:
        QMessageBox.information(
            self, "Bulk Import Complete",
            f"Imported {sheets:,} sheet{'s' if sheets != 1 else ''} ({rows:,} rows).",
        )
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()
        self._tabs.set_index(2)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        upload_id = str(upload_doc.get("_id") or "")
        if not upload_id:
            return
        filename = upload_doc.get("source_filename") or "Unknown file"
        sheet_label = upload_doc.get("sheet_label") or "—"
        count = int(upload_doc.get("record_count", 0))
        if QMessageBox.question(
            self,
            "Delete Upload",
            f"Delete sheet \"{sheet_label}\" from \"{filename}\" "
            f"and all {count:,} records?\n\nThis cannot be undone.",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(upload_id))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            deleted = await svc.delete_zambia_parking_upload(upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        if deleted <= 0:
            QMessageBox.warning(self, "Delete Upload", "No records were deleted.")
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._money_in.refresh()
        self._show_browse()


# ═══════════════════════════════════════════════════════════════════════════════
#  Placeholder widget factory  (Afritrack, Third Party, COMESA)
# ═══════════════════════════════════════════════════════════════════════════════

class _PlaceholderExpenseWidget(QWidget):
    def __init__(
        self,
        title: str,
        icon_name: str,
        hint: str = "Import functionality coming soon.",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader(title, icon_name)
        import_btn = _btn("Import", "mdi.upload-outline")
        import_btn.setEnabled(False)
        header.add_right(import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        # Coming Soon card
        card = QFrame()
        card.setStyleSheet(
            f"QFrame{{background:{_WHITE};border:1px solid {_BORDER};"
            "border-radius:8px;}}"
        )
        card_vl = QVBoxLayout(card)
        card_vl.setContentsMargins(0, 60, 0, 60)
        card_vl.setAlignment(Qt.AlignCenter)

        try:
            icon_lbl = QLabel()
            icon_lbl.setPixmap(qta.icon(icon_name, color="#D1D5DB").pixmap(56, 56))
            icon_lbl.setAlignment(Qt.AlignCenter)
            icon_lbl.setStyleSheet("background:transparent;border:none;")
            card_vl.addWidget(icon_lbl)
        except Exception:
            pass

        card_vl.addSpacing(16)
        t_lbl = _lbl("Coming Soon", size=16, weight=700, color=_TM)
        t_lbl.setAlignment(Qt.AlignCenter)
        card_vl.addWidget(t_lbl)

        card_vl.addSpacing(6)
        h_lbl = _lbl(hint, size=12, color=_T2)
        h_lbl.setAlignment(Qt.AlignCenter)
        card_vl.addWidget(h_lbl)

        vl.addWidget(card, 1)

    def refresh(self) -> None:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
#  Afritrack — Schedule of Differences  (Excel-style interactive grid)
# ═══════════════════════════════════════════════════════════════════════════════

_AF_HEADERS = [
    "S/NO", "TRUCKS", "NO OF DAYS", "NON-TRANS DAYS",
    "TRANS DAYS", "RATE / DAY", "TOTAL (TAHMEED)",
    "TOTAL (INVOICE)", "VARIANCE", "REMARKS",
]
_AF_NCOLS       = len(_AF_HEADERS)
_AF_COL_SNO     = 0
_AF_COL_TRUCK   = 1
_AF_COL_DAYS    = 2
_AF_COL_NTRANS  = 3
_AF_COL_TRANS   = 4   # formula: DAYS − NTRANS
_AF_COL_RATE    = 5
_AF_COL_TOTAL_T = 6   # formula: TRANS × RATE
_AF_COL_TOTAL_I = 7
_AF_COL_VAR     = 8   # formula: TOTAL_T − TOTAL_I
_AF_COL_REMARKS = 9
_AF_FORMULA_COLS = frozenset({_AF_COL_TRANS, _AF_COL_TOTAL_T, _AF_COL_VAR})
_AF_NUM_COLS     = frozenset(range(2, 9))    # cols 2-8 are numeric

_AF_FOOTER_DEFS: List[Tuple[str, bool, bool, bool, bool]] = [
    # (key,        bold,   red,    show_i, show_var)
    ("sub",        False,  False,  True,   True),
    ("inst",       False,  False,  True,   True),
    ("sub2",       False,  False,  True,   True),
    ("vat",        False,  False,  True,   True),
    ("sub3",       False,  False,  True,   True),
    ("wht",        False,  False,  True,   True),
    ("payable",    True,   False,  True,   True),
    ("bal",        False,  True,   False,  False),
    ("total",      True,   True,   False,  False),
]


# ── tiny helpers ───────────────────────────────────────────────────────────────

def _af_flt(text: str) -> float:
    if not text or text in ("-", "—"):
        return 0.0
    try:
        return float(text.replace(",", ""))
    except Exception:
        return 0.0


def _af_fmt(val: float, decimals: int = 2) -> str:
    if val == 0.0:
        return "-"
    return f"{val:,.{decimals}f}"


# ── Excel-style reader ─────────────────────────────────────────────────────────

def _read_afritrack_file(path: str) -> Tuple[List[List[str]], float, float, float]:
    """Parse Afritrack xlsx.  Returns (data_rows, inst_t, inst_i, bal_mar)."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Afritrack files.")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    raw = list(ws.iter_rows(values_only=True))
    if not raw:
        return [], 0.0, 0.0, 0.0

    data_rows: List[List[str]] = []
    inst_t = inst_i = bal_mar = 0.0

    def _v(row, idx, fallback=0.0):
        try:
            v = row[idx] if len(row) > idx else None
            return float(v) if v is not None else fallback
        except Exception:
            return fallback

    def _s(val, dec=2):
        if val is None or val == 0.0:
            return "-"
        try:
            f = float(val)
            return f"{f:,.{dec}f}" if f != 0 else "-"
        except Exception:
            return str(val).strip()

    for row in raw[1:]:           # skip header row
        sno = row[0] if row else None
        if sno is None:
            label = " ".join(str(c or "").lower() for c in row[:9])
            if "installation fee" in label:
                inst_t = _v(row, 6)
                inst_i = _v(row, 7)
            elif "bal mar" in label or "balance mar" in label:
                bal_mar = _v(row, 6)
            continue
        try:
            sno = int(sno)
        except (TypeError, ValueError):
            continue

        days    = _v(row, 2)
        ntrans  = _v(row, 3)
        trans   = _v(row, 4) if (len(row) > 4 and row[4] is not None) else days - ntrans
        rate    = _v(row, 5)
        total_t = _v(row, 6) if (len(row) > 6 and row[6] is not None) else trans * rate
        total_i = _v(row, 7)
        if len(row) > 8 and row[8] is not None:
            variance = _v(row, 8)
        else:
            variance = total_t - total_i
        remarks = str(row[9] or "").strip() if len(row) > 9 else ""

        def _int_or_dec(f):
            return _s(f, 0) if f == int(f) else _s(f)

        data_rows.append([
            str(sno),
            str(row[1] or "").strip(),
            _int_or_dec(days),
            _int_or_dec(ntrans),
            _int_or_dec(trans),
            _s(rate, 6),
            _s(total_t),
            _s(total_i),
            _s(variance),
            remarks,
        ])

    return data_rows, inst_t, inst_i, bal_mar


# ── Excel exporter ─────────────────────────────────────────────────────────────

def _export_afritrack_xlsx(
    path: str,
    grid: "_AfritrackGrid",
    inst_t: float,
    inst_i: float,
    bal_mar: float,
    vat_rate: float,
    period: str,
) -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export.")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"Afritrack {period}"

    thin  = Side(border_style="thin", color="E5E7EB")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    r_aln = Alignment(horizontal="right",  vertical="center")
    l_aln = Alignment(horizontal="left",   vertical="center")
    c_aln = Alignment(horizontal="center", vertical="center")
    w_aln = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hdr_fill = PatternFill("solid", fgColor="EFF6FF")
    hdr_font = Font(name="Calibri", bold=True, size=11, color="1E3A5F")
    frm_fill = PatternFill("solid", fgColor="EFF6FF")
    frm_font = Font(name="Calibri", size=11, color="1D4ED8")
    alt_fill = PatternFill("solid", fgColor="F9FAFB")
    red_fill = PatternFill("solid", fgColor="FEF2F2")
    red_font = Font(name="Calibri", bold=True, size=11, color="B91C1C")
    bld_font = Font(name="Calibri", bold=True, size=11)
    nrm_font = Font(name="Calibri", size=11)

    col_widths_ch = [6, 16, 10, 12, 10, 10, 15, 15, 10, 20]
    for c, w in enumerate(col_widths_ch, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w

    # header row
    ws.row_dimensions[1].height = 34
    col_labels = [
        "S/NO", "TRUCKS", f"NO OF DAYS\n{period.upper()}",
        "NON-TRANS\nDAYS", "TRANS\nDAYS", "RATE/DAY",
        "TOTAL AS PER\nTAHMEED", "TOTAL AS PER\nINVOICE", "VARIANCE", "REMARKS",
    ]
    for c, label in enumerate(col_labels, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.border = bdr;   cell.alignment = w_aln

    # data rows
    data = grid.get_all_data()
    for ri, row in enumerate(data, 2):
        ws.row_dimensions[ri].height = 18
        is_alt = (ri % 2 == 0)
        for ci, val in enumerate(row, 1):
            is_formula = (ci - 1) in _AF_FORMULA_COLS
            num = None
            if 3 <= ci <= 9 and val not in ("", "-", "—"):
                try:
                    num = float(val.replace(",", ""))
                except Exception:
                    pass
            cell = ws.cell(row=ri, column=ci,
                           value=num if num is not None else (val or None))
            cell.border    = bdr
            cell.alignment = r_aln if 3 <= ci <= 9 else (c_aln if ci == 1 else l_aln)
            if is_formula:
                cell.font = frm_font; cell.fill = frm_fill
            elif is_alt:
                cell.font = nrm_font; cell.fill = alt_fill
            else:
                cell.font = nrm_font

    # footer rows
    st_t, st_i, _ = grid.get_col_totals()
    it = inst_t; ii = inst_i; bm = bal_mar; vr = vat_rate
    s2t  = st_t + it;     s2i = st_i + ii
    vat_t = s2t * vr / 100; vat_i = s2i * vr / 100
    s3t  = s2t + vat_t;   s3i = s2i + vat_i
    wt   = s3t * 0.05;    wi  = s3i * 0.05
    pt   = s3t - wt;      pi  = s3i - wi
    tot  = pt + bm

    fr = len(data) + 2
    footer_rows = [
        (fr,   "",                                    st_t,  st_i,  st_t-st_i, False, False),
        (fr+1, "Installation Fees",                   it,    ii,    it-ii,     False, False),
        (fr+2, "",                                    s2t,   s2i,   s2t-s2i,   False, False),
        (fr+3, f"VAT @ {vr:.0f}%",                   vat_t, vat_i, vat_t-vat_i, False, False),
        (fr+4, "",                                    s3t,   s3i,   s3t-s3i,   False, False),
        (fr+5, "Less WHT @ 5%",                      wt,    wi,    wt-wi,     False, False),
        (fr+6, "Amount Payable",                      pt,    pi,    pt-pi,     True,  False),
        (fr+7, "Add Bal. Mar (WHT Calculations)",     bm,    None,  None,      False, True),
        (fr+8, f"Total Payable {period} Account",     tot,   None,  None,      True,  True),
    ]
    for frow, label, vt_, vi_, vvar, bold, red in footer_rows:
        ws.row_dimensions[frow].height = 18
        if label:
            lc = ws.cell(row=frow, column=5, value=label)
            lc.alignment = r_aln
            lc.font = red_font if red else (bld_font if bold else nrm_font)
            if red:
                lc.fill = red_fill
        for xc, xv in [(_AF_COL_TOTAL_T+1, vt_),
                        (_AF_COL_TOTAL_I+1, vi_),
                        (_AF_COL_VAR+1,     vvar)]:
            if xv is not None:
                cell = ws.cell(row=frow, column=xc, value=xv)
                cell.alignment = r_aln; cell.border = bdr
                cell.font = red_font if red else (bld_font if bold else nrm_font)
                if red:
                    cell.fill = red_fill

    wb.save(path)


# ── Editable grid ─────────────────────────────────────────────────────────────

class _AfritrackGrid(QTableWidget):
    """
    Editable Excel-like grid.  Formula cols (TRANS DAYS, TOTAL TAHMEED,
    VARIANCE) auto-compute and are read-only.  All other cols are editable.
    Supports Ctrl+C/X/V, Ctrl+Z/Y, Delete, multi-select, right-click menu.
    """
    data_changed = Signal()

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(0, _AF_NCOLS, parent)
        self._block          = False
        self._undo_stack: List[List[List[str]]] = []
        self._redo_stack: List[List[List[str]]] = []
        self._pre_edit_snap: Optional[List[List[str]]] = None
        self._setup()

    # ── initialisation ────────────────────────────────────────────────────────

    def _setup(self) -> None:
        self.setHorizontalHeaderLabels(_AF_HEADERS)
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setEditTriggers(
            QAbstractItemView.DoubleClicked | QAbstractItemView.AnyKeyPressed
        )
        self.setAlternatingRowColors(False)
        self.verticalHeader().setVisible(False)
        self.verticalHeader().setMinimumSectionSize(18)
        self.verticalHeader().setDefaultSectionSize(_ROW_H)
        self.setShowGrid(True)
        self.setSortingEnabled(False)

        hdr = self.horizontalHeader()
        hdr.setHighlightSections(False)
        hdr.setMinimumSectionSize(36)
        hdr.setStretchLastSection(True)
        col_widths = [38, 110, 60, 76, 60, 62, 96, 96, 66, 120]
        for c, w in enumerate(col_widths):
            self.setColumnWidth(c, w)

        self.setStyleSheet(self._grid_ss())
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._ctx_menu)
        self.currentItemChanged.connect(self._capture_pre_edit)
        self.itemChanged.connect(self._on_item_changed)

    def _grid_ss(self) -> str:
        return (
            f"QTableWidget{{background:{_WHITE};gridline-color:{_BORDER};"
            "border:none;font-size:12px;font-family:'Segoe UI';}}"
            f"QTableWidget::item{{padding:0 6px;color:{_T1};}}"
            f"QTableWidget::item:selected{{background:{_QB_SEL_BG};color:{_QB_SEL_FG};}}"
            f"QHeaderView::section{{background:{_QB_HDR_BG};color:{_QB_HDR_FG};"
            "font-size:10px;font-weight:700;font-family:'Segoe UI';"
            f"border:none;border-right:1px solid {_BORDER};"
            f"border-bottom:2px solid {_BLUE};padding:0 6px;height:{_HDR_H}px;}}"
            "QScrollBar:vertical{width:8px;background:transparent;}"
            f"QScrollBar::handle:vertical{{background:#D1D5DB;border-radius:4px;}}"
            "QScrollBar:horizontal{height:8px;background:transparent;}"
            f"QScrollBar::handle:horizontal{{background:#D1D5DB;border-radius:4px;}}"
        )

    # ── keyboard ──────────────────────────────────────────────────────────────

    def keyPressEvent(self, event) -> None:
        mod = event.modifiers()
        key = event.key()
        if mod == Qt.ControlModifier:
            if key == Qt.Key_C: self._do_copy();  return
            if key == Qt.Key_X: self._do_cut();   return
            if key == Qt.Key_V: self._do_paste();  return
            if key == Qt.Key_Z: self._do_undo();   return
            if key == Qt.Key_Y: self._do_redo();   return
        # Tab on last column → wrap to TRUCKS col of next row (add row if needed)
        if key == Qt.Key_Tab and not (mod & Qt.ShiftModifier):
            row, col = self.currentRow(), self.currentColumn()
            if col == _AF_NCOLS - 1:
                if row >= self.rowCount() - 1:
                    self.add_row()
                self.setCurrentCell(min(row + 1, self.rowCount() - 1), _AF_COL_TRUCK)
                return
        if key == Qt.Key_Delete:
            self.delete_selected_rows(); return
        if key == Qt.Key_Backspace:
            self._do_clear(); return
        super().keyPressEvent(event)

    def closeEditor(self, editor, hint) -> None:
        """Tab while editing the last column → commit + move to TRUCKS of next row."""
        from PySide6.QtWidgets import QAbstractItemDelegate
        if hint == QAbstractItemDelegate.EditNextItem:
            col = self.currentColumn()
            row = self.currentRow()
            if col == _AF_NCOLS - 1:
                super().closeEditor(editor, QAbstractItemDelegate.SubmitModelCache)
                if row >= self.rowCount() - 1:
                    self.add_row()
                self.setCurrentCell(min(row + 1, self.rowCount() - 1), _AF_COL_TRUCK)
                return
        super().closeEditor(editor, hint)

    # ── copy / cut / paste ────────────────────────────────────────────────────

    def _sel_rows_cols(self) -> Tuple[List[int], List[int]]:
        idxs = self.selectedIndexes()
        if not idxs:
            return [], []
        return (
            sorted({i.row() for i in idxs}),
            sorted({i.column() for i in idxs}),
        )

    def _do_copy(self) -> None:
        rows, cols = self._sel_rows_cols()
        if not rows:
            return
        lines = []
        for r in rows:
            lines.append("\t".join(
                (self.item(r, c).text() if self.item(r, c) else "") for c in cols
            ))
        QApplication.clipboard().setText("\n".join(lines))

    def _do_cut(self) -> None:
        self._do_copy()
        self._do_clear()

    def _do_paste(self) -> None:
        text = QApplication.clipboard().text()
        if not text:
            return
        paste_rows = [line.split("\t") for line in text.splitlines()]
        while paste_rows and all(c == "" for c in paste_rows[-1]):
            paste_rows.pop()
        sel_rows, sel_cols = self._sel_rows_cols()
        r0 = sel_rows[0] if sel_rows else max(self.currentRow(), 0)
        c0 = sel_cols[0] if sel_cols else max(self.currentColumn(), 0)
        self._push_undo()
        self._block = True
        for dr, row_data in enumerate(paste_rows):
            r = r0 + dr
            if r >= self.rowCount():
                self._insert_blank_row(r)
            for dc, val in enumerate(row_data):
                c = c0 + dc
                if c >= self.columnCount() or c in _AF_FORMULA_COLS:
                    continue
                item = self.item(r, c) or QTableWidgetItem()
                item.setText(val.strip())
                self._style_item(item, r, c)
                self.setItem(r, c, item)
        self._block = False
        self._recalc_all()
        self.data_changed.emit()

    def _do_clear(self) -> None:
        self._push_undo()
        self._block = True
        for item in self.selectedItems():
            if item.column() not in (_AF_FORMULA_COLS | {_AF_COL_SNO}):
                item.setText("")
        self._block = False
        self._recalc_all()
        self.data_changed.emit()

    # ── undo / redo ───────────────────────────────────────────────────────────

    def _snapshot(self) -> List[List[str]]:
        return [
            [(self.item(r, c).text() if self.item(r, c) else "")
             for c in range(self.columnCount())]
            for r in range(self.rowCount())
        ]

    def _push_undo(self) -> None:
        self._undo_stack.append(self._snapshot())
        if len(self._undo_stack) > 50:
            self._undo_stack.pop(0)
        self._redo_stack.clear()

    def _restore(self, snap: List[List[str]]) -> None:
        self._block = True
        self.setRowCount(len(snap))
        for r, row_data in enumerate(snap):
            for c, text in enumerate(row_data):
                item = self.item(r, c) or QTableWidgetItem()
                item.setText(text)
                self._style_item(item, r, c)
                self.setItem(r, c, item)
        self._block = False
        self._recalc_all()
        self.data_changed.emit()

    def _do_undo(self) -> None:
        if not self._undo_stack:
            return
        self._redo_stack.append(self._snapshot())
        self._restore(self._undo_stack.pop())

    def _do_redo(self) -> None:
        if not self._redo_stack:
            return
        self._push_undo()
        self._restore(self._redo_stack.pop())

    # ── formula engine ────────────────────────────────────────────────────────

    def _capture_pre_edit(self, _curr, _prev) -> None:
        if not self._block:
            self._pre_edit_snap = self._snapshot()

    def _on_item_changed(self, item: QTableWidgetItem) -> None:
        if self._block:
            return
        col = item.column()
        if col not in _AF_FORMULA_COLS:
            if self._pre_edit_snap is not None:
                self._undo_stack.append(self._pre_edit_snap)
                if len(self._undo_stack) > 50:
                    self._undo_stack.pop(0)
                self._redo_stack.clear()
                self._pre_edit_snap = None
        self._recalc_row(item.row())
        self.data_changed.emit()

    def _recalc_all(self) -> None:
        for r in range(self.rowCount()):
            self._recalc_row(r, renumber=False)
        self._renumber()

    def _recalc_row(self, row: int, renumber: bool = True) -> None:
        self._block = True

        def _txt(c):
            it = self.item(row, c)
            return it.text() if it else ""

        days    = _af_flt(_txt(_AF_COL_DAYS))
        ntrans  = _af_flt(_txt(_AF_COL_NTRANS))
        rate    = _af_flt(_txt(_AF_COL_RATE))
        total_i = _af_flt(_txt(_AF_COL_TOTAL_I))
        trans   = days - ntrans
        total_t = trans * rate
        var     = total_t - total_i

        self._set_fcell(row, _AF_COL_TRANS,   trans,  is_var=False)
        self._set_fcell(row, _AF_COL_TOTAL_T, total_t, is_var=False)
        self._set_fcell(row, _AF_COL_VAR,     var,    is_var=True)

        self._block = False
        if renumber:
            self._renumber()

    def _set_fcell(self, row: int, col: int, val: float,
                   is_var: bool = False) -> None:
        item = self.item(row, col)
        if item is None:
            item = QTableWidgetItem()
            self.setItem(row, col, item)
        item.setText(_af_fmt(val))
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        item.setFlags(
            (item.flags() & ~Qt.ItemIsEditable) | Qt.ItemIsSelectable | Qt.ItemIsEnabled
        )
        if is_var:
            if val < -0.005:
                item.setForeground(QColor(_QB_VNEG_FG))
                item.setBackground(QColor(_QB_VNEG_BG))
            elif val > 0.005:
                item.setForeground(QColor(_QB_VPOS_FG))
                item.setBackground(QColor(_QB_FORM_BG))
            else:
                item.setForeground(QColor(_QB_VZRO_FG))
                item.setBackground(QColor(_QB_FORM_BG))
        else:
            item.setForeground(QColor(_QB_FORM_FG))
            item.setBackground(QColor(_QB_FORM_BG))

    def _style_item(self, item: QTableWidgetItem, row: int, col: int) -> None:
        if col in _AF_FORMULA_COLS:
            item.setBackground(QColor(_QB_FORM_BG))
            item.setForeground(QColor(_QB_FORM_FG))
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        elif col in _AF_NUM_COLS:
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item.setBackground(QColor(_stripe_bg(row)))
        elif col == _AF_COL_SNO:
            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            item.setForeground(QColor(_T2))
            item.setBackground(QColor(_stripe_bg(row)))
        else:
            item.setBackground(QColor(_stripe_bg(row)))

    def _renumber(self) -> None:
        self._block = True
        for r in range(self.rowCount()):
            item = self.item(r, _AF_COL_SNO)
            if item is None:
                item = QTableWidgetItem()
                self.setItem(r, _AF_COL_SNO, item)
            item.setText(str(r + 1))
            item.setForeground(QColor(_T2))
            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            item.setFlags(
                (item.flags() & ~Qt.ItemIsEditable) | Qt.ItemIsSelectable | Qt.ItemIsEnabled
            )
        self._block = False

    # ── row operations ────────────────────────────────────────────────────────

    def _insert_blank_row(self, r: int) -> None:
        self.insertRow(r)
        self._block = True
        for c in range(_AF_NCOLS):
            item = QTableWidgetItem("")
            self._style_item(item, r, c)
            self.setItem(r, c, item)
        self.setRowHeight(r, _ROW_H)
        self._block = False

    def add_row(self, values: List[str] = None) -> None:
        r = self.rowCount()
        self._insert_blank_row(r)
        if values:
            self._block = True
            for c in range(min(len(values), _AF_NCOLS)):
                if c in (_AF_FORMULA_COLS | {_AF_COL_SNO}):
                    continue
                item = self.item(r, c) or QTableWidgetItem()
                item.setText(str(values[c]) if values[c] is not None else "")
                self._style_item(item, r, c)
                self.setItem(r, c, item)
            self._block = False
        self._recalc_row(r)
        self.scrollToBottom()

    def insert_row_above(self) -> None:
        sel_rows, _ = self._sel_rows_cols()
        r = sel_rows[0] if sel_rows else self.rowCount()
        self._push_undo()
        self._insert_blank_row(r)
        self._recalc_row(r)
        self._renumber()
        self.data_changed.emit()

    def delete_selected_rows(self) -> None:
        rows = sorted({i.row() for i in self.selectedIndexes()}, reverse=True)
        if not rows:
            return
        self._push_undo()
        for r in rows:
            self.removeRow(r)
        self._renumber()
        self.data_changed.emit()

    # ── data access ───────────────────────────────────────────────────────────

    def get_all_data(self) -> List[List[str]]:
        return [
            [(self.item(r, c).text() if self.item(r, c) else "")
             for c in range(self.columnCount())]
            for r in range(self.rowCount())
        ]

    def get_col_totals(self) -> Tuple[float, float, float]:
        def _sum(col):
            return sum(
                _af_flt(self.item(r, col).text() if self.item(r, col) else "")
                for r in range(self.rowCount())
                if not self.isRowHidden(r)
            )
        t = _sum(_AF_COL_TOTAL_T)
        i = _sum(_AF_COL_TOTAL_I)
        return t, i, t - i

    def filter_rows(self, text: str) -> None:
        text = text.lower().strip()
        for r in range(self.rowCount()):
            match = any(
                text in (self.item(r, c).text().lower() if self.item(r, c) else "")
                for c in range(self.columnCount())
            )
            self.setRowHidden(r, bool(text) and not match)

    # ── context menu ──────────────────────────────────────────────────────────

    def _ctx_menu(self, pos) -> None:
        menu = QMenu(self)
        menu.setStyleSheet(
            f"QMenu{{background:{_WHITE};border:1px solid {_BORDER};"
            f"border-radius:6px;padding:4px;"
            f"font-size:12px;font-family:'Segoe UI';}}"
            f"QMenu::item{{padding:6px 20px;border-radius:4px;color:{_T1};}}"
            f"QMenu::item:selected{{background:{_BLUE_L};}}"
            f"QMenu::separator{{height:1px;background:{_BORDER};margin:4px 8px;}}"
        )
        menu.addAction("Copy\t\tCtrl+C",        self._do_copy)
        menu.addAction("Cut\t\tCtrl+X",         self._do_cut)
        menu.addAction("Paste\t\tCtrl+V",        self._do_paste)
        menu.addAction("Clear Cells\tBackspace",  self._do_clear)
        menu.addSeparator()
        menu.addAction("Insert Row Above",              self.insert_row_above)
        menu.addAction("Add Row at Bottom\tTab",        self.add_row)
        menu.addAction("Delete Row(s)\t\tDel",          self.delete_selected_rows)
        menu.addSeparator()
        menu.addAction("Undo\t\tCtrl+Z",  self._do_undo)
        menu.addAction("Redo\t\tCtrl+Y",  self._do_redo)
        menu.exec(self.viewport().mapToGlobal(pos))


# ── Pinned footer ─────────────────────────────────────────────────────────────

class _AfritrackFooter(QWidget):
    """
    Non-scrolling totals section that mirrors the Excel footer layout:
    Subtotal → Installation Fees → Sub2 → VAT → Total+VAT →
    Less WHT @5% → Amount Payable → Add Bal. Mar → Total Payable Account
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._inst_t   = 0.0
        self._inst_i   = 0.0
        self._bal_mar  = 0.0
        self._vat_rate = 15.0
        self._build()

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(0)

        sep = QFrame()
        sep.setFixedHeight(2)
        sep.setStyleSheet(f"background:{_BLUE};border:none;")
        vl.addWidget(sep)

        self._table = QTableWidget(len(_AF_FOOTER_DEFS), _AF_NCOLS)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.horizontalHeader().setVisible(False)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(24)
        self._table.setShowGrid(True)
        self._table.setStyleSheet(self._foot_ss())
        self._table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self._table.setFocusPolicy(Qt.NoFocus)

        for ri in range(len(_AF_FOOTER_DEFS)):
            self._table.setSpan(ri, 0, 1, 6)   # cols 0-5 → merged label cell

        self._fix_height()
        vl.addWidget(self._table)

    def _foot_ss(self) -> str:
        return (
            f"QTableWidget{{background:{_QB_FOOT_BG};gridline-color:{_BORDER};"
            "border:none;font-size:12px;font-family:'Segoe UI';}}"
            f"QTableWidget::item{{padding:0 8px;color:{_T1};}}"
        )

    def _fix_height(self) -> None:
        h = sum(self._table.rowHeight(r) for r in range(len(_AF_FOOTER_DEFS))) + 2
        self._table.setFixedHeight(h)

    def sync_columns(self, grid: "_AfritrackGrid") -> None:
        for c in range(_AF_NCOLS):
            self._table.setColumnWidth(c, grid.columnWidth(c))

    def set_params(self, inst_t: float, inst_i: float,
                   bal_mar: float, vat_rate: float) -> None:
        self._inst_t   = inst_t
        self._inst_i   = inst_i
        self._bal_mar  = bal_mar
        self._vat_rate = vat_rate

    def refresh(self, grid: "_AfritrackGrid", month: str) -> None:
        sum_t, sum_i, _ = grid.get_col_totals()
        it   = self._inst_t;  ii  = self._inst_i
        s2t  = sum_t + it;    s2i = sum_i + ii
        vt   = s2t * self._vat_rate / 100
        vi   = s2i * self._vat_rate / 100
        s3t  = s2t + vt;      s3i = s2i + vi
        wt   = s3t * 0.05;    wi  = s3i * 0.05
        pt   = s3t - wt;      pi  = s3i - wi
        tot  = pt + self._bal_mar

        vals: Dict[str, Tuple[str, float, Optional[float], Optional[float]]] = {
            "sub":     ("",                                    sum_t,          sum_i,    sum_t - sum_i),
            "inst":    ("Installation Fees",                   it,             ii,       it - ii),
            "sub2":    ("",                                    s2t,            s2i,      s2t - s2i),
            "vat":     (f"VAT @ {self._vat_rate:.0f}%",       vt,             vi,       vt - vi),
            "sub3":    ("",                                    s3t,            s3i,      s3t - s3i),
            "wht":     ("Less WHT @ 5%",                      wt,             wi,       wt - wi),
            "payable": ("Amount Payable",                      pt,             pi,       pt - pi),
            "bal":     (f"Add Bal. Mar (WHT Calculations)",    self._bal_mar,  None,     None),
            "total":   (f"Total Payable {month} Account",      tot,            None,     None),
        }

        t = self._table
        for ri, (key, bold, red, show_i, show_var) in enumerate(_AF_FOOTER_DEFS):
            label, vt_, vi_, vvar = vals[key]

            # label cell (merged cols 0-5)
            lbl = t.item(ri, 0) or QTableWidgetItem()
            lbl.setText(label)
            lbl.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            fn = lbl.font(); fn.setBold(bold); lbl.setFont(fn)
            if red:
                lbl.setForeground(QColor(_QB_RED_DARK))
                lbl.setBackground(QColor(_QB_RED_BG))
            elif bold:
                lbl.setForeground(QColor(_QB_BOLD_FG))
                lbl.setBackground(QColor(_QB_FOOT_BG))
            else:
                lbl.setForeground(QColor(_T2))
                lbl.setBackground(QColor(_QB_FOOT_BG))
            t.setItem(ri, 0, lbl)

            def _put(col: int, val: Optional[float]) -> None:
                cell = t.item(ri, col) or QTableWidgetItem()
                if val is None:
                    cell.setText(""); cell.setBackground(QColor(_QB_FOOT_BG))
                    t.setItem(ri, col, cell); return
                cell.setText(_af_fmt(val))
                cell.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                fn2 = cell.font(); fn2.setBold(bold); cell.setFont(fn2)
                if red:
                    cell.setForeground(QColor(_QB_RED_DARK))
                    cell.setBackground(QColor(_QB_RED_BG))
                elif col == _AF_COL_VAR and val < -0.005:
                    cell.setForeground(QColor(_QB_VNEG_FG))
                    cell.setBackground(QColor(_QB_VNEG_BG))
                elif col == _AF_COL_VAR and val > 0.005:
                    cell.setForeground(QColor(_QB_VPOS_FG))
                    cell.setBackground(QColor(_QB_FOOT_BG))
                else:
                    cell.setForeground(QColor(_QB_BOLD_FG if bold else _T1))
                    cell.setBackground(QColor(_QB_FOOT_BG))
                t.setItem(ri, col, cell)

            _put(_AF_COL_TOTAL_T, vt_)
            _put(_AF_COL_TOTAL_I, vi_  if show_i   else None)
            _put(_AF_COL_VAR,     vvar if show_var  else None)

        self._fix_height()

    def get_export_data(self) -> Dict[str, float]:
        return {
            "inst_t":   self._inst_t,
            "inst_i":   self._inst_i,
            "bal_mar":  self._bal_mar,
            "vat_rate": self._vat_rate,
        }


# ── QB Bill-card field helpers ────────────────────────────────────────────────

def _qb_field_widget(label_text: str, input_widget: QWidget) -> QWidget:
    """Small ALL-CAPS gray label above an input — matches QB Bill field style."""
    w = QWidget()
    w.setStyleSheet("background:transparent;")
    vl = QVBoxLayout(w)
    vl.setContentsMargins(0, 0, 0, 0)
    vl.setSpacing(3)
    lbl = QLabel(label_text)
    lbl.setStyleSheet(
        f"color:{_T2};font-size:9px;font-weight:600;letter-spacing:0.8px;"
        f"font-family:'Segoe UI';background:transparent;"
    )
    vl.addWidget(lbl)
    vl.addWidget(input_widget)
    return w


def _qb_amount_widget(
    label_text: str,
    badge: str = "USD",
    red: bool = False,
) -> Tuple[QWidget, QLabel]:
    """
    QB-style amount display: currency badge on left + large bold value on right.
    Returns (container_widget, value_label).
    """
    w = QWidget()
    w.setStyleSheet("background:transparent;")
    vl = QVBoxLayout(w)
    vl.setContentsMargins(0, 0, 0, 0)
    vl.setSpacing(3)

    lbl = QLabel(label_text)
    lbl.setStyleSheet(
        f"color:{_T2};font-size:9px;font-weight:600;letter-spacing:0.8px;"
        f"font-family:'Segoe UI';background:transparent;"
    )
    vl.addWidget(lbl)

    box = QFrame()
    box_bg = _QB_RED_BG if red else _BG
    box.setStyleSheet(
        f"QFrame{{background:{box_bg};border:1px solid {_BORDER};border-radius:4px;}}"
    )
    box.setFixedHeight(34)
    hl = QHBoxLayout(box)
    hl.setContentsMargins(0, 0, 0, 0)
    hl.setSpacing(0)

    badge_lbl = QLabel(badge)
    badge_lbl.setFixedWidth(40)
    badge_lbl.setAlignment(Qt.AlignCenter)
    badge_bg = _QB_RED_BG if red else _QB_HDR_BG
    badge_fg = _QB_RED_DARK if red else _QB_HDR_FG
    badge_lbl.setStyleSheet(
        f"QLabel{{background:{badge_bg};border:none;"
        f"border-right:1px solid {_BORDER};"
        f"border-top-left-radius:4px;border-bottom-left-radius:4px;"
        f"color:{badge_fg};font-size:10px;font-weight:700;"
        f"font-family:'Segoe UI';}}"
    )
    hl.addWidget(badge_lbl)

    val_lbl = QLabel("0.00")
    val_lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
    val_fg = _QB_RED_DARK if red else _QB_HDR_FG
    val_lbl.setStyleSheet(
        f"QLabel{{color:{val_fg};font-size:14px;font-weight:700;"
        f"font-family:'Segoe UI';background:transparent;padding:0 10px;}}"
    )
    hl.addWidget(val_lbl, 1)
    vl.addWidget(box)
    return w, val_lbl


# ── Main widget / persisted upload flow ───────────────────────────────────────

_AF_DETAIL_HEADERS = [
    "S/NO", "TRUCK", "DAYS", "NON-TRANS", "TRANS", "RATE/DAY",
    "TOTAL TAHMEED", "TOTAL INVOICE", "VARIANCE", "REMARKS",
]
_AF_TEMPLATE_TITLE = "Afritrack Schedule — Schedule of Differences"
_AF_TEMPLATE_FILENAME = "Afritrack_Import_Template.xlsx"


def _afritrack_fill_row(t: QTableWidget, r: int, rec: dict) -> None:
    variance = float(rec.get("variance", 0) or 0)
    values = [
        str(rec.get("row_index", r + 1)),
        rec.get("truck", ""),
        _fmt_num(rec.get("days"), decimals=0) if rec.get("days") is not None else "—",
        _fmt_num(rec.get("non_trans_days"), decimals=0) if rec.get("non_trans_days") is not None else "—",
        _fmt_num(rec.get("trans_days"), decimals=0) if rec.get("trans_days") is not None else "—",
        _fmt_num(rec.get("rate_per_day"), decimals=6) if rec.get("rate_per_day") is not None else "—",
        _fmt_num(rec.get("total_tahmeed"), decimals=2) if rec.get("total_tahmeed") is not None else "—",
        _fmt_num(rec.get("total_invoice"), decimals=2) if rec.get("total_invoice") is not None else "—",
        _fmt_num(variance, decimals=2),
        rec.get("remarks", ""),
    ]
    aligns = [
        Qt.AlignCenter | Qt.AlignVCenter, Qt.AlignLeft | Qt.AlignVCenter,
        Qt.AlignRight | Qt.AlignVCenter, Qt.AlignRight | Qt.AlignVCenter,
        Qt.AlignRight | Qt.AlignVCenter, Qt.AlignRight | Qt.AlignVCenter,
        Qt.AlignRight | Qt.AlignVCenter, Qt.AlignRight | Qt.AlignVCenter,
        Qt.AlignRight | Qt.AlignVCenter, Qt.AlignLeft | Qt.AlignVCenter,
    ]
    for c, val in enumerate(values):
        color = ""
        if c == 8:
            color = _RED if variance < -0.005 else (_GREEN if variance > 0.005 else "")
        t.setItem(r, c, _cell(val, aligns[c], mono=False, color=color))
    _finish_table_row(t, r)


class _AfritrackAllEntries(QWidget):
    """Flat, filterable list of every Afritrack record with infinite scroll."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._search = ""
        self._year = 0
        self._month = 0
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search truck, remarks, period, file…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        tbl.addStretch()
        vl.addWidget(tb)

        self._totals = _TotalsBar([
            ("t", "Tahmeed "),
            ("i", "Invoice "),
            ("v", "Variance "),
            ("c", "Rows "),
        ])
        vl.addWidget(self._totals)

        self._table = _make_table(_AF_DETAIL_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        widths = [56, 120, 70, 90, 70, 90, 120, 120, 100, 180]
        for i, w in enumerate(widths):
            self._table.setColumnWidth(i, w)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            years = await svc.get_afritrack_available_years()
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_afritrack_all_totals(self._search, self._year, month, **self._date_kw()),
                svc.get_afritrack_all_records(self._search, self._year, month, limit=_SCROLL_CHUNK, skip=0, **self._date_kw()),
                svc.count_afritrack_all_records(self._search, self._year, month, **self._date_kw()),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return

        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("t", float(totals.get("total_tahmeed", 0) or 0), "USD ")
        self._totals.set_total("i", float(totals.get("total_invoice", 0) or 0), "USD ")
        self._totals.set_total("v", float(totals.get("total_variance", 0) or 0), "USD ")
        self._totals.set_total("c", int(totals.get("count", 0) or 0), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_afritrack_all_records(
                self._search, self._year, month,
                limit=_SCROLL_CHUNK, skip=self._loaded, **self._date_kw(),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _afritrack_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()


def _afritrack_rows_to_records(
    rows: List[List[str]],
    period: str,
    source_filename: str,
    upload_id: str,
    inst_t: float,
    inst_i: float,
    bal_mar: float,
    vat_rate: float = 15.0,
) -> List[dict]:
    records: List[dict] = []
    for row_idx, row in enumerate(rows, 1):
        truck = str(row[_AF_COL_TRUCK] or "").strip() if len(row) > _AF_COL_TRUCK else ""
        if not truck:
            continue
        total_t = _af_flt(row[_AF_COL_TOTAL_T]) if len(row) > _AF_COL_TOTAL_T else 0.0
        total_i = _af_flt(row[_AF_COL_TOTAL_I]) if len(row) > _AF_COL_TOTAL_I else 0.0
        if len(row) > _AF_COL_VAR and str(row[_AF_COL_VAR] or "").strip() not in ("", "-", "—"):
            variance = _af_flt(row[_AF_COL_VAR])
        else:
            variance = total_t - total_i
        records.append({
            "feed_type": "afritrack",
            "upload_id": upload_id,
            "source_filename": source_filename,
            "period": period,
            "row_index": row_idx,
            "truck": truck,
            "days": _af_flt(row[_AF_COL_DAYS]) if len(row) > _AF_COL_DAYS else 0.0,
            "non_trans_days": _af_flt(row[_AF_COL_NTRANS]) if len(row) > _AF_COL_NTRANS else 0.0,
            "trans_days": _af_flt(row[_AF_COL_TRANS]) if len(row) > _AF_COL_TRANS else 0.0,
            "rate_per_day": _af_flt(row[_AF_COL_RATE]) if len(row) > _AF_COL_RATE else 0.0,
            "total_tahmeed": total_t,
            "total_invoice": total_i,
            "variance": variance,
            "remarks": str(row[_AF_COL_REMARKS] or "").strip() if len(row) > _AF_COL_REMARKS else "",
            "installation_tahmeed": inst_t,
            "installation_invoice": inst_i,
            "balance_mar": bal_mar,
            "vat_rate": vat_rate,
        })
    return records


class _AfritrackUploadBrowse(QWidget):
    upload_clicked = Signal(object)
    delete_clicked = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        headers = ["UPLOAD DATE", "PERIOD", "FILE", "ROWS", "TAHMEED", "INVOICE", "VARIANCE"]
        self._table = _make_table(headers)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 150)
        self._table.setColumnWidth(1, 95)
        self._table.setColumnWidth(2, 220)
        self._table.setColumnWidth(3, 70)
        self._table.setColumnWidth(4, 110)
        self._table.setColumnWidth(5, 110)
        self._table.setColumnWidth(6, 110)
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([
            ("t", "Tahmeed "),
            ("i", "Invoice "),
            ("v", "Variance "),
            ("c", "Rows "),
        ])
        vl.addWidget(self._totals)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        uploads = await svc.get_afritrack_uploads()
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        total_t = total_i = total_v = 0.0
        total_c = 0
        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)
            import_dt = up.get("import_date")
            date_str = import_dt.strftime("%d %b %Y  %H:%M") if isinstance(import_dt, datetime) else "—"
            count = int(up.get("record_count", 0))
            tah = float(up.get("total_tahmeed", 0) or 0)
            inv = float(up.get("total_invoice", 0) or 0)
            var = float(up.get("total_variance", 0) or 0)
            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("period") or "—"))
            t.setItem(r, 2, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 3, _cell(f"{count:,}", Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 4, _cell(_fmt_num(tah, decimals=2), Qt.AlignRight | Qt.AlignVCenter, mono=True))
            t.setItem(r, 5, _cell(_fmt_num(inv, decimals=2), Qt.AlignRight | Qt.AlignVCenter, mono=True))
            t.setItem(r, 6, _cell(
                _fmt_num(var, decimals=2), Qt.AlignRight | Qt.AlignVCenter, mono=True,
                color=_RED if var < 0 else (_GREEN if var > 0 else ""),
            ))
            _finish_table_row(t, r)
            total_t += tah; total_i += inv; total_v += var; total_c += count
        self._totals.set_total("t", total_t, "USD ")
        self._totals.set_total("i", total_i, "USD ")
        self._totals.set_total("v", total_v, "USD ")
        self._totals.set_total("c", total_c, "")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if 0 <= row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


class _AfritrackUploadDetail(QWidget):
    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._search = ""
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        nh = QHBoxLayout(nav)
        nh.setContentsMargins(0, 0, 0, 0)
        nh.setSpacing(8)
        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        nh.addWidget(back_btn)
        self._crumb_lbl = _lbl("", size=12, color=_T2)
        nh.addWidget(self._crumb_lbl)
        nh.addStretch()
        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        nh.addWidget(delete_btn)
        self._export_btn = _btn("Export Upload", "mdi.download-outline", primary=False, height=30)
        self._export_btn.clicked.connect(self._export_current_upload)
        nh.addWidget(self._export_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tl = QHBoxLayout(tb)
        tl.setContentsMargins(0, 0, 0, 0)
        tl.setSpacing(8)
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search truck or remarks…")
        self._search_edit.setFixedWidth(280)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tl.addWidget(self._search_edit)
        tl.addStretch()
        vl.addWidget(tb)

        self._table = _make_table(_AF_DETAIL_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        widths = [56, 120, 70, 90, 70, 90, 120, 120, 100, 180]
        for i, w in enumerate(widths):
            self._table.setColumnWidth(i, w)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([
            ("t", "Tahmeed "),
            ("i", "Invoice "),
            ("v", "Variance "),
            ("c", "Rows "),
        ])
        vl.addWidget(self._totals)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        period = upload_doc.get("period") or "—"
        filename = upload_doc.get("source_filename") or "Unknown file"
        count = int(upload_doc.get("record_count", 0))
        self._crumb_lbl.setText(f"Uploads  ›  {period}")
        self._info_lbl.setText(f"{filename}   •   {period}   •   {count:,} rows")
        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._reset_and_load()

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_initial(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            recs, total, totals = await asyncio.gather(
                svc.get_afritrack_upload_records(self._upload_id, self._search, _SCROLL_CHUNK, 0),
                svc.count_afritrack_upload_records(self._upload_id, self._search),
                svc.get_afritrack_upload_totals(self._upload_id, self._search),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("t", float(totals.get("total_tahmeed", 0) or 0), "USD ")
        self._totals.set_total("i", float(totals.get("total_invoice", 0) or 0), "USD ")
        self._totals.set_total("v", float(totals.get("total_variance", 0) or 0), "USD ")
        self._totals.set_total("c", int(totals.get("count", total) or 0), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            recs = await svc.get_afritrack_upload_records(
                self._upload_id, self._search, _SCROLL_CHUNK, self._loaded,
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _afritrack_fill_row(self._table, r, rec)

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _export_current_upload(self) -> None:
        asyncio.ensure_future(self._do_export_current_upload())

    async def _do_export_current_upload(self) -> None:
        from tahmeed.services import accountant_service as svc
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Afritrack Upload",
            f"Afritrack_{(self._upload_doc.get('period') or 'upload').replace(' ', '_')}.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        recs = await svc.get_afritrack_upload_records(self._upload_id, "", 10000, 0)
        grid = _AfritrackGrid()
        grid.setRowCount(0)
        for rec in recs:
            grid.add_row([
                str(rec.get("row_index", "")),
                rec.get("truck", ""),
                _af_fmt(float(rec.get("days", 0) or 0), 0),
                _af_fmt(float(rec.get("non_trans_days", 0) or 0), 0),
                _af_fmt(float(rec.get("trans_days", 0) or 0), 0),
                _af_fmt(float(rec.get("rate_per_day", 0) or 0), 6),
                _af_fmt(float(rec.get("total_tahmeed", 0) or 0)),
                _af_fmt(float(rec.get("total_invoice", 0) or 0)),
                _af_fmt(float(rec.get("variance", 0) or 0)),
                rec.get("remarks", ""),
            ])
        _export_afritrack_xlsx(
            path,
            grid,
            float(self._upload_doc.get("installation_tahmeed", 0) or 0),
            float(self._upload_doc.get("installation_invoice", 0) or 0),
            float(self._upload_doc.get("balance_mar", 0) or 0),
            15.0,
            self._upload_doc.get("period") or "Upload",
        )
        QMessageBox.information(self, "Export Complete", f"Saved:\n{path}")


class AfritrackWidget(QWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader("Afritrack Schedule", "mdi.satellite-variant")
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        import_btn = _btn("Import from Excel", "mdi.upload-outline")
        import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Uploads"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _AfritrackAllEntries()
        self._main_stack.addWidget(self._all_entries)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _AfritrackUploadBrowse()
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _AfritrackUploadDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        self._tabs.tab_changed.connect(self._main_stack.setCurrentIndex)
        vl.addWidget(self._main_stack, 1)

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._show_browse()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(1, emit=False)
        self._main_stack.setCurrentIndex(1)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            _AF_TEMPLATE_FILENAME,
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, _AF_TEMPLATE_TITLE, _AF_DETAIL_HEADERS)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Afritrack Schedule", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)",
        )
        if not path:
            return
        asyncio.ensure_future(self._do_import(path))

    async def _do_import(self, path: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            rows, inst_t, inst_i, bal_mar = _read_afritrack_file(path)
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            return
        if not rows:
            QMessageBox.warning(self, "Nothing Found", "No Afritrack truck rows were found in that file.")
            return
        source_filename = Path(path).name
        period = Path(path).stem.replace("_", " ")
        upload_id = str(uuid.uuid4())
        records = _afritrack_rows_to_records(
            rows, period, source_filename, upload_id, inst_t, inst_i, bal_mar, 15.0
        )
        try:
            saved = await svc.save_imported_feed(records)
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            return
        QMessageBox.information(self, "Import Complete", f"Imported {saved:,} Afritrack rows.")
        self._all_entries.refresh()
        uploads = await svc.get_afritrack_uploads()
        doc = next((u for u in uploads if str(u.get("_id")) == upload_id), None)
        if doc is not None:
            self._show_detail(doc)
        else:
            self._show_browse()

    def _on_delete_upload(self, upload_doc: dict) -> None:
        count = int(upload_doc.get("record_count", 0))
        if QMessageBox.question(
            self, "Delete upload",
            f"Delete this Afritrack upload and its {count:,} rows?",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(str(upload_doc.get("_id") or "")))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            deleted = await svc.delete_afritrack_upload(upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._show_browse()


# ═══════════════════════════════════════════════════════════════════════════════
#  Insurance helpers — QB-style table, file readers, import dialog, exporters
# ═══════════════════════════════════════════════════════════════════════════════

_INS_MONTHS = [
    "All Months", "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY",
    "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
]


def _ins_make_table(headers: List[str]) -> QTableWidget:
    """Insurance / COMESA tables — same compact scheme as other separate expenses."""
    return _make_table(headers)


class _InsTotalsBar(QFrame):
    """Flexible totals footer bar for insurance widgets."""

    def __init__(
        self,
        labels: List[Tuple[str, str, str]],   # (key, display_label, num_prefix)
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.setFixedHeight(38)
        self.setStyleSheet(
            f"QFrame{{background:{_QB_HDR_BG};border-top:2px solid #BFDBFE;"
            "border-bottom-left-radius:6px;border-bottom-right-radius:6px;}}"
        )
        hl = QHBoxLayout(self)
        hl.setContentsMargins(16, 0, 16, 0)
        hl.setSpacing(32)

        self._map: Dict[str, Tuple[QLabel, str]] = {}
        for key, display_label, num_prefix in labels:
            sub = QWidget()
            sub.setStyleSheet("background:transparent;")
            sub_hl = QHBoxLayout(sub)
            sub_hl.setContentsMargins(0, 0, 0, 0)
            sub_hl.setSpacing(4)
            sub_hl.addWidget(_lbl(display_label + ":", size=11, weight=600,
                                   color=_QB_HDR_FG))
            val = _lbl("—", size=12, weight=700, color=_QB_BOLD_FG)
            sub_hl.addWidget(val)
            self._map[key] = (val, num_prefix)
            hl.addWidget(sub)
        hl.addStretch()

    def set_value(self, key: str, value: float, decimals: int = 0) -> None:
        if key in self._map:
            lbl, prefix = self._map[key]
            lbl.setText(_fmt_num(value, prefix=prefix, decimals=decimals))

    def set_text(self, key: str, text: str) -> None:
        if key in self._map:
            self._map[key][0].setText(text)


# ── COMESA file reader ─────────────────────────────────────────────────────────

def _read_comesa_rows(path: str) -> List[dict]:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import COMESA files.")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next(
        (wb[n] for n in wb.sheetnames if "comesa" in n.lower()),
        wb.active,
    )
    records: List[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            int(row[0])
        except (TypeError, ValueError):
            continue

        card_no = str(row[2] or "").strip() if len(row) > 2 else ""
        if not card_no:
            continue

        name      = str(row[1] or "").strip() if len(row) > 1 else ""
        vf        = _fmt_date(row[3]) if (len(row) > 3 and row[3]) else ""
        vt        = _fmt_date(row[4]) if (len(row) > 4 and row[4]) else ""
        truck_reg = str(row[5] or "").strip() if len(row) > 5 else ""
        raw_p     = row[6] if len(row) > 6 else None
        try:
            premium = float(raw_p) if raw_p is not None else 0.0
        except (TypeError, ValueError):
            premium = 0.0
        month = str(row[7] or "").strip().upper() if len(row) > 7 else ""

        records.append({
            "feed_type":  "comesa",
            "name":       name,
            "card_no":    card_no,
            "valid_from": vf,
            "valid_to":   vt,
            "truck_reg":  truck_reg,
            "premium":    premium,
            "month":      month,
            "dedup_id":   card_no,
        })
    return records


# ── Third Party file reader ────────────────────────────────────────────────────

def _read_third_party_rows(path: str) -> List[dict]:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required to import Third Party files.")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next(
        (wb[n] for n in wb.sheetnames
         if "third" in n.lower() or "party" in n.lower()),
        wb.active,
    )
    records: List[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        # ── PAID block: cols A(0)–H(7) ──────────────────────────────────────
        if row[0] is not None:
            try:
                int(row[0])
            except (TypeError, ValueError):
                pass
            else:
                reg_no = str(row[3] or row[2] or "").strip() if len(row) > 3 else ""
                name   = str(row[1] or "").strip() if len(row) > 1 else ""
                raw_p  = row[4] if len(row) > 4 else None
                month  = str(row[7] or "").strip().upper() if len(row) > 7 else ""
                if reg_no:
                    try:
                        prem = float(raw_p) if raw_p is not None else 0.0
                    except (TypeError, ValueError):
                        prem = 0.0
                    vat   = round(prem * 0.18, 2)
                    total = round(prem + vat, 2)
                    records.append({
                        "feed_type":     "third_party",
                        "name":          name,
                        "reg_no":        reg_no,
                        "premium":       prem,
                        "vat":           vat,
                        "total_premium": total,
                        "month":         month,
                        "status":        "PAID",
                        "dedup_id":      f"{reg_no}|{month}|PAID",
                    })

        # ── UNPAID block: cols L(11)–Q(16) ──────────────────────────────────
        if len(row) > 11 and row[11] is not None:
            name_u  = str(row[11] or "").strip()
            reg_u   = str(row[12] or "").strip() if len(row) > 12 else ""
            raw_pu  = row[13] if len(row) > 13 else None
            month_u = str(row[16] or "").strip().upper() if len(row) > 16 else ""
            if reg_u:
                try:
                    prem_u = float(raw_pu) if raw_pu is not None else 0.0
                except (TypeError, ValueError):
                    prem_u = 0.0
                vat_u   = round(prem_u * 0.18, 2)
                total_u = round(prem_u + vat_u, 2)
                records.append({
                    "feed_type":     "third_party",
                    "name":          name_u,
                    "reg_no":        reg_u,
                    "premium":       prem_u,
                    "vat":           vat_u,
                    "total_premium": total_u,
                    "month":         month_u,
                    "status":        "UNPAID",
                    "dedup_id":      f"{reg_u}|{month_u}|UNPAID",
                })
    return records


# ── Shared insurance import dialog ────────────────────────────────────────────

class _InsuranceImportDialog(QDialog):
    imported = Signal(int)

    def __init__(
        self,
        feed_type: str,
        reader_fn,
        preview_headers: List[str],
        preview_keys: List[str],
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self._feed_type      = feed_type
        self._reader_fn      = reader_fn
        self._preview_hdrs   = preview_headers
        self._preview_keys   = preview_keys
        self._new_rows: List[dict] = []
        self.setWindowTitle(f"Import — {feed_type.replace('_', ' ').title()}")
        self.setMinimumWidth(760)
        self.setStyleSheet(f"background:{_WHITE};")
        self._build()

    def _build(self) -> None:
        vl = QVBoxLayout(self)
        vl.setSpacing(12)
        vl.setContentsMargins(20, 20, 20, 20)

        self._drop = _DropZone()
        self._drop.file_dropped.connect(self._on_file)
        vl.addWidget(self._drop)

        browse_row = QWidget()
        browse_row.setStyleSheet("background:transparent;")
        brl = QHBoxLayout(browse_row)
        brl.setContentsMargins(0, 0, 0, 0)
        brl.setSpacing(8)
        browse_btn = _btn("Browse File", "mdi.folder-open-outline", primary=False)
        browse_btn.clicked.connect(self._browse)
        brl.addStretch()
        brl.addWidget(browse_btn)
        vl.addWidget(browse_row)

        self._stats_lbl = _lbl("No file loaded.", size=12, color=_T2)
        vl.addWidget(self._stats_lbl)

        vl.addWidget(_hsep())
        vl.addWidget(_lbl("Preview (first 10 rows)", size=12, weight=600))

        self._preview_tbl = _ins_make_table(self._preview_hdrs)
        self._preview_tbl.setMinimumHeight(200)
        self._preview_tbl.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeToContents
        )
        vl.addWidget(self._preview_tbl)

        vl.addWidget(_hsep())

        btn_row = QWidget()
        btn_row.setStyleSheet("background:transparent;")
        bbl = QHBoxLayout(btn_row)
        bbl.setContentsMargins(0, 0, 0, 0)
        bbl.addStretch()
        cancel_btn = _btn("Cancel", primary=False)
        cancel_btn.clicked.connect(self.reject)
        bbl.addWidget(cancel_btn)
        self._import_btn = _btn("Import Records", "mdi.check-circle-outline")
        self._import_btn.setEnabled(False)
        self._import_btn.clicked.connect(self._do_import)
        bbl.addWidget(self._import_btn)
        vl.addWidget(btn_row)

    def _browse(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self, "Open File", "", "Excel (*.xlsx *.xls)"
        )
        if path:
            self._drop.set_path(path)
            self._on_file(path)

    def _on_file(self, path: str) -> None:
        self._stats_lbl.setText("Reading file…")
        try:
            records = self._reader_fn(path)
        except Exception as exc:
            self._stats_lbl.setText(f"Error reading file: {exc}")
            return
        keys = [r["dedup_id"] for r in records if r.get("dedup_id")]
        asyncio.ensure_future(self._check_dupes(records, keys))

    async def _check_dupes(self, records: List[dict], keys: List[str]) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            existing = await svc.get_existing_insurance_keys(self._feed_type, keys)
        except Exception:
            existing = set()
        self._new_rows = [r for r in records if r.get("dedup_id") not in existing]
        dupes = len(records) - len(self._new_rows)
        self._stats_lbl.setText(
            f"New records: {len(self._new_rows):,}  ·  Duplicates (skipped): {dupes:,}"
        )
        self._import_btn.setEnabled(bool(self._new_rows))
        self._import_btn.setText(f"Import {len(self._new_rows):,} Records")
        self._fill_preview(self._new_rows[:10])

    def _fill_preview(self, rows: List[dict]) -> None:
        t = self._preview_tbl
        t.setRowCount(0)
        for row in rows:
            r = t.rowCount()
            t.insertRow(r)
            for c, key in enumerate(self._preview_keys):
                if c >= t.columnCount():
                    break
                val = row.get(key, "")
                if isinstance(val, float):
                    val = f"{val:,.0f}" if val else ""
                t.setItem(r, c, _cell(str(val)))
            _finish_table_row(t, r)

    def _do_import(self) -> None:
        self._import_btn.setEnabled(False)
        self._import_btn.setText("Importing…")
        asyncio.ensure_future(self._async_import())

    async def _async_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            saved = await svc.save_imported_feed(self._new_rows)
            self.imported.emit(saved)
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Import Error", str(exc))
            self._import_btn.setEnabled(True)
            self._import_btn.setText(f"Import {len(self._new_rows):,} Records")


# ── QB-style info card shared by both insurance widgets ───────────────────────

def _ins_info_card(title: str, icon_name: str) -> Tuple[QFrame, QLabel]:
    card = QFrame()
    card.setObjectName("ins_info_card")
    card.setStyleSheet(
        f"QFrame#ins_info_card{{background:{_QB_HDR_BG};"
        "border:1px solid #BFDBFE;border-radius:8px;}}"
    )
    hl = QHBoxLayout(card)
    hl.setContentsMargins(20, 14, 20, 14)
    hl.setSpacing(14)

    try:
        icon_lbl = QLabel()
        icon_lbl.setPixmap(qta.icon(icon_name, color=_QB_HDR_FG).pixmap(32, 32))
        icon_lbl.setFixedSize(32, 32)
        icon_lbl.setStyleSheet("background:transparent;")
        hl.addWidget(icon_lbl)
    except Exception:
        pass

    vl2 = QVBoxLayout()
    vl2.setSpacing(2)
    t_lbl = QLabel(title)
    t_lbl.setStyleSheet(
        f"color:{_QB_HDR_FG};font-size:18px;font-weight:700;"
        "font-family:'Segoe UI';background:transparent;"
    )
    vl2.addWidget(t_lbl)
    s_lbl = _lbl("Import records to get started", size=11, color=_T2)
    vl2.addWidget(s_lbl)
    hl.addLayout(vl2)
    hl.addStretch()
    return card, s_lbl


# ── COMESA Excel exporter ──────────────────────────────────────────────────────

def _export_comesa_xlsx(path: str, recs: List[dict]) -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export.")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "COMESA Covers"

    thin     = Side(border_style="thin", color="E5E7EB")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="EFF6FF")
    hdr_font = Font(name="Calibri", bold=True, size=11, color="1E3A5F")
    alt_fill = PatternFill("solid", fgColor="F9FAFB")
    nrm_font = Font(name="Calibri", size=11)

    col_widths = [6, 30, 14, 14, 16, 16, 14, 12]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w

    ws.row_dimensions[1].height = 30
    hdrs = ["S/NO", "NAME", "CARD NO.", "VALID FROM", "VALID TO",
            "TRUCK REG", "PREMIUM", "MONTH"]
    for c, label in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, rec in enumerate(recs, 2):
        ws.row_dimensions[ri].height = 18
        data = [
            ri - 1,
            rec.get("name", ""),
            rec.get("card_no", ""),
            rec.get("valid_from", ""),
            rec.get("valid_to", ""),
            rec.get("truck_reg", ""),
            rec.get("premium") or None,
            rec.get("month", ""),
        ]
        aligns = ["center", "left", "center", "center", "center",
                  "center", "right", "center"]
        for ci, (val, aln) in enumerate(zip(data, aligns), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = nrm_font
            cell.border = bdr
            cell.alignment = Alignment(horizontal=aln, vertical="center")
            if ri % 2 == 0:
                cell.fill = alt_fill

    wb.save(path)


# ── Third Party Excel exporter ────────────────────────────────────────────────

def _export_third_party_xlsx(path: str, recs: List[dict]) -> None:
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is required for Excel export.")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Third Party Covers"

    thin      = Side(border_style="thin", color="E5E7EB")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", fgColor="EFF6FF")
    hdr_font  = Font(name="Calibri", bold=True, size=11, color="1E3A5F")
    alt_fill  = PatternFill("solid", fgColor="F9FAFB")
    paid_fill = PatternFill("solid", fgColor="DCFCE7")
    upd_fill  = PatternFill("solid", fgColor="FEF3C7")
    nrm_font  = Font(name="Calibri", size=11)

    col_widths = [6, 30, 16, 14, 14, 16, 12, 10]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w

    ws.row_dimensions[1].height = 30
    hdrs = ["S/NO", "NAME", "REG. NO.", "PREMIUM", "VAT 18%",
            "TOTAL PREMIUM", "MONTH", "STATUS"]
    for c, label in enumerate(hdrs, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = bdr
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for ri, rec in enumerate(recs, 2):
        ws.row_dimensions[ri].height = 18
        status = rec.get("status", "")
        row_fill = (
            paid_fill if status == "PAID" else
            upd_fill  if status == "UNPAID" else
            (alt_fill if ri % 2 == 0 else None)
        )
        data = [
            ri - 1,
            rec.get("name", ""),
            rec.get("reg_no", ""),
            rec.get("premium") or None,
            rec.get("vat") or None,
            rec.get("total_premium") or None,
            rec.get("month", ""),
            status,
        ]
        aligns = ["center", "left", "center", "right", "right",
                  "right", "center", "center"]
        for ci, (val, aln) in enumerate(zip(data, aligns), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = nrm_font
            cell.border = bdr
            cell.alignment = Alignment(horizontal=aln, vertical="center")
            if row_fill:
                cell.fill = row_fill

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
#  COMESA Covers
# ═══════════════════════════════════════════════════════════════════════════════

_COMESA_HEADERS         = ["S/NO", "NAME", "CARD NO.", "VALID FROM",
                            "VALID TO", "TRUCK REG", "PREMIUM", "MONTH"]
_COMESA_PREVIEW_HEADERS = ["NAME", "CARD NO.", "VALID FROM", "VALID TO",
                            "TRUCK REG", "PREMIUM", "MONTH"]
_COMESA_PREVIEW_KEYS    = ["name", "card_no", "valid_from", "valid_to",
                            "truck_reg", "premium", "month"]


class ComesaWidget(QWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._page      = 1
        self._page_size = 50
        self._total     = 0
        self._search    = ""
        self._month     = ""
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 16, 20, 12)
        vl.setSpacing(8)

        # QB info card
        self._info_card, self._summary_lbl = _ins_info_card(
            "COMESA Covers", "mdi.certificate"
        )
        vl.addWidget(self._info_card)

        # Toolbar
        tb  = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search card no., truck, name…")
        self._search_edit.setFixedWidth(250)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._month_cb = QComboBox()
        for m in _INS_MONTHS:
            self._month_cb.addItem(m, "" if m == "All Months" else m)
        self._month_cb.setFixedWidth(155)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)
        tbl.addStretch()

        self._import_btn = _btn("Import", "mdi.upload-outline", height=32)
        self._export_btn = _btn("Export", "mdi.download-outline",
                                primary=False, height=32)
        self._import_btn.clicked.connect(self._do_import)
        self._export_btn.clicked.connect(self._do_export)
        tbl.addWidget(self._import_btn)
        tbl.addWidget(self._export_btn)
        vl.addWidget(tb)

        # Table card
        card = QFrame()
        card.setStyleSheet(
            f"QFrame{{background:{_WHITE};border:1px solid {_BORDER};"
            "border-radius:6px;}}"
        )
        cl = QVBoxLayout(card)
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(0)

        self._table = _ins_make_table(_COMESA_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeToContents
        )
        self._table.horizontalHeader().setStretchLastSection(True)
        cl.addWidget(self._table, 1)

        self._totals_bar = _InsTotalsBar([
            ("premium", "TOTAL PREMIUM", "TSh "),
            ("records", "RECORDS", ""),
        ])
        cl.addWidget(self._totals_bar)
        vl.addWidget(card, 1)

        self._pager = _PaginationBar()
        self._pager.page_changed.connect(self._go_page)
        self._pager.size_changed.connect(self._on_page_size)
        vl.addWidget(self._pager)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        skip = (self._page - 1) * self._page_size
        recs, total, totals = await asyncio.gather(
            svc.get_insurance_feed(
                "comesa", self._search, self._month, "",
                self._page_size, skip,
            ),
            svc.count_insurance_feed("comesa", self._search, self._month),
            svc.get_insurance_totals("comesa", self._month),
        )
        self._total = total
        self._fill_table(recs)
        self._pager.set_total(total, self._page_size, self._page)

        prem = totals.get("premium", 0.0)
        self._totals_bar.set_value("premium", prem)
        self._totals_bar.set_text("records", f"{total:,}")
        self._summary_lbl.setText(
            f"{total:,} records  ·  Total Premium: TSh {prem:,.0f}"
        )

    def _fill_table(self, recs: List[dict]) -> None:
        t   = self._table
        off = (self._page - 1) * self._page_size
        t.setRowCount(0)
        for idx, rec in enumerate(recs, off + 1):
            r = t.rowCount()
            t.insertRow(r)
            t.setItem(r, 0, _cell(str(idx),
                                   Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 1, _cell(rec.get("name", "")))
            t.setItem(r, 2, _cell(rec.get("card_no", ""), mono=True))
            t.setItem(r, 3, _cell(rec.get("valid_from", "")))
            t.setItem(r, 4, _cell(rec.get("valid_to", "")))
            t.setItem(r, 5, _cell(rec.get("truck_reg", ""), mono=True))
            prem = rec.get("premium", 0.0)
            t.setItem(r, 6, _cell(
                _fmt_num(prem, decimals=0) if prem else "—",
                Qt.AlignRight | Qt.AlignVCenter, mono=True,
            ))
            t.setItem(r, 7, _cell(rec.get("month", ""),
                                   Qt.AlignCenter | Qt.AlignVCenter))
            _finish_table_row(t, r)

    def _do_import(self) -> None:
        dlg = _InsuranceImportDialog(
            "comesa", _read_comesa_rows,
            _COMESA_PREVIEW_HEADERS, _COMESA_PREVIEW_KEYS,
            parent=self,
        )
        dlg.imported.connect(lambda n: (
            QMessageBox.information(
                self, "Import Complete", f"Imported {n:,} new records."
            ),
            self.refresh(),
        ))
        dlg.exec()

    def _do_export(self) -> None:
        if self._total == 0:
            QMessageBox.information(self, "Export", "No records to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export COMESA Covers", "COMESA_Covers.xlsx",
            "Excel (*.xlsx)",
        )
        if path:
            asyncio.ensure_future(self._async_export(path))

    async def _async_export(self, path: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            recs = await svc.get_insurance_feed(
                "comesa", self._search, self._month, "", 10000, 0
            )
            _export_comesa_xlsx(path, recs)
            QMessageBox.information(self, "Export Complete", f"Saved:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))

    def _on_search(self, text: str) -> None:
        self._search = text
        self._page   = 1
        asyncio.ensure_future(self._load())

    def _on_month(self) -> None:
        self._month = self._month_cb.currentData() or ""
        self._page  = 1
        asyncio.ensure_future(self._load())

    def _go_page(self, page: int) -> None:
        self._page = page
        asyncio.ensure_future(self._load())

    def _on_page_size(self, size: int) -> None:
        self._page_size = size
        self._page      = 1
        asyncio.ensure_future(self._load())


# ═══════════════════════════════════════════════════════════════════════════════
#  Third Party Covers
# ═══════════════════════════════════════════════════════════════════════════════

_TP_HEADERS         = ["S/NO", "NAME", "REG. NO.", "PREMIUM",
                        "VAT 18%", "TOTAL PREMIUM", "MONTH", "STATUS"]
_TP_PREVIEW_HEADERS = ["NAME", "REG. NO.", "PREMIUM", "VAT 18%",
                        "TOTAL PREMIUM", "MONTH", "STATUS"]
_TP_PREVIEW_KEYS    = ["name", "reg_no", "premium", "vat",
                        "total_premium", "month", "status"]

_STATUS_COLORS = {
    "PAID":   (_GREEN, _GREEN_L),
    "UNPAID": (_AMBER, _AMBER_L),
}


class ThirdPartyWidget(QWidget):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._page      = 1
        self._page_size = 50
        self._total     = 0
        self._search    = ""
        self._month     = ""
        self._status    = ""
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 16, 20, 12)
        vl.setSpacing(8)

        # QB info card
        self._info_card, self._summary_lbl = _ins_info_card(
            "Third Party Covers", "mdi.shield-account"
        )
        vl.addWidget(self._info_card)

        # Toolbar
        tb  = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search reg. no., name…")
        self._search_edit.setFixedWidth(220)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._month_cb = QComboBox()
        for m in _INS_MONTHS:
            self._month_cb.addItem(m, "" if m == "All Months" else m)
        self._month_cb.setFixedWidth(155)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._status_cb = QComboBox()
        for label, val in [("All Status", ""), ("PAID", "PAID"),
                            ("UNPAID", "UNPAID")]:
            self._status_cb.addItem(label, val)
        self._status_cb.setFixedWidth(120)
        self._status_cb.setStyleSheet(_input_ss())
        self._status_cb.currentIndexChanged.connect(self._on_status)
        tbl.addWidget(self._status_cb)
        tbl.addStretch()

        self._import_btn = _btn("Import", "mdi.upload-outline", height=32)
        self._export_btn = _btn("Export", "mdi.download-outline",
                                primary=False, height=32)
        self._import_btn.clicked.connect(self._do_import)
        self._export_btn.clicked.connect(self._do_export)
        tbl.addWidget(self._import_btn)
        tbl.addWidget(self._export_btn)
        vl.addWidget(tb)

        # Table card
        card = QFrame()
        card.setStyleSheet(
            f"QFrame{{background:{_WHITE};border:1px solid {_BORDER};"
            "border-radius:6px;}}"
        )
        cl = QVBoxLayout(card)
        cl.setContentsMargins(0, 0, 0, 0)
        cl.setSpacing(0)

        self._table = _ins_make_table(_TP_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeToContents
        )
        self._table.horizontalHeader().setStretchLastSection(True)
        cl.addWidget(self._table, 1)

        self._totals_bar = _InsTotalsBar([
            ("premium", "PREMIUM",      "TSh "),
            ("vat",     "VAT 18%",      "TSh "),
            ("total",   "TOTAL",        "TSh "),
            ("records", "RECORDS",      ""),
        ])
        cl.addWidget(self._totals_bar)
        vl.addWidget(card, 1)

        self._pager = _PaginationBar()
        self._pager.page_changed.connect(self._go_page)
        self._pager.size_changed.connect(self._on_page_size)
        vl.addWidget(self._pager)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        skip = (self._page - 1) * self._page_size
        recs, total, totals = await asyncio.gather(
            svc.get_insurance_feed(
                "third_party", self._search, self._month, self._status,
                self._page_size, skip,
            ),
            svc.count_insurance_feed(
                "third_party", self._search, self._month, self._status
            ),
            svc.get_insurance_totals(
                "third_party", self._month, self._status
            ),
        )
        self._total = total
        self._fill_table(recs)
        self._pager.set_total(total, self._page_size, self._page)

        prem = totals.get("premium", 0.0)
        vat  = totals.get("vat",     0.0)
        tot  = totals.get("total_premium", 0.0)
        self._totals_bar.set_value("premium", prem)
        self._totals_bar.set_value("vat",     vat)
        self._totals_bar.set_value("total",   tot)
        self._totals_bar.set_text("records",  f"{total:,}")
        self._summary_lbl.setText(
            f"{total:,} records  ·  Premium: TSh {prem:,.0f}"
            f"  ·  Total inc. VAT: TSh {tot:,.0f}"
        )

    def _fill_table(self, recs: List[dict]) -> None:
        t   = self._table
        off = (self._page - 1) * self._page_size
        t.setRowCount(0)
        for idx, rec in enumerate(recs, off + 1):
            r = t.rowCount()
            t.insertRow(r)
            t.setItem(r, 0, _cell(str(idx),
                                   Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 1, _cell(rec.get("name", "")))
            t.setItem(r, 2, _cell(rec.get("reg_no", ""), mono=True))
            prem = rec.get("premium", 0.0)
            vat  = rec.get("vat",     0.0)
            tot  = rec.get("total_premium", 0.0)
            t.setItem(r, 3, _cell(
                _fmt_num(prem, decimals=0) if prem else "—",
                Qt.AlignRight | Qt.AlignVCenter, mono=True,
            ))
            t.setItem(r, 4, _cell(
                _fmt_num(vat, decimals=0) if vat else "—",
                Qt.AlignRight | Qt.AlignVCenter, mono=True,
            ))
            t.setItem(r, 5, _cell(
                _fmt_num(tot, decimals=0) if tot else "—",
                Qt.AlignRight | Qt.AlignVCenter, mono=True,
            ))
            t.setItem(r, 6, _cell(rec.get("month", ""),
                                   Qt.AlignCenter | Qt.AlignVCenter))

            status = rec.get("status", "")
            s_item = _cell(status, Qt.AlignCenter | Qt.AlignVCenter)
            if status in _STATUS_COLORS:
                fg, bg = _STATUS_COLORS[status]
                s_item.setForeground(QColor(fg))
                s_item.setBackground(QColor(bg))
            t.setItem(r, 7, s_item)
            _finish_table_row(t, r)
            if status in _STATUS_COLORS:
                fg, bg = _STATUS_COLORS[status]
                item = t.item(r, 7)
                if item:
                    item.setForeground(QColor(fg))
                    item.setBackground(QColor(bg))

    def _do_import(self) -> None:
        dlg = _InsuranceImportDialog(
            "third_party", _read_third_party_rows,
            _TP_PREVIEW_HEADERS, _TP_PREVIEW_KEYS,
            parent=self,
        )
        dlg.imported.connect(lambda n: (
            QMessageBox.information(
                self, "Import Complete", f"Imported {n:,} new records."
            ),
            self.refresh(),
        ))
        dlg.exec()

    def _do_export(self) -> None:
        if self._total == 0:
            QMessageBox.information(self, "Export", "No records to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Third Party Covers", "ThirdParty_Covers.xlsx",
            "Excel (*.xlsx)",
        )
        if path:
            asyncio.ensure_future(self._async_export(path))

    async def _async_export(self, path: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            recs = await svc.get_insurance_feed(
                "third_party", self._search, self._month, self._status,
                10000, 0,
            )
            _export_third_party_xlsx(path, recs)
            QMessageBox.information(self, "Export Complete", f"Saved:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))

    def _on_search(self, text: str) -> None:
        self._search = text
        self._page   = 1
        asyncio.ensure_future(self._load())

    def _on_month(self) -> None:
        self._month = self._month_cb.currentData() or ""
        self._page  = 1
        asyncio.ensure_future(self._load())

    def _on_status(self) -> None:
        self._status = self._status_cb.currentData() or ""
        self._page   = 1
        asyncio.ensure_future(self._load())

    def _go_page(self, page: int) -> None:
        self._page = page
        asyncio.ensure_future(self._load())

    def _on_page_size(self, size: int) -> None:
        self._page_size = size
        self._page      = 1
        asyncio.ensure_future(self._load())


# ═══════════════════════════════════════════════════════════════════════════════
#  RahnTech — Transacted Devices import
# ═══════════════════════════════════════════════════════════════════════════════

_RAHNTECH_HEADERS = [
    "S/N", "SALES DATE", "TRIP NUMBER", "DEVICE NUMBER",
    "TRUCK NUMBER", "DRIVER NAME", "DO",
]
_RAHNTECH_TEMPLATE_TITLE = "RahnTech — Transacted Devices"
_RAHNTECH_TEMPLATE_FILENAME = "RahnTech_Import_Template.xlsx"
_RAHNTECH_BROWSE_HEADERS = [
    "UPLOAD DATE", "FILE NAME", "RECORDS", "DATE RANGE",
]
_RAHNTECH_COL_MAP = {
    "sn":            ["s/n", "sn", "#"],
    "sales_date":    ["sales date", "date", "transaction date"],
    "trip_number":   ["trip number", "trip no", "trip"],
    "device_number": ["device number", "device no", "device"],
    "truck_number":  ["truck number", "truck no", "truck"],
    "driver_name":   ["driver name", "driver"],
    "do_number":     ["do", "do number", "delivery order"],
}


def _read_rahntech_rows(path: str) -> Tuple[List[str], List[List[Any]]]:
    """RahnTech xlsx: row 0 is a title banner; row 1 holds the real headers."""
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xls") and _HAS_OPENPYXL:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return [], []
        headers = [str(c) if c is not None else "" for c in rows[1]]
        data = [
            [str(c) if c is not None else "" for c in r]
            for r in rows[2:]
            if any(c is not None for c in r)
        ]
        return headers, data
    return _read_file_rows(path)


class _RahnTechImportDialog(ImportDialog):
    """ImportDialog that uses the RahnTech-specific row reader (skips title row)."""

    def _on_file(self, path: str) -> None:
        self._stats_lbl.setText("Reading file…")
        self._source_filename = Path(path).name
        try:
            headers, rows = _read_rahntech_rows(path)
        except Exception as exc:
            self._stats_lbl.setText(f"Error reading file: {exc}")
            return

        self._raw_headers = headers
        hdr_lower = {h.strip().lower(): i for i, h in enumerate(headers)}

        def _find(candidates: List[str]) -> Optional[int]:
            for c in candidates:
                idx = hdr_lower.get(c.lower())
                if idx is not None:
                    return idx
            return None

        field_idxs: Dict[str, Optional[int]] = {
            key: _find(cands) for key, cands in self._col_map.items()
        }

        records: List[dict] = []
        for row in rows:
            rec: dict = {
                "_raw": row,
                "feed_type":       self._feed_type,
                "upload_id":       self._upload_id,
                "source_filename": self._source_filename,
            }
            for key, idx in field_idxs.items():
                rec[key] = str(row[idx]).strip() if (idx is not None and idx < len(row)) else ""
            records.append(rec)

        self._all_rows = records
        dedup_vals = [r.get(self._dedup_key, "") for r in records if r.get(self._dedup_key)]
        asyncio.ensure_future(self._check_dupes(records, dedup_vals))


def _rahntech_fill_row(t: QTableWidget, r: int, rec: dict) -> None:
    t.setItem(r, 0, _cell(rec.get("sn", "")))
    t.setItem(r, 1, _cell(rec.get("sales_date", "")))
    t.setItem(r, 2, _cell(rec.get("trip_number", "")))
    t.setItem(r, 3, _cell(rec.get("device_number", "")))
    t.setItem(r, 4, _cell(rec.get("truck_number", "")))
    t.setItem(r, 5, _cell(rec.get("driver_name", "")))
    t.setItem(r, 6, _cell(rec.get("do_number", "")))
    _finish_table_row(t, r)


class _RahnTechAllEntries(QWidget):
    """Flat, filterable list of every RahnTech record with infinite scroll."""

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._search = ""
        self._year = 0
        self._month = 0
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)

        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search truck, driver, trip, device…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tbl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _TOLL_MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tbl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tbl, self._reset_and_load, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        tbl.addStretch()
        vl.addWidget(tb)

        self._totals = _TotalsBar([("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_table(_RAHNTECH_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def refresh(self) -> None:
        asyncio.ensure_future(self._reload_years_and_data())

    async def _reload_years_and_data(self) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            years = await svc.get_rahntech_available_years()
        except Exception:
            years = []
        self._year = _populate_year_combo(self._year_cb, years, self._year)
        if self._year <= 0:
            self._month = 0
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._month_cb.blockSignals(False)
        self._reset_and_load()

    def _effective_month(self) -> int:
        return self._month if self._year > 0 else 0

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_initial(self) -> None:
        if self._loading:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            totals, recs, total = await asyncio.gather(
                svc.get_rahntech_all_totals(self._search, self._year, month, **self._date_kw()),
                svc.get_rahntech_all_records(self._search, self._year, month, limit=_SCROLL_CHUNK, skip=0, **self._date_kw()),
                svc.count_rahntech_all_records(self._search, self._year, month, **self._date_kw()),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("count", int(totals.get("count", 0)), "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        if self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        from tahmeed.services import accountant_service as svc
        month = self._effective_month()
        try:
            recs = await svc.get_rahntech_all_records(
                self._search, self._year, month,
                limit=_SCROLL_CHUNK, skip=self._loaded, **self._date_kw(),
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _rahntech_fill_row(self._table, r, rec)

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _date_kw(self) -> dict:
        if not hasattr(self, "_from_date"):
            return {}
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        if hasattr(self, "_from_date"):
            sync_from_to(
                self._from_date, self._to_date, self._year, self._month, optional=True,
            )
        self._reset_and_load()


class _RahnTechUploadBrowse(QWidget):
    """Table of every RahnTech import batch. Clicking a row drills into it."""

    upload_clicked = Signal(object)
    delete_clicked = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._uploads: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        self._totals = _TotalsBar([("count", "Total records: ")])
        vl.addWidget(self._totals)

        self._table = _make_table(_RAHNTECH_BROWSE_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.setColumnWidth(0, 160)
        self._table.setColumnWidth(1, 260)
        self._table.setColumnWidth(2, 80)
        self._table.setStyleSheet(_table_style())
        self._table.setCursor(Qt.PointingHandCursor)
        self._table.cellClicked.connect(self._on_row_clicked)
        self._table.setContextMenuPolicy(Qt.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_menu)
        vl.addWidget(self._table, 1)

        hint = _lbl("Click any row to view its records · right-click to delete an upload.", size=11, color=_TM)
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        from tahmeed.services import accountant_service as svc
        uploads = await svc.get_rahntech_uploads()
        self._uploads = uploads
        self._fill(uploads)

    def _fill(self, uploads: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        total_recs = 0
        for up in uploads:
            r = t.rowCount()
            t.insertRow(r)
            import_dt = up.get("import_date")
            date_str = (
                import_dt.strftime("%d %b %Y  %H:%M")
                if isinstance(import_dt, datetime)
                else (str(import_dt) if import_dt else "—")
            )
            count = int(up.get("record_count", 0))
            min_d = str(up.get("min_sales_date", "") or "").strip()
            max_d = str(up.get("max_sales_date", "") or "").strip()
            date_range = f"{min_d[:10]} — {max_d[:10]}" if min_d else "—"

            t.setItem(r, 0, _cell(date_str))
            t.setItem(r, 1, _cell(up.get("source_filename") or "Unknown"))
            t.setItem(r, 2, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 3, _cell(date_range))
            _finish_table_row(t, r)
            total_recs += count

        self._totals.set_total("count", total_recs, "")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if row < len(self._uploads):
            self.upload_clicked.emit(self._uploads[row])

    def _on_menu(self, pos) -> None:
        row = self._table.rowAt(pos.y())
        if not (0 <= row < len(self._uploads)):
            return
        menu = QMenu(self)
        act = menu.addAction("Delete this upload")
        if menu.exec(self._table.viewport().mapToGlobal(pos)) == act:
            self.delete_clicked.emit(self._uploads[row])


class _RahnTechUploadDetail(QWidget):
    """Full record table for a single RahnTech upload batch."""

    back_requested = Signal()
    delete_requested = Signal(object)

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._upload_id = ""
        self._upload_doc: dict = {}
        self._search = ""
        self._loaded = 0
        self._total = 0
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background:transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background:transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)
        back_btn = _btn("← All Uploads", primary=False, height=30)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)
        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()
        delete_btn = _btn("Delete Upload", "mdi.trash-can-outline", danger=True, height=30)
        delete_btn.clicked.connect(self._request_delete)
        navl.addWidget(delete_btn)
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600, color=_T1)
        vl.addWidget(self._info_lbl)

        tb = QWidget()
        tb.setStyleSheet("background:transparent;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(0, 0, 0, 0)
        tbl.setSpacing(8)
        self._search_edit = QLineEdit()
        self._search_edit.setPlaceholderText("Search truck, driver, trip, device…")
        self._search_edit.setFixedWidth(300)
        self._search_edit.setStyleSheet(_input_ss())
        self._search_edit.textChanged.connect(self._on_search)
        tbl.addWidget(self._search_edit)
        tbl.addStretch()
        vl.addWidget(tb)

        self._table = _make_table(_RAHNTECH_HEADERS)
        self._table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        vl.addWidget(self._table, 1)

        self._totals = _TotalsBar([("count", "Records: ")])
        vl.addWidget(self._totals)

        self._status_lbl = _lbl("", size=11, color=_TM)
        self._status_lbl.setAlignment(Qt.AlignCenter)
        vl.addWidget(self._status_lbl)

    def load_upload(self, upload_doc: dict) -> None:
        self._upload_doc = upload_doc
        self._upload_id = str(upload_doc.get("_id") or "")
        filename = upload_doc.get("source_filename") or "Unknown file"
        count = int(upload_doc.get("record_count", 0))
        import_dt = upload_doc.get("import_date")
        date_str = (
            import_dt.strftime("%d %b %Y")
            if isinstance(import_dt, datetime) else ""
        )
        self._crumb_lbl.setText(f"Uploads  ›  {filename}")
        self._info_lbl.setText(
            f"{filename}   •   {count:,} records   •   {date_str}"
        )
        self._totals.set_total("count", count, "")

        self._search = ""
        self._search_edit.blockSignals(True)
        self._search_edit.setText("")
        self._search_edit.blockSignals(False)
        self._reset_and_load()

    def _request_delete(self) -> None:
        if self._upload_doc:
            self.delete_requested.emit(self._upload_doc)

    def _reset_and_load(self) -> None:
        self._loaded = 0
        self._total = 0
        self._table.setRowCount(0)
        asyncio.ensure_future(self._load_initial())

    def _update_status(self) -> None:
        if self._loading:
            suffix = "  •  Loading…"
        elif self._loaded >= self._total:
            suffix = ""
        else:
            suffix = "  •  Scroll down for more"
        self._status_lbl.setText(f"Showing {self._loaded:,} of {self._total:,}{suffix}")

    async def _load_initial(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading:
            return
        self._loading = True
        self._update_status()
        try:
            recs, total = await asyncio.gather(
                svc.get_rahntech_upload_records(
                    self._upload_id, self._search, _SCROLL_CHUNK, 0
                ),
                svc.count_rahntech_upload_records(self._upload_id, self._search),
            )
        except Exception:
            self._loading = False
            self._status_lbl.setText("Failed to load records.")
            return
        self._total = total
        self._append_rows(recs)
        self._loaded = len(recs)
        self._totals.set_total("count", total, "")
        self._loading = False
        self._update_status()

    async def _load_more(self) -> None:
        from tahmeed.services import accountant_service as svc
        if not self._upload_id or self._loading or self._loaded >= self._total:
            return
        self._loading = True
        self._update_status()
        try:
            recs = await svc.get_rahntech_upload_records(
                self._upload_id, self._search, _SCROLL_CHUNK, self._loaded
            )
        except Exception:
            self._loading = False
            self._update_status()
            return
        if recs:
            self._append_rows(recs)
            self._loaded += len(recs)
        self._loading = False
        self._update_status()

    def _append_rows(self, recs: List[dict]) -> None:
        for rec in recs:
            r = self._table.rowCount()
            self._table.insertRow(r)
            _rahntech_fill_row(self._table, r, rec)

    def _on_search(self, text: str) -> None:
        self._search = text
        self._reset_and_load()

    def _on_scroll(self, value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if value >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())


class RahnTechWidget(QWidget):
    """
    RahnTech main page — browse uploads then drill into per-record detail.
    """

    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background:{_BG};")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(20, 20, 20, 16)
        vl.setSpacing(12)

        header = _PageHeader("RahnTech", "mdi.devices")
        tmpl_btn = _btn("Download Template", "mdi.download-outline", primary=False)
        tmpl_btn.clicked.connect(self._download_template)
        self._import_btn = _btn("Import Transacted Devices", "mdi.upload-outline")
        self._import_btn.clicked.connect(self._open_import)
        header.add_right(tmpl_btn)
        header.add_right(self._import_btn)
        vl.addWidget(header)
        vl.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Uploads"])
        vl.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background:transparent;")

        self._all_entries = _RahnTechAllEntries()
        self._main_stack.addWidget(self._all_entries)

        upload_host = QWidget()
        upload_host.setStyleSheet("background:transparent;")
        upload_vl = QVBoxLayout(upload_host)
        upload_vl.setContentsMargins(0, 0, 0, 0)
        upload_vl.setSpacing(0)

        self._upload_stack = QStackedWidget()
        self._upload_stack.setStyleSheet("background:transparent;")

        self._browse = _RahnTechUploadBrowse()
        self._browse.upload_clicked.connect(self._show_detail)
        self._browse.delete_clicked.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._browse)

        self._detail = _RahnTechUploadDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._detail.delete_requested.connect(self._on_delete_upload)
        self._upload_stack.addWidget(self._detail)

        upload_vl.addWidget(self._upload_stack, 1)
        self._main_stack.addWidget(upload_host)

        self._tabs.tab_changed.connect(self._main_stack.setCurrentIndex)
        vl.addWidget(self._main_stack, 1)

    def refresh(self) -> None:
        self._all_entries.refresh()
        self._show_browse()

    def _show_browse(self) -> None:
        self._upload_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, upload_doc: dict) -> None:
        self._tabs.set_index(1, emit=False)
        self._main_stack.setCurrentIndex(1)
        self._upload_stack.setCurrentIndex(1)
        self._detail.load_upload(upload_doc)

    def _download_template(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Import Template",
            _RAHNTECH_TEMPLATE_FILENAME,
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        try:
            _write_xlsx_template(path, _RAHNTECH_TEMPLATE_TITLE, _RAHNTECH_HEADERS)
            QMessageBox.information(
                self, "Template Saved",
                f"Empty template saved to:\n{path}\n\n"
                "Fill in your data using the column headers shown, then import the file.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Template Error", str(exc))

    def _open_import(self) -> None:
        from tahmeed.services import accountant_service as svc
        dlg = _RahnTechImportDialog(
            feed_type="rahntech",
            dedup_key="trip_number",
            preview_headers=_RAHNTECH_HEADERS,
            col_map=_RAHNTECH_COL_MAP,
            save_fn=svc.save_imported_feed,
            exist_fn=svc.get_existing_feed_keys,
            parent=self,
        )
        dlg.imported.connect(self._on_imported)
        dlg.exec()

    def _on_imported(self, n: int) -> None:
        QMessageBox.information(self, "Import Complete", f"Imported {n:,} new records.")
        self._all_entries.refresh()
        self._show_browse()
        self._tabs.set_index(1)

    def _on_delete_upload(self, upload_doc: dict) -> None:
        upload_id = str(upload_doc.get("_id") or "")
        if not upload_id:
            return
        filename = upload_doc.get("source_filename") or "Unknown file"
        count = int(upload_doc.get("record_count", 0))
        if QMessageBox.question(
            self,
            "Delete Upload",
            f"Delete upload from \"{filename}\" and all {count:,} records?\n\nThis cannot be undone.",
        ) != QMessageBox.Yes:
            return
        asyncio.ensure_future(self._delete_upload(upload_id))

    async def _delete_upload(self, upload_id: str) -> None:
        from tahmeed.services import accountant_service as svc
        try:
            deleted = await svc.delete_rahntech_upload(upload_id)
        except Exception as exc:
            QMessageBox.critical(self, "Delete Error", str(exc))
            return
        if deleted <= 0:
            QMessageBox.warning(self, "Delete Upload", "No records were deleted.")
            return
        QMessageBox.information(
            self, "Upload Deleted", f"Removed {deleted:,} record{'s' if deleted != 1 else ''}."
        )
        self._all_entries.refresh()
        self._show_browse()
