"""Diesel Cash — verified daily transactions for selected diesel items.

Toll Plaza–style layout:
  All Entries tab  — flat list with year/month + From/To date filters
  Transactions tab — FY month browse → drill-down, also with month + From/To
  Items tab        — pick which item names this view catches
"""

from __future__ import annotations

import asyncio
from datetime import datetime
from typing import List, Optional

import qtawesome as qta

from PySide6.QtCore import Qt, QSize, QTimer, Signal, QDate
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFrame, QLabel,
    QHeaderView,
    QLineEdit, QComboBox, QPushButton, QMessageBox, QFileDialog,
    QStackedWidget, QListWidget, QListWidgetItem, QAbstractItemView,
)

from tahmeed.app_state import app_state
from tahmeed.models.transaction import Transaction
from tahmeed.services.diesel_liters import format_liters, parse_liters_from_description
from tahmeed.ui.accountant.category_tables import (
    _WHITE, _BG, _BORDER, _BLUE, _GREEN,
    _T1, _T2, _TM, _RED, _lbl, _input_ss, _btn, _receipt_text,
)
from tahmeed.ui.accountant.date_filters import (
    add_from_to_editors, read_from_to, sync_from_to,
)
from tahmeed.ui.accountant.separate_expenses import (
    _make_table, _cell, _finish_table_row, _SegmentTabBar, _hsep,
)
from tahmeed.ui.accountant.feed_sort_helpers import (
    DIESEL_CASH_SORT, wire_feed_table_sort, sort_kw, reset_feed_sort,
)

_MONTH_SHORT = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

_BROWSE_HEADERS = ["MONTH", "RECORDS", "TOTAL (TZS)", "DATE RANGE"]
_PAGE_SIZES = [25, 50, 100]
_MONTHS = [
    ("All Months", 0),
    ("January", 1), ("February", 2), ("March", 3), ("April", 4),
    ("May", 5), ("June", 6), ("July", 7), ("August", 8),
    ("September", 9), ("October", 10), ("November", 11), ("December", 12),
]

# Local column set — LTRS sits after DESCRIPTION. Parsed from the description
# at display time; the stored transaction is not rewritten.
_DIESEL_COLS = [
    ("S/NO",        52,  "center", False),
    ("DATE",        100, "left",   False),
    ("ITEM",        110, "left",   False),
    ("DESCRIPTION", 260, "left",   False),
    ("LTRS",        72,  "right",  True),
    ("TRUCK NO.",   95,  "left",   False),
    ("MEMO",        140, "left",   False),
    ("NOTES",       56,  "center", False),
    ("TZS",         120, "right",  True),
    ("RECEIPT",     110, "center", False),
    ("OWNERSHIP",   95,  "left",   False),
    ("APR BY",      90,  "left",   False),
]
_EXPORT_SPAN = "A1:L1"


def _fmt_date_range(min_d, max_d) -> str:
    if not min_d or not max_d:
        return "—"
    if isinstance(min_d, datetime) and isinstance(max_d, datetime):
        a = min_d.strftime("%d %b %Y")
        b = max_d.strftime("%d %b %Y")
        return a if a == b else f"{a} — {b}"
    return "—"


def _fill_diesel_entry_row(table, r: int, tx: Transaction, sno: int) -> None:
    """Paint one Diesel Cash ledger row, including litres parsed from description."""
    liters = parse_liters_from_description(tx.description or "")
    liters_txt = format_liters(liters)
    table.setItem(r, 0, _cell(str(sno), align=Qt.AlignCenter | Qt.AlignVCenter))
    table.setItem(r, 1, _cell(
        tx.date.strftime("%d %b %Y") if tx.date else "—",
    ))
    table.setItem(r, 2, _cell(tx.item or "—"))
    table.setItem(r, 3, _cell(tx.description or "—"))
    table.setItem(r, 4, _cell(
        liters_txt,
        align=Qt.AlignRight | Qt.AlignVCenter,
        color=_TM if liters is None else _T1,
    ))
    table.setItem(r, 5, _cell(tx.truck_number or "—"))
    table.setItem(r, 6, _cell(tx.memo or "—"))
    table.setItem(r, 7, _cell(
        "✓" if tx.notes_flag else "—",
        align=Qt.AlignCenter | Qt.AlignVCenter,
        color=_BLUE if tx.notes_flag else _TM,
    ))
    if tx.currency == "TZS":
        tzs_txt = f"{tx.amount:,.0f}"
        tzs_col = _RED if tx.amount < 0 else _T1
    else:
        tzs_txt, tzs_col = "—", _TM
    table.setItem(r, 8, _cell(
        tzs_txt, align=Qt.AlignRight | Qt.AlignVCenter, color=tzs_col,
    ))
    rcpt_text, rcpt_fg = _receipt_text(tx.receipt_status or "pending")
    table.setItem(r, 9, _cell(
        rcpt_text, align=Qt.AlignCenter | Qt.AlignVCenter, color=rcpt_fg,
    ))
    table.setItem(r, 10, _cell(tx.ownership or "—"))
    table.setItem(r, 11, _cell(tx.approver or "—"))
    _finish_table_row(table, r)


def _diesel_export_row(tx: Transaction) -> list:
    liters = parse_liters_from_description(tx.description or "")
    tzs_val = tx.amount if tx.currency == "TZS" else None
    receipt_str = {"received": "Received", "pending": "Pending",
                   "missing": "No Receipt"}.get(tx.receipt_status or "", "—")
    return [
        None,
        tx.date.strftime("%d %b %Y") if tx.date else "",
        tx.item or "",
        tx.description or "",
        liters if liters is not None else None,
        tx.truck_number or "",
        tx.memo or "",
        "Yes" if tx.notes_flag else "",
        tzs_val,
        receipt_str,
        tx.ownership or "",
        tx.approver or "",
    ]


# ═══════════════════════════════════════════════════════════════════════════════
#  All Entries — flat filtered list with From/To
# ═══════════════════════════════════════════════════════════════════════════════

class _DieselCashAllEntries(QWidget):
    """All verified Diesel Cash transactions with date filters."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._year = 0
        self._month = 0
        self._page = 0
        self._total = 0
        self._loading = False

        self._debounce = QTimer(self)
        self._debounce.setSingleShot(True)
        self._debounce.setInterval(350)
        self._debounce.timeout.connect(lambda: asyncio.ensure_future(self._reload()))

        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background: transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        toolbar = QWidget()
        toolbar.setStyleSheet("background: transparent;")
        tl = QHBoxLayout(toolbar)
        tl.setContentsMargins(0, 0, 0, 0)
        tl.setSpacing(10)

        self._search = QLineEdit()
        self._search.setPlaceholderText("Search description or truck…")
        self._search.setFixedWidth(240)
        self._search.setStyleSheet(_input_ss())
        self._search.textChanged.connect(self._on_filter_changed)
        tl.addWidget(self._search)

        self._year_cb = QComboBox()
        self._year_cb.addItem("All Years", 0)
        current_yr = datetime.now().year
        for yr in range(current_yr - 3, current_yr + 2):
            self._year_cb.addItem(str(yr), yr)
        self._year_cb.setFixedWidth(110)
        self._year_cb.setStyleSheet(_input_ss())
        self._year_cb.currentIndexChanged.connect(self._on_year)
        tl.addWidget(self._year_cb)

        self._month_cb = QComboBox()
        for label, val in _MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.setEnabled(False)
        self._month_cb.currentIndexChanged.connect(self._on_month)
        tl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tl, self._on_filter_changed, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        self._rcpt_cb = QComboBox()
        for label, val in [("All Receipts", "all"), ("Received", "received"),
                           ("Pending", "pending"), ("No Receipt", "missing")]:
            self._rcpt_cb.addItem(label, val)
        self._rcpt_cb.setFixedWidth(128)
        self._rcpt_cb.setStyleSheet(_input_ss())
        self._rcpt_cb.currentIndexChanged.connect(self._on_filter_changed)
        tl.addWidget(self._rcpt_cb)

        clear_btn = _btn("Clear", "mdi.filter-remove-outline", primary=False)
        clear_btn.setToolTip("Clear search, date, receipt filters, and column sort.")
        clear_btn.clicked.connect(self._clear_filters)
        tl.addWidget(clear_btn)
        tl.addStretch()

        export_btn = _btn("Export Excel", "mdi.microsoft-excel")
        export_btn.clicked.connect(self._on_export)
        tl.addWidget(export_btn)
        vl.addWidget(toolbar)

        self._table = _make_table([c[0] for c in _DIESEL_COLS])
        self._table.setShowGrid(True)
        hdr = self._table.horizontalHeader()
        hdr.setSectionsMovable(False)
        hdr.setStretchLastSection(True)
        for i, (_, width, _a, _m) in enumerate(_DIESEL_COLS):
            self._table.setColumnWidth(i, width)
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        self._sort_state = wire_feed_table_sort(
            self._table,
            DIESEL_CASH_SORT,
            default_field="date",
            default_asc=False,
            on_sort_changed=self._on_sort_changed,
        )
        vl.addWidget(self._table, 1)

        footer = QFrame()
        footer.setFixedHeight(40)
        footer.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 2px solid {_BORDER}; }}"
        )
        fol = QHBoxLayout(footer)
        fol.setContentsMargins(16, 0, 16, 0)
        fol.setSpacing(20)
        fol.addWidget(_lbl("TOTAL (filtered view)", size=11, weight=700, color=_T2))
        self._tzs_lbl = _lbl("TZS  —", size=13, weight=700)
        self._tzs_lbl.setStyleSheet(
            f"color: {_T1}; font-size: 13px; font-weight: 700;"
            " font-family: 'Cascadia Code', 'Consolas', monospace; background: transparent;"
        )
        fol.addWidget(self._tzs_lbl)
        fol.addStretch()
        self._footer_count = _lbl("", size=11, color=_T2)
        fol.addWidget(self._footer_count)
        vl.addWidget(footer)

        pager = QFrame()
        pager.setFixedHeight(44)
        pager.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 1px solid {_BORDER}; }}"
        )
        pl = QHBoxLayout(pager)
        pl.setContentsMargins(16, 0, 16, 0)
        pl.setSpacing(10)
        self._size_cb = QComboBox()
        for sz in _PAGE_SIZES:
            self._size_cb.addItem(f"Show {sz}", sz)
        self._size_cb.setFixedWidth(100)
        self._size_cb.setStyleSheet(_input_ss())
        self._size_cb.currentIndexChanged.connect(self._on_size_changed)
        pl.addWidget(self._size_cb)
        self._page_info = _lbl("—", size=12, color=_T2)
        pl.addWidget(self._page_info)
        pl.addStretch()
        self._prev_btn = _btn("← Prev", "mdi.chevron-left")
        self._prev_btn.setFixedWidth(88)
        self._prev_btn.clicked.connect(self._on_prev)
        pl.addWidget(self._prev_btn)
        self._next_btn = _btn("Next →", "mdi.chevron-right")
        self._next_btn.setFixedWidth(88)
        self._next_btn.clicked.connect(self._on_next)
        pl.addWidget(self._next_btn)
        vl.addWidget(pager)

    def refresh(self) -> None:
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _search_text(self) -> str:
        return self._search.text().strip()

    def _receipt_filter(self) -> str:
        return self._rcpt_cb.currentData() or "all"

    def _page_size(self) -> int:
        return self._size_cb.currentData() or _PAGE_SIZES[0]

    def _on_sort_changed(self, field: str, asc: bool) -> None:
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _date_kw(self) -> dict:
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _filter_kw(self) -> dict:
        month = self._month if self._year > 0 else 0
        return dict(
            year=self._year or None,
            month=month,
            search=self._search_text(),
            truck="",
            receipt=self._receipt_filter(),
            **self._date_kw(),
        )

    def _on_filter_changed(self) -> None:
        self._page = 0
        self._debounce.start()

    def _clear_filters(self) -> None:
        self._search.blockSignals(True)
        self._year_cb.blockSignals(True)
        self._month_cb.blockSignals(True)
        self._rcpt_cb.blockSignals(True)
        try:
            self._search.clear()
            self._year_cb.setCurrentIndex(0)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.setEnabled(False)
            self._rcpt_cb.setCurrentIndex(0)
        finally:
            self._search.blockSignals(False)
            self._year_cb.blockSignals(False)
            self._month_cb.blockSignals(False)
            self._rcpt_cb.blockSignals(False)
        self._year = 0
        self._month = 0
        sync_from_to(self._from_date, self._to_date, 0, 0, optional=True)
        reset_feed_sort(self._sort_state)
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _on_year(self, _idx: int) -> None:
        self._year = int(self._year_cb.currentData() or 0)
        has_year = self._year > 0
        self._month_cb.setEnabled(has_year)
        if not has_year:
            self._month_cb.blockSignals(True)
            self._month_cb.setCurrentIndex(0)
            self._month_cb.blockSignals(False)
            self._month = 0
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _on_month(self, _idx: int) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _on_size_changed(self) -> None:
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _on_prev(self) -> None:
        if self._page > 0:
            self._page -= 1
            asyncio.ensure_future(self._reload())

    def _on_next(self) -> None:
        size = self._page_size()
        max_pg = max(0, (self._total - 1) // size) if self._total else 0
        if self._page < max_pg:
            self._page += 1
            asyncio.ensure_future(self._reload())

    async def _reload(self) -> None:
        if self._loading:
            return
        self._loading = True
        try:
            from tahmeed.services.accountant_service import (
                get_diesel_cash_transactions, count_diesel_cash_transactions,
                get_diesel_cash_totals,
            )
            size = self._page_size()
            skip = self._page * size
            kw = self._filter_kw()
            txs, total, totals = await asyncio.gather(
                get_diesel_cash_transactions(
                    **kw, **sort_kw(self._sort_state), limit=size, skip=skip,
                ),
                count_diesel_cash_transactions(**kw),
                get_diesel_cash_totals(**kw),
            )
            self._total = total
            self._populate(txs, skip)
            self._tzs_lbl.setText(f"TZS  {totals['tzs']:,.0f}")
            self._tzs_lbl.setStyleSheet(
                f"color: {_RED if totals['tzs'] < 0 else _T1};"
                " font-size: 13px; font-weight: 700;"
                " font-family: 'Cascadia Code', 'Consolas', monospace; background: transparent;"
            )
            self._footer_count.setText(f"{total:,} records")
            self._update_pager(total, size)
        except Exception as exc:
            self._table.setRowCount(0)
            self._page_info.setText(f"Failed to load: {exc}")
        finally:
            self._loading = False

    def _populate(self, txs: List[Transaction], skip: int) -> None:
        t = self._table
        t.setRowCount(0)
        for i, tx in enumerate(txs):
            r = t.rowCount()
            t.insertRow(r)
            _fill_diesel_entry_row(t, r, tx, skip + i + 1)

    def _update_pager(self, total: int, size: int) -> None:
        max_pg = max(0, (total - 1) // size) if total else 0
        self._prev_btn.setEnabled(self._page > 0)
        self._next_btn.setEnabled(self._page < max_pg)
        start = self._page * size + 1 if total else 0
        end = min((self._page + 1) * size, total)
        self._page_info.setText(
            f"Showing {start:,}–{end:,} of {total:,}  ·  Page {self._page + 1} of {max_pg + 1}"
        )

    def _on_export(self) -> None:
        asyncio.ensure_future(self._do_export())

    async def _do_export(self) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n\nRun: pip install openpyxl",
            )
            return

        from tahmeed.services.accountant_service import get_diesel_cash_transactions
        kw = self._filter_kw()
        try:
            txs = await get_diesel_cash_transactions(
                **kw, **sort_kw(self._sort_state), limit=10_000, skip=0,
            )
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", f"Failed to fetch data: {exc}")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diesel Cash"
        ws.merge_cells(_EXPORT_SPAN)
        ws["A1"] = "TAHMEED COACH TZ LTD"
        ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1B2B4B")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:L2")
        ws["A2"] = "Diesel Cash — All Entries"
        ws["A2"].font = Font(name="Segoe UI", bold=True, size=11, color="374151")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append([c[0] for c in _DIESEL_COLS])
        hdr_row = ws.max_row
        grey = PatternFill("solid", fgColor="F1F5F9")
        for cell in ws[hdr_row]:
            cell.font = Font(name="Segoe UI", bold=True, size=10, color="6B7280")
            cell.fill = grey
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for tx in txs:
            ws.append(_diesel_export_row(tx))

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Diesel Cash",
            "DieselCash_AllEntries.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"
        try:
            wb.save(path)
            QMessageBox.information(self, "Export Complete", f"Saved to:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))


# ═══════════════════════════════════════════════════════════════════════════════
#  Transactions — month browse → detail (with year/month + From/To)
# ═══════════════════════════════════════════════════════════════════════════════

class _DieselCashMonthBrowse(QWidget):
    """One row per calendar month — click to drill into that month's entries."""

    month_clicked = Signal(object)

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._year = app_state.fiscal_year
        self._month = 0
        self._summaries: List[dict] = []
        self._all_summaries: List[dict] = []
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background: transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        toolbar = QWidget()
        toolbar.setStyleSheet("background: transparent;")
        tl = QHBoxLayout(toolbar)
        tl.setContentsMargins(0, 0, 0, 0)
        tl.setSpacing(10)
        tl.addWidget(_lbl("FY", size=12, color=_T2))
        self._fy_cb = QComboBox()
        current_yr = datetime.now().year
        for yr in range(current_yr - 3, current_yr + 2):
            self._fy_cb.addItem(str(yr), yr)
        idx = self._fy_cb.findData(self._year)
        self._fy_cb.setCurrentIndex(idx if idx >= 0 else self._fy_cb.count() - 2)
        self._fy_cb.setFixedWidth(80)
        self._fy_cb.setStyleSheet(_input_ss())
        self._fy_cb.currentIndexChanged.connect(self._on_year_changed)
        tl.addWidget(self._fy_cb)

        self._month_cb = QComboBox()
        for label, val in _MONTHS:
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.currentIndexChanged.connect(self._on_month_changed)
        tl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tl, self._on_date_changed, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)

        tl.addStretch()
        self._summary_lbl = _lbl("", size=12, color=_T2)
        tl.addWidget(self._summary_lbl)
        vl.addWidget(toolbar)

        self._table = _make_table(_BROWSE_HEADERS)
        self._table.setShowGrid(True)
        self._table.setCursor(Qt.PointingHandCursor)
        hdr = self._table.horizontalHeader()
        hdr.setStretchLastSection(True)
        self._table.setColumnWidth(0, 140)
        self._table.setColumnWidth(1, 100)
        self._table.setColumnWidth(2, 140)
        self._table.cellClicked.connect(self._on_row_clicked)
        vl.addWidget(self._table, 1)

        hint = _lbl("Click a month to view verified cashier entries.", size=11, color=_TM)
        hint.setAlignment(Qt.AlignCenter)
        vl.addWidget(hint)

    def year(self) -> int:
        return self._year

    def date_filters(self):
        return read_from_to(self._from_date, self._to_date, optional=True)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    def _on_year_changed(self) -> None:
        yr = self._fy_cb.currentData()
        if yr and yr != self._year:
            self._year = yr
            app_state.fiscal_year = yr
            sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)
            asyncio.ensure_future(self._load())

    def _on_month_changed(self) -> None:
        self._month = int(self._month_cb.currentData() or 0)
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)
        self._apply_filters()

    def _on_date_changed(self) -> None:
        self._apply_filters()

    async def _load(self) -> None:
        from tahmeed.services.accountant_service import get_diesel_cash_month_summaries
        try:
            summaries = await get_diesel_cash_month_summaries(self._year)
        except Exception as exc:
            self._table.setRowCount(0)
            self._all_summaries = []
            self._summaries = []
            self._summary_lbl.setText(f"Failed to load: {exc}")
            return
        self._all_summaries = summaries
        self._apply_filters()

    def _apply_filters(self) -> None:
        date_from, date_to = self.date_filters()
        rows = list(self._all_summaries)
        if self._month and 1 <= self._month <= 12:
            rows = [s for s in rows if int(s.get("month", 0)) == self._month]
        if date_from or date_to:
            filtered = []
            for s in rows:
                min_d = s.get("min_date")
                max_d = s.get("max_date")
                if not isinstance(min_d, datetime) or not isinstance(max_d, datetime):
                    continue
                if date_from and max_d < date_from:
                    continue
                if date_to and min_d > date_to:
                    continue
                filtered.append(s)
            rows = filtered
        self._summaries = rows
        self._fill(rows)

    def _fill(self, summaries: List[dict]) -> None:
        t = self._table
        t.setRowCount(0)
        total_recs = 0
        total_tzs = 0.0
        for i, s in enumerate(summaries):
            r = t.rowCount()
            t.insertRow(r)
            month_idx = int(s.get("month", 0))
            month_label = s.get("month_name") or _MONTH_SHORT[month_idx]
            count = int(s.get("record_count", 0))
            tzs = float(s.get("tzs_total", 0))
            date_range = _fmt_date_range(s.get("min_date"), s.get("max_date"))

            t.setItem(r, 0, _cell(month_label))
            t.setItem(r, 1, _cell(f"{count:,}", align=Qt.AlignCenter | Qt.AlignVCenter))
            t.setItem(r, 2, _cell(f"{tzs:,.0f}", align=Qt.AlignRight | Qt.AlignVCenter, mono=True))
            t.setItem(r, 3, _cell(date_range))
            _finish_table_row(t, r)
            total_recs += count
            total_tzs += tzs

        if summaries:
            self._summary_lbl.setText(
                f"{len(summaries)} months  ·  {total_recs:,} records  ·  TZS {total_tzs:,.0f}"
            )
        else:
            self._summary_lbl.setText("No Diesel Cash entries for this filter.")

    def _on_row_clicked(self, row: int, _col: int) -> None:
        if 0 <= row < len(self._summaries):
            doc = dict(self._summaries[row])
            doc["year"] = self._year
            date_from, date_to = self.date_filters()
            doc["date_from"] = date_from
            doc["date_to"] = date_to
            self.month_clicked.emit(doc)


class _DieselCashMonthDetail(QWidget):
    """Verified Diesel Cash transactions for one calendar month (with From/To)."""

    back_requested = Signal()

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._year = app_state.fiscal_year
        self._month = 0
        self._month_name = ""
        self._page = 0
        self._total = 0
        self._loading = False

        self._debounce = QTimer(self)
        self._debounce.setSingleShot(True)
        self._debounce.setInterval(350)
        self._debounce.timeout.connect(lambda: asyncio.ensure_future(self._reload()))

        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background: transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(8)

        nav = QWidget()
        nav.setStyleSheet("background: transparent;")
        navl = QHBoxLayout(nav)
        navl.setContentsMargins(0, 0, 0, 0)
        navl.setSpacing(8)
        back_btn = _btn("← All Months", "mdi.chevron-left")
        back_btn.setFixedWidth(120)
        back_btn.clicked.connect(self.back_requested)
        navl.addWidget(back_btn)
        self._crumb_lbl = _lbl("", size=12, color=_T2)
        navl.addWidget(self._crumb_lbl)
        navl.addStretch()
        vl.addWidget(nav)

        self._info_lbl = _lbl("", size=12, weight=600)
        vl.addWidget(self._info_lbl)

        toolbar = QWidget()
        toolbar.setStyleSheet("background: transparent;")
        tl = QHBoxLayout(toolbar)
        tl.setContentsMargins(0, 0, 0, 0)
        tl.setSpacing(10)
        self._search = QLineEdit()
        self._search.setPlaceholderText("Search description or truck…")
        self._search.setFixedWidth(220)
        self._search.setStyleSheet(_input_ss())
        self._search.textChanged.connect(self._on_filter_changed)
        tl.addWidget(self._search)

        self._month_cb = QComboBox()
        for label, val in _MONTHS:
            if val == 0:
                continue  # detail is always a specific month
            self._month_cb.addItem(label, val)
        self._month_cb.setFixedWidth(130)
        self._month_cb.setStyleSheet(_input_ss())
        self._month_cb.currentIndexChanged.connect(self._on_month_changed)
        tl.addWidget(self._month_cb)

        self._from_date, self._to_date = add_from_to_editors(
            tl, self._on_filter_changed, input_ss=_input_ss(), lbl_factory=_lbl,
            optional=True,
        )

        self._rcpt_cb = QComboBox()
        for label, val in [("All Receipts", "all"), ("Received", "received"),
                           ("Pending", "pending"), ("No Receipt", "missing")]:
            self._rcpt_cb.addItem(label, val)
        self._rcpt_cb.setFixedWidth(128)
        self._rcpt_cb.setStyleSheet(_input_ss())
        self._rcpt_cb.currentIndexChanged.connect(self._on_filter_changed)
        tl.addWidget(self._rcpt_cb)

        clear_btn = _btn("Clear", "mdi.filter-remove-outline", primary=False)
        clear_btn.setToolTip("Clear search, date, receipt filters, and column sort.")
        clear_btn.clicked.connect(self._clear_filters)
        tl.addWidget(clear_btn)
        tl.addStretch()

        export_btn = _btn("Export Excel", "mdi.microsoft-excel")
        export_btn.clicked.connect(self._on_export)
        tl.addWidget(export_btn)
        vl.addWidget(toolbar)

        self._table = _make_table([c[0] for c in _DIESEL_COLS])
        self._table.setShowGrid(True)
        hdr = self._table.horizontalHeader()
        hdr.setSectionsMovable(False)
        hdr.setStretchLastSection(True)
        for i, (_, width, _a, _m) in enumerate(_DIESEL_COLS):
            self._table.setColumnWidth(i, width)
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        self._sort_state = wire_feed_table_sort(
            self._table,
            DIESEL_CASH_SORT,
            default_field="date",
            default_asc=False,
            on_sort_changed=self._on_sort_changed,
        )
        vl.addWidget(self._table, 1)

        footer = QFrame()
        footer.setFixedHeight(40)
        footer.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 2px solid {_BORDER}; }}"
        )
        fol = QHBoxLayout(footer)
        fol.setContentsMargins(16, 0, 16, 0)
        fol.setSpacing(20)
        fol.addWidget(_lbl("TOTAL (filtered view)", size=11, weight=700, color=_T2))
        self._tzs_lbl = _lbl("TZS  —", size=13, weight=700)
        self._tzs_lbl.setStyleSheet(
            f"color: {_T1}; font-size: 13px; font-weight: 700;"
            " font-family: 'Cascadia Code', 'Consolas', monospace; background: transparent;"
        )
        fol.addWidget(self._tzs_lbl)
        fol.addStretch()
        self._footer_count = _lbl("", size=11, color=_T2)
        fol.addWidget(self._footer_count)
        vl.addWidget(footer)

        pager = QFrame()
        pager.setFixedHeight(44)
        pager.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 1px solid {_BORDER}; }}"
        )
        pl = QHBoxLayout(pager)
        pl.setContentsMargins(16, 0, 16, 0)
        pl.setSpacing(10)
        self._size_cb = QComboBox()
        for sz in _PAGE_SIZES:
            self._size_cb.addItem(f"Show {sz}", sz)
        self._size_cb.setFixedWidth(100)
        self._size_cb.setStyleSheet(_input_ss())
        self._size_cb.currentIndexChanged.connect(self._on_size_changed)
        pl.addWidget(self._size_cb)
        self._page_info = _lbl("—", size=12, color=_T2)
        pl.addWidget(self._page_info)
        pl.addStretch()
        self._prev_btn = _btn("← Prev", "mdi.chevron-left")
        self._prev_btn.setFixedWidth(88)
        self._prev_btn.clicked.connect(self._on_prev)
        pl.addWidget(self._prev_btn)
        self._next_btn = _btn("Next →", "mdi.chevron-right")
        self._next_btn.setFixedWidth(88)
        self._next_btn.clicked.connect(self._on_next)
        pl.addWidget(self._next_btn)
        vl.addWidget(pager)

    def load_month(self, summary: dict) -> None:
        self._year = int(summary.get("year") or app_state.fiscal_year)
        self._month = int(summary.get("month", 0))
        self._month_name = summary.get("month_name") or _MONTH_SHORT[self._month]
        count = int(summary.get("record_count", 0))
        tzs = float(summary.get("tzs_total", 0))
        date_range = _fmt_date_range(summary.get("min_date"), summary.get("max_date"))

        self._crumb_lbl.setText(f"FY {self._year}  ›  {self._month_name}")
        self._info_lbl.setText(
            f"{self._month_name} {self._year}   •   {count:,} records   •   "
            f"TZS {tzs:,.0f}   •   {date_range}"
        )
        self._search.blockSignals(True)
        self._search.setText("")
        self._search.blockSignals(False)
        self._rcpt_cb.setCurrentIndex(0)

        idx = self._month_cb.findData(self._month)
        self._month_cb.blockSignals(True)
        self._month_cb.setCurrentIndex(idx if idx >= 0 else 0)
        self._month_cb.blockSignals(False)

        # Prefer From/To carried from browse; otherwise sync to the month window
        date_from = summary.get("date_from")
        date_to = summary.get("date_to")
        if date_from or date_to:
            self._from_date.blockSignals(True)
            self._to_date.blockSignals(True)
            if date_from:
                self._from_date.setDate(QDate(date_from.year, date_from.month, date_from.day))
            else:
                sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)
            if date_to:
                self._to_date.setDate(QDate(date_to.year, date_to.month, date_to.day))
            self._from_date.blockSignals(False)
            self._to_date.blockSignals(False)
        else:
            sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)

        self._page = 0
        asyncio.ensure_future(self._reload())

    def _search_text(self) -> str:
        return self._search.text().strip()

    def _receipt_filter(self) -> str:
        return self._rcpt_cb.currentData() or "all"

    def _page_size(self) -> int:
        return self._size_cb.currentData() or _PAGE_SIZES[0]

    def _on_sort_changed(self, field: str, asc: bool) -> None:
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _date_kw(self) -> dict:
        df, dt = read_from_to(self._from_date, self._to_date, optional=True)
        return {"date_from": df, "date_to": dt}

    def _on_filter_changed(self) -> None:
        self._page = 0
        self._debounce.start()

    def _clear_filters(self) -> None:
        self._search.blockSignals(True)
        self._month_cb.blockSignals(True)
        self._rcpt_cb.blockSignals(True)
        try:
            self._search.clear()
            self._rcpt_cb.setCurrentIndex(0)
        finally:
            self._search.blockSignals(False)
            self._month_cb.blockSignals(False)
            self._rcpt_cb.blockSignals(False)
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)
        reset_feed_sort(self._sort_state)
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _on_month_changed(self) -> None:
        month = int(self._month_cb.currentData() or 0)
        if not month or month == self._month:
            return
        self._month = month
        self._month_name = _MONTH_SHORT[month]
        self._crumb_lbl.setText(f"FY {self._year}  ›  {self._month_name}")
        sync_from_to(self._from_date, self._to_date, self._year, self._month, optional=True)
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _on_size_changed(self) -> None:
        self._page = 0
        asyncio.ensure_future(self._reload())

    def _on_prev(self) -> None:
        if self._page > 0:
            self._page -= 1
            asyncio.ensure_future(self._reload())

    def _on_next(self) -> None:
        size = self._page_size()
        max_pg = max(0, (self._total - 1) // size) if self._total else 0
        if self._page < max_pg:
            self._page += 1
            asyncio.ensure_future(self._reload())

    async def _reload(self) -> None:
        if self._loading or not self._month:
            return
        self._loading = True
        try:
            from tahmeed.services.accountant_service import (
                get_diesel_cash_transactions, count_diesel_cash_transactions,
                get_diesel_cash_totals,
            )
            size = self._page_size()
            skip = self._page * size
            kw = dict(
                year=self._year, month=self._month,
                search=self._search_text(), truck="",
                receipt=self._receipt_filter(),
                **self._date_kw(),
            )
            txs, total, totals = await asyncio.gather(
                get_diesel_cash_transactions(
                    **kw, **sort_kw(self._sort_state), limit=size, skip=skip,
                ),
                count_diesel_cash_transactions(**kw),
                get_diesel_cash_totals(**kw),
            )
            self._total = total
            self._populate(txs, skip)
            self._tzs_lbl.setText(f"TZS  {totals['tzs']:,.0f}")
            self._tzs_lbl.setStyleSheet(
                f"color: {_RED if totals['tzs'] < 0 else _T1};"
                " font-size: 13px; font-weight: 700;"
                " font-family: 'Cascadia Code', 'Consolas', monospace; background: transparent;"
            )
            self._footer_count.setText(f"{total:,} records")
            self._info_lbl.setText(
                f"{self._month_name} {self._year}   •   {total:,} records   •   "
                f"TZS {totals['tzs']:,.0f}"
            )
            self._update_pager(total, size)
        except Exception as exc:
            self._table.setRowCount(0)
            self._page_info.setText(f"Failed to load: {exc}")
        finally:
            self._loading = False

    def _populate(self, txs: List[Transaction], skip: int) -> None:
        t = self._table
        t.setRowCount(0)
        for i, tx in enumerate(txs):
            r = t.rowCount()
            t.insertRow(r)
            _fill_diesel_entry_row(t, r, tx, skip + i + 1)

    def _update_pager(self, total: int, size: int) -> None:
        max_pg = max(0, (total - 1) // size) if total else 0
        self._prev_btn.setEnabled(self._page > 0)
        self._next_btn.setEnabled(self._page < max_pg)
        start = self._page * size + 1 if total else 0
        end = min((self._page + 1) * size, total)
        self._page_info.setText(
            f"Showing {start:,}–{end:,} of {total:,}  ·  Page {self._page + 1} of {max_pg + 1}"
        )

    def _on_export(self) -> None:
        asyncio.ensure_future(self._do_export())

    async def _do_export(self) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n\nRun: pip install openpyxl",
            )
            return

        from tahmeed.services.accountant_service import get_diesel_cash_transactions
        kw = dict(
            year=self._year, month=self._month,
            search=self._search_text(), truck="",
            receipt=self._receipt_filter(),
            **self._date_kw(),
        )
        try:
            txs = await get_diesel_cash_transactions(
                **kw, **sort_kw(self._sort_state), limit=10_000, skip=0,
            )
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", f"Failed to fetch data: {exc}")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diesel Cash"
        ws.merge_cells(_EXPORT_SPAN)
        ws["A1"] = "TAHMEED COACH TZ LTD"
        ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1B2B4B")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:L2")
        ws["A2"] = f"Diesel Cash — {self._month_name} {self._year}"
        ws["A2"].font = Font(name="Segoe UI", bold=True, size=11, color="374151")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.append([])
        ws.append([c[0] for c in _DIESEL_COLS])
        hdr_row = ws.max_row
        grey = PatternFill("solid", fgColor="F1F5F9")
        for cell in ws[hdr_row]:
            cell.font = Font(name="Segoe UI", bold=True, size=10, color="6B7280")
            cell.fill = grey
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for tx in txs:
            ws.append(_diesel_export_row(tx))

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Diesel Cash",
            f"DieselCash_{self._month_name}_{self._year}.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return
        if not path.endswith(".xlsx"):
            path += ".xlsx"
        try:
            wb.save(path)
            QMessageBox.information(self, "Export Complete", f"Saved to:\n{path}")
        except Exception as exc:
            QMessageBox.critical(self, "Export Error", str(exc))


# ═══════════════════════════════════════════════════════════════════════════════
#  Shell — All Entries + Transactions
# ═══════════════════════════════════════════════════════════════════════════════

class DieselCashWidget(QWidget):
    """Diesel Cash — All Entries, month browse, and item-name configuration."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._build()
        self.refresh()

    def _build(self) -> None:
        self.setStyleSheet(f"background: {_BG};")
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 20, 20, 16)
        root.setSpacing(12)

        title_bar = QFrame()
        title_bar.setFixedHeight(52)
        title_bar.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-bottom: 1px solid {_BORDER}; }}"
        )
        tb = QHBoxLayout(title_bar)
        tb.setContentsMargins(0, 0, 0, 0)
        tb.setSpacing(12)
        try:
            icon_lbl = QLabel()
            icon_lbl.setFixedSize(22, 22)
            icon_lbl.setPixmap(
                qta.icon("mdi.gas-station-outline", color=_BLUE).pixmap(22, 22)
            )
            icon_lbl.setStyleSheet("background: transparent;")
            tb.addWidget(icon_lbl)
        except Exception:
            pass
        tb.addWidget(_lbl("Diesel Cash", size=16, weight=700))
        tb.addWidget(_lbl("Verified daily transactions for selected items", size=12, color=_T2))
        tb.addStretch()

        refresh_btn = QPushButton()
        refresh_btn.setFixedSize(32, 32)
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.setStyleSheet(
            "QPushButton { background: transparent; border: none; border-radius: 4px; }"
            f"QPushButton:hover {{ background: {_BG}; }}"
        )
        try:
            refresh_btn.setIcon(qta.icon("mdi.refresh", color=_T2))
            refresh_btn.setIconSize(QSize(18, 18))
        except Exception:
            refresh_btn.setText("↻")
        refresh_btn.clicked.connect(self.refresh)
        tb.addWidget(refresh_btn)
        root.addWidget(title_bar)
        root.addWidget(_hsep())

        self._tabs = _SegmentTabBar(["All Entries", "Transactions", "Items"])
        root.addWidget(self._tabs)

        self._main_stack = QStackedWidget()
        self._main_stack.setStyleSheet("background: transparent;")

        self._all_entries = _DieselCashAllEntries()
        self._main_stack.addWidget(self._all_entries)

        month_host = QWidget()
        month_host.setStyleSheet("background: transparent;")
        month_vl = QVBoxLayout(month_host)
        month_vl.setContentsMargins(0, 0, 0, 0)
        month_vl.setSpacing(0)

        self._month_stack = QStackedWidget()
        self._month_stack.setStyleSheet("background: transparent;")

        self._browse = _DieselCashMonthBrowse()
        self._browse.month_clicked.connect(self._show_detail)
        self._month_stack.addWidget(self._browse)

        self._detail = _DieselCashMonthDetail()
        self._detail.back_requested.connect(self._show_browse)
        self._month_stack.addWidget(self._detail)

        month_vl.addWidget(self._month_stack, 1)
        self._main_stack.addWidget(month_host)

        self._config = _DieselCashItemsConfig()
        self._config.saved.connect(self._on_items_saved)
        self._main_stack.addWidget(self._config)

        self._tabs.tab_changed.connect(self._on_tab)
        root.addWidget(self._main_stack, 1)

    def _on_tab(self, idx: int) -> None:
        self._main_stack.setCurrentIndex(idx)
        if idx == 2:
            self._config.refresh()

    def refresh(self) -> None:
        self._all_entries.refresh()
        if self._tabs.current_index() == 2:
            self._config.refresh()
        else:
            self._show_browse()

    def _on_items_saved(self) -> None:
        self._all_entries.refresh()
        self._browse.refresh()

    def _show_browse(self) -> None:
        self._month_stack.setCurrentIndex(0)
        self._browse.refresh()

    def _show_detail(self, summary: dict) -> None:
        self._tabs.set_index(1, emit=False)
        self._main_stack.setCurrentIndex(1)
        self._month_stack.setCurrentIndex(1)
        self._detail.load_month(summary)


# ═══════════════════════════════════════════════════════════════════════════════
#  Items — which item names this tab catches
# ═══════════════════════════════════════════════════════════════════════════════

_LIST_SS = (
    f"QListWidget {{"
    f"  background: {_WHITE}; border: 1px solid {_BORDER}; border-radius: 6px;"
    f"  font-size: 13px; font-family:'Segoe UI'; color: {_T1}; padding: 4px;"
    f"}}"
    f"QListWidget::item {{ padding: 6px 10px; border-radius: 4px; }}"
    f"QListWidget::item:selected {{ background: {_BLUE}1A; color: {_T1}; }}"
    f"QListWidget::item:hover {{ background: {_BG}; }}"
    f"QScrollBar:vertical {{ background: {_BG}; width: 8px; margin: 0; }}"
    f"QScrollBar::handle:vertical {{ background: #D1D5DB; border-radius: 4px; min-height: 24px; }}"
    f"QScrollBar::add-line, QScrollBar::sub-line {{ width: 0; height: 0; }}"
)


class _DieselCashItemsConfig(QWidget):
    """Pick one or many items whose verified transactions appear in Diesel Cash."""

    saved = Signal()

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._loading = False
        self._build()

    def _build(self) -> None:
        self.setStyleSheet("background: transparent;")
        vl = QVBoxLayout(self)
        vl.setContentsMargins(0, 0, 0, 0)
        vl.setSpacing(12)

        intro = _lbl(
            "Tick every item that should appear in Diesel Cash. "
            "Any verified daily transaction with one of those item names is included — "
            "cashier-created and daily imports alike. Matching is the item name, "
            "not the description (so DIESEL in Excel only lands here if it was mapped "
            "to a ticked item).",
            size=12, color=_T2,
        )
        intro.setWordWrap(True)
        vl.addWidget(intro)

        toolbar = QWidget()
        toolbar.setStyleSheet("background: transparent;")
        tl = QHBoxLayout(toolbar)
        tl.setContentsMargins(0, 0, 0, 0)
        tl.setSpacing(10)

        self._search = QLineEdit()
        self._search.setPlaceholderText("Search items…")
        self._search.setFixedWidth(260)
        self._search.setStyleSheet(_input_ss())
        self._search.textChanged.connect(self._apply_search)
        tl.addWidget(self._search)

        select_all = _btn("Select all")
        select_all.clicked.connect(self._select_visible)
        tl.addWidget(select_all)
        clear_btn = _btn("Clear")
        clear_btn.clicked.connect(self._clear_visible)
        tl.addWidget(clear_btn)
        tl.addStretch()
        self._count_lbl = _lbl("", size=12, color=_T2)
        tl.addWidget(self._count_lbl)
        vl.addWidget(toolbar)

        self._list = QListWidget()
        self._list.setStyleSheet(_LIST_SS)
        self._list.setSelectionMode(QAbstractItemView.NoSelection)
        self._list.setFocusPolicy(Qt.NoFocus)
        self._list.itemChanged.connect(self._on_item_changed)
        vl.addWidget(self._list, 1)

        footer = QFrame()
        footer.setFixedHeight(48)
        footer.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 1px solid {_BORDER}; }}"
        )
        fol = QHBoxLayout(footer)
        fol.setContentsMargins(0, 8, 0, 0)
        fol.setSpacing(12)
        self._status = _lbl("", size=12, color=_T2)
        fol.addWidget(self._status)
        fol.addStretch()
        save_btn = _btn("Save items", "mdi.content-save-outline", primary=True)
        save_btn.clicked.connect(lambda: asyncio.ensure_future(self._save()))
        fol.addWidget(save_btn)
        vl.addWidget(footer)

    def refresh(self) -> None:
        asyncio.ensure_future(self._load())

    async def _load(self) -> None:
        if self._loading:
            return
        self._loading = True
        try:
            from tahmeed.services.accountant_service import get_diesel_cash_item_names
            from tahmeed.services.category_service import get_all_categories

            selected, categories = await asyncio.gather(
                get_diesel_cash_item_names(),
                get_all_categories(),
            )
            selected_keys = {n.casefold() for n in selected}
            by_key = {c.name.casefold(): c.name for c in categories if (c.name or "").strip()}
            rows: List[tuple[str, bool, bool]] = []
            for name in selected:
                key = name.casefold()
                display = by_key.get(key, name)
                missing = key not in by_key
                rows.append((display, True, missing))
            for cat in categories:
                name = (cat.name or "").strip()
                if not name or name.casefold() in selected_keys:
                    continue
                rows.append((name, False, False))
            self._populate(rows)
            self._status.setText("")
            self._status.setStyleSheet(
                f"color: {_T2}; font-size: 12px; font-family:'Segoe UI';"
                " background: transparent;"
            )
        except Exception as exc:
            self._list.clear()
            self._status.setText(f"Could not load items: {exc}")
            self._status.setStyleSheet(
                f"color: {_RED}; font-size: 12px; font-family:'Segoe UI';"
                " background: transparent;"
            )
        finally:
            self._loading = False

    def _populate(self, rows: List[tuple[str, bool, bool]]) -> None:
        self._list.blockSignals(True)
        self._list.clear()
        for name, checked, missing in rows:
            label = f"{name}  (not in items list)" if missing else name
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, name)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if checked else Qt.Unchecked)
            if missing:
                item.setForeground(Qt.gray)
            self._list.addItem(item)
        self._list.blockSignals(False)
        self._apply_search()
        self._update_count()

    def _apply_search(self) -> None:
        needle = self._search.text().strip().casefold()
        for i in range(self._list.count()):
            item = self._list.item(i)
            name = str(item.data(Qt.UserRole) or item.text()).casefold()
            item.setHidden(bool(needle) and needle not in name)

    def _visible_items(self) -> List[QListWidgetItem]:
        return [
            self._list.item(i)
            for i in range(self._list.count())
            if not self._list.item(i).isHidden()
        ]

    def _select_visible(self) -> None:
        self._list.blockSignals(True)
        for item in self._visible_items():
            item.setCheckState(Qt.Checked)
        self._list.blockSignals(False)
        self._update_count()

    def _clear_visible(self) -> None:
        self._list.blockSignals(True)
        for item in self._visible_items():
            item.setCheckState(Qt.Unchecked)
        self._list.blockSignals(False)
        self._update_count()

    def _on_item_changed(self, _item: QListWidgetItem) -> None:
        self._update_count()
        self._status.setText("Unsaved changes")
        self._status.setStyleSheet(
            f"color: {_T2}; font-size: 12px; font-family:'Segoe UI';"
            " background: transparent;"
        )

    def _checked_names(self) -> List[str]:
        names: List[str] = []
        for i in range(self._list.count()):
            item = self._list.item(i)
            if item.checkState() != Qt.Checked:
                continue
            name = str(item.data(Qt.UserRole) or "").strip()
            if name:
                names.append(name)
        return names

    def _update_count(self) -> None:
        n = len(self._checked_names())
        if n == 0:
            self._count_lbl.setText("No items selected — this tab will be empty")
        elif n == 1:
            self._count_lbl.setText("1 item selected")
        else:
            self._count_lbl.setText(f"{n} items selected")

    async def _save(self) -> None:
        from tahmeed.services.accountant_service import set_diesel_cash_item_names

        names = self._checked_names()
        try:
            saved = await set_diesel_cash_item_names(names)
        except Exception as exc:
            QMessageBox.critical(self, "Save items", f"Could not save:\n{exc}")
            return
        n = len(saved)
        if n == 0:
            msg = "Saved — no items selected, so Diesel Cash will show no rows."
        elif n == 1:
            msg = f"Saved — catching “{saved[0]}”."
        else:
            msg = f"Saved — catching {n} items."
        self._status.setText(msg)
        self._status.setStyleSheet(
            f"color: {_GREEN}; font-size: 12px; font-weight: 600;"
            " font-family:'Segoe UI'; background: transparent;"
        )
        self.saved.emit()

