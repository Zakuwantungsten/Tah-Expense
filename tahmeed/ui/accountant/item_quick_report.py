"""Item Account QuickReport — verified/master transactions for one catalog item.

Column set matches Master Expenses. Filter bar: Dates preset, From/To,
description multi-select (from DB), Excel column filters, header sort, export.
"""

from __future__ import annotations

import asyncio
import calendar
from datetime import date, datetime
from typing import Dict, List, Optional, Set, Tuple

from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QBrush, QColor, QFont
from PySide6.QtWidgets import (
    QAbstractItemView, QComboBox, QFrame, QHBoxLayout, QHeaderView, QLabel,
    QPushButton, QSizePolicy, QTableWidget, QTableWidgetItem, QVBoxLayout,
    QWidget,
)

from tahmeed.models.transaction import Transaction
from tahmeed.ui.accountant.date_filters import (
    add_from_to_editors, read_from_to, sync_from_to,
)
from tahmeed.ui.widgets.checkable_multi_combo import CheckableMultiCombo
from tahmeed.ui.widgets.column_persistence import bind_column_width_persistence
from tahmeed.ui.widgets.excel_column_filter import (
    ExcelFilterHeaderView, SORT_ASC, cascade_column_values,
)
from tahmeed.ui.widgets.export_runner import (
    FAST_STYLE_ROW_LIMIT, PROGRESS_EVERY,
    attach_export_overlay, export_file_ready, fetch_records_with_progress,
    hide_export_busy, normalize_xlsx_path, notify_export_error,
    notify_export_info, pick_export_path, run_export_write, show_export_busy,
)
from tahmeed.ui.widgets.loading_overlay import LoadingOverlay
from tahmeed.ui.widgets.split_export_button import make_export_menu_btn

_WHITE = "#FFFFFF"
_BG = "#F4F6F8"
_BORDER = "#E5E7EB"
_T1 = "#111827"
_T2 = "#6B7280"
_TM = "#9CA3AF"
_HDR_BG = "#F1F5F9"
_STRIPE = "#F1F5F9"
_BLUE_L = "#E8F4FD"
_ROW_H = 28
_PAGE_SIZE = 100

# Master Expenses columns (no running Balance on this tab)
# (label, default width, align, sort_field)
_COLS = [
    ("S/NO", 52, "center", None),
    ("DATE", 72, "left", "date"),
    ("ITEM", 110, "left", "item"),
    ("DESCRIPTION", 200, "left", "description"),
    ("TRUCK NO", 95, "left", "truck_number"),
    ("MEMO", 120, "left", "memo"),
    ("REF_FLOAT", 110, "left", "ref_float"),
    ("TZS", 110, "right", "amount"),
    ("USD", 100, "right", "amount"),
    ("RECEIPT", 100, "center", "receipt_status"),
    ("OWNERSHIP", 90, "left", "ownership"),
    ("APPROVED BY", 100, "left", "approver"),
    ("CASHIER", 100, "left", None),
]
_COL_DEFAULTS = [c[1] for c in _COLS]
_DESC_COL = 3
_FILTERABLE_COLS: Set[int] = set(range(1, len(_COLS)))
_SORT_KINDS = {1: "date", 7: "number", 8: "number", 4: "truck"}
_COL_FIELD: Dict[int, str] = {
    1: "date",
    2: "item",
    3: "description",
    4: "truck_number",
    5: "memo",
    6: "ref_float",
    7: "tzs",
    8: "usd",
    9: "receipt_status",
    10: "ownership",
    11: "approver",
    12: "cashier",
}

_RECEIPT_LABELS = {
    "received": "Received",
    "pending": "Pending",
    "missing": "No Receipt",
    "no_receipt": "No Receipt",
}

_TABLE_SS = (
    f"QTableWidget {{"
    f"  background: {_WHITE}; gridline-color: {_BORDER};"
    f"  font-size: 11px; font-family:'Segoe UI';"
    f"  color: {_T1}; border: 1px solid {_BORDER};"
    f"}}"
    f"QTableWidget::item {{ padding: 2px 8px; border: none; }}"
    f"QTableWidget::item:selected {{ background: {_BLUE_L}; color: {_T1}; }}"
    f"QHeaderView::section {{"
    f"  background: {_HDR_BG}; color: {_T2};"
    f"  font-size: 10px; font-weight: 600; font-family:'Segoe UI';"
    f"  border: none; border-bottom: 1px solid {_BORDER};"
    f"  border-right: 1px solid {_BORDER};"
    f"  padding: 0 18px 0 8px; min-height: 28px;"
    f"}}"
    f"QHeaderView::section:hover {{ background: #E2E8F0; }}"
    f"QScrollBar:vertical {{ background: {_BG}; width: 8px; margin: 0; }}"
    f"QScrollBar::handle:vertical {{ background: #D1D5DB; border-radius: 4px; min-height: 24px; }}"
    f"QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ width: 0; height: 0; }}"
)

_INPUT_SS = (
    f"QComboBox, QDateEdit {{"
    f"  border: 1px solid {_BORDER}; border-radius: 5px;"
    f"  background: {_WHITE}; color: {_T1}; font-size: 12px;"
    "  font-family:'Segoe UI'; padding: 0 8px;"
    "  min-height: 30px; max-height: 30px; }}"
    f"QComboBox:focus, QDateEdit:focus {{ border-color: #0077C5; }}"
    f"QComboBox::drop-down {{ border: none; width: 22px; }}"
)


def _lbl(text: str = "", size: int = 12, weight: int = 400, color: str = _T1) -> QLabel:
    w = QLabel(text)
    w.setStyleSheet(
        f"color: {color}; font-size: {size}px; font-weight: {weight};"
        " font-family:'Segoe UI'; background: transparent;"
    )
    return w


def _cell_font(*, mono: bool = False, bold: bool = False) -> QFont:
    f = QFont("Cascadia Code" if mono else "Segoe UI")
    if mono:
        f.setStyleHint(QFont.Monospace)
    f.setPixelSize(11)
    f.setBold(bold)
    return f


def _set_cell(
    table: QTableWidget,
    row: int,
    col: int,
    text: str,
    align: str,
    row_bg: str,
    *,
    color: str = _T1,
    mono: bool = False,
) -> None:
    it = QTableWidgetItem(text)
    flag = Qt.AlignVCenter
    if align == "center":
        flag |= Qt.AlignHCenter
    elif align == "right":
        flag |= Qt.AlignRight
    else:
        flag |= Qt.AlignLeft
    it.setTextAlignment(flag)
    it.setBackground(QBrush(QColor(row_bg)))
    it.setForeground(QBrush(QColor(color)))
    it.setFont(_cell_font(mono=mono))
    it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
    table.setItem(row, col, it)


def _short_name(name: str) -> str:
    parts = (name or "").strip().split()
    if len(parts) >= 2:
        return f"{parts[0]} {parts[1][0]}."
    return name or "—"


def _ref_float_display(tx: Transaction) -> str:
    text = (getattr(tx, "ref_float", None) or "").strip()
    if text:
        return text.upper()
    if tx.notes_flag:
        return "REFUND TO FLOAT"
    return ""


def _receipt_label(status: str) -> str:
    key = (status or "pending").strip().lower().replace(" ", "_")
    if key in ("no", "n/a", "none"):
        key = "no_receipt"
    return _RECEIPT_LABELS.get(key, "—")


def _used_amount(value: float) -> float:
    """Amount used on this tab — always positive (negatives are float/use signs)."""
    return abs(float(value or 0.0))


def _fmt_report_date(dt, *, all_years: bool) -> str:
    if not dt:
        return "—"
    if hasattr(dt, "date") and not isinstance(dt, date):
        try:
            dt = dt.date()
        except Exception:
            pass
    if isinstance(dt, datetime):
        dt = dt.date()
    if not isinstance(dt, date):
        return "—"
    return dt.strftime("%d %b %Y" if all_years else "%d %b")


class ItemQuickReportView(QWidget):
    """Read-only master-column table for one item's verified transactions."""

    header_context_changed = Signal(str, str)

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self._category = ""
        self._page = 0
        self._total = 0
        self._loading = False
        self._scroll_loading = False
        self._loaded = 0
        self._syncing_dates = False
        self._export_in_flight = False
        self._reload_generation = 0
        self._sort_field = "date"
        self._sort_asc = False
        self._col_filters: Dict[int, Set[str]] = {}
        self._col_value_cache: Dict[int, Set[str]] = {}
        self._filter_debounce = QTimer(self)
        self._filter_debounce.setSingleShot(True)
        self._filter_debounce.setInterval(300)
        self._filter_debounce.timeout.connect(self._on_filter_commit)
        self._build()

    def _build(self) -> None:
        self.setStyleSheet(f"background: {_WHITE};")
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        root.addWidget(self._build_filter_bar())

        self._account_bar = QFrame()
        self._account_bar.setFixedHeight(28)
        self._account_bar.setStyleSheet(
            f"QFrame {{ background: {_HDR_BG}; border: 1px solid {_BORDER};"
            f" border-bottom: none; }}"
        )
        abl = QHBoxLayout(self._account_bar)
        abl.setContentsMargins(10, 0, 10, 0)
        self._account_bar_lbl = _lbl("", size=12, weight=700, color=_T1)
        abl.addWidget(self._account_bar_lbl)
        abl.addStretch()
        root.addWidget(self._account_bar)

        self._table = QTableWidget(0, len(_COLS))
        self._table.setHorizontalHeaderLabels([c[0] for c in _COLS])
        self._table.setStyleSheet(_TABLE_SS)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(_ROW_H)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SingleSelection)
        self._table.setShowGrid(True)
        self._table.setAlternatingRowColors(False)
        self._table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        filter_hdr = ExcelFilterHeaderView(
            self._table,
            filterable_columns=_FILTERABLE_COLS,
            sort_kinds=_SORT_KINDS,
        )
        filter_hdr.set_value_provider(self._filter_menu_values)
        filter_hdr.set_label_provider(
            lambda c: _COLS[c][0] if 0 <= c < len(_COLS) else "",
        )
        filter_hdr.filter_changed.connect(self._on_col_filter_changed)
        filter_hdr.sort_requested.connect(self._on_excel_sort)
        self._table.setHorizontalHeader(filter_hdr)

        hdr = self._table.horizontalHeader()
        hdr.setSectionsMovable(False)
        hdr.setStretchLastSection(True)
        hdr.setSortIndicatorShown(True)
        for i, (_, width, _, _sf) in enumerate(_COLS):
            self._table.setColumnWidth(i, width)
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        bind_column_width_persistence(
            self._table, "item_quick_report", _COL_DEFAULTS,
        )
        hdr.sectionClicked.connect(self._on_header_click)
        self._update_sort_indicator()
        self._table.verticalScrollBar().valueChanged.connect(self._on_scroll)
        root.addWidget(self._table, 1)

        footer = QFrame()
        footer.setFixedHeight(40)
        footer.setStyleSheet(
            f"QFrame {{ background: {_WHITE}; border-top: 2px solid {_BORDER}; }}"
        )
        fl = QHBoxLayout(footer)
        fl.setContentsMargins(12, 0, 12, 0)
        fl.setSpacing(16)

        fl.addWidget(_lbl("TOTAL", size=11, weight=700, color=_T2))
        self._tzs_lbl = _lbl("TZS  —", size=13, weight=700)
        self._tzs_lbl.setStyleSheet(
            f"color: {_T1}; font-size: 13px; font-weight: 700;"
            " font-family:'Cascadia Code','Consolas',monospace;"
            " background: transparent;"
        )
        fl.addWidget(self._tzs_lbl)
        self._usd_lbl = _lbl("USD  —", size=12, weight=600, color=_T1)
        fl.addWidget(self._usd_lbl)
        self._count_lbl = _lbl("", size=11, color=_T2)
        fl.addWidget(self._count_lbl)
        fl.addStretch()

        self._page_info = _lbl("—", size=11, color=_T2)
        fl.addWidget(self._page_info)

        root.addWidget(footer)
        self._loading_overlay = LoadingOverlay(self, "Loading report…")

    def _build_filter_bar(self) -> QFrame:
        bar = QFrame()
        bar.setFixedHeight(48)
        bar.setStyleSheet(
            f"QFrame {{ background: {_BG}; border: 1px solid {_BORDER};"
            f" border-bottom: none; }}"
        )
        hl = QHBoxLayout(bar)
        hl.setContentsMargins(10, 0, 10, 0)
        hl.setSpacing(8)

        hl.addWidget(_lbl("Dates", size=11, color=_T2))
        self._dates_cb = QComboBox()
        self._dates_cb.setFixedWidth(130)
        self._dates_cb.setStyleSheet(_INPUT_SS)
        self._dates_cb.currentIndexChanged.connect(self._on_dates_preset)
        hl.addWidget(self._dates_cb)

        self._from_date, self._to_date = add_from_to_editors(
            hl, self._on_from_to_changed, input_ss=_INPUT_SS, lbl_factory=_lbl,
            optional=True, width=118,
        )

        hl.addWidget(_lbl("Description", size=11, color=_T2))
        self._desc_cb = CheckableMultiCombo(
            "All Descriptions", noun_plural="descriptions", parent=self,
        )
        self._desc_cb.setFixedWidth(180)
        self._desc_cb.setStyleSheet(_INPUT_SS)
        self._desc_cb.selectionChanged.connect(self._on_filter_changed)
        hl.addWidget(self._desc_cb)

        hl.addStretch()

        self._export_btn = make_export_menu_btn(
            self._on_export_filtered,
            self._on_export_all,
            parent=self,
            height=30,
            btn_tip=(
                "Export Filtered — current dates, descriptions, column filters, and sort.\n"
                "Use the ▾ menu for Export All (full item report, no extra filters)."
            ),
            filtered_tip="Export rows matching the current filters and sort order.",
            all_tip="Export every verified record for this item (ignores toolbar and column filters).",
        )
        hl.addWidget(self._export_btn)

        clear_btn = QPushButton("Clear")
        clear_btn.setFixedHeight(30)
        clear_btn.setCursor(Qt.PointingHandCursor)
        clear_btn.setStyleSheet(
            f"QPushButton {{ background: {_WHITE}; color: {_T2};"
            f" border: 1px solid {_BORDER}; border-radius: 5px;"
            " font-size: 12px; font-family:'Segoe UI'; padding: 0 12px; }}"
            f"QPushButton:hover {{ background: {_WHITE}; color: {_T1}; }}"
        )
        clear_btn.clicked.connect(self._clear_filters)
        hl.addWidget(clear_btn)

        self._rebuild_dates_options([])
        return bar

    def clear(self) -> None:
        self._category = ""
        self._page = 0
        self._total = 0
        self._loaded = 0
        self._table.setRowCount(0)
        self._account_bar_lbl.setText("")
        self._tzs_lbl.setText("TZS  —")
        self._usd_lbl.setText("USD  —")
        self._count_lbl.setText("")
        self._page_info.setText("—")
        self._reset_filters_ui()

    def load(self, category_name: str) -> None:
        self._category = (category_name or "").strip()
        self._page = 0
        self._account_bar_lbl.setText(self._category)
        self._reset_filters_ui()
        asyncio.ensure_future(self._prepare_and_reload())

    def _reset_filters_ui(self) -> None:
        self._syncing_dates = True
        try:
            self._dates_cb.blockSignals(True)
            self._desc_cb.blockSignals(True)
            self._rebuild_dates_options([])
            self._dates_cb.setCurrentIndex(0)
            sync_from_to(self._from_date, self._to_date, 0, 0, optional=True)
            self._desc_cb.reset_to_all(emit=False)
            self._sort_field = "date"
            self._sort_asc = False
            self._clear_column_filters_ui()
            self._update_sort_indicator()
        finally:
            self._dates_cb.blockSignals(False)
            self._desc_cb.blockSignals(False)
            self._syncing_dates = False

    def _clear_column_filters_ui(self) -> None:
        self._col_filters.clear()
        self._col_value_cache.clear()
        hdr = self._table.horizontalHeader()
        if isinstance(hdr, ExcelFilterHeaderView):
            hdr.clear_filters()

    def _rebuild_dates_options(self, years: List[int]) -> None:
        current = self._dates_cb.currentData()
        self._dates_cb.blockSignals(True)
        self._dates_cb.clear()
        self._dates_cb.addItem("All", ("all", None))
        self._dates_cb.addItem("This Month", ("this_month", None))
        self._dates_cb.addItem("This Year", ("this_year", None))
        self._dates_cb.addItem("Last Year", ("last_year", None))
        for yr in years:
            self._dates_cb.addItem(str(yr), ("year", yr))
        self._dates_cb.addItem("Custom", ("custom", None))
        idx = 0
        if current is not None:
            for i in range(self._dates_cb.count()):
                if self._dates_cb.itemData(i) == current:
                    idx = i
                    break
        self._dates_cb.setCurrentIndex(idx)
        self._dates_cb.blockSignals(False)

    async def _prepare_and_reload(self) -> None:
        if not self._category:
            return
        try:
            from tahmeed.services.accountant_service import (
                get_category_report_descriptions,
                get_master_available_years,
            )

            date_from, date_to = read_from_to(
                self._from_date, self._to_date, optional=True,
            )
            years, descs = await asyncio.gather(
                get_master_available_years(),
                get_category_report_descriptions(
                    self._category, date_from=date_from, date_to=date_to,
                ),
            )
            years = sorted({int(y) for y in years}, reverse=True)
            self._rebuild_dates_options(years)
            self._desc_cb.set_options(descs, keep_selected=True, emit=False)
        except Exception:
            pass
        await self._reload()

    async def _refresh_description_options(self) -> None:
        if not self._category:
            return
        try:
            from tahmeed.services.accountant_service import get_category_report_descriptions

            date_from, date_to = read_from_to(
                self._from_date, self._to_date, optional=True,
            )
            descs = await get_category_report_descriptions(
                self._category, date_from=date_from, date_to=date_to,
            )
            self._desc_cb.set_options(descs, keep_selected=True, emit=False)
        except Exception:
            pass

    def _on_filter_changed(self) -> None:
        self._page = 0
        self._filter_debounce.start()

    def _on_filter_commit(self) -> None:
        asyncio.ensure_future(self._reload())

    def _on_scroll(self, _value: int) -> None:
        bar = self._table.verticalScrollBar()
        if bar.maximum() <= 0:
            return
        if bar.value() >= bar.maximum() - 24:
            asyncio.ensure_future(self._load_more())

    def _fill_if_needed(self) -> None:
        bar = self._table.verticalScrollBar()
        if (
            not self._loading
            and not self._scroll_loading
            and self._loaded < self._total
            and bar.maximum() <= 0
        ):
            asyncio.ensure_future(self._load_more())

    def _on_dates_preset(self) -> None:
        if self._syncing_dates:
            return
        data = self._dates_cb.currentData()
        if not data:
            return
        kind, year = data
        today = date.today()
        self._syncing_dates = True
        try:
            if kind == "all":
                sync_from_to(self._from_date, self._to_date, 0, 0, optional=True)
            elif kind == "this_month":
                sync_from_to(
                    self._from_date, self._to_date, today.year, today.month,
                    optional=True,
                )
            elif kind == "this_year":
                sync_from_to(
                    self._from_date, self._to_date, today.year, 0, optional=True,
                )
            elif kind == "last_year":
                sync_from_to(
                    self._from_date, self._to_date, today.year - 1, 0, optional=True,
                )
            elif kind == "year" and year:
                sync_from_to(self._from_date, self._to_date, int(year), 0, optional=True)
            elif kind == "custom":
                pass
        finally:
            self._syncing_dates = False
        self._on_filter_changed()

    def _on_from_to_changed(self) -> None:
        if self._syncing_dates:
            return
        self._syncing_dates = True
        try:
            for i in range(self._dates_cb.count()):
                data = self._dates_cb.itemData(i)
                if data and data[0] == "custom":
                    self._dates_cb.blockSignals(True)
                    self._dates_cb.setCurrentIndex(i)
                    self._dates_cb.blockSignals(False)
                    break
        finally:
            self._syncing_dates = False
        self._on_filter_changed()

    def _clear_filters(self) -> None:
        self._page = 0
        self._syncing_dates = True
        try:
            self._dates_cb.blockSignals(True)
            self._desc_cb.blockSignals(True)
            self._dates_cb.setCurrentIndex(0)
            sync_from_to(self._from_date, self._to_date, 0, 0, optional=True)
            self._desc_cb.reset_to_all(emit=False)
            self._sort_field = "date"
            self._sort_asc = False
            self._clear_column_filters_ui()
            self._update_sort_indicator()
        finally:
            self._dates_cb.blockSignals(False)
            self._desc_cb.blockSignals(False)
            self._syncing_dates = False
        asyncio.ensure_future(self._prepare_and_reload())

    def _all_years_scope(self) -> bool:
        date_from, date_to = read_from_to(
            self._from_date, self._to_date, optional=True,
        )
        return date_from is None and date_to is None

    def _description_filter(self) -> List[str]:
        return self._desc_cb.selected_values()

    def column_filters_for_query(self) -> Dict[str, List[str]]:
        out: Dict[str, List[str]] = {}
        for col, accepted in self._col_filters.items():
            field = _COL_FIELD.get(col)
            if not field or not accepted:
                continue
            vals = sorted(v for v in accepted if v and v != "—")
            if vals:
                out[field] = vals
        return out

    def _query_kwargs(self, *, all_records: bool = False) -> dict:
        if all_records:
            return {
                "descriptions": None,
                "date_from": None,
                "date_to": None,
                "column_filters": None,
                "sort_field": self._sort_field,
                "sort_asc": bool(self._sort_asc),
            }
        date_from, date_to = read_from_to(
            self._from_date, self._to_date, optional=True,
        )
        descs = self._description_filter()
        col_filters = self.column_filters_for_query()
        kw = {
            "descriptions": descs or None,
            "date_from": date_from,
            "date_to": date_to,
            "sort_field": self._sort_field,
            "sort_asc": bool(self._sort_asc),
        }
        if col_filters:
            kw["column_filters"] = col_filters
        return kw

    def _header_context(self) -> Tuple[str, str]:
        date_from, date_to = read_from_to(
            self._from_date, self._to_date, optional=True,
        )
        descs = self._description_filter()
        if not descs:
            desc_scope = "All Transactions"
        elif len(descs) == 1:
            desc_scope = descs[0]
        else:
            desc_scope = f"{len(descs)} descriptions"

        if date_from is None and date_to is None:
            return str(date.today().year), desc_scope

        if date_from and date_to:
            if (
                date_from.year == date_to.year
                and date_from.month == 1 and date_from.day == 1
                and date_to.month == 12 and date_to.day == 31
            ):
                return str(date_from.year), desc_scope
            if (
                date_from.year == date_to.year
                and date_from.month == date_to.month
                and date_from.day == 1
                and date_to.day == calendar.monthrange(date_to.year, date_to.month)[1]
            ):
                month_name = date_from.strftime("%b %Y")
                return str(date_from.year), f"{desc_scope}  ·  {month_name}"
            if date_from.year == date_to.year:
                return (
                    str(date_from.year),
                    f"{desc_scope}  ·  {date_from.strftime('%d %b')}–{date_to.strftime('%d %b %Y')}",
                )
            return (
                f"{date_from.year}–{date_to.year}",
                f"{desc_scope}  ·  {date_from.strftime('%d %b %Y')}–{date_to.strftime('%d %b %Y')}",
            )

        if date_from:
            return str(date_from.year), f"{desc_scope}  ·  From {date_from.strftime('%d %b %Y')}"
        assert date_to is not None
        return str(date_to.year), f"{desc_scope}  ·  To {date_to.strftime('%d %b %Y')}"

    def _cell_text(self, row: int, col: int) -> str:
        it = self._table.item(row, col)
        raw = (it.text() if it else "").strip()
        if col == 9 and raw:
            return _receipt_label(raw)
        return raw

    def _filter_source_rows(self) -> List[dict]:
        rows: List[dict] = []
        for r in range(self._table.rowCount()):
            row: dict = {}
            for c in _FILTERABLE_COLS:
                txt = self._cell_text(r, c)
                if txt and txt != "—":
                    row[c] = txt
            if row:
                rows.append(row)
        return rows

    def _filter_menu_values(self, col: int) -> set:
        if col in self._col_value_cache:
            return set(self._col_value_cache.get(col) or set())
        return cascade_column_values(
            self._filter_source_rows(),
            target_col=col,
            active_filters=self._col_filters,
        )

    def _on_col_filter_changed(self, col: int, accepted) -> None:
        accepted = set(accepted or [])
        if accepted:
            self._col_filters[col] = accepted
        else:
            self._col_filters.pop(col, None)
        hdr = self._table.horizontalHeader()
        if isinstance(hdr, ExcelFilterHeaderView):
            hdr.sync_active(self._col_filters)
        self._page = 0
        asyncio.ensure_future(self._reload(keep_col_filters=True))

    def _on_header_click(self, col: int) -> None:
        sort_field = _COLS[col][3] if 0 <= col < len(_COLS) else None
        if sort_field is None:
            return
        if self._sort_field == sort_field:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_field = sort_field
            self._sort_asc = False
        self._update_sort_indicator()
        self._page = 0
        asyncio.ensure_future(self._reload(keep_col_filters=True))

    def _on_excel_sort(self, col: int, mode: str) -> None:
        sort_field = _COLS[col][3] if 0 <= col < len(_COLS) else None
        if sort_field is None:
            return
        self._sort_field = sort_field
        self._sort_asc = mode == SORT_ASC
        self._update_sort_indicator()
        self._page = 0
        asyncio.ensure_future(self._reload(keep_col_filters=True))

    def _update_sort_indicator(self) -> None:
        hdr = self._table.horizontalHeader()
        col_idx = next(
            (i for i, c in enumerate(_COLS) if c[3] == self._sort_field),
            1,
        )
        order = Qt.AscendingOrder if self._sort_asc else Qt.DescendingOrder
        hdr.setSortIndicator(col_idx, order)

    async def _reload(self, *, keep_col_filters: bool = False) -> None:
        if self._loading or not self._category:
            return
        self._loading = True
        self._page = 0
        self._loaded = 0
        self._reload_generation += 1
        generation = self._reload_generation
        if not keep_col_filters:
            self._clear_column_filters_ui()
        self._loading_overlay.show_loading("Loading report…")
        try:
            from tahmeed.services.accountant_service import (
                get_category_report_transactions,
                get_category_report_totals,
                get_cashier_names,
            )

            kw = self._query_kwargs()
            year_lbl, scope_lbl = self._header_context()
            self.header_context_changed.emit(year_lbl, scope_lbl)
            self._account_bar_lbl.setText(self._category)

            txs, totals = await asyncio.gather(
                get_category_report_transactions(
                    self._category,
                    descriptions=kw["descriptions"],
                    date_from=kw["date_from"],
                    date_to=kw["date_to"],
                    column_filters=kw.get("column_filters"),
                    sort_field=kw["sort_field"],
                    sort_asc=kw["sort_asc"],
                    limit=_PAGE_SIZE,
                    skip=0,
                ),
                get_category_report_totals(
                    self._category,
                    descriptions=kw["descriptions"],
                    date_from=kw["date_from"],
                    date_to=kw["date_to"],
                    column_filters=kw.get("column_filters"),
                ),
            )
            if generation != self._reload_generation:
                return
            self._total = int(totals.get("count") or 0)
            self._loaded = len(txs)
            cashier_ids = [tx.cashier_id for tx in txs if tx.cashier_id]
            names = await get_cashier_names(cashier_ids) if cashier_ids else {}
            self._populate(txs, 0, names, append=False)
            self._update_footer(totals)
            self._fill_if_needed()
            asyncio.ensure_future(self._load_column_values(generation, kw))
            asyncio.ensure_future(self._refresh_description_options())
        except Exception as exc:
            self._table.setRowCount(0)
            self._loaded = 0
            self._page_info.setText(f"Failed to load: {exc}")
        finally:
            self._loading = False
            self._loading_overlay.hide_loading()
            self._update_footer_status()

    async def _load_column_values(self, generation: int, kw: dict) -> None:
        try:
            from tahmeed.services.accountant_service import get_category_report_column_values

            base_kw = {
                k: v for k, v in kw.items()
                if k not in ("column_filters", "sort_field", "sort_asc")
            }
            col_filters = kw.get("column_filters") or {}
            fields = sorted(set(_COL_FIELD.values()))
            results = await asyncio.gather(*[
                get_category_report_column_values(
                    field,
                    self._category,
                    descriptions=base_kw.get("descriptions"),
                    date_from=base_kw.get("date_from"),
                    date_to=base_kw.get("date_to"),
                    column_filters=col_filters or None,
                )
                for field in fields
            ])
            if generation != self._reload_generation:
                return
            field_to_col = {f: c for c, f in _COL_FIELD.items()}
            self._col_value_cache = {
                field_to_col[field]: set(vals)
                for field, vals in zip(fields, results)
                if field in field_to_col
            }
        except Exception:
            pass

    async def _load_more(self) -> None:
        if self._scroll_loading or self._loading or not self._category:
            return
        if self._loaded >= self._total:
            return
        self._scroll_loading = True
        self._update_footer_status()
        try:
            from tahmeed.services.accountant_service import (
                get_category_report_transactions,
                get_cashier_names,
            )

            kw = self._query_kwargs()
            skip = self._loaded
            txs = await get_category_report_transactions(
                self._category,
                descriptions=kw["descriptions"],
                date_from=kw["date_from"],
                date_to=kw["date_to"],
                column_filters=kw.get("column_filters"),
                sort_field=kw["sort_field"],
                sort_asc=kw["sort_asc"],
                limit=_PAGE_SIZE,
                skip=skip,
            )
            if not txs:
                return
            cashier_ids = [tx.cashier_id for tx in txs if tx.cashier_id]
            names = await get_cashier_names(cashier_ids) if cashier_ids else {}
            self._populate(txs, skip, names, append=True)
            self._loaded += len(txs)
            self._fill_if_needed()
        except Exception as exc:
            self._page_info.setText(f"Failed to load more: {exc}")
        finally:
            self._scroll_loading = False
            self._update_footer_status()

    def _update_footer(self, totals: dict) -> None:
        tzs = _used_amount(totals.get("tzs") or 0.0)
        usd = _used_amount(totals.get("usd") or 0.0)
        count = int(totals.get("count") or 0)
        self._total = count
        self._tzs_lbl.setText(f"TZS  {tzs:,.0f}")
        self._usd_lbl.setText(f"USD  ${usd:,.2f}" if usd else "USD  —")
        self._count_lbl.setText(f"{count:,} entries")
        self._update_footer_status()

    def _update_footer_status(self) -> None:
        loaded = self._loaded
        total = self._total
        filtered = bool(self.column_filters_for_query())
        if self._loading or self._scroll_loading:
            suffix = "  ·  Loading…"
        elif loaded >= total and total:
            suffix = "  ·  Column filters on" if filtered else ""
        elif total:
            extra = "  ·  Column filters on" if filtered else ""
            suffix = f"  ·  Scroll for more{extra}"
        else:
            suffix = ""
        self._page_info.setText(f"Showing {loaded:,} of {total:,}{suffix}")

    def _populate(
        self,
        txs: List[Transaction],
        skip: int,
        cashier_names: Dict,
        *,
        append: bool = False,
    ) -> None:
        all_years = self._all_years_scope()
        start = self._table.rowCount() if append else 0
        if not append:
            self._table.setRowCount(len(txs))
        else:
            self._table.setRowCount(start + len(txs))

        for i, tx in enumerate(txs):
            r = start + i
            row_bg = _STRIPE if r % 2 else _WHITE
            tzs_raw, usd_raw = tx.money_parts()
            tzs = _used_amount(tzs_raw) if tzs_raw else 0.0
            usd = _used_amount(usd_raw) if usd_raw else 0.0
            ref = _ref_float_display(tx)
            rcpt_txt = _receipt_label(tx.receipt_status or "pending")
            cashier = (
                _short_name(cashier_names.get(tx.cashier_id, ""))
                if tx.cashier_id else "—"
            )
            item_str = tx.item or tx.category_name or "—"
            date_txt = _fmt_report_date(tx.date, all_years=all_years)

            _set_cell(self._table, r, 0, str(skip + i + 1), "center", row_bg)
            _set_cell(self._table, r, 1, date_txt, "left", row_bg)
            _set_cell(self._table, r, 2, item_str, "left", row_bg)
            _set_cell(self._table, r, 3, tx.description or "—", "left", row_bg)
            _set_cell(self._table, r, 4, tx.truck_number or "—", "left", row_bg)
            _set_cell(self._table, r, 5, tx.memo or "—", "left", row_bg)
            _set_cell(self._table, r, 6, ref or "—", "left", row_bg)
            _set_cell(
                self._table, r, 7,
                f"{tzs:,.0f}" if tzs_raw else "—", "right", row_bg, mono=True,
            )
            _set_cell(
                self._table, r, 8,
                f"{usd:,.2f}" if usd_raw else "—", "right", row_bg, mono=True,
            )
            _set_cell(self._table, r, 9, rcpt_txt, "center", row_bg)
            _set_cell(self._table, r, 10, tx.ownership or "—", "left", row_bg)
            _set_cell(self._table, r, 11, tx.approver or "—", "left", row_bg)
            _set_cell(self._table, r, 12, cashier, "left", row_bg)
            self._table.setRowHeight(r, _ROW_H)

    # ── Export ───────────────────────────────────────────────────────────────

    def _on_export_filtered(self) -> None:
        if self._export_in_flight:
            return
        asyncio.ensure_future(self._do_export(all_records=False))

    def _on_export_all(self) -> None:
        if self._export_in_flight:
            return
        asyncio.ensure_future(self._do_export(all_records=True))

    async def _fetch_export_page(self, kw: dict, *, limit: int, skip: int) -> List[Transaction]:
        from tahmeed.services.accountant_service import get_category_report_transactions

        return await get_category_report_transactions(
            self._category,
            descriptions=kw.get("descriptions"),
            date_from=kw.get("date_from"),
            date_to=kw.get("date_to"),
            column_filters=kw.get("column_filters"),
            sort_field=kw["sort_field"],
            sort_asc=kw["sort_asc"],
            limit=limit,
            skip=skip,
        )

    async def _do_export(self, *, all_records: bool) -> None:
        if self._export_in_flight or not self._category:
            return
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            await notify_export_error(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n\nRun: pip install openpyxl",
            )
            return

        from tahmeed.services.accountant_service import get_cashier_names

        self._export_in_flight = True
        overlay = attach_export_overlay(self)
        kw = self._query_kwargs(all_records=all_records)
        mode_label = "All" if all_records else "Filtered"
        safe_cat = self._category.replace(" ", "_").replace("&", "and").replace("/", "-")
        try:
            show_export_busy(
                overlay, f"Loading {mode_label.lower()} report…", maximum=0,
            )
            try:
                txs = await fetch_records_with_progress(
                    overlay,
                    lambda *, limit, skip: self._fetch_export_page(
                        kw, limit=limit, skip=skip,
                    ),
                    phase=f"Loading {mode_label.lower()} records",
                )
            except Exception as exc:
                await notify_export_error(self, "Export Error", f"Failed to fetch data: {exc}")
                return
            finally:
                hide_export_busy(self)

            if not txs:
                await notify_export_info(self, "Export", "No records to export.")
                return

            export_cashier_ids = [tx.cashier_id for tx in txs if tx.cashier_id]
            export_cashier_names = (
                await get_cashier_names(export_cashier_ids) if export_cashier_ids else {}
            )

            default = f"{safe_cat}_QuickReport_{mode_label}.xlsx"
            path = await pick_export_path(
                self, f"Export {self._category} ({mode_label})", default,
            )
            if not path:
                return
            path = normalize_xlsx_path(path)

            scope_note = (
                f"All verified records — {self._category}"
                if all_records
                else f"Filtered view — {self._category}"
            )
            total = len(txs)
            fast = total >= FAST_STYLE_ROW_LIMIT
            category = self._category

            def _write(progress_cb) -> None:
                from openpyxl.styles import Font, PatternFill, Alignment

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = category[:28]

                ws.merge_cells("A1:M1")
                ws["A1"] = "TAHMEED COACH TZ LTD"
                ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1B2B4B")
                ws["A1"].alignment = Alignment(horizontal="center")
                ws.merge_cells("A2:M2")
                ws["A2"] = f"Account QuickReport — {scope_note}"
                ws["A2"].font = Font(name="Segoe UI", bold=True, size=11, color="374151")
                ws["A2"].alignment = Alignment(horizontal="center")
                ws.merge_cells("A3:M3")
                ws["A3"] = f"Exported: {datetime.now().strftime('%d %b %Y  %H:%M')}"
                ws["A3"].font = Font(name="Segoe UI", italic=True, size=9, color="9CA3AF")
                ws["A3"].alignment = Alignment(horizontal="center")
                ws.append([])

                headers = [c[0] for c in _COLS]
                ws.append(headers)
                hdr_row = ws.max_row
                grey = PatternFill("solid", fgColor="F1F5F9")
                for cell in ws[hdr_row]:
                    cell.font = Font(name="Segoe UI", bold=True, size=10, color="6B7280")
                    cell.fill = grey
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                stripe = PatternFill("solid", fgColor="F1F5F9")
                white = PatternFill("solid", fgColor="FFFFFF")
                mono = Font(name="Cascadia Code", size=10)
                tzs_total = usd_total = 0.0

                for i, tx in enumerate(txs):
                    tzs_raw, usd_raw = tx.money_parts()
                    tzs_val = _used_amount(tzs_raw) if tzs_raw else None
                    usd_val = _used_amount(usd_raw) if usd_raw else None
                    if tzs_val:
                        tzs_total += tzs_val
                    if usd_val:
                        usd_total += usd_val
                    ref = _ref_float_display(tx)
                    rcpt = _receipt_label(tx.receipt_status or "pending")
                    cashier = (
                        _short_name(export_cashier_names.get(tx.cashier_id, ""))
                        if tx.cashier_id else ""
                    )
                    date_str = tx.date.strftime("%d-%b-%Y") if tx.date else ""
                    ws.append([
                        i + 1, date_str, tx.item or tx.category_name or "",
                        tx.description or "", tx.truck_number or "", tx.memo or "",
                        ref or "", tzs_val or "", usd_val or "", rcpt,
                        tx.ownership or "", tx.approver or "", cashier,
                    ])
                    if not fast:
                        r = ws.max_row
                        fill = stripe if i % 2 else white
                        for cell in ws[r]:
                            cell.fill = fill
                            cell.alignment = Alignment(vertical="center")
                        if tzs_val is not None:
                            c = ws.cell(r, 8)
                            c.font = mono
                            c.number_format = "#,##0"
                            c.alignment = Alignment(horizontal="right", vertical="center")
                        if usd_val is not None:
                            c = ws.cell(r, 9)
                            c.font = mono
                            c.number_format = "#,##0.00"
                            c.alignment = Alignment(horizontal="right", vertical="center")

                    if progress_cb and (
                        (i + 1) % PROGRESS_EVERY == 0 or i + 1 == total
                    ):
                        progress_cb(i + 1, "Writing rows")

                ws.append([])
                ws.append(["", "", "", "TOTAL", "", "", "", tzs_total or "", usd_total or ""])
                total_r = ws.max_row
                ws.cell(total_r, 4).font = Font(name="Segoe UI", bold=True, size=11)
                if tzs_total:
                    c = ws.cell(total_r, 8)
                    c.font = Font(name="Cascadia Code", bold=True, size=11)
                    c.number_format = "#,##0"
                    c.alignment = Alignment(horizontal="right", vertical="center")
                if usd_total:
                    c = ws.cell(total_r, 9)
                    c.font = Font(name="Cascadia Code", bold=True, size=11)
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right", vertical="center")

                widths = [6, 12, 14, 34, 12, 20, 16, 15, 12, 13, 13, 12, 12]
                for idx, w in enumerate(widths, 1):
                    ws.column_dimensions[ws.cell(1, idx).column_letter].width = w
                ws.freeze_panes = ws.cell(hdr_row + 1, 1)
                if progress_cb:
                    progress_cb(total, "Saving file")
                wb.save(path)

            try:
                await run_export_write(overlay, total, _write)
            except Exception as exc:
                await notify_export_error(self, "Save Error", f"Could not save file:\n{exc}")
                return
            finally:
                hide_export_busy(self)

            if not export_file_ready(path):
                await notify_export_error(
                    self, "Save Error", f"The file could not be saved:\n{path}",
                )
                return

            await notify_export_info(
                self, "Export Complete",
                f"Exported {len(txs):,} records to:\n{path}",
            )
        finally:
            hide_export_busy(self)
            self._export_in_flight = False
