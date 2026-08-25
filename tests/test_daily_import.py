"""Tests for daily Excel import helpers (receipt, amount, date detection)."""

from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest
from bson import ObjectId

from tahmeed.models.transaction import Transaction
from tahmeed.services.daily_import_service import (
    DailyImportCancelled,
    DailyImportPreview,
    DailyImportRow,
    REASON_MISSING_DATE,
    _looks_like_classic_matumizi,
    absorb_import_problems,
    analyze_date_allocation,
    apply_date_policy,
    detect_date_from_name,
    normalize_receipt,
    parse_amount,
    parse_date_value,
    parse_daily_expenses_excel,
    pick_primary_date,
    problem_to_import_row,
    staged_row_payload,
    suggested_reconciled_date,
)


def test_normalize_receipt_preserves_exact_text() -> None:
    assert normalize_receipt("RECEIPT") == "RECEIPT"
    assert normalize_receipt("receipt") == "receipt"
    assert normalize_receipt("  Receipt  ") == "Receipt"
    assert normalize_receipt("NO RECEIPT") == "NO RECEIPT"
    assert normalize_receipt("no receipt") == "no receipt"
    assert normalize_receipt("pending") == "pending"
    assert normalize_receipt("missing") == "missing"
    assert normalize_receipt("1") == "1"
    assert normalize_receipt("yes") == "yes"
    assert normalize_receipt("") == ""


def test_parse_preserves_exact_receipt_text(tmp_path: Path) -> None:
    path = tmp_path / "matumizi_rcpt.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "TZS", "Receipt"])
    ws.append([datetime(2026, 8, 20), "DIESEL", 1000, "RECEIPT"])
    ws.append([datetime(2026, 8, 20), "TOLLS", 500, "NO RECEIPT"])
    ws.append([datetime(2026, 8, 20), "PARKING", 200, "yes"])
    ws.append([datetime(2026, 8, 20), "MISC", 100, None])
    wb.save(path)
    wb.close()

    result = parse_daily_expenses_excel(path)
    assert [r.receipt_status for r in result.rows] == [
        "RECEIPT",
        "NO RECEIPT",
        "yes",
        "",
    ]


def test_parse_amount_plain_and_negative() -> None:
    assert parse_amount(10000) == 10000.0
    assert parse_amount(-10000) == -10000.0
    assert parse_amount("10,000") == 10000.0
    assert parse_amount("-10,000") == -10000.0


def test_parse_amount_parentheses() -> None:
    assert parse_amount("(10000)") == -10000.0
    assert parse_amount("(10,000)") == -10000.0
    assert parse_amount("( 1,234.50 )") == -1234.5


def test_parse_amount_blank() -> None:
    assert parse_amount(None) is None
    assert parse_amount("") is None
    assert parse_amount("—") is None


def test_detect_date_from_filename() -> None:
    assert detect_date_from_name("MATUMIZI YA 23-07-2026.xlsx") == date(2026, 7, 23)
    assert detect_date_from_name("23-07-2026") == date(2026, 7, 23)
    assert detect_date_from_name("no-date-here") is None


def test_pick_primary_date_majority() -> None:
    dates = [date(2026, 7, 21)] * 10 + [date(2026, 6, 28)] * 2
    assert pick_primary_date(dates, filename="MATUMIZI YA 23-07-2026.xlsx") == date(
        2026, 7, 21
    )


def test_analyze_date_allocation_clear_majority() -> None:
    dates = [date(2025, 12, 31)] * 300 + [date(2025, 12, 30)] * 120
    alloc = analyze_date_allocation(dates)
    assert alloc.clear_majority is True
    assert alloc.primary == date(2025, 12, 31)
    assert alloc.candidates == (date(2025, 12, 31),)


def test_analyze_date_allocation_tie_is_unclear() -> None:
    dates = [date(2026, 7, 21)] * 5 + [date(2026, 7, 22)] * 5
    alloc = analyze_date_allocation(dates)
    assert alloc.clear_majority is False
    assert alloc.primary is None
    assert alloc.candidates == (date(2026, 7, 21), date(2026, 7, 22))


def test_pick_primary_date_filename_fallback() -> None:
    assert pick_primary_date(
        [], filename="MATUMIZI YA 23-07-2026.xlsx", sheet_name="Sheet1"
    ) == date(2026, 7, 23)


def test_suggested_reconciled_date_uses_primary() -> None:
    preview = DailyImportPreview(
        source_filename="x.xlsx",
        source_path="x.xlsx",
        primary_date=date(2026, 7, 21),
        date_counts={date(2026, 7, 21): 10, date(2026, 6, 28): 2},
        date_majority_clear=True,
    )
    assert suggested_reconciled_date(preview) == date(2026, 7, 21)


def test_suggested_reconciled_date_tie_picks_earliest_top() -> None:
    preview = DailyImportPreview(
        source_filename="x.xlsx",
        source_path="x.xlsx",
        primary_date=None,
        date_counts={date(2026, 7, 22): 5, date(2026, 7, 21): 5},
        date_majority_clear=False,
    )
    assert suggested_reconciled_date(preview) == date(2026, 7, 21)


def test_resolve_import_date_policy_always_prompts(monkeypatch) -> None:
    pytest.importorskip("PySide6")
    from tahmeed.ui.dialogs import date_outlier_dialog as mod

    rows = [
        DailyImportRow(
            serial=1,
            date=datetime(2026, 7, 21),
            description="Fuel",
            truck_number="",
            lpo_do="",
            do_number="",
            memo="",
            notes="",
            amount=100.0,
            currency="TZS",
            receipt_status="pending",
            ownership="",
            approver="",
        ),
    ]
    preview = DailyImportPreview(
        source_filename="x.xlsx",
        source_path="x.xlsx",
        rows=rows,
        primary_date=date(2026, 7, 21),
        date_counts={date(2026, 7, 21): 1},
        date_majority_clear=True,
        outlier_count=0,
    )
    seen: dict = {}

    class _FakeDlg:
        Accepted = 1

        def __init__(self, *args, parent=None, default_date=None, **kwargs):
            seen["default"] = default_date

        def exec(self):
            return 1

        def chosen_date(self):
            return date(2026, 8, 1)

    monkeypatch.setattr(mod, "DateAllocationDialog", _FakeDlg)
    assert mod.resolve_import_date_policy(preview) is True
    assert seen["default"] == date(2026, 7, 21)
    assert preview.primary_date == date(2026, 8, 1)
    assert preview.outlier_count == 1


def test_resolve_import_date_policy_cancel(monkeypatch) -> None:
    pytest.importorskip("PySide6")
    from tahmeed.ui.dialogs import date_outlier_dialog as mod

    rows = [
        DailyImportRow(
            serial=1,
            date=datetime(2026, 7, 21),
            description="Fuel",
            truck_number="",
            lpo_do="",
            do_number="",
            memo="",
            notes="",
            amount=100.0,
            currency="TZS",
            receipt_status="pending",
            ownership="",
            approver="",
        ),
    ]
    preview = DailyImportPreview(
        source_filename="x.xlsx",
        source_path="x.xlsx",
        rows=rows,
        primary_date=date(2026, 7, 21),
        date_counts={date(2026, 7, 21): 1},
        date_majority_clear=True,
    )

    class _CancelDlg:
        Accepted = 1

        def __init__(self, *args, parent=None, default_date=None, **kwargs):
            pass

        def exec(self):
            return 0

        def chosen_date(self):
            return None

    monkeypatch.setattr(mod, "DateAllocationDialog", _CancelDlg)
    assert mod.resolve_import_date_policy(preview) is False
    assert preview.primary_date == date(2026, 7, 21)


def test_apply_date_policy_keeps_excel_row_dates() -> None:
    rows = [
        DailyImportRow(
            serial=1,
            date=datetime(2025, 12, 31),
            description="Fuel",
            truck_number="",
            lpo_do="",
            do_number="",
            memo="",
            notes="",
            amount=100.0,
            currency="TZS",
            receipt_status="pending",
            ownership="",
            approver="",
        ),
        DailyImportRow(
            serial=2,
            date=datetime(2025, 12, 30),
            description="Oil",
            truck_number="",
            lpo_do="",
            do_number="",
            memo="",
            notes="",
            amount=50.0,
            currency="TZS",
            receipt_status="pending",
            ownership="",
            approver="",
        ),
    ]
    preview = DailyImportPreview(
        source_filename="x.xlsx",
        source_path="x.xlsx",
        rows=rows,
        primary_date=date(2025, 12, 31),
        outlier_count=1,
        date_majority_clear=True,
    )
    apply_date_policy(preview, force_primary=True, flag_discrepancy=True)
    assert rows[0].date.date() == date(2025, 12, 31)
    assert rows[1].date.date() == date(2025, 12, 30)
    assert preview.force_primary_date is False
    assert preview.flag_date_discrepancy is False
    payload = staged_row_payload(rows[1], preview)
    assert payload["date"].date() == date(2025, 12, 30)
    assert payload["import_primary_date"].date() == date(2025, 12, 31)
    assert payload["date_discrepancy"] is False


def test_parse_date_value() -> None:
    assert parse_date_value(datetime(2026, 7, 21)).date() == date(2026, 7, 21)
    assert parse_date_value("21/07/2026").date() == date(2026, 7, 21)
    assert parse_date_value("21-07-2026").date() == date(2026, 7, 21)
    assert parse_date_value(46146).date() == date(2026, 5, 4)
    assert parse_date_value("46146").date() == date(2026, 5, 4)


def test_looks_like_classic_matumizi() -> None:
    assert _looks_like_classic_matumizi(
        ["S/NO", "DATE", "X", "DESCRIPTION", "TRUCK NO."]
    )
    assert not _looks_like_classic_matumizi(["Name", "Amount", "Notes"])
    assert not _looks_like_classic_matumizi([])


def test_reject_wrong_format_workbook(tmp_path: Path) -> None:
    path = tmp_path / "parking.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Ticket", "Plate", "Fee"])
    ws.append(["1", "T123 ABC", "5000"])
    wb.save(path)
    wb.close()

    with pytest.raises(ValueError, match="does not match the Daily Register format"):
        parse_daily_expenses_excel(path)


def test_accept_header_mapped_workbook(tmp_path: Path) -> None:
    path = tmp_path / "matumizi.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "TZS", "Truck No."])
    ws.append([datetime(2026, 7, 21), "DIESEL", 10000, "T688 EAF"])
    wb.save(path)
    wb.close()

    result = parse_daily_expenses_excel(path)
    assert result.sheet_name
    assert result.skipped_blank == 0
    assert len(result.rows) == 1
    assert result.rows[0].description == "DIESEL"
    assert result.rows[0].amount == 10000.0
    assert result.rows[0].currency == "TZS"
    assert result.rows[0].amount_usd is None
    assert result.problems == []

    result2 = parse_daily_expenses_excel(path, should_cancel=lambda: False)
    assert len(result2.rows) == 1


def test_parse_dual_tzs_and_usd_when_both_columns_present(tmp_path: Path) -> None:
    path = tmp_path / "matumizi_dual.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "TZS", "USD", "Truck No."])
    ws.append([datetime(2026, 7, 21), "PARTS", 50000, 40.5, "T688 EAF"])
    ws.append([datetime(2026, 7, 21), "USD ONLY", None, 12.0, ""])
    wb.save(path)
    wb.close()

    result = parse_daily_expenses_excel(path)
    rows = result.rows
    assert result.skipped_blank == 0
    assert len(rows) == 2
    assert rows[0].amount == 50000.0
    assert rows[0].amount_usd == 40.5
    assert rows[0].currency == "TZS"
    assert rows[1].amount == 12.0
    assert rows[1].amount_usd == 12.0
    assert rows[1].currency == "USD"

    preview = DailyImportPreview(
        source_filename=path.name,
        source_path=str(path),
        rows=rows,
        primary_date=date(2026, 7, 21),
    )
    payload = staged_row_payload(rows[0], preview)
    assert payload["amount"] == 50000.0
    assert payload["amount_usd"] == 40.5
    assert payload["currency"] == "TZS"


def test_pack_money_and_money_parts_roundtrip() -> None:
    from tahmeed.models.transaction import pack_money

    assert pack_money(1000.0, None) == (1000.0, None, "TZS")
    assert pack_money(None, 25.0) == (25.0, 25.0, "USD")
    assert pack_money(1000.0, 25.0) == (1000.0, 25.0, "TZS")

    dual = Transaction(
        date=datetime(2026, 7, 21),
        description="X",
        truck_number="",
        amount=1000.0,
        currency="TZS",
        amount_usd=25.0,
    )
    assert dual.money_parts() == (1000.0, 25.0)
    doc = dual.to_doc()
    assert doc["amount_usd"] == 25.0
    assert Transaction.from_doc(doc).money_parts() == (1000.0, 25.0)

    legacy_usd = Transaction(
        date=datetime(2026, 7, 21),
        description="Y",
        truck_number="",
        amount=25.0,
        currency="USD",
    )
    assert legacy_usd.money_parts() == (0.0, 25.0)


def test_parse_daily_expenses_excel_can_cancel(tmp_path: Path) -> None:
    path = tmp_path / "matumizi.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "TZS"])
    for i in range(40):
        ws.append([datetime(2026, 7, 21), f"ITEM {i}", 1000])
    wb.save(path)
    wb.close()

    with pytest.raises(DailyImportCancelled):
        parse_daily_expenses_excel(path, should_cancel=lambda: True)


def test_parse_missing_date_becomes_problem_not_silent_skip(tmp_path: Path) -> None:
    path = tmp_path / "matumizi_missing_date.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Description", "TZS", "Truck No."])
    ws.append([datetime(2026, 8, 20), "DIESEL", 10000, "T688 EAF"])
    ws.append([None, "TOLLS", 5000, "T100 ABC"])  # missing date
    ws.append(["", "PARKING", 2000, ""])  # empty date
    ws.append([None, "TOTAL", 17000, ""])  # total line — auto-skip
    wb.save(path)
    wb.close()

    result = parse_daily_expenses_excel(path)
    assert len(result.rows) == 1
    assert result.rows[0].description == "DIESEL"
    assert len(result.problems) == 2
    assert all(p.reason == REASON_MISSING_DATE for p in result.problems)
    assert {p.description for p in result.problems} == {"TOLLS", "PARKING"}
    assert result.skip_reasons.get(REASON_MISSING_DATE) == 2
    assert result.skip_reasons.get("total_row") == 1
    assert result.skipped_blank == 1  # only TOTAL auto-skipped


@pytest.mark.asyncio
async def test_absorb_import_problems_promotes_fixed_and_counts_skips(
    monkeypatch,
) -> None:
    from tahmeed.services.daily_import_service import DailyImportProblemRow
    from tahmeed.services.description_mapping_service import normalize_description

    preview = DailyImportPreview(
        source_filename="x.xlsx",
        source_path="x.xlsx",
        rows=[
            DailyImportRow(
                serial=1,
                date=datetime(2026, 8, 20),
                description="DIESEL",
                truck_number="",
                lpo_do="",
                do_number="",
                memo="",
                notes="",
                amount=1000.0,
                currency="TZS",
                receipt_status="pending",
                ownership="",
                approver="",
            )
        ],
        primary_date=date(2026, 8, 20),
        problem_rows=[
            DailyImportProblemRow(
                excel_row=3,
                reason=REASON_MISSING_DATE,
                description="TOLLS",
                amount=5000.0,
                date=datetime(2026, 8, 20),
            ),
            DailyImportProblemRow(
                excel_row=4,
                reason=REASON_MISSING_DATE,
                description="SKIP ME",
                amount=100.0,
                skipped=True,
            ),
        ],
    )

    async def _no_maps(_descs):
        return {}

    monkeypatch.setattr(
        "tahmeed.services.daily_import_service.get_mappings_for_descriptions",
        _no_maps,
    )
    await absorb_import_problems(preview)

    assert len(preview.problem_rows) == 0
    assert len(preview.rows) == 2
    assert preview.rows[1].description == "TOLLS"
    assert preview.rows[1].date == datetime(2026, 8, 20)
    assert preview.skipped_blank == 1
    assert preview.skip_reasons.get("missing_date_user_skip") == 1
    assert normalize_description("TOLLS") in preview.unmapped


def test_problem_to_import_row_requires_date() -> None:
    from tahmeed.services.daily_import_service import DailyImportProblemRow

    problem = DailyImportProblemRow(
        excel_row=2,
        reason=REASON_MISSING_DATE,
        description="X",
    )
    with pytest.raises(ValueError):
        problem_to_import_row(problem)


def test_transaction_to_doc_omits_null_import_id() -> None:
    tx = Transaction(
        date=datetime(2026, 7, 21),
        description="TEST",
        truck_number="",
        amount=100.0,
    )
    doc = tx.to_doc()
    assert "daily_import_id" not in doc
    assert "daily_import_source" not in doc

    tx.daily_import_id = "batch-1"
    tx.daily_import_source = "file.xlsx"
    doc2 = tx.to_doc()
    assert doc2["daily_import_id"] == "batch-1"
    assert doc2["daily_import_source"] == "file.xlsx"


def test_browse_match_upload_id_skips_date_filter() -> None:
    from tahmeed.services.cashier_service import _browse_match

    match = _browse_match(
        date_from=date(2025, 1, 1),
        date_to=date(2025, 1, 31),
        daily_import_id="batch-abc",
    )
    assert match["daily_import_id"] == "batch-abc"
    assert "date" not in match


def test_register_day_clause_keeps_prior_excel_on_primary_day() -> None:
    """Aug 18 Excel rows filed under Aug 19 belong only to Aug 19."""
    from tahmeed.services.cashier_service import _register_day_clause

    aug18 = _register_day_clause(date(2026, 8, 18))
    aug19 = _register_day_clause(date(2026, 8, 19))

    prior = {
        "date": datetime(2026, 8, 18, 10, 0, 0),
        "import_primary_date": datetime(2026, 8, 19, 12, 0, 0),
    }
    primary = {
        "date": datetime(2026, 8, 19, 10, 0, 0),
        "import_primary_date": datetime(2026, 8, 19, 12, 0, 0),
    }
    manual = {
        "date": datetime(2026, 8, 18, 9, 0, 0),
    }

    def matches(clause: dict, doc: dict) -> bool:
        def _one(q: dict) -> bool:
            if "$or" in q:
                return any(_one(s) for s in q["$or"])
            if "$and" in q:
                return all(_one(s) for s in q["$and"])
            for k, exp in q.items():
                actual = doc.get(k)
                if isinstance(exp, dict):
                    if "$gte" in exp and not (
                        actual is not None and actual >= exp["$gte"]
                    ):
                        return False
                    if "$lte" in exp and not (
                        actual is not None and actual <= exp["$lte"]
                    ):
                        return False
                elif actual != exp:
                    return False
            return True

        return _one(clause)

    assert matches(aug19, prior)
    assert matches(aug19, primary)
    assert not matches(aug18, prior)
    assert matches(aug18, manual)
    assert not matches(aug19, manual)


def test_register_day_expr_used_in_daily_summaries_pipeline() -> None:
    """Simple day list must group by coalesce(import_primary_date, date)."""
    import inspect
    from tahmeed.services import cashier_service

    src = inspect.getsource(cashier_service.get_daily_summaries)
    assert "_REGISTER_DAY_EXPR" in src
    assert "_register_day_range_clause" in src


@pytest.mark.asyncio
async def test_apply_mapping_keeps_assignment_when_save_fails(monkeypatch) -> None:
    async def boom(*_args, **_kwargs):
        raise RuntimeError("Your role cannot perform this action")

    monkeypatch.setattr(
        "tahmeed.services.daily_import_service.save_mapping",
        boom,
    )
    cat_id = ObjectId()
    row = DailyImportRow(
        serial=1,
        date=datetime(2026, 7, 21),
        description="DIESEL",
        truck_number="",
        lpo_do="",
        do_number="",
        memo="",
        notes="",
        amount=100.0,
        currency="TZS",
        receipt_status="pending",
        ownership="",
        approver="",
    )
    preview = DailyImportPreview(
        source_filename="x.xlsx",
        source_path="x.xlsx",
        rows=[row],
        unmapped={"DIESEL": 1},
    )
    from tahmeed.services.daily_import_service import apply_mapping_to_preview

    await apply_mapping_to_preview(preview, "DIESEL", cat_id, "Fuel")
    assert row.category_id == cat_id
    assert row.category_name == "Fuel"
    assert row.skipped_item is False
    assert "DIESEL" not in preview.unmapped


def test_master_query_filters_by_excel_date() -> None:
    from tahmeed.services.accountant_service import _build_master_query

    q = _build_master_query(2025, 12, "", "", "", "", "")
    assert q["verified"] is True
    assert "date" in q
    assert q["date"]["$gte"].month == 12
    assert q["date"]["$lte"].month == 12
    assert "import_primary_date" not in q
    assert "$or" not in q

